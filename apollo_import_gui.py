from __future__ import annotations

import csv
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher, get_close_matches
from html import unescape
import io
import json
import os
from pathlib import Path
import random
import re
import ssl
import string
import subprocess
import tempfile
import threading
import sys
import tkinter as tk
import unicodedata
from tkinter import filedialog, font as tkfont, messagebox, ttk
from tkinter.scrolledtext import ScrolledText
from typing import Callable
from urllib import error as urllib_error
from urllib import request as urllib_request
from urllib.parse import parse_qs, quote, unquote, urlencode, urlparse
import xml.etree.ElementTree as ET
import webbrowser

from openpyxl import Workbook, load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

try:
    from PIL import Image, ImageTk
except ImportError:  # pragma: no cover - optional preview dependency
    Image = None
    ImageTk = None

try:
    import pymupdf
except ImportError:  # pragma: no cover - optional preview dependency
    try:
        import fitz as pymupdf  # type: ignore[no-redef]
    except ImportError:  # pragma: no cover - optional preview dependency
        pymupdf = None


APP_TITLE = "Apollo Import GUI Prototype"
APP_VERSION = "0.1.14"
APP_VERSION_TAG = f"v{APP_VERSION}"

# Zentrale UI-Palette: helles, neutrales Design mit blauem Akzent.
UI_FONT_FAMILY = "Segoe UI"
UI_BACKGROUND = "#F5F6FA"
UI_SURFACE = "#FFFFFF"
UI_BORDER = "#D5DAE1"
UI_TEXT = "#1F2933"
UI_TEXT_MUTED = "#5E6472"
UI_ACCENT = "#2563EB"
UI_ACCENT_DARK = "#1D4ED8"
UI_ACCENT_SOFT = "#DCE7FB"
UI_HEADING_BG = "#EDF0F4"
UI_TAB_INACTIVE = "#E4E8EE"
DEFAULT_IMPORT_DIR = Path(r"C:\Users\heimbuchner\Desktop\Apollo Import App\Aktuelle Import Datein")
DEFAULT_OUTPUT_DIR = Path.cwd() / "output"
DEFAULT_GENART_SOURCE = Path(r"C:\Users\heimbuchner\Downloads\Genarten.xlsx")
DEFAULT_COMPETITOR_SOURCE = Path(r"C:\Users\heimbuchner\Downloads\KHer.csv")
DEFAULT_ATTRIBUTE_SOURCE = Path(r"G:\Apollo\Export aus SQL\Attribute alle.xlsx")
DEFAULT_ATTRIBUTE_KEY_VALUE_SOURCE = Path("G:/Apollo/Export aus SQL/Schl\u00fcsselwerte.xlsx")
DEFAULT_VEHICLE_SOURCE = Path(r"G:\Apollo\KTyp.xlsx")
DEFAULT_ATTRIBUTE_MAPPING_SOURCE = Path(r"G:\Apollo\Attribut_Zuordnung.xlsx")
DEEPL_DEFAULT_BASE_URL = "https://api.deepl.com"
ID_ALPHABET = string.ascii_uppercase + string.digits
ID_LENGTH = 6
SHORT_TEXT_MAX_LENGTH = 60
ATTACHMENT_FORMAT_TYPE_HEADER = "TecDoc Anhangsformattyp ID"
LAST_WRITTEN_HEADER = "Zuletzt geschrieben am"
# TecDoc Verknuepfungstyp IDs (Fahrzeugtyp) aus dem Apollo-Import. PKW = 2.
VEHICLE_TYPE_CHOICES = [
    ("2", "PKW"),
    ("16", "NKW"),
    ("14", "Motor"),
    ("19", "Achse"),
]
DEFAULT_VEHICLE_TYPE_ID = "2"


def vehicle_type_label_for_id(vehicle_type_id: str) -> str:
    key = str(vehicle_type_id or "").strip()
    for type_id, label in VEHICLE_TYPE_CHOICES:
        if type_id == key:
            return label
    return ""


def vehicle_type_id_for_label(label: str) -> str:
    key = str(label or "").strip().casefold()
    for type_id, type_label in VEHICLE_TYPE_CHOICES:
        if type_label.casefold() == key:
            return type_id
    return ""


def format_vehicle_type_choice(type_id: str, label: str) -> str:
    label = label.strip() or vehicle_type_label_for_id(type_id)
    return f"{label} ({type_id})" if type_id else label


def normalize_motorcode(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().upper())


def split_code_list_input(raw: str) -> list[str]:
    """Zerlegt eine eingefügte Code-/Wortliste in einzelne Einträge.

    Enthält der Text ein Semikolon, gilt nur das Semikolon als Trenner;
    Zeilenumbrüche und Tabs innerhalb der Teile (harte Umbrüche aus der
    Kopierquelle, z. B. ``B 5254<Umbruch>S``) werden zu Leerzeichen
    zusammengefasst. Ohne Semikolon trennt wie bisher jede neue Zeile
    bzw. jeder Tab.
    """
    text = str(raw or "")
    if ";" in text:
        parts = text.split(";")
    else:
        parts = re.split(r"[\n\r\t]+", text)
    return [" ".join(part.split()) for part in parts if part.strip()]


def motorcode_keys(value: object) -> list[str]:
    """Robuste Suchschlüssel für einen Motorcode.

    Berücksichtigt Gross-/Kleinschreibung, Leerzeichen und Klammerzusätze,
    z. B. ``108C (XV8)`` -> ``108C``.
    """
    collapsed = normalize_motorcode(value)
    if not collapsed:
        return []
    keys = {collapsed, collapsed.replace(" ", "")}
    base = collapsed.split("(", 1)[0].strip()
    if base:
        keys.add(base)
        keys.add(base.replace(" ", ""))
    return [key for key in keys if key]


def format_year_range(year_from: str, year_to: str) -> str:
    year_from = str(year_from or "").strip()
    year_to = str(year_to or "").strip()
    if year_from and year_to:
        return f"{year_from} - {year_to}"
    return year_from or year_to
SHORT_TEXT_UMLAUT_REPLACEMENTS = str.maketrans(
    {
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "Ä": "Ae",
        "Ö": "Oe",
        "Ü": "Ue",
        "ß": "ss",
    }
)
LOOKUP_TEXT_REPLACEMENTS = (
    ("\u00c3\u00a4", "ae"),
    ("\u00c3\u00b6", "oe"),
    ("\u00c3\u00bc", "ue"),
    ("\u00c3\u0084", "Ae"),
    ("\u00c3\u0096", "Oe"),
    ("\u00c3\u009c", "Ue"),
    ("\u00c3\u009f", "ss"),
    ("\u00e4", "ae"),
    ("\u00f6", "oe"),
    ("\u00fc", "ue"),
    ("\u00c4", "Ae"),
    ("\u00d6", "Oe"),
    ("\u00dc", "Ue"),
    ("\u00df", "ss"),
)
ATTRIBUTE_KEY_VALUE_FORMAT_KEYS = {"schluesselwert"}
SESSION_STATE_FILE = Path.cwd() / "apollo_import_gui_state.json"
APP_ICON_PNG_RELATIVE_PATH = Path("assets") / "apollo_import_logo.png"
APP_ICON_ICO_RELATIVE_PATH = Path("assets") / "apollo_import_logo.ico"
GITHUB_REPO_FULL_NAME = "GlnSevi/Apollo-Import"
GITHUB_RELEASES_LATEST_API_URL = f"https://api.github.com/repos/{GITHUB_REPO_FULL_NAME}/releases/latest"
GITHUB_RELEASES_PAGE_URL = f"https://github.com/{GITHUB_REPO_FULL_NAME}/releases"


def path_exists_safe(path: Path) -> bool:
    try:
        return path.exists()
    except OSError:
        return False


SHORT_TEXT_HEADERS = [
    "Artikelnummer",
    "Text Modul ID",
    "Modultyp ID",
    "Infotyp ID",
    "Language ID1",
    "TEXT UNI",
    "Language ID2",
    "TEXT DE",
    "Language ID3",
    "TEXT EN",
    "Language ID4",
    "TEXT CZ",
    "Language ID5",
    "TEXT FR",
    "Language ID6",
    "TEXT IT",
    "Language ID7",
    "TEXT NL",
    LAST_WRITTEN_HEADER,
]

SHORT_TEXT_FILE = ("Kurzbezeichnung-NEU.xlsx", "Import")
SHORT_MAPPING_FILE = ("Kurzbezeichnung_zu_ID.xlsx", "Zuordnung")
LONG_TEXT_FILE = ("Text-NEU.xlsx", "Import")
IMAGE_FILE = ("Bilder.xlsx", "Sheet1")
DOCUMENT_FILE = ("Dokumente.xlsx", "Sheet1")
VIDEO_FILE = ("Videos.xlsx", "Tabelle1")
WEB_LINK_FILE = ("Web Link.xlsx", "Tabelle1")
GENART_FILE = ("GenArt_Artikel.xlsx", "GenArt")
OE_FILE = ("OE-Nummern.xlsx", "Sheet1")
COMPARISON_FILE = ("Vergleichsnummern.xlsx", "Sheet1")
ATTRIBUTE_FILE = ("Attribute.xlsx", "Sheet1")
VEHICLE_LINK_FILE = ("Fahrzeugverknuepfungen.xlsx", "Sheet1")

IMAGE_HEADERS = ["Artikelnummer", "BILDPFAD", "Art", "Sprache", ATTACHMENT_FORMAT_TYPE_HEADER, LAST_WRITTEN_HEADER]
DOCUMENT_HEADERS = ["Artikelnummer", "Pfad zum Dokument", "Sprache", "Art", ATTACHMENT_FORMAT_TYPE_HEADER, LAST_WRITTEN_HEADER]
VIDEO_HEADERS = ["Produktnummer", "Link", ATTACHMENT_FORMAT_TYPE_HEADER, LAST_WRITTEN_HEADER]
WEB_HEADERS = ["Artikelnummer ", "Link", ATTACHMENT_FORMAT_TYPE_HEADER, LAST_WRITTEN_HEADER]
SHORT_MAPPING_HEADERS = ["Artikelnummer", "Text Modul ID", LAST_WRITTEN_HEADER]
GENART_HEADERS = ["Artikelnummer", "GenArt ID", "GenArt Bezeichnung", LAST_WRITTEN_HEADER]
LEGACY_OE_HEADERS = ["Artikelnummer", "TecDoc ID", "OE-Nummer", LAST_WRITTEN_HEADER]
OE_HEADERS = ["Artikelnummer", "TecDoc ID", "KHerNr", "OE-Nummer", LAST_WRITTEN_HEADER]
COMPARISON_HEADERS = ["Artikelnummer", "TecDoc ID", "Mitbewerber ID", "Vergleichsnummer", LAST_WRITTEN_HEADER]
LEGACY_ATTRIBUTE_HEADERS = ["Artikelnummer", "TecDoc Kriterien ID", "Attribut Bezeichnung", "Format", "Wert", LAST_WRITTEN_HEADER]
ATTRIBUTE_HEADERS_WITH_VALUE_FROM = [
    "Artikelnummer",
    "TecDoc Kriterien ID",
    "Attribut Bezeichnung",
    "Format",
    "Wert",
    "Wert von",
    "Wert bis",
    LAST_WRITTEN_HEADER,
]
ATTRIBUTE_HEADERS = [
    "Artikelnummer",
    "TecDoc Kriterien ID",
    "Attribut Bezeichnung",
    "Format",
    "Wert",
    "Wert bis",
    LAST_WRITTEN_HEADER,
]
# Fahrzeugverknüpfungen für den Apollo-Import.
# Fahrzeugtyp  -> TecDoc Verknuepfungstyp ID (PKW = 2)
# KTypNr       -> TecDoc Verknuepfungs ID (pro Zeile eine KTyp-Nummer)
# KTyp-System  -> Info, aus welchem Nummernkreis die KTypNr stammt (Topmotive/TecDoc)
# GenArt ID    -> GenArt des Artikels; bei mehreren GenArts eine Zeile je GenArt und KTyp
LEGACY_VEHICLE_LINK_HEADERS = ["Artikelnummer", "Fahrzeugtyp", "KTypNr", "KTyp-System", LAST_WRITTEN_HEADER]
VEHICLE_LINK_HEADERS = ["Artikelnummer", "Fahrzeugtyp", "KTypNr", "KTyp-System", "GenArt ID", "GenArt Bezeichnung", LAST_WRITTEN_HEADER]
# Zuordnung von Text-Labels (aus gescrapten Produkttexten) zu TecDoc Kriterien IDs.
ATTRIBUTE_MAPPING_SHEET = "Zuordnung"
ATTRIBUTE_MAPPING_HEADERS = ["Text-Label", "TecDoc Kriterien ID", "Hinweis"]
# Suchwörter werden als Attribut 'Zusatzbezeichnung' (TOPMOTIVE, alphanumerisch)
# exportiert, weil Apollo keine nativen Keywords kennt. Pro Suchwort eine Attributzeile.
SEARCH_TERM_CRITERIA_ID = "9595"
SEARCH_TERM_ATTRIBUTE_LABEL = "Zusatzbezeichnung"
SEARCH_TERM_ATTRIBUTE_FORMAT = "Alphanumerisch"
SEARCH_TERM_MAX_LENGTH = 20
CSV_ARTICLE_HEADER_ALIASES = {
    "artikelnummer",
    "artikelnr",
    "artikelnr",
    "artnr",
    "produktnummer",
    "sku",
    "article",
}
CSV_URL_HEADER_ALIASES = {
    "kunzerprodukturl",
    "kunzerurl",
    "produkturl",
    "produktlink",
    "produktseite",
    "produktseitenurl",
    "url",
    "link",
}

EXPORT_LANGUAGE_LAYOUT = [
    ("255", "uni"),
    ("1", "de"),
    ("4", "en"),
    ("18", "cz"),
    ("6", "fr"),
    ("7", "it"),
    ("9", "nl"),
]

UI_LANGUAGE_ORDER = [
    ("de", "Deutsch"),
    ("en", "Englisch"),
    ("cz", "Tschechisch"),
    ("fr", "Französisch"),
    ("it", "Italienisch"),
    ("nl", "Niederländisch"),
    ("uni", "UNI"),
]

DEEPL_TARGET_LANGUAGES = {
    "en": "EN-GB",
    "cz": "CS",
    "fr": "FR",
    "it": "IT",
    "nl": "NL",
}

BROWSER_EXECUTABLE_CANDIDATES = [
    Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
    Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"),
]

KUNZER_GENERIC_YOUTUBE_LINKS = {
    "https://www.youtube.com/channel/UCUpBSumIkYx8-C6QYZe9Z0Q",
}

ATTACHMENT_FORMAT_TYPE_BY_EXTENSION = {
    ".jpg": "3",
    ".jpeg": "3",
    ".png": "6",
    ".pdf": "2",
    ".gif": "7",
}

GENART_FAMILY_TERMS = {
    "light": frozenset(
        {
            "lampe",
            "leuchte",
            "licht",
            "lamp",
            "light",
            "worklight",
            "flashlight",
            "headlamp",
            "inspection",
            "arbeitslampe",
            "meisterlampe",
            "prueflampe",
            "pruefleuchte",
            "stablampe",
            "taschenlampe",
            "stirnlampe",
            "inspectionlamp",
            "inspectionlight",
            "werkzeugleuchte",
        }
    ),
    "pin": frozenset(
        {
            "stift",
            "splint",
            "pin",
            "peg",
            "stiftschraube",
            "zylinderstift",
            "zentrierstift",
            "verriegelungsstift",
            "haltestift",
        }
    ),
}

GENART_COMPOUND_SUFFIXES = frozenset(
    {
        *{term for terms in GENART_FAMILY_TERMS.values() for term in terms},
        "zange",
        "hammer",
        "buerste",
        "schluessel",
    }
)

@dataclass
class TranslationSet:
    de: str = ""
    en: str = ""
    cz: str = ""
    fr: str = ""
    it: str = ""
    nl: str = ""
    uni: str = ""

    def normalized(self) -> "TranslationSet":
        return TranslationSet(
            de=self.de.strip(),
            en=self.en.strip(),
            cz=self.cz.strip(),
            fr=self.fr.strip(),
            it=self.it.strip(),
            nl=self.nl.strip(),
            uni=self.uni.strip(),
        )

    def export_values(self, auto_uni: bool) -> list[str]:
        values = self.normalized()
        if auto_uni:
            values.uni = values.de
        export = []
        for language_id, key in EXPORT_LANGUAGE_LAYOUT:
            export.extend([language_id, getattr(values, key)])
        return export

    def populated_count(self, auto_uni: bool) -> int:
        values = self.normalized()
        if auto_uni:
            values.uni = values.de
        return sum(1 for value in vars(values).values() if value)


@dataclass
class MediaRow:
    path_or_link: str
    art: str = ""
    sprache: str = ""


@dataclass(frozen=True)
class OeNumberRow:
    value: str = ""
    manufacturer_id: str = ""
    manufacturer_code: str = ""
    manufacturer_name: str = ""

    def display_manufacturer_label(self) -> str:
        parts = [self.manufacturer_id.strip(), self.manufacturer_code.strip(), self.manufacturer_name.strip()]
        return " | ".join(part for part in parts if part) or "-"


@dataclass(frozen=True)
class ComparisonNumberRow:
    competitor_id: str = ""
    competitor_code: str = ""
    competitor_name: str = ""
    reference_number: str = ""

    def display_competitor_label(self) -> str:
        parts = [self.competitor_id.strip(), self.competitor_code.strip(), self.competitor_name.strip()]
        return " | ".join(part for part in parts if part) or "-"


@dataclass(frozen=True)
class CompetitorOption:
    competitor_id: str
    code: str = ""
    name: str = ""

    @property
    def search_blob(self) -> str:
        return " ".join(part for part in [self.competitor_id, self.code, self.name] if part).casefold()

    def display_label(self) -> str:
        parts = [self.competitor_id.strip(), self.code.strip(), self.name.strip()]
        return " | ".join(part for part in parts if part)


@dataclass(frozen=True)
class VehicleLinkRow:
    """Eine Fahrzeugverknüpfung eines Artikels (ein Fahrzeug / KTyp).

    Ein Fahrzeug trägt beide KTyp-Nummern (Topmotive + TecDoc); beim Export
    wird daraus je Nummer eine Zeile erzeugt.
    """

    vehicle_type_id: str = DEFAULT_VEHICLE_TYPE_ID
    vehicle_type_label: str = "PKW"
    motorcode: str = ""
    ktyp_topmotive: str = ""
    ktyp_tecdoc: str = ""
    manufacturer: str = ""
    model: str = ""
    description: str = ""
    year_from: str = ""
    year_to: str = ""
    power: str = ""

    def has_ktyp(self) -> bool:
        return bool(self.ktyp_topmotive.strip() or self.ktyp_tecdoc.strip())

    def display_vehicle_label(self) -> str:
        parts = [self.manufacturer.strip(), self.model.strip(), self.description.strip()]
        return " ".join(part for part in parts if part) or self.motorcode.strip() or "-"


@dataclass(frozen=True)
class VehicleMatch:
    """Ein Treffer aus den KTyp-Stammdaten für einen Motorcode."""

    motorcode: str = ""
    ktyp_topmotive: str = ""
    ktyp_tecdoc: str = ""
    manufacturer: str = ""
    model: str = ""
    description: str = ""
    year_from: str = ""
    year_to: str = ""
    power: str = ""

    def identity(self) -> tuple[str, str]:
        return (self.ktyp_topmotive, self.ktyp_tecdoc)


@dataclass(frozen=True)
class AttributeOption:
    criteria_id: str
    label: str = ""
    value_format: str = ""
    max_length: int | None = None
    source: str = ""
    type_name: str = ""

    @property
    def search_blob(self) -> str:
        return " ".join(
            part
            for part in [
                self.criteria_id,
                self.label,
                self.value_format,
                self.source,
                self.type_name,
            ]
            if part
        ).casefold()

    def display_label(self) -> str:
        parts = [self.criteria_id.strip(), self.label.strip()]
        return " | ".join(part for part in parts if part) or self.criteria_id

    def format_summary(self) -> str:
        details = [self.value_format.strip() or "-"]
        if self.max_length is not None:
            details.append(f"max. {self.max_length}")
        return " | ".join(details)

    def key_value_group_candidates(self) -> list[str]:
        candidates: list[str] = []
        for value in [self.type_name.strip(), self.label.strip()]:
            if not value:
                continue
            normalized = normalize_attribute_key_value_group(value)
            if normalized in candidates:
                continue
            candidates.append(normalized)
        return candidates


@dataclass(frozen=True)
class AttributeKeyValueOption:
    key_value_id: str
    label: str = ""
    attribute_group: str = ""
    source: str = ""

    @property
    def search_blob(self) -> str:
        return " ".join(
            part
            for part in [self.key_value_id, self.label, self.attribute_group, self.source]
            if part
        ).casefold()

    def display_label(self) -> str:
        parts = [self.key_value_id.strip(), self.label.strip()]
        return " | ".join(part for part in parts if part) or self.key_value_id


@dataclass(frozen=True)
class AttributeRow:
    criteria_id: str = ""
    label: str = ""
    value_format: str = ""
    max_length: int | None = None
    type_name: str = ""
    value: str = ""
    value_to: str = ""

    def display_label(self) -> str:
        parts = [self.criteria_id.strip(), self.label.strip()]
        return " | ".join(part for part in parts if part) or "-"

    def display_value(self) -> str:
        value = self.value.strip()
        value_to = self.value_to.strip()
        if value and value_to:
            return f"{value} bis {value_to}"
        if value:
            return value
        if value_to:
            return f"bis {value_to}"
        if self.value_format.strip().casefold() == "kein wert":
            return "(kein Wert)"
        return "-"

    def key_value_group_candidates(self) -> list[str]:
        candidates: list[str] = []
        for value in [self.type_name.strip(), self.label.strip()]:
            normalized = normalize_attribute_key_value_group(value)
            if not normalized or normalized in candidates:
                continue
            candidates.append(normalized)
        return candidates


@dataclass(frozen=True)
class ExtractedSpecLine:
    """Eine erkannte 'Label: Wert'-Zeile aus einem Produkttext."""

    label: str = ""
    raw_value: str = ""
    source_line: str = ""


@dataclass
class AttributeSuggestion:
    """Ein Attribut-Vorschlag aus einem Produkttext für den Bestätigungsdialog."""

    option: AttributeOption
    value: str = ""
    value_to: str = ""
    source_line: str = ""
    confident: bool = True
    note: str = ""


@dataclass
class GenArtOption:
    id: str
    bezeichnung: str
    genart: str
    normalized_bezeichnung: str = ""
    normalized_genart: str = ""
    tokens: frozenset[str] = field(default_factory=frozenset)
    families: frozenset[str] = field(default_factory=frozenset)

    def display_label(self) -> str:
        if self.genart and self.bezeichnung and self.genart != self.bezeichnung:
            return f"{self.id} | {self.genart} | {self.bezeichnung}"
        if self.bezeichnung:
            return f"{self.id} | {self.bezeichnung}"
        return self.id


@dataclass(frozen=True)
class GenArtSelection:
    id: str = ""
    bezeichnung: str = ""

    def display_label(self) -> str:
        option_id = self.id.strip()
        bezeichnung = self.bezeichnung.strip()
        if option_id and bezeichnung and option_id != bezeichnung:
            return f"{option_id} | {bezeichnung}"
        return option_id or bezeichnung


@dataclass
class GitHubReleaseAsset:
    name: str
    download_url: str
    content_type: str = ""
    size: int = 0

    @property
    def suffix(self) -> str:
        return Path(self.name).suffix.lower()


@dataclass
class GitHubReleaseInfo:
    tag_name: str
    name: str = ""
    html_url: str = ""
    body: str = ""
    published_at: str = ""
    assets: list[GitHubReleaseAsset] = field(default_factory=list)


@dataclass
class ExportBundle:
    article_number: str
    short_module_id: str
    long_module_id: str
    short_texts: TranslationSet
    short_auto_uni: bool
    long_texts: TranslationSet
    long_auto_uni: bool
    genart_selections: list[GenArtSelection] = field(default_factory=list)
    genart_id: str = ""
    genart_bezeichnung: str = ""
    image_rows: list[MediaRow] = field(default_factory=list)
    document_rows: list[MediaRow] = field(default_factory=list)
    video_rows: list[MediaRow] = field(default_factory=list)
    web_rows: list[MediaRow] = field(default_factory=list)
    oe_number_rows: list[OeNumberRow] = field(default_factory=list)
    comparison_number_rows: list[ComparisonNumberRow] = field(default_factory=list)
    vehicle_link_rows: list[VehicleLinkRow] = field(default_factory=list)
    attribute_rows: list[AttributeRow] = field(default_factory=list)
    search_terms: list[str] = field(default_factory=list)
    attribute_key_values_by_group: dict[str, list[AttributeKeyValueOption]] = field(default_factory=dict)
    include_short_text: bool = True
    include_long_text: bool = True
    include_images: bool = True
    include_documents: bool = True
    include_videos: bool = True
    include_web_links: bool = True

    def __post_init__(self) -> None:
        self.sync_genart_fields()
        self.oe_number_rows = normalize_oe_number_rows(self.oe_number_rows)
        self.comparison_number_rows = normalize_comparison_number_rows(self.comparison_number_rows)
        self.vehicle_link_rows = normalize_vehicle_link_rows(self.vehicle_link_rows)
        self.attribute_rows = normalize_attribute_rows(self.attribute_rows)
        self.search_terms = normalize_search_terms(self.search_terms)

    def sync_genart_fields(self) -> None:
        self.genart_selections = normalize_genart_selections(
            self.genart_selections,
            fallback_id=self.genart_id,
            fallback_bezeichnung=self.genart_bezeichnung,
        )
        if self.genart_selections:
            self.genart_id = self.genart_selections[0].id
            self.genart_bezeichnung = self.genart_selections[0].bezeichnung
        else:
            self.genart_id = ""
            self.genart_bezeichnung = ""


@dataclass
class StoredArticleSnapshot:
    article_number: str
    source_label: str
    source_folder: Path
    short_module_id: str = ""
    long_module_id: str = ""
    genart_selections: list[GenArtSelection] = field(default_factory=list)
    genart_id: str = ""
    genart_bezeichnung: str = ""
    short_texts: TranslationSet = field(default_factory=TranslationSet)
    short_auto_uni: bool = True
    long_texts: TranslationSet = field(default_factory=TranslationSet)
    long_auto_uni: bool = True
    image_rows: list[MediaRow] = field(default_factory=list)
    document_rows: list[MediaRow] = field(default_factory=list)
    video_rows: list[MediaRow] = field(default_factory=list)
    web_rows: list[MediaRow] = field(default_factory=list)
    oe_number_rows: list[OeNumberRow] = field(default_factory=list)
    comparison_number_rows: list[ComparisonNumberRow] = field(default_factory=list)
    vehicle_link_rows: list[VehicleLinkRow] = field(default_factory=list)
    attribute_rows: list[AttributeRow] = field(default_factory=list)
    search_terms: list[str] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.sync_genart_fields()
        self.oe_number_rows = normalize_oe_number_rows(self.oe_number_rows)
        self.comparison_number_rows = normalize_comparison_number_rows(self.comparison_number_rows)
        self.vehicle_link_rows = normalize_vehicle_link_rows(self.vehicle_link_rows)
        self.attribute_rows = normalize_attribute_rows(self.attribute_rows)
        self.search_terms = normalize_search_terms(self.search_terms)

    def sync_genart_fields(self) -> None:
        self.genart_selections = normalize_genart_selections(
            self.genart_selections,
            fallback_id=self.genart_id,
            fallback_bezeichnung=self.genart_bezeichnung,
        )
        if self.genart_selections:
            self.genart_id = self.genart_selections[0].id
            self.genart_bezeichnung = self.genart_selections[0].bezeichnung
        else:
            self.genart_id = ""
            self.genart_bezeichnung = ""


def parse_genart_selection_label(raw: str) -> GenArtSelection | None:
    value = raw.strip()
    if not value:
        return None
    parts = [part.strip() for part in value.split("|") if part.strip()]
    if not parts:
        return None
    if len(parts) == 1:
        return GenArtSelection(id=parts[0], bezeichnung=parts[0])
    return GenArtSelection(id=parts[0], bezeichnung=parts[-1])


def normalize_genart_selections(
    selections: list[GenArtSelection],
    fallback_id: str = "",
    fallback_bezeichnung: str = "",
) -> list[GenArtSelection]:
    normalized: list[GenArtSelection] = []
    seen_keys: set[str] = set()
    candidates = list(selections)
    if fallback_id.strip() or fallback_bezeichnung.strip():
        candidates.append(GenArtSelection(id=fallback_id, bezeichnung=fallback_bezeichnung))

    for candidate in candidates:
        option_id = candidate.id.strip()
        bezeichnung = candidate.bezeichnung.strip()
        if not option_id and not bezeichnung:
            continue
        if not bezeichnung:
            bezeichnung = option_id
        if not option_id:
            option_id = bezeichnung
        dedupe_key = option_id.casefold() or bezeichnung.casefold()
        if dedupe_key in seen_keys:
            continue
        seen_keys.add(dedupe_key)
        normalized.append(GenArtSelection(id=option_id, bezeichnung=bezeichnung))
    return normalized


def summarize_genart_selections(
    selections: list[GenArtSelection],
    *,
    empty_label: str = "-",
    limit: int = 2,
) -> str:
    normalized = normalize_genart_selections(selections)
    if not normalized:
        return empty_label
    labels = [selection.display_label() for selection in normalized]
    if len(labels) <= limit:
        return ", ".join(labels)
    return ", ".join(labels[:limit]) + f" (+{len(labels) - limit} weitere)"


def format_genart_selections(label: str, selections: list[GenArtSelection]) -> str:
    normalized = normalize_genart_selections(selections)
    if not normalized:
        return f"{label}: -"
    lines = [f"{label}:"]
    for selection in normalized:
        lines.append(f"- {selection.display_label()}")
    return "\n".join(lines)


def normalize_oe_number_rows(rows: list[OeNumberRow]) -> list[OeNumberRow]:
    normalized: list[OeNumberRow] = []
    seen_values: set[tuple[str, str]] = set()
    for row in rows:
        value = row.value.strip()
        manufacturer_id = row.manufacturer_id.strip()
        manufacturer_code = row.manufacturer_code.strip()
        manufacturer_name = row.manufacturer_name.strip()
        if not value:
            continue
        dedupe_key = (manufacturer_id.casefold(), value.casefold())
        if dedupe_key in seen_values:
            continue
        seen_values.add(dedupe_key)
        normalized.append(
            OeNumberRow(
                value=value,
                manufacturer_id=manufacturer_id,
                manufacturer_code=manufacturer_code,
                manufacturer_name=manufacturer_name,
            )
        )
    return normalized


def normalize_comparison_number_rows(rows: list[ComparisonNumberRow]) -> list[ComparisonNumberRow]:
    normalized: list[ComparisonNumberRow] = []
    seen_values: set[tuple[str, str]] = set()
    for row in rows:
        competitor_id = row.competitor_id.strip()
        competitor_code = row.competitor_code.strip()
        competitor_name = row.competitor_name.strip()
        reference_number = row.reference_number.strip()
        if not competitor_id or not reference_number:
            continue
        dedupe_key = (competitor_id.casefold(), reference_number.casefold())
        if dedupe_key in seen_values:
            continue
        seen_values.add(dedupe_key)
        normalized.append(
            ComparisonNumberRow(
                competitor_id=competitor_id,
                competitor_code=competitor_code,
                competitor_name=competitor_name,
                reference_number=reference_number,
            )
        )
    return normalized


def normalize_vehicle_link_rows(rows: list[VehicleLinkRow]) -> list[VehicleLinkRow]:
    normalized: list[VehicleLinkRow] = []
    seen_values: set[tuple[str, str, str]] = set()
    for row in rows:
        vehicle_type_id = row.vehicle_type_id.strip() or DEFAULT_VEHICLE_TYPE_ID
        vehicle_type_label = row.vehicle_type_label.strip() or vehicle_type_label_for_id(vehicle_type_id)
        ktyp_topmotive = row.ktyp_topmotive.strip()
        ktyp_tecdoc = row.ktyp_tecdoc.strip()
        if not ktyp_topmotive and not ktyp_tecdoc:
            continue
        dedupe_key = (vehicle_type_id, ktyp_topmotive, ktyp_tecdoc)
        if dedupe_key in seen_values:
            continue
        seen_values.add(dedupe_key)
        normalized.append(
            VehicleLinkRow(
                vehicle_type_id=vehicle_type_id,
                vehicle_type_label=vehicle_type_label,
                motorcode=row.motorcode.strip(),
                ktyp_topmotive=ktyp_topmotive,
                ktyp_tecdoc=ktyp_tecdoc,
                manufacturer=row.manufacturer.strip(),
                model=row.model.strip(),
                description=row.description.strip(),
                year_from=row.year_from.strip(),
                year_to=row.year_to.strip(),
                power=row.power.strip(),
            )
        )
    return normalized


def normalize_search_terms(terms: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for term in terms:
        cleaned = " ".join(str(term or "").split())
        if not cleaned:
            continue
        dedupe_key = cleaned.casefold()
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        normalized.append(cleaned)
    return normalized


def normalize_mapping_label(value: str) -> str:
    """Normalisiert ein Text-Label für die Attribut-Zuordnung (Umlaute, Klammern, Spaces)."""
    text = str(value or "").strip()
    text = re.sub(r"\([^)]*\)", " ", text)  # Klammerzusätze wie (LxBxH) entfernen
    text = text.translate(SHORT_TEXT_UMLAUT_REPLACEMENTS)
    text = re.sub(r"[^a-z0-9]+", " ", text.casefold())
    return " ".join(text.split())


# Einheiten-Normalisierung: Schreibweisen im Text -> kanonische Einheit
_UNIT_ALIASES = {
    "mm": "mm", "cm": "cm", "m": "m", "meter": "m",
    "g": "g", "kg": "kg", "t": "t", "gramm": "g", "kilogramm": "kg",
    "n": "n", "kn": "kn", "nm": "nm",
    "v": "v", "volt": "v", "a": "a", "ampere": "a",
    "ah": "ah", "mah": "mah",
    "w": "w", "kw": "kw",
    "lm": "lm", "lumen": "lm",
    "std": "std", "std.": "std", "h": "std", "stunden": "std", "stunde": "std",
    "min": "min", "minuten": "min", "s": "s", "sek": "s",
    "bar": "bar", "l": "l", "liter": "l", "ml": "ml",
    "zoll": "zoll", '"': "zoll",
    "u/min": "1/min", "1/min": "1/min", "upm": "1/min",
    "st": "st", "st.": "st", "stk": "st", "stk.": "st", "tlg": "tlg", "tlg.": "tlg", "teile": "st",
    "%": "%", "grad": "grad",
}

# Umrechnung (von, nach) -> Faktor
_UNIT_CONVERSIONS = {
    ("mm", "cm"): 0.1, ("cm", "mm"): 10.0,
    ("mm", "m"): 0.001, ("m", "mm"): 1000.0,
    ("cm", "m"): 0.01, ("m", "cm"): 100.0,
    ("zoll", "mm"): 25.4, ("mm", "zoll"): 1 / 25.4,
    ("g", "kg"): 0.001, ("kg", "g"): 1000.0,
    ("t", "kg"): 1000.0, ("kg", "t"): 0.001,
    ("mah", "ah"): 0.001, ("ah", "mah"): 1000.0,
    ("min", "std"): 1 / 60.0, ("std", "min"): 60.0,
    ("ml", "l"): 0.001, ("l", "ml"): 1000.0,
    ("kw", "w"): 1000.0, ("w", "kw"): 0.001,
    ("kn", "n"): 1000.0, ("n", "kn"): 0.001,
}


def normalize_unit_text(value: str) -> str:
    return _UNIT_ALIASES.get(str(value or "").strip().casefold(), str(value or "").strip().casefold())


def extract_attribute_unit(attribute_label: str) -> str:
    """Liest die Zieleinheit aus einer Attribut-Bezeichnung wie 'Gewicht [kg]'."""
    match = re.search(r"\[([^\]]+)\]", attribute_label or "")
    return normalize_unit_text(match.group(1)) if match else ""


def parse_german_number(text: str) -> float | None:
    value = str(text or "").strip().replace(" ", "").replace(" ", "")
    if not value:
        return None
    value = re.sub(r"\.(?=\d{3}(\D|$))", "", value)  # Tausenderpunkte
    value = value.replace(",", ".")
    try:
        return float(value)
    except ValueError:
        return None


def format_german_number(value: float) -> str:
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{round(value, 3):g}".replace(".", ",")


def convert_unit_value(value: float, from_unit: str, to_unit: str) -> float | None:
    from_unit = normalize_unit_text(from_unit)
    to_unit = normalize_unit_text(to_unit)
    if not to_unit or from_unit == to_unit:
        return value
    factor = _UNIT_CONVERSIONS.get((from_unit, to_unit))
    if factor is None:
        return None
    return value * factor


_NUMBER_PATTERN = r"\d+(?:[.,]\d+)*"
_UNIT_PATTERN = r"[A-Za-z/%\".]{0,6}"


def extract_spec_lines_from_text(text: str) -> list[ExtractedSpecLine]:
    """Findet 'Label<Tab>Wert'- und 'Label: Wert'-Zeilen in einem Produkttext."""
    specs: list[ExtractedSpecLine] = []
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip().lstrip("•·*–-• ").strip()
        if not line or len(line) > 160:
            continue
        if "\t" in line:
            label, _, value = line.partition("\t")
        elif ":" in line:
            label, _, value = line.partition(":")
        else:
            continue
        label = " ".join(label.split())
        value = value.strip()
        if not label or not value or len(label) > 45:
            continue
        if "." in label and len(label.split()) > 4:
            continue  # Fliesstext, kein Label
        specs.append(ExtractedSpecLine(label=label, raw_value=value, source_line=line))
    return specs


def _strip_trailing_parenthetical(value: str) -> tuple[str, str]:
    match = re.match(r"^(.*?)\s*\(([^)]*)\)\s*$", value.strip())
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return value.strip(), ""


def parse_dimension_value(raw_value: str) -> tuple[list[float], str] | None:
    """Erkennt Abmessungen wie '155 x 17 x 13,5 mm' -> ([155, 17, 13.5], 'mm')."""
    value, _note = _strip_trailing_parenthetical(raw_value)
    match = re.match(
        rf"^({_NUMBER_PATTERN})\s*[x×]\s*({_NUMBER_PATTERN})(?:\s*[x×]\s*({_NUMBER_PATTERN}))?\s*({_UNIT_PATTERN})?\s*$",
        value,
    )
    if not match:
        return None
    numbers = [parse_german_number(part) for part in match.groups()[:3] if part]
    if any(number is None for number in numbers) or len(numbers) < 2:
        return None
    unit = normalize_unit_text(match.group(4) or "mm")
    return [float(n) for n in numbers if n is not None], unit


def parse_numeric_value(raw_value: str) -> tuple[float, float | None, str] | None:
    """Erkennt Einzelwerte und Bereiche: '612 g' / '150 - 530 mm' -> (Wert, Wert-bis, Einheit)."""
    value, _note = _strip_trailing_parenthetical(raw_value)
    range_match = re.match(
        rf"^({_NUMBER_PATTERN})\s*(?:-|–|bis)\s*({_NUMBER_PATTERN})\s*({_UNIT_PATTERN})?\s*$",
        value,
    )
    if range_match:
        first = parse_german_number(range_match.group(1))
        second = parse_german_number(range_match.group(2))
        if first is not None and second is not None:
            return first, second, normalize_unit_text(range_match.group(3) or "")
    single_match = re.match(rf"^(?:ca\.?\s*|~\s*)?({_NUMBER_PATTERN})\s*({_UNIT_PATTERN})?\s*$", value)
    if single_match:
        number = parse_german_number(single_match.group(1))
        if number is not None:
            return number, None, normalize_unit_text(single_match.group(2) or "")
    return None


_DIMENSION_ORDER_LABELS = {"l": "laenge", "b": "breite", "h": "hoehe", "t": "tiefe"}


def parse_dimension_order_hint(label: str) -> list[str]:
    """Liest die Reihenfolge aus Labels wie 'Abmessungen (LxBxH)' oder '... (HxBxT)'."""
    match = re.search(r"\(([lbht])\s*x\s*([lbht])\s*(?:x\s*([lbht]))?\)", label, re.IGNORECASE)
    if match:
        parts = [part for part in match.groups() if part]
        return [_DIMENSION_ORDER_LABELS[part.casefold()] for part in parts]
    return ["laenge", "breite", "hoehe"]


def is_dimension_spec_label(label: str) -> bool:
    return normalize_mapping_label(label) in {"abmessung", "abmessungen", "masse", "abmessungen ca"}


def build_attribute_suggestions_from_text(
    text: str,
    mapping: dict[str, str],
    attribute_options_by_id: dict[str, AttributeOption],
    attribute_key_values_by_group: dict[str, list[AttributeKeyValueOption]],
) -> tuple[list[AttributeSuggestion], list[ExtractedSpecLine]]:
    """Erzeugt Attribut-Vorschläge aus einem Produkttext.

    Rückgabe: (Vorschläge, nicht zugeordnete Spezifikationszeilen).
    """
    suggestions: list[AttributeSuggestion] = []
    unmatched: list[ExtractedSpecLine] = []
    seen: set[tuple[str, str, str]] = set()

    def add_suggestion(suggestion: AttributeSuggestion) -> None:
        key = (suggestion.option.criteria_id, suggestion.value.casefold(), suggestion.value_to.casefold())
        if key in seen:
            return
        seen.add(key)
        suggestions.append(suggestion)

    def suggest_for_option(
        option: AttributeOption,
        raw_value: str,
        source_line: str,
        parsed: tuple[float, float | None, str] | None = None,
    ) -> AttributeSuggestion | None:
        value_format = option.value_format.strip().casefold()
        if value_format == "kein wert":
            return None
        if parsed is None:
            parsed = parse_numeric_value(raw_value)

        if value_format == "numerisch":
            if parsed is None:
                return AttributeSuggestion(
                    option=option, value=raw_value.strip(), source_line=source_line,
                    confident=False, note="Wert nicht als Zahl erkannt",
                )
            number, number_to, unit = parsed
            target_unit = extract_attribute_unit(option.label)
            note = ""
            if unit and target_unit and unit != target_unit:
                converted = convert_unit_value(number, unit, target_unit)
                converted_to = convert_unit_value(number_to, unit, target_unit) if number_to is not None else None
                if converted is None:
                    return AttributeSuggestion(
                        option=option, value=format_german_number(number), source_line=source_line,
                        confident=False, note=f"Einheit '{unit}' passt nicht zu [{target_unit}]",
                    )
                note = f"umgerechnet aus {format_german_number(number)} {unit}"
                number, number_to = converted, converted_to
            return AttributeSuggestion(
                option=option,
                value=format_german_number(number),
                value_to=format_german_number(number_to) if number_to is not None else "",
                source_line=source_line,
                confident=True,
                note=note,
            )

        if is_attribute_key_value_format(option.value_format):
            candidates = [raw_value.strip()]
            if parsed is not None:
                number, _number_to, unit = parsed
                target_unit = extract_attribute_unit(option.label)
                if unit and target_unit and unit != target_unit:
                    converted = convert_unit_value(number, unit, target_unit)
                    if converted is not None:
                        number = converted
                candidates.insert(0, format_german_number(number))
                candidates.append(format_german_number(number).replace(",", "."))
            probe = AttributeRow(
                criteria_id=option.criteria_id, label=option.label,
                value_format=option.value_format, type_name=option.type_name,
            )
            for candidate in candidates:
                resolved = resolve_attribute_key_value_option(probe, candidate, attribute_key_values_by_group)
                if resolved is not None:
                    return AttributeSuggestion(
                        option=option, value=resolved.display_label(), source_line=source_line,
                        confident=True, note="Schlüsselwert",
                    )
            return AttributeSuggestion(
                option=option, value=raw_value.strip(), source_line=source_line,
                confident=False, note="Schlüsselwert nicht in Werteliste gefunden",
            )

        # Alphanumerisch und alles Übrige: Text direkt übernehmen
        cleaned = " ".join(raw_value.split())
        max_length = option.max_length if option.max_length else None
        if max_length and len(cleaned) > max_length:
            return AttributeSuggestion(
                option=option, value=cleaned, source_line=source_line,
                confident=False, note=f"länger als {max_length} Zeichen",
            )
        return AttributeSuggestion(option=option, value=cleaned, source_line=source_line, confident=True)

    for spec in extract_spec_lines_from_text(text):
        base_label = normalize_mapping_label(spec.label)
        if not base_label:
            continue

        # Abmessungen wie '390 x 450 x 130 mm' in Länge/Breite/Höhe/Tiefe aufteilen
        dimensions = parse_dimension_value(spec.raw_value)
        if dimensions is not None and (is_dimension_spec_label(spec.label) or base_label in mapping):
            numbers, unit = dimensions
            order = parse_dimension_order_hint(spec.label + " " + spec.raw_value)
            if is_dimension_spec_label(spec.label):
                handled = False
                for dimension_label, number in zip(order, numbers):
                    criteria_id = mapping.get(dimension_label)
                    option = attribute_options_by_id.get(criteria_id or "")
                    if option is None:
                        continue
                    suggestion = suggest_for_option(
                        option, f"{format_german_number(number)} {unit}", spec.source_line,
                        parsed=(number, None, unit),
                    )
                    if suggestion is not None:
                        add_suggestion(suggestion)
                        handled = True
                if handled:
                    continue

        criteria_id = mapping.get(base_label)
        option = attribute_options_by_id.get(criteria_id or "") if criteria_id else None
        if option is None:
            unmatched.append(spec)
            continue

        raw_value, _paren_note = _strip_trailing_parenthetical(spec.raw_value)
        # Mehrfachwerte ('80.000 / 40.000 kg', '8 - 9 - 10 mm' mit Aufzählungszeichen) nur als unsicher vorschlagen
        if re.search(rf"{_NUMBER_PATTERN}\s*[/|•]\s*{_NUMBER_PATTERN}", raw_value):
            first_number = parse_numeric_value(re.split(r"[/|•]", raw_value, 1)[0].strip() + " " + (re.search(r"([A-Za-z/%\".]{1,6})\s*$", raw_value).group(1) if re.search(r"([A-Za-z/%\".]{1,6})\s*$", raw_value) else ""))
            suggestion = AttributeSuggestion(
                option=option,
                value=format_german_number(first_number[0]) if first_number else raw_value,
                source_line=spec.source_line,
                confident=False,
                note="Mehrere Werte in der Zeile - bitte prüfen",
            )
            add_suggestion(suggestion)
            continue

        suggestion = suggest_for_option(option, raw_value, spec.source_line)
        if suggestion is not None:
            add_suggestion(suggestion)
        else:
            unmatched.append(spec)

    return suggestions, unmatched


def normalize_attribute_rows(rows: list[AttributeRow]) -> list[AttributeRow]:
    normalized: list[AttributeRow] = []
    seen_values: set[tuple[str, str, str]] = set()
    for row in rows:
        criteria_id = row.criteria_id.strip()
        label = row.label.strip()
        value_format = row.value_format.strip()
        value = row.value.strip()
        value_to = row.value_to.strip()
        if not criteria_id:
            continue
        if value_format.casefold() == "kein wert":
            value = ""
            value_to = ""
        dedupe_key = (criteria_id.casefold(), value.casefold(), value_to.casefold())
        if dedupe_key in seen_values:
            continue
        seen_values.add(dedupe_key)
        normalized.append(
            AttributeRow(
                criteria_id=criteria_id,
                label=label,
                value_format=value_format,
                max_length=row.max_length,
                type_name=row.type_name.strip(),
                value=value,
                value_to=value_to,
            )
        )
    return normalized


def format_oe_number_rows(label: str, rows: list[OeNumberRow]) -> str:
    normalized = normalize_oe_number_rows(rows)
    lines = [label]
    if not normalized:
        lines.append("-")
        return "\n".join(lines)
    for index, row in enumerate(normalized, start=1):
        lines.append(f"{index}. {row.value}")
        lines.append(f"   Hersteller: {row.display_manufacturer_label()}")
    return "\n".join(lines)


def format_comparison_number_rows(label: str, rows: list[ComparisonNumberRow]) -> str:
    normalized = normalize_comparison_number_rows(rows)
    lines = [label]
    if not normalized:
        lines.append("-")
        return "\n".join(lines)
    for index, row in enumerate(normalized, start=1):
        competitor_label = row.display_competitor_label()
        lines.append(f"{index}. {row.reference_number}")
        lines.append(f"   Mitbewerber: {competitor_label}")
    return "\n".join(lines)


def format_attribute_rows(label: str, rows: list[AttributeRow]) -> str:
    normalized = normalize_attribute_rows(rows)
    lines = [label]
    if not normalized:
        lines.append("-")
        return "\n".join(lines)
    for index, row in enumerate(normalized, start=1):
        lines.append(f"{index}. {row.display_label()}")
        lines.append(f"   Format: {row.value_format or '-'} | Wert: {row.display_value()}")
    return "\n".join(lines)


def split_search_terms(value: str) -> list[str]:
    return [part for part in re.split(r"\s+", value.strip().casefold()) if part]


def score_attribute_option_match(option: AttributeOption, query: str) -> float:
    query_text = query.strip().casefold()
    if not query_text:
        return 0.0

    label = option.label.casefold()
    criteria_id = option.criteria_id.casefold()
    display_label = option.display_label().casefold()
    tokens = split_search_terms(query_text)
    if not tokens:
        return 0.0

    score = 0.0
    for token in tokens:
        if token == criteria_id:
            score += 5000
        elif token == label:
            score += 4500
        elif criteria_id.startswith(token):
            score += 2400
        elif label.startswith(token):
            score += 2200
        elif any(word.startswith(token) for word in re.split(r"[^a-z0-9]+", display_label) if word):
            score += 1500
        elif token in display_label:
            score += 900
        elif token in option.search_blob:
            score += 450
        else:
            score -= 200

    score += SequenceMatcher(None, query_text, display_label).ratio() * 200
    return score


def normalize_lookup_text(value: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return ""
    for source, replacement in LOOKUP_TEXT_REPLACEMENTS:
        normalized = normalized.replace(source, replacement)
    normalized = normalized.casefold()
    normalized = unicodedata.normalize("NFKD", normalized)
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def normalize_attribute_key_value_group(value: str) -> str:
    return normalize_lookup_text(value)


def is_attribute_key_value_format(value: str) -> bool:
    return normalize_lookup_text(value) in ATTRIBUTE_KEY_VALUE_FORMAT_KEYS


def parse_attribute_key_value_display(value: str) -> tuple[str, str]:
    parts = [part.strip() for part in value.split("|", 1)]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def score_attribute_key_value_match(option: AttributeKeyValueOption, query: str) -> float:
    query_text = normalize_lookup_text(query)
    if not query_text:
        return 0.0

    label = normalize_lookup_text(option.label)
    key_value_id = normalize_lookup_text(option.key_value_id)
    display_label = normalize_lookup_text(option.display_label())
    tokens = split_search_terms(query_text)
    if not tokens:
        return 0.0

    score = 0.0
    for token in tokens:
        if token == key_value_id:
            score += 5000
        elif token == label:
            score += 4500
        elif key_value_id.startswith(token):
            score += 2400
        elif label.startswith(token):
            score += 2200
        elif any(word.startswith(token) for word in re.split(r"[^a-z0-9]+", display_label) if word):
            score += 1500
        elif token in display_label:
            score += 900
        elif token in option.search_blob:
            score += 450
        else:
            score -= 200

    score += SequenceMatcher(None, query_text, display_label).ratio() * 200
    return score


def resolve_attribute_key_value_option(
    row: AttributeRow,
    raw_value: str,
    attribute_key_values_by_group: dict[str, list[AttributeKeyValueOption]],
) -> AttributeKeyValueOption | None:
    value = raw_value.strip()
    if not value:
        return None

    group_options: list[AttributeKeyValueOption] = []
    for group_key in row.key_value_group_candidates():
        group_options = attribute_key_values_by_group.get(group_key, [])
        if group_options:
            break
    if not group_options:
        return None

    value_key = normalize_lookup_text(value)
    key_value_id, key_value_label = parse_attribute_key_value_display(value)
    key_value_id_key = normalize_lookup_text(key_value_id)
    key_value_label_key = normalize_lookup_text(key_value_label)

    for option in group_options:
        if value_key and value_key == normalize_lookup_text(option.display_label()):
            return option
    for option in group_options:
        if key_value_id_key and key_value_id_key == normalize_lookup_text(option.key_value_id):
            return option

    exact_label_matches = [option for option in group_options if value_key and value_key == normalize_lookup_text(option.label)]
    if len(exact_label_matches) == 1:
        return exact_label_matches[0]

    parsed_label_matches = [option for option in group_options if key_value_label_key and key_value_label_key == normalize_lookup_text(option.label)]
    if len(parsed_label_matches) == 1:
        return parsed_label_matches[0]

    partial_matches = [
        option
        for option in group_options
        if value_key
        and (
            value_key in normalize_lookup_text(option.display_label())
            or value_key in normalize_lookup_text(option.label)
            or value_key in normalize_lookup_text(option.key_value_id)
        )
    ]
    if len(partial_matches) == 1:
        return partial_matches[0]
    return None


def resolve_attribute_export_value(
    row: AttributeRow,
    attribute_key_values_by_group: dict[str, list[AttributeKeyValueOption]],
) -> str:
    raw_value = row.value.strip()
    if not raw_value:
        return ""
    option = resolve_attribute_key_value_option(row, raw_value, attribute_key_values_by_group)
    if option is not None:
        return option.key_value_id
    key_value_id, _key_value_label = parse_attribute_key_value_display(raw_value)
    return key_value_id or raw_value


def resolve_attribute_key_value_display_value(
    row: AttributeRow,
    attribute_key_values_by_group: dict[str, list[AttributeKeyValueOption]],
) -> str:
    raw_value = row.value.strip()
    if not raw_value:
        return ""
    option = resolve_attribute_key_value_option(row, raw_value, attribute_key_values_by_group)
    return option.display_label() if option is not None else raw_value


def build_attribute_key_value_group_index(
    options: list[AttributeKeyValueOption],
) -> dict[str, list[AttributeKeyValueOption]]:
    grouped: dict[str, list[AttributeKeyValueOption]] = {}
    for option in options:
        group_key = normalize_attribute_key_value_group(option.attribute_group)
        if not group_key:
            continue
        grouped.setdefault(group_key, []).append(option)
    return grouped


def normalize_article_number(value: str) -> str:
    return " ".join(value.strip().split())


def normalize_header_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def parse_csv_flag(value: object) -> bool:
    return str(value or "").strip().casefold() in {"1", "true", "wahr", "yes", "ja", "x"}


def parse_bool_flag(value: object) -> bool:
    return str(value or "").strip().casefold() in {"1", "true", "wahr", "yes", "ja", "x"}


def load_competitor_options(csv_path: Path, *, comparison_only: bool = True) -> list[CompetitorOption]:
    if not path_exists_safe(csv_path):
        return []

    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with csv_path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle, delimiter=";")
                if not reader.fieldnames:
                    raise ValueError("Die KHer-CSV hat keine Kopfzeile.")
                header_map = {normalize_header_key(field_name): field_name for field_name in reader.fieldnames if field_name}
                id_key = header_map.get("khernr")
                code_key = header_map.get("khkz")
                name_key = header_map.get("bez")
                compare_key = header_map.get("vgl")
                if id_key is None:
                    raise ValueError("In der KHer-CSV fehlt die Spalte 'KHerNr'.")
                if name_key is None:
                    raise ValueError("In der KHer-CSV fehlt die Spalte 'Bez'.")

                options: list[CompetitorOption] = []
                seen_ids: set[str] = set()
                for row in reader:
                    competitor_id = str(row.get(id_key, "")).strip()
                    if not competitor_id or competitor_id in seen_ids:
                        continue
                    if comparison_only and compare_key is not None and not parse_csv_flag(row.get(compare_key)):
                        continue
                    seen_ids.add(competitor_id)
                    options.append(
                        CompetitorOption(
                            competitor_id=competitor_id,
                            code=str(row.get(code_key, "")).strip() if code_key else "",
                            name=str(row.get(name_key, "")).strip(),
                        )
                    )

                options.sort(key=lambda option: (option.name.casefold(), option.code.casefold(), option.competitor_id.casefold()))
                return options
        except UnicodeError as exc:
            last_error = exc
            continue
        except Exception as exc:
            last_error = exc
            break

    if last_error is not None:
        raise ValueError(f"KHer-CSV konnte nicht geladen werden: {last_error}") from last_error
    return []


class VehicleCatalog:
    """Nachschlagewerk für KTyp-Stammdaten (Motorcode -> Fahrzeuge)."""

    def __init__(self) -> None:
        self.by_motorcode: dict[str, list[VehicleMatch]] = {}
        self.by_ktyp_topmotive: dict[str, VehicleMatch] = {}
        self.by_ktyp_tecdoc: dict[str, VehicleMatch] = {}
        self.count = 0
        self._motorcode_set: set[str] = set()
        self._sorted_motorcodes: list[str] | None = None

    def _register(self, match: VehicleMatch) -> None:
        for key in motorcode_keys(match.motorcode):
            bucket = self.by_motorcode.setdefault(key, [])
            if not any(existing.identity() == match.identity() for existing in bucket):
                bucket.append(match)
        if match.motorcode:
            self._motorcode_set.add(match.motorcode)
        if match.ktyp_topmotive:
            self.by_ktyp_topmotive.setdefault(match.ktyp_topmotive, match)
        if match.ktyp_tecdoc:
            self.by_ktyp_tecdoc.setdefault(match.ktyp_tecdoc, match)
        self.count += 1

    def sorted_motorcodes(self) -> list[str]:
        if self._sorted_motorcodes is None:
            self._sorted_motorcodes = sorted(self._motorcode_set)
        return self._sorted_motorcodes

    def suggest(self, query: str, limit: int = 12) -> list[str]:
        """Vorschläge für einen (Teil-)Motorcode inkl. Fuzzy-Nähe."""
        normalized = normalize_motorcode(query)
        if not normalized:
            return []
        codes = self.sorted_motorcodes()
        query_ns = normalized.replace(" ", "")
        prefix: list[str] = []
        substring: list[str] = []
        for code in codes:
            code_ns = code.replace(" ", "")
            if code.startswith(normalized) or code_ns.startswith(query_ns):
                prefix.append(code)
            elif normalized in code or query_ns in code_ns:
                substring.append(code)
        results: list[str] = prefix[:limit]
        for code in substring:
            if len(results) >= limit:
                break
            results.append(code)
        if len(results) < limit and len(query_ns) >= 2:
            for code in get_close_matches(normalized, codes, n=limit, cutoff=0.6):
                if code not in results:
                    results.append(code)
                    if len(results) >= limit:
                        break
        return results[:limit]

    def lookup(self, motorcode: str) -> list[VehicleMatch]:
        seen: set[tuple[str, str]] = set()
        results: list[VehicleMatch] = []
        for key in motorcode_keys(motorcode):
            for match in self.by_motorcode.get(key, []):
                identity = match.identity()
                if identity in seen:
                    continue
                seen.add(identity)
                results.append(match)
        return results

    def lookup_by_ktyp(self, ktyp_topmotive: str = "", ktyp_tecdoc: str = "") -> VehicleMatch | None:
        ktyp_topmotive = str(ktyp_topmotive or "").strip()
        ktyp_tecdoc = str(ktyp_tecdoc or "").strip()
        if ktyp_topmotive and ktyp_topmotive in self.by_ktyp_topmotive:
            return self.by_ktyp_topmotive[ktyp_topmotive]
        if ktyp_tecdoc and ktyp_tecdoc in self.by_ktyp_tecdoc:
            return self.by_ktyp_tecdoc[ktyp_tecdoc]
        return None


def _clean_ktyp_value(value: object) -> str:
    text = str(value if value is not None else "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _format_vehicle_date(value: object) -> str:
    """Reduziert ein BJ-Datum auf ``MM.YYYY`` (aus ``01.09.1997 00:00:00``)."""
    text = str(value if value is not None else "").strip()
    if not text:
        return ""
    date_part = text.split(" ", 1)[0]
    match = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", date_part)
    if match:
        return f"{match.group(2).zfill(2)}.{match.group(3)}"
    iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", date_part)
    if iso_match:
        return f"{iso_match.group(2)}.{iso_match.group(1)}"
    return date_part


def load_vehicle_catalog(workbook_path: Path) -> VehicleCatalog:
    catalog = VehicleCatalog()
    if not path_exists_safe(workbook_path):
        return catalog

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"KTyp-Datei konnte nicht geladen werden: {exc}") from exc

    try:
        sheet_name = "Daten" if "Daten" in workbook.sheetnames else workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        row_iter = worksheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            raise ValueError("Die KTyp-Datei hat keine Kopfzeile.")
        header_map = {
            normalize_header_key(str(value)): index
            for index, value in enumerate(header_row)
            if str(value or "").strip()
        }
        idx_motor = header_map.get("motorcode")
        idx_topmotive = header_map.get("ktypnrtopmotive")
        idx_tecdoc = header_map.get("ktypnrtecdoc")
        idx_manufacturer = header_map.get("herstellerbezeichnung")
        idx_model = header_map.get("modellbezeichnung")
        idx_description = header_map.get("bezeichnung")
        idx_year_from = header_map.get("bjvon")
        idx_year_to = header_map.get("bjbis")
        idx_kw = header_map.get("lw") or header_map.get("kw")
        idx_ps = header_map.get("ps")
        idx_deleted = header_map.get("gelscht") or header_map.get("geloescht")
        if idx_motor is None or (idx_topmotive is None and idx_tecdoc is None):
            raise ValueError("In der KTyp-Datei fehlen die Spalten 'Motorcode' und 'KTypNr'.")

        def cell(values: tuple, index: int | None) -> str:
            if index is None or index >= len(values):
                return ""
            value = values[index]
            return str(value).strip() if value is not None else ""

        for values in row_iter:
            motorcode = cell(values, idx_motor)
            if not motorcode:
                continue
            if idx_deleted is not None and parse_bool_flag(values[idx_deleted] if idx_deleted < len(values) else ""):
                continue
            ktyp_topmotive = _clean_ktyp_value(values[idx_topmotive]) if idx_topmotive is not None and idx_topmotive < len(values) else ""
            ktyp_tecdoc = _clean_ktyp_value(values[idx_tecdoc]) if idx_tecdoc is not None and idx_tecdoc < len(values) else ""
            if not ktyp_topmotive and not ktyp_tecdoc:
                continue
            power_kw = cell(values, idx_kw)
            power_ps = cell(values, idx_ps)
            if power_kw and power_ps:
                power = f"{power_kw} kW / {power_ps} PS"
            elif power_kw:
                power = f"{power_kw} kW"
            elif power_ps:
                power = f"{power_ps} PS"
            else:
                power = ""
            catalog._register(
                VehicleMatch(
                    motorcode=normalize_motorcode(motorcode),
                    ktyp_topmotive=ktyp_topmotive,
                    ktyp_tecdoc=ktyp_tecdoc,
                    manufacturer=cell(values, idx_manufacturer),
                    model=cell(values, idx_model),
                    description=cell(values, idx_description),
                    year_from=_format_vehicle_date(values[idx_year_from]) if idx_year_from is not None and idx_year_from < len(values) else "",
                    year_to=_format_vehicle_date(values[idx_year_to]) if idx_year_to is not None and idx_year_to < len(values) else "",
                    power=power,
                )
            )
    finally:
        workbook.close()

    return catalog


def load_attribute_options(workbook_path: Path) -> list[AttributeOption]:
    if not path_exists_safe(workbook_path):
        return []

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Attributdatei konnte nicht geladen werden: {exc}") from exc

    try:
        worksheet = workbook[workbook.sheetnames[0]]
        row_iter = worksheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            raise ValueError("Die Attributdatei hat keine Kopfzeile.")
        header_map = {
            normalize_header_key(str(value)): index
            for index, value in enumerate(header_row)
            if str(value or "").strip()
        }
        criteria_index = header_map.get("tecdockriterienid")
        label_index = header_map.get("bezeichnung")
        format_index = header_map.get("format")
        max_length_index = header_map.get("maxlnge") or header_map.get("maxlnge".replace("ä", "a"))
        source_index = header_map.get("herkunft")
        type_index = header_map.get("typ")
        deleted_index = header_map.get("gelscht") or header_map.get("geloescht")
        deletion_date_index = header_map.get("lschdatum") or header_map.get("loeschdatum")
        if criteria_index is None or label_index is None or format_index is None:
            raise ValueError("In der Attributdatei fehlen benötigte Spalten.")

        options: list[AttributeOption] = []
        seen_ids: set[str] = set()
        for row in row_iter:
            values = list(row)
            criteria_id = str(values[criteria_index] or "").strip() if criteria_index < len(values) else ""
            if not criteria_id or criteria_id in seen_ids:
                continue
            deleted = parse_bool_flag(values[deleted_index]) if deleted_index is not None and deleted_index < len(values) else False
            deletion_date = str(values[deletion_date_index] or "").strip() if deletion_date_index is not None and deletion_date_index < len(values) else ""
            if deleted or deletion_date:
                continue
            raw_max_length = values[max_length_index] if max_length_index is not None and max_length_index < len(values) else None
            max_length: int | None = None
            if raw_max_length not in {None, ""}:
                try:
                    max_length = int(float(str(raw_max_length).strip()))
                except Exception:
                    max_length = None

            seen_ids.add(criteria_id)
            options.append(
                AttributeOption(
                    criteria_id=criteria_id,
                    label=str(values[label_index] or "").strip() if label_index < len(values) else "",
                    value_format=str(values[format_index] or "").strip() if format_index < len(values) else "",
                    max_length=max_length,
                    source=str(values[source_index] or "").strip() if source_index is not None and source_index < len(values) else "",
                    type_name=str(values[type_index] or "").strip() if type_index is not None and type_index < len(values) else "",
                )
            )

        options.sort(key=lambda option: (option.label.casefold(), option.criteria_id.casefold()))
        return options
    finally:
        workbook.close()


def load_attribute_mapping(workbook_path: Path) -> dict[str, str]:
    """Lädt die Text-Label -> TecDoc Kriterien ID Zuordnung für Attributvorschläge."""
    if not path_exists_safe(workbook_path):
        return {}

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Attribut-Zuordnung konnte nicht geladen werden: {exc}") from exc

    try:
        sheet_name = ATTRIBUTE_MAPPING_SHEET if ATTRIBUTE_MAPPING_SHEET in workbook.sheetnames else workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        row_iter = worksheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            raise ValueError("Die Attribut-Zuordnung hat keine Kopfzeile.")
        header_map = {
            normalize_header_key(str(value)): index
            for index, value in enumerate(header_row)
            if str(value or "").strip()
        }
        label_index = header_map.get("textlabel")
        criteria_index = header_map.get("tecdockriterienid")
        if label_index is None or criteria_index is None:
            raise ValueError("In der Attribut-Zuordnung fehlen die Spalten 'Text-Label' und 'TecDoc Kriterien ID'.")

        mapping: dict[str, str] = {}
        for values in row_iter:
            label = str(values[label_index] or "").strip() if label_index < len(values) else ""
            criteria_id = str(values[criteria_index] or "").strip() if criteria_index < len(values) else ""
            if criteria_id.endswith(".0") and criteria_id[:-2].isdigit():
                criteria_id = criteria_id[:-2]
            key = normalize_mapping_label(label)
            if key and criteria_id:
                mapping[key] = criteria_id
        return mapping
    finally:
        workbook.close()


def append_attribute_mapping_entries(workbook_path: Path, entries: list[tuple[str, str, str]]) -> None:
    """Hängt gelernte Zuordnungen (Text-Label, Kriterien-ID, Hinweis) an die Zuordnungsdatei an."""
    if not entries:
        return
    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
        worksheet = workbook[ATTRIBUTE_MAPPING_SHEET] if ATTRIBUTE_MAPPING_SHEET in workbook.sheetnames else workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = ATTRIBUTE_MAPPING_SHEET
        worksheet.append(ATTRIBUTE_MAPPING_HEADERS)
    for label, criteria_id, note in entries:
        worksheet.append([label, criteria_id, note])
    workbook.save(workbook_path)


def load_attribute_key_value_options(workbook_path: Path) -> list[AttributeKeyValueOption]:
    if not path_exists_safe(workbook_path):
        return []

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Schlüsselwertdatei konnte nicht geladen werden: {exc}") from exc

    try:
        worksheet = workbook[workbook.sheetnames[0]]
        row_iter = worksheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            raise ValueError("Die Schlüsselwertdatei hat keine Kopfzeile.")

        header_map = {
            normalize_header_key(str(value)): index
            for index, value in enumerate(header_row)
            if str(value or "").strip()
        }
        key_value_index = header_map.get("tecdocschlsselwert") or header_map.get("tecdocschluesselwert")
        label_index = header_map.get("bezeichnung")
        group_index = header_map.get("attributtabellenbezeichnung")
        source_index = header_map.get("herkunft")
        deleted_index = header_map.get("gelscht") or header_map.get("geloescht")
        deletion_date_index = header_map.get("lschdatum") or header_map.get("loeschdatum")
        if key_value_index is None or label_index is None or group_index is None:
            raise ValueError("In der Schlüsselwertdatei fehlen benötigte Spalten.")

        options: list[AttributeKeyValueOption] = []
        seen_keys: set[tuple[str, str]] = set()
        for row in row_iter:
            values = list(row)
            key_value_id = str(values[key_value_index] or "").strip() if key_value_index < len(values) else ""
            label = str(values[label_index] or "").strip() if label_index < len(values) else ""
            attribute_group = str(values[group_index] or "").strip() if group_index < len(values) else ""
            if not key_value_id or not attribute_group:
                continue
            deleted = parse_bool_flag(values[deleted_index]) if deleted_index is not None and deleted_index < len(values) else False
            deletion_date = str(values[deletion_date_index] or "").strip() if deletion_date_index is not None and deletion_date_index < len(values) else ""
            if deleted or deletion_date:
                continue
            dedupe_key = (normalize_attribute_key_value_group(attribute_group), key_value_id.casefold())
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)
            options.append(
                AttributeKeyValueOption(
                    key_value_id=key_value_id,
                    label=label,
                    attribute_group=attribute_group,
                    source=str(values[source_index] or "").strip() if source_index is not None and source_index < len(values) else "",
                )
            )

        options.sort(key=lambda option: (option.attribute_group.casefold(), option.label.casefold(), option.key_value_id.casefold()))
        return options
    finally:
        workbook.close()


def replace_short_text_umlauts(value: str) -> str:
    return value.translate(SHORT_TEXT_UMLAUT_REPLACEMENTS)


def get_application_base_path() -> Path:
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", "")
        if meipass:
            return Path(meipass)
    return Path(__file__).resolve().parent


def resolve_application_asset_path(relative_path: str | Path) -> Path:
    return get_application_base_path() / Path(relative_path)


def parse_version_parts(value: str) -> tuple[int, ...]:
    parts = [int(part) for part in re.findall(r"\d+", value)]
    return tuple(parts) if parts else (0,)


def is_newer_release_tag(current_tag: str, candidate_tag: str) -> bool:
    current_parts = parse_version_parts(current_tag)
    candidate_parts = parse_version_parts(candidate_tag)
    max_length = max(len(current_parts), len(candidate_parts))
    current_padded = current_parts + (0,) * (max_length - len(current_parts))
    candidate_padded = candidate_parts + (0,) * (max_length - len(candidate_parts))
    return candidate_padded > current_padded


def fetch_latest_github_release(timeout_seconds: int = 20) -> GitHubReleaseInfo:
    request = urllib_request.Request(
        GITHUB_RELEASES_LATEST_API_URL,
        headers={
            "Accept": "application/vnd.github+json",
            "User-Agent": f"{APP_TITLE}/{APP_VERSION}",
        },
    )
    try:
        with urllib_request.urlopen(request, timeout=timeout_seconds) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib_error.HTTPError as exc:
        raise ValueError(f"GitHub Release konnte nicht geladen werden ({exc.code}).") from exc
    except urllib_error.URLError as exc:
        raise ValueError(f"GitHub Release konnte nicht geladen werden: {exc.reason}") from exc

    assets = [
        GitHubReleaseAsset(
            name=str(item.get("name", "")).strip(),
            download_url=str(item.get("browser_download_url", "")).strip(),
            content_type=str(item.get("content_type", "")).strip(),
            size=int(item.get("size", 0) or 0),
        )
        for item in payload.get("assets", [])
        if str(item.get("browser_download_url", "")).strip()
    ]
    return GitHubReleaseInfo(
        tag_name=str(payload.get("tag_name", "")).strip(),
        name=str(payload.get("name", "")).strip(),
        html_url=str(payload.get("html_url", "")).strip(),
        body=str(payload.get("body", "") or ""),
        published_at=str(payload.get("published_at", "")).strip(),
        assets=assets,
    )


def choose_release_asset(release: GitHubReleaseInfo) -> GitHubReleaseAsset | None:
    def score(asset: GitHubReleaseAsset) -> tuple[int, int]:
        name = asset.name.casefold()
        if asset.suffix == ".exe" and ("setup" in name or "installer" in name):
            return (4, len(asset.name))
        if asset.suffix == ".exe" and "onefile" in name:
            return (3, len(asset.name))
        if asset.suffix == ".exe":
            return (2, len(asset.name))
        if asset.suffix == ".zip":
            return (1, len(asset.name))
        return (0, len(asset.name))

    ranked = sorted(release.assets, key=score, reverse=True)
    best = ranked[0] if ranked else None
    if best is None or score(best)[0] <= 0:
        return None
    return best


def download_release_asset(asset: GitHubReleaseAsset, target_path: Path, timeout_seconds: int = 120) -> Path:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    request = urllib_request.Request(
        asset.download_url,
        headers={"User-Agent": f"{APP_TITLE}/{APP_VERSION}"},
    )
    try:
        with urllib_request.urlopen(request, timeout=timeout_seconds) as response, target_path.open("wb") as handle:
            while True:
                chunk = response.read(1024 * 64)
                if not chunk:
                    break
                handle.write(chunk)
    except urllib_error.HTTPError as exc:
        raise ValueError(f"Release-Datei konnte nicht geladen werden ({exc.code}).") from exc
    except urllib_error.URLError as exc:
        raise ValueError(f"Release-Datei konnte nicht geladen werden: {exc.reason}") from exc
    return target_path


def _certifi_ssl_context() -> ssl.SSLContext | None:
    try:
        import certifi
    except ImportError:  # pragma: no cover - optional dependency
        return None
    try:
        return ssl.create_default_context(cafile=certifi.where())
    except Exception:  # pragma: no cover - defensive
        return None


def fetch_preview_bytes_from_url(url: str, timeout_seconds: int, max_bytes: int | None = None) -> bytes:
    """Lädt Vorschau-Inhalte per HTTP(S) mit tolerantem SSL-Fallback.

    Manche Server (z. B. restapi.kunzer.de seit 11.07.2026) liefern eine
    unvollständige Zertifikatskette aus; Browser reparieren das selbst
    (AIA-Fetching), Pythons SSL nicht. Für reine Vorschau-Inhalte wird bei
    Zertifikatsfehlern deshalb erst das certifi-Bundle und zuletzt eine
    unverifizierte Verbindung genutzt. Update- und API-Downloads bleiben
    strikt verifiziert.
    """
    request = urllib_request.Request(url, headers={"User-Agent": "ApolloImportGui/1.0"})
    contexts: list[ssl.SSLContext | None] = [None]
    certifi_context = _certifi_ssl_context()
    if certifi_context is not None:
        contexts.append(certifi_context)
    contexts.append(ssl._create_unverified_context())
    last_error: Exception | None = None
    for context in contexts:
        try:
            with urllib_request.urlopen(request, timeout=timeout_seconds, context=context) as response:
                return response.read(max_bytes) if max_bytes is not None else response.read()
        except urllib_error.URLError as exc:
            reason = getattr(exc, "reason", None)
            if isinstance(reason, ssl.SSLCertVerificationError):
                last_error = exc
                continue
            raise
    raise last_error if last_error is not None else ValueError("Download fehlgeschlagen.")


def sanitize_short_translation_set(translations: TranslationSet) -> TranslationSet:
    values: dict[str, str] = {}
    for code in [language_code for language_code, _label in UI_LANGUAGE_ORDER]:
        sanitized = replace_short_text_umlauts(getattr(translations, code))
        values[code] = sanitized[:SHORT_TEXT_MAX_LENGTH]
    return TranslationSet(**values)


def normalize_match_text(value: str) -> str:
    normalized = (
        value.casefold()
        .replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
    )
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return " ".join(normalized.split())


def build_match_tokens(value: str) -> frozenset[str]:
    normalized = normalize_match_text(value)
    base_tokens = {token for token in normalized.split() if token}
    expanded_tokens = set(base_tokens)

    for token in list(base_tokens):
        for suffix in GENART_COMPOUND_SUFFIXES:
            if token != suffix and len(token) > len(suffix) + 2 and token.endswith(suffix):
                expanded_tokens.add(suffix)

    return frozenset(expanded_tokens)


def infer_match_families_from_tokens(tokens: set[str] | frozenset[str]) -> frozenset[str]:
    token_set = set(tokens)
    families = set()
    for family_name, terms in GENART_FAMILY_TERMS.items():
        if token_set & terms:
            families.add(family_name)
    return frozenset(families)


def infer_document_art(document_path: str) -> str:
    parsed = urlparse(document_path.strip())
    raw_name = unquote(parsed.path if parsed.scheme else document_path)
    lowered = Path(raw_name).name.lower()
    normalized = (
        lowered.replace("_", " ")
        .replace("-", " ")
        .replace(".", " ")
        .replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
    )
    if any(token in normalized for token in ["bedienungsanleitung", "betriebsanleitung", "einbauanleitung", "montageanleitung"]):
        return "14"
    if any(token in normalized for token in ["produktinfo", "zubehoer", "zubehor", "datenblatt", "broschuere", "flyer"]):
        return "17"
    return "17"


def safe_folder_name(article_number: str) -> str:
    cleaned = article_number.strip().replace(" ", "_")
    invalid = '<>:"/\\|?*'
    for char in invalid:
        cleaned = cleaned.replace(char, "_")
    return cleaned or "ohne_artikelnummer"


def normalize_csv_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


@dataclass
class CsvImportItem:
    article_number: str = ""
    product_url: str = ""


def _choose_csv_delimiter(sample: str) -> str:
    delimiters = [";", ",", "\t", "|"]
    counts = {delimiter: sample.count(delimiter) for delimiter in delimiters}
    best = max(counts, key=counts.get)
    return best if counts[best] else ";"


def _read_import_items_from_rows(header_values: list[object], data_rows: list[list[object]], source_label: str) -> list[CsvImportItem]:
    header_map = {
        normalize_csv_header(str(field_name)): index
        for index, field_name in enumerate(header_values)
        if field_name is not None and str(field_name).strip()
    }
    article_index = next((header_map[key] for key in CSV_ARTICLE_HEADER_ALIASES if key in header_map), None)
    url_index = next((header_map[key] for key in CSV_URL_HEADER_ALIASES if key in header_map), None)
    if article_index is None and url_index is None:
        raise ValueError(f"Die {source_label} braucht mindestens eine Spalte für Artikelnummer oder Produkt-URL.")

    items: list[CsvImportItem] = []
    for row in data_rows:
        article_value = ""
        if article_index is not None and article_index < len(row) and row[article_index] is not None:
            article_value = normalize_article_number(str(row[article_index]))

        url_value = ""
        if url_index is not None and url_index < len(row) and row[url_index] is not None:
            url_value = str(row[url_index]).strip()

        if not article_value and not url_value:
            continue
        items.append(CsvImportItem(article_number=article_value, product_url=url_value))

    if not items:
        raise ValueError(f"In der {source_label} wurden keine importierbaren Produkte gefunden.")
    return items


def read_product_import_items(path: Path) -> list[CsvImportItem]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not rows:
            raise ValueError("Die XLSX-Datei ist leer.")
        return _read_import_items_from_rows(list(rows[0]), [list(row) for row in rows[1:]], "XLSX-Datei")

    if suffix != ".csv":
        raise ValueError("Unterstützt werden derzeit CSV- und XLSX-Dateien.")

    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                delimiter = _choose_csv_delimiter(sample)
                reader = csv.DictReader(handle, delimiter=delimiter)
                if not reader.fieldnames:
                    raise ValueError("Die CSV-Datei hat keine Kopfzeile.")
                rows = [[row.get(field_name, "") for field_name in reader.fieldnames] for row in reader]
                return _read_import_items_from_rows(list(reader.fieldnames), rows, "CSV-Datei")
        except UnicodeDecodeError as exc:
            last_error = exc

    if last_error is not None:
        raise ValueError(f"CSV-Datei konnte nicht gelesen werden: {last_error}") from last_error
    raise ValueError("CSV-Datei konnte nicht gelesen werden.")


def normalize_youtube_url_for_embed(url: str) -> str:
    value = url.strip()
    parsed = urlparse(value)
    host = parsed.netloc.lower()
    path = parsed.path.strip("/")
    query = parse_qs(parsed.query)
    video_id = ""

    if host in {"youtu.be", "www.youtu.be"}:
        video_id = path.split("/", 1)[0]
    elif host.endswith("youtube.com") or host.endswith("youtube-nocookie.com"):
        parts = [part for part in path.split("/") if part]
        if len(parts) >= 2 and parts[0] in {"embed", "shorts", "live"}:
            video_id = parts[1]
        elif parts[:1] == ["watch"]:
            video_id = (query.get("v") or [""])[0]

    if not video_id:
        return value

    passthrough: dict[str, str] = {}
    for key in ["start", "t"]:
        if query.get(key):
            passthrough[key] = query[key][0]

    if "t" in passthrough and "start" not in passthrough:
        passthrough["start"] = passthrough.pop("t")

    query_string = f"?{urlencode(passthrough)}" if passthrough else ""
    return f"https://www.youtube.com/embed/{video_id}{query_string}"


def extract_youtube_video_id(url: str) -> str:
    value = url.strip()
    parsed = urlparse(value)
    host = parsed.netloc.lower()
    path = parsed.path.strip("/")
    query = parse_qs(parsed.query)

    if host in {"youtu.be", "www.youtu.be"}:
        return path.split("/", 1)[0]
    if host.endswith("youtube.com") or host.endswith("youtube-nocookie.com"):
        parts = [part for part in path.split("/") if part]
        if len(parts) >= 2 and parts[0] in {"embed", "shorts", "live"}:
            return parts[1]
        if parts[:1] == ["watch"]:
            return (query.get("v") or [""])[0]
    return ""


def detect_browser_path() -> Path | None:
    for candidate in BROWSER_EXECUTABLE_CANDIDATES:
        if candidate.exists():
            return candidate
    return None


def infer_attachment_format_type_id(path_or_link: str) -> str:
    value = path_or_link.strip()
    parsed = urlparse(value)
    if parsed.scheme in {"http", "https"}:
        return "4"
    path_part = parsed.path if parsed.scheme else value
    suffix = Path(unquote(path_part)).suffix.lower()
    if suffix in ATTACHMENT_FORMAT_TYPE_BY_EXTENSION:
        return ATTACHMENT_FORMAT_TYPE_BY_EXTENSION[suffix]
    return ""


class IdRegistry:
    def __init__(self) -> None:
        self.used_ids: set[str] = set()
        self.short_ids_by_article: dict[str, str] = {}
        self.long_ids_by_article: dict[str, str] = {}
        self.random = random.SystemRandom()

    def load_from_folder(self, folder: Path, clear_existing: bool = True) -> tuple[int, list[str]]:
        if clear_existing:
            self.used_ids.clear()
            self.short_ids_by_article.clear()
            self.long_ids_by_article.clear()
        warnings: list[str] = []
        candidates = [
            "Kurzbezeichnung-NEU.xlsx",
            "Text-NEU.xlsx",
            "Text_zu_ID.xlsx",
            "Kurzbezeichnung_zu_ID.xlsx",
        ]

        for file_name in candidates:
            workbook_path = folder / file_name
            if not workbook_path.exists():
                continue
            try:
                workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            except Exception as exc:  # pragma: no cover - defensive user feedback
                warnings.append(f"{file_name}: {exc}")
                continue

            worksheet = workbook[workbook.sheetnames[0]]
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                workbook.close()
                continue
            header_map = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
            if "Text Modul ID" not in header_map:
                workbook.close()
                continue
            id_index = header_map["Text Modul ID"]
            article_index = None
            for candidate_name in ["Artikelnummer", "Artikelnummer ", "Produktnummer"]:
                if candidate_name in header_map:
                    article_index = header_map[candidate_name]
                    break
            for row in rows:
                if row is None or id_index >= len(row):
                    continue
                value = row[id_index]
                if value is None:
                    continue
                candidate = str(value).strip()
                if candidate:
                    self.used_ids.add(candidate)
                    if article_index is not None and article_index < len(row) and row[article_index] is not None:
                        article_number = normalize_article_number(str(row[article_index]))
                        if article_number:
                            if file_name in {"Kurzbezeichnung-NEU.xlsx", "Kurzbezeichnung_zu_ID.xlsx"}:
                                self.short_ids_by_article[article_number] = candidate
                            elif file_name in {"Text-NEU.xlsx", "Text_zu_ID.xlsx"}:
                                self.long_ids_by_article[article_number] = candidate
            workbook.close()

        return len(self.used_ids), warnings

    def reserve(self, value: str) -> None:
        if value:
            self.used_ids.add(value)

    def remember_article_ids(self, article_number: str, short_id: str, long_id: str) -> None:
        article_key = normalize_article_number(article_number)
        if not article_key:
            return
        if short_id:
            self.short_ids_by_article[article_key] = short_id
            self.used_ids.add(short_id)
        if long_id:
            self.long_ids_by_article[article_key] = long_id
            self.used_ids.add(long_id)

    def get_ids_for_article(self, article_number: str) -> tuple[str, str]:
        article_key = normalize_article_number(article_number)
        return (
            self.short_ids_by_article.get(article_key, ""),
            self.long_ids_by_article.get(article_key, ""),
        )

    def generate_unique(self) -> str:
        while True:
            candidate = "".join(self.random.choice(ID_ALPHABET) for _ in range(ID_LENGTH))
            if candidate not in self.used_ids:
                self.used_ids.add(candidate)
                return candidate


class GenArtRegistry:
    def __init__(self) -> None:
        self.options: list[GenArtOption] = []
        self.options_by_display: dict[str, GenArtOption] = {}
        self.options_by_id: dict[str, GenArtOption] = {}
        self.source_path: Path | None = None

    def load_from_workbook(self, path: Path) -> int:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook["Genarten"] if "Genarten" in workbook.sheetnames else workbook.active
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                raise ValueError("Die GenArt-Datei ist leer.")
            header_map = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
            required_headers = {"ID", "Bezeichnung", "Genart"}
            if not required_headers.issubset(header_map):
                raise ValueError("Die GenArt-Datei braucht die Spalten ID, Bezeichnung und Genart.")

            options: list[GenArtOption] = []
            for row in rows:
                if row is None:
                    continue
                option_id = str(row[header_map["ID"]]).strip() if header_map["ID"] < len(row) and row[header_map["ID"]] is not None else ""
                bezeichnung = (
                    str(row[header_map["Bezeichnung"]]).strip()
                    if header_map["Bezeichnung"] < len(row) and row[header_map["Bezeichnung"]] is not None
                    else ""
                )
                genart = (
                    str(row[header_map["Genart"]]).strip()
                    if header_map["Genart"] < len(row) and row[header_map["Genart"]] is not None
                    else ""
                )
                if not option_id or not (bezeichnung or genart):
                    continue

                normalized_bezeichnung = normalize_match_text(bezeichnung)
                normalized_genart = normalize_match_text(genart)
                tokens = build_match_tokens(f"{bezeichnung} {genart}")
                families = infer_match_families_from_tokens(tokens)
                option = GenArtOption(
                    id=option_id,
                    bezeichnung=bezeichnung,
                    genart=genart,
                    normalized_bezeichnung=normalized_bezeichnung,
                    normalized_genart=normalized_genart,
                    tokens=tokens,
                    families=families,
                )
                options.append(option)
        finally:
            workbook.close()

        self.options = options
        self.options_by_display = {option.display_label(): option for option in options}
        self.options_by_id = {option.id: option for option in options}
        self.source_path = path
        return len(options)

    def list_display_values(self) -> list[str]:
        return [option.display_label() for option in self.options]

    def search_options(self, query: str, limit: int = 200) -> list[tuple[GenArtOption, float]]:
        if not query.strip():
            return [(option, 0.0) for option in self.options[:limit]]

        raw_query = query.strip()
        raw_query_casefold = raw_query.casefold()
        normalized_query = normalize_match_text(raw_query)
        query_tokens = set(build_match_tokens(raw_query))
        query_families = infer_match_families_from_tokens(query_tokens)
        matches: list[tuple[GenArtOption, float]] = []

        for option in self.options:
            score = 0.0
            searchable_fields = [
                option.id.casefold(),
                option.normalized_genart,
                option.normalized_bezeichnung,
            ]
            searchable_text = " ".join(part for part in searchable_fields if part)

            if raw_query_casefold == option.id.casefold():
                score += 1200
            elif option.id.casefold().startswith(raw_query_casefold):
                score += 850
            elif raw_query_casefold in option.id.casefold():
                score += 560

            if searchable_text and raw_query_casefold == searchable_text:
                score += 900

            if normalized_query:
                if option.normalized_bezeichnung.startswith(normalized_query):
                    score += 520
                elif normalized_query in option.normalized_bezeichnung:
                    score += 340

                if option.normalized_genart.startswith(normalized_query):
                    score += 500
                elif normalized_query in option.normalized_genart:
                    score += 320

                if normalized_query in searchable_text:
                    score += 220

                overlap = query_tokens & option.tokens
                if overlap:
                    score += len(overlap) * 46
                    score += (len(overlap) / max(1, len(query_tokens))) * 60

                family_overlap = query_families & option.families
                if family_overlap:
                    score += 180 * len(family_overlap)

                if option.normalized_bezeichnung:
                    score += SequenceMatcher(None, normalized_query, option.normalized_bezeichnung).ratio() * 55
                if option.normalized_genart:
                    score += SequenceMatcher(None, normalized_query, option.normalized_genart).ratio() * 50

            matched_terms = 0
            term_prefix_matches = 0
            for token in query_tokens:
                if token in searchable_text:
                    matched_terms += 1
                if any(field.startswith(token) for field in searchable_fields if field):
                    term_prefix_matches += 1

            if matched_terms:
                coverage = matched_terms / max(1, len(query_tokens))
                score += matched_terms * 85
                score += coverage * 260
                if len(query_tokens) > 1 and matched_terms == len(query_tokens):
                    score += 420
            if term_prefix_matches:
                score += term_prefix_matches * 70

            display_label = option.display_label().casefold()
            if raw_query_casefold and raw_query_casefold in display_label:
                score += 80

            if score > 0:
                matches.append((option, score))

        matches.sort(key=lambda item: (-item[1], item[0].display_label()))
        return matches[:limit]

    def search_display_values(self, query: str, limit: int = 200) -> list[str]:
        return [option.display_label() for option, _score in self.search_options(query, limit=limit)]

    def resolve(self, value: str) -> GenArtOption | None:
        query = value.strip()
        if not query:
            return None
        if query in self.options_by_display:
            return self.options_by_display[query]
        if query in self.options_by_id:
            return self.options_by_id[query]
        normalized_query = normalize_match_text(query)
        for option in self.options:
            if normalized_query in {option.normalized_bezeichnung, option.normalized_genart}:
                return option
        return None


class DeepLTranslationError(RuntimeError):
    pass


class DeepLClient:
    def __init__(self, auth_key: str, base_url: str, timeout_seconds: int = 45) -> None:
        self.auth_key = auth_key.strip()
        self.base_url = base_url.rstrip("/")
        self.timeout_seconds = timeout_seconds

    def translate_texts(self, texts: list[str], target_lang: str, source_lang: str = "DE") -> list[str]:
        payload = {
            "text": texts,
            "source_lang": source_lang,
            "target_lang": target_lang,
            "preserve_formatting": True,
        }
        request = urllib_request.Request(
            url=f"{self.base_url}/v2/translate",
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "Authorization": f"DeepL-Auth-Key {self.auth_key}",
                "Content-Type": "application/json",
                "User-Agent": "ApolloImportGui/1.0",
            },
            method="POST",
        )
        try:
            with urllib_request.urlopen(request, timeout=self.timeout_seconds) as response:
                body = response.read().decode("utf-8")
        except urllib_error.HTTPError as exc:
            details = exc.read().decode("utf-8", errors="replace")
            raise DeepLTranslationError(f"HTTP {exc.code}: {details or exc.reason}") from exc
        except urllib_error.URLError as exc:
            raise DeepLTranslationError(f"Verbindungsfehler: {exc.reason}") from exc

        try:
            parsed = json.loads(body)
            return [item["text"] for item in parsed["translations"]]
        except (KeyError, TypeError, json.JSONDecodeError) as exc:
            raise DeepLTranslationError(f"Unerwartete Antwort von DeepL: {body[:400]}") from exc

    def translate_from_german(self, german_text: str) -> dict[str, str]:
        text = german_text.strip()
        if not text:
            raise DeepLTranslationError("Kein deutscher Quelltext zum Übersetzen vorhanden.")

        translations: dict[str, str] = {}
        for ui_code, deepl_code in DEEPL_TARGET_LANGUAGES.items():
            translations[ui_code] = self.translate_texts([text], target_lang=deepl_code, source_lang="DE")[0]
        return translations


class KunzerScrapeError(RuntimeError):
    pass


@dataclass
class KunzerScrapeResult:
    article_number: str
    product_url: str
    title: str
    breadcrumb_text: str
    short_text_de: str
    long_text_de: str
    image_links: list[str]
    document_links: list[str]
    video_links: list[str]


class KunzerScraper:
    def __init__(self) -> None:
        self._sitemap_index: dict[str, str] | None = None
        self._result_cache: dict[str, KunzerScrapeResult] = {}

    def resolve_product_url(self, article_number_or_url: str) -> str:
        value = article_number_or_url.strip()
        if not value:
            raise KunzerScrapeError("Bitte eine Artikelnummer oder Kunzer Produkt-URL eingeben.")
        if value.lower().startswith(("http://", "https://")):
            return value

        sitemap = self._load_sitemap_index()
        normalized = normalize_article_number(value)
        if normalized in sitemap:
            return sitemap[normalized]

        encoded = quote(normalized, safe=".-")
        return f"https://www.kunzer.de/shop/p/{encoded}"

    def scrape_product(self, article_number_or_url: str) -> KunzerScrapeResult:
        target_url = self.resolve_product_url(article_number_or_url)
        cached = self._result_cache.get(target_url)
        if cached is not None:
            return self._clone_result(cached)
        browser_path = self._detect_browser_path()

        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                executable_path=str(browser_path) if browser_path else None,
                headless=True,
            )
            page = browser.new_page(viewport={"width": 1440, "height": 2400})
            try:
                page.goto(target_url, wait_until="domcontentloaded", timeout=90000)
                self._accept_cookies_if_present(page)
                try:
                    page.get_by_role("heading").first.wait_for(timeout=12000)
                except Exception:
                    pass
                try:
                    page.wait_for_load_state("networkidle", timeout=8000)
                except Exception:
                    pass
                self._expand_all_documents(page)
                page.wait_for_timeout(250)

                title = self._clean_single_line(page.title().replace("| Kunzer", ""))
                heading = self._safe_first_text(page.get_by_role("heading").all_inner_texts())
                short_text_de = self._clean_single_line(heading or title)
                breadcrumb_text = self._extract_breadcrumb_text(page)

                description_text = self._extract_tab_content(page, "Artikelbeschreibung")
                info_text = self._extract_tab_content(page, "Artikelinformationen")
                long_text_de = self._select_long_text(short_text_de, description_text, info_text)

                links = self._extract_links(page)
                image_links = [href for href in links if self._is_image_link(href)]
                document_links = [href for href in links if href.lower().endswith(".pdf")]
                video_links = [href for href in links if self._is_video_link(href)]

                article_number = normalize_article_number(
                    self._extract_article_number(page.url, page.locator("body").inner_text(), short_text_de)
                )

                result = KunzerScrapeResult(
                    article_number=article_number,
                    product_url=page.url,
                    title=short_text_de,
                    breadcrumb_text=breadcrumb_text,
                    short_text_de=short_text_de,
                    long_text_de=long_text_de,
                    image_links=image_links,
                    document_links=document_links,
                    video_links=video_links,
                )
                self._result_cache[target_url] = result
                self._result_cache[result.product_url] = result
                return self._clone_result(result)
            except PlaywrightTimeoutError as exc:
                raise KunzerScrapeError(f"Timeout beim Laden der Kunzer-Seite: {target_url}") from exc
            except Exception as exc:  # pragma: no cover - runtime GUI feedback
                raise KunzerScrapeError(f"Kunzer-Seite konnte nicht gelesen werden: {exc}") from exc
            finally:
                browser.close()

    def _clone_result(self, result: KunzerScrapeResult) -> KunzerScrapeResult:
        return KunzerScrapeResult(
            article_number=result.article_number,
            product_url=result.product_url,
            title=result.title,
            breadcrumb_text=result.breadcrumb_text,
            short_text_de=result.short_text_de,
            long_text_de=result.long_text_de,
            image_links=list(result.image_links),
            document_links=list(result.document_links),
            video_links=list(result.video_links),
        )

    def _load_sitemap_index(self) -> dict[str, str]:
        if self._sitemap_index is not None:
            return self._sitemap_index

        request = urllib_request.Request(
            url="https://www.kunzer.de/products-sitemap.xml",
            headers={"User-Agent": "ApolloImportGui/1.0"},
            method="GET",
        )
        try:
            with urllib_request.urlopen(request, timeout=60) as response:
                xml_text = response.read().decode("utf-8")
        except urllib_error.URLError as exc:
            raise KunzerScrapeError(f"Kunzer Sitemap konnte nicht geladen werden: {exc.reason}") from exc

        try:
            root = ET.fromstring(xml_text)
        except ET.ParseError as exc:
            raise KunzerScrapeError("Kunzer Sitemap ist nicht lesbar.") from exc

        namespace = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
        entries: dict[str, str] = {}
        for loc in root.findall(".//sm:loc", namespace):
            if loc.text:
                url = loc.text.strip()
                segment = unquote(url.rstrip("/").rsplit("/", 1)[-1])
                entries[normalize_article_number(segment)] = url

        self._sitemap_index = entries
        return entries

    def _detect_browser_path(self) -> Path | None:
        return detect_browser_path()

    def _accept_cookies_if_present(self, page: object) -> None:
        try:
            page.get_by_role("button", name="Alle zulassen").click(timeout=4000)
            page.wait_for_timeout(800)
        except Exception:
            return

    def _expand_all_documents(self, page: object) -> None:
        try:
            page.get_by_role("button", name="Alle Dokumente anzeigen").click(timeout=4000)
            page.wait_for_timeout(800)
        except Exception:
            return

    def _extract_links(self, page: object) -> list[str]:
        hrefs = []
        for href in page.locator("a").evaluate_all("(nodes) => nodes.map((node) => node.href || '')"):
            cleaned = href.strip()
            if cleaned and cleaned not in hrefs:
                hrefs.append(cleaned)
        return hrefs

    def _extract_breadcrumb_text(self, page: object) -> str:
        script = """
        () => {
            const clean = (value) => (value || "").replace(/\\s+/g, " ").trim();
            const selectors = [
                '[aria-label*="breadcrumb" i]',
                '[class*="breadcrumb" i]',
                '[data-testid*="breadcrumb" i]',
                'nav',
            ];
            const candidates = [];
            const seen = new Set();

            for (const selector of selectors) {
                for (const node of document.querySelectorAll(selector)) {
                    const parts = [];
                    const descendants = node.querySelectorAll('a, span, li, p, div');
                    for (const descendant of descendants) {
                        const text = clean(descendant.textContent);
                        if (!text || text.length > 50) {
                            continue;
                        }
                        if (parts.length && parts[parts.length - 1] === text) {
                            continue;
                        }
                        parts.push(text);
                    }

                    if (parts.length < 2 || parts.length > 6) {
                        continue;
                    }

                    const text = parts.join(' > ');
                    if (!text || text.length > 140 || seen.has(text)) {
                        continue;
                    }
                    seen.add(text);

                    let score = 0;
                    if (/produkte/i.test(text)) score += 120;
                    if (/breadcrumb/i.test(selector)) score += 80;
                    if (/shop|downloads|kontakt|über uns/i.test(text)) score -= 60;
                    score += Math.max(0, 40 - parts.length * 6);
                    score += Math.max(0, 50 - text.length / 3);
                    candidates.push({ text, score });
                }
            }

            candidates.sort((left, right) => right.score - left.score);
            return candidates.length ? candidates[0].text : '';
        }
        """
        try:
            breadcrumb_text = str(page.evaluate(script)).strip()
        except Exception:
            return ""
        return self._clean_single_line(breadcrumb_text)

    def _extract_tab_content(self, page: object, button_label: str) -> str:
        try:
            button = page.get_by_role("button", name=button_label)
            button.click(timeout=4000)
            page.wait_for_timeout(400)
        except Exception:
            return ""

        try:
            controls_id = button.get_attribute("aria-controls")
            if controls_id:
                panel = page.locator(f"#{controls_id}")
                if panel.count() > 0:
                    content = panel.inner_text().strip()
                    if content:
                        return self._normalize_multiline_text(content)
        except Exception:
            pass

        body_text = page.locator("body").inner_text()
        sections = re.split(r"\n(?=Artikelbeschreibung|Artikelinformationen)\n", body_text)
        for section in sections:
            if section.startswith(button_label):
                lines = [line.strip() for line in section.splitlines()[1:]]
                collected: list[str] = []
                for line in lines:
                    if not line or line in {"Artikelbeschreibung", "Artikelinformationen", "Anfrage senden", "Verfügbar"}:
                        break
                    collected.append(line)
                if collected:
                    return self._normalize_multiline_text("\n".join(collected))
        return ""

    def _select_long_text(self, title: str, description_text: str, info_text: str) -> str:
        description = self._normalize_multiline_text(description_text)
        if description and description != title:
            return description

        info = self._normalize_multiline_text(info_text)
        if info and info != title:
            return info

        return title

    def _extract_article_number(self, page_url: str, body_text: str, title: str) -> str:
        url_segment = unquote(urlparse(page_url).path.rstrip("/").rsplit("/", 1)[-1])
        if url_segment and url_segment.lower() not in {"p", "shop"}:
            return url_segment

        candidates = body_text.splitlines()
        for line in candidates:
            cleaned = self._clean_single_line(line)
            if cleaned and cleaned != title and len(cleaned) <= 40 and re.fullmatch(r"[A-Z0-9 .\\-]+", cleaned):
                return cleaned
        return title

    def _is_image_link(self, href: str) -> bool:
        lower = href.lower()
        return lower.endswith((".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff"))

    def _is_video_link(self, href: str) -> bool:
        lower = href.lower()
        if href in KUNZER_GENERIC_YOUTUBE_LINKS:
            return False
        return "youtube.com" in lower or "youtu.be" in lower

    def _safe_first_text(self, values: list[str]) -> str:
        for value in values:
            cleaned = self._clean_single_line(value)
            if cleaned:
                return cleaned
        return ""

    def _clean_single_line(self, value: str) -> str:
        return " ".join(value.replace("\r", " ").replace("\n", " ").split()).strip()

    def _normalize_multiline_text(self, value: str) -> str:
        lines = [line.strip() for line in value.replace("\r", "\n").split("\n")]
        compacted: list[str] = []
        blank_pending = False
        for line in lines:
            if line:
                if blank_pending and compacted:
                    compacted.append("")
                compacted.append(line)
                blank_pending = False
            else:
                blank_pending = True
        return "\n".join(compacted).strip()


def write_workbook(path: Path, sheet_name: str, headers: list[str], rows: list[list[str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def read_workbook_headers(path: Path, sheet_name: str) -> list[str]:
    if not path.exists():
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return ["" if value is None else str(value) for value in first_row]
    finally:
        workbook.close()


def is_legacy_oe_workbook(path: Path, sheet_name: str) -> bool:
    header_keys = [normalize_header_key(value) for value in read_workbook_headers(path, sheet_name)]
    if not header_keys:
        return False
    has_oe_number = "oenummer" in header_keys
    has_manufacturer = any(key in header_keys for key in ("khernr", "herstellerid", "herstellernr"))
    return has_oe_number and not has_manufacturer


def is_legacy_attribute_workbook(path: Path, sheet_name: str) -> bool:
    header_keys = [normalize_header_key(value) for value in read_workbook_headers(path, sheet_name)]
    if not header_keys:
        return False
    has_value = "wert" in header_keys
    has_range = any(key in header_keys for key in ("wertvon", "wertbis", "von", "bis"))
    return has_value and not has_range


def is_attribute_workbook_with_value_from(path: Path, sheet_name: str) -> bool:
    header_keys = [normalize_header_key(value) for value in read_workbook_headers(path, sheet_name)]
    return "wertvon" in header_keys and "wertbis" in header_keys


def is_legacy_vehicle_link_workbook(path: Path, sheet_name: str) -> bool:
    header_keys = [normalize_header_key(value) for value in read_workbook_headers(path, sheet_name)]
    if not header_keys:
        return False
    has_ktyp = "ktypnr" in header_keys
    has_genart = any(key in header_keys for key in ("genartid", "genart"))
    return has_ktyp and not has_genart


def read_workbook_rows(path: Path, sheet_name: str, expected_width: int) -> list[list[str]]:
    if not path.exists():
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        existing_rows: list[list[str]] = []
        row_iter = worksheet.iter_rows(values_only=True)
        next(row_iter, None)
        for values in row_iter:
            row = ["" if value is None else str(value) for value in values[:expected_width]]
            if len(row) < expected_width:
                row.extend([""] * (expected_width - len(row)))
            if any(cell.strip() for cell in row):
                existing_rows.append(row)
        return existing_rows
    finally:
        workbook.close()


def prepare_existing_rows_for_write(path: Path, sheet_name: str, headers: list[str]) -> list[list[str]]:
    if headers == OE_HEADERS and is_legacy_oe_workbook(path, sheet_name):
        legacy_rows = read_workbook_rows(path, sheet_name, len(LEGACY_OE_HEADERS))
        existing_rows = []
        for row in legacy_rows:
            padded = list(row) + [""] * max(0, len(LEGACY_OE_HEADERS) - len(row))
            existing_rows.append([padded[0], padded[1], "", padded[2], padded[3]])
    elif headers == ATTRIBUTE_HEADERS and is_attribute_workbook_with_value_from(path, sheet_name):
        legacy_rows = read_workbook_rows(path, sheet_name, len(ATTRIBUTE_HEADERS_WITH_VALUE_FROM))
        existing_rows = []
        for row in legacy_rows:
            padded = list(row) + [""] * max(0, len(ATTRIBUTE_HEADERS_WITH_VALUE_FROM) - len(row))
            value = padded[4].strip() or padded[5].strip()
            existing_rows.append([padded[0], padded[1], padded[2], padded[3], value, padded[6], padded[7]])
    elif headers == ATTRIBUTE_HEADERS and is_legacy_attribute_workbook(path, sheet_name):
        legacy_rows = read_workbook_rows(path, sheet_name, len(LEGACY_ATTRIBUTE_HEADERS))
        existing_rows = []
        for row in legacy_rows:
            padded = list(row) + [""] * max(0, len(LEGACY_ATTRIBUTE_HEADERS) - len(row))
            existing_rows.append([padded[0], padded[1], padded[2], padded[3], padded[4], "", padded[5]])
    elif headers == VEHICLE_LINK_HEADERS and is_legacy_vehicle_link_workbook(path, sheet_name):
        legacy_rows = read_workbook_rows(path, sheet_name, len(LEGACY_VEHICLE_LINK_HEADERS))
        existing_rows = []
        for row in legacy_rows:
            padded = list(row) + [""] * max(0, len(LEGACY_VEHICLE_LINK_HEADERS) - len(row))
            existing_rows.append([padded[0], padded[1], padded[2], padded[3], "", "", padded[4]])
    else:
        existing_rows = read_workbook_rows(path, sheet_name, len(headers))
    if headers and headers[-1] == LAST_WRITTEN_HEADER and path.exists():
        fallback_written_at = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        for row in existing_rows:
            if row and not row[-1].strip() and any(cell.strip() for cell in row[:-1]):
                row[-1] = fallback_written_at
    return existing_rows


def merge_rows_by_article(
    existing_rows: list[list[str]],
    new_rows: list[list[str]],
    key_index: int = 0,
    replace_article_keys: set[str] | None = None,
) -> list[list[str]]:
    article_keys = set(replace_article_keys or set())
    article_keys.update(
        normalize_article_number(str(row[key_index]))
        for row in new_rows
        if len(row) > key_index and normalize_article_number(str(row[key_index]))
    )
    if not article_keys:
        return existing_rows

    merged_rows = [
        row
        for row in existing_rows
        if len(row) <= key_index or normalize_article_number(str(row[key_index])) not in article_keys
    ]
    merged_rows.extend(new_rows)
    return merged_rows


def write_workbook_with_upsert(
    path: Path,
    sheet_name: str,
    headers: list[str],
    rows: list[list[str]],
    key_index: int = 0,
    replace_article_keys: set[str] | None = None,
) -> None:
    existing_rows = prepare_existing_rows_for_write(path, sheet_name, headers)
    merged_rows = merge_rows_by_article(existing_rows, rows, key_index=key_index, replace_article_keys=replace_article_keys)
    write_workbook(path, sheet_name, headers, merged_rows)


def remove_article_rows_from_workbook(
    path: Path,
    sheet_name: str,
    headers: list[str],
    article_numbers: set[str],
    key_index: int = 0,
) -> None:
    if not path.exists():
        return

    normalized_article_numbers = {normalize_article_number(article_number) for article_number in article_numbers if normalize_article_number(article_number)}
    if not normalized_article_numbers:
        return

    existing_rows = prepare_existing_rows_for_write(path, sheet_name, headers)
    remaining_rows = [
        row
        for row in existing_rows
        if len(row) <= key_index or normalize_article_number(str(row[key_index])) not in normalized_article_numbers
    ]
    write_workbook(path, sheet_name, headers, remaining_rows)


def replace_article_rows_preserving_timestamps(
    path: Path,
    sheet_name: str,
    headers: list[str],
    article_number: str,
    rows: list[list[str]],
    key_index: int = 0,
) -> None:
    existing_rows = prepare_existing_rows_for_write(path, sheet_name, headers)
    article_key = normalize_article_number(article_number)
    has_last_written = bool(headers) and headers[-1] == LAST_WRITTEN_HEADER
    written_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    rebuilt_rows = rows
    if has_last_written:
        existing_timestamps_by_signature: dict[tuple[str, ...], list[str]] = {}
        signature_width = len(headers) - 1
        for row in existing_rows:
            if len(row) <= key_index or normalize_article_number(str(row[key_index])) != article_key:
                continue
            signature = tuple(str(cell) for cell in row[:signature_width])
            existing_timestamps_by_signature.setdefault(signature, []).append(row[-1].strip() or written_at)

        rebuilt_rows = []
        for row in rows:
            signature_values = [str(cell) for cell in row[:signature_width]]
            if len(signature_values) < signature_width:
                signature_values.extend([""] * (signature_width - len(signature_values)))
            signature = tuple(signature_values)
            known_timestamps = existing_timestamps_by_signature.get(signature, [])
            timestamp = known_timestamps.pop(0) if known_timestamps else written_at
            rebuilt_rows.append(signature_values + [timestamp])

    merged_rows: list[list[str]] = []
    inserted = False
    for existing_row in existing_rows:
        if len(existing_row) > key_index and normalize_article_number(str(existing_row[key_index])) == article_key:
            if not inserted:
                merged_rows.extend(rebuilt_rows)
                inserted = True
            continue
        merged_rows.append(existing_row)

    if not inserted:
        merged_rows.extend(rebuilt_rows)

    write_workbook(path, sheet_name, headers, merged_rows)


def build_short_text_export_row(bundle: ExportBundle) -> list[str]:
    short_texts = sanitize_short_translation_set(bundle.short_texts)
    return [
        bundle.article_number,
        bundle.short_module_id,
        "1",
        "1",
        *short_texts.export_values(bundle.short_auto_uni),
    ]


def build_long_text_export_row(bundle: ExportBundle) -> list[str]:
    return [
        bundle.article_number,
        bundle.long_module_id,
        "2",
        "1",
        *bundle.long_texts.export_values(bundle.long_auto_uni),
    ]


def build_short_mapping_export_row(bundle: ExportBundle) -> list[str]:
    return [bundle.article_number, bundle.short_module_id]


def build_genart_export_rows(bundle: ExportBundle) -> list[list[str]]:
    export_rows: list[list[str]] = []
    for selection in bundle.genart_selections:
        genart_id = selection.id.strip()
        genart_bezeichnung = selection.bezeichnung.strip()
        if not genart_id and not genart_bezeichnung:
            continue
        export_rows.append([bundle.article_number, genart_id, genart_bezeichnung])
    return export_rows


def build_image_export_rows(bundle: ExportBundle) -> list[list[str]]:
    export_rows: list[list[str]] = []
    for row in bundle.image_rows:
        format_type_id = infer_attachment_format_type_id(row.path_or_link)
        if not format_type_id:
            raise ValueError(f"Kein TecDoc Anhangsformattyp für Bild ableitbar: {row.path_or_link}")
        export_rows.append([bundle.article_number, row.path_or_link, row.art or "5", row.sprache or "255", format_type_id])
    return export_rows


def build_document_export_rows(bundle: ExportBundle) -> list[list[str]]:
    export_rows: list[list[str]] = []
    for row in bundle.document_rows:
        format_type_id = infer_attachment_format_type_id(row.path_or_link)
        if not format_type_id:
            raise ValueError(f"Kein TecDoc Anhangsformattyp für Dokument ableitbar: {row.path_or_link}")
        export_rows.append([bundle.article_number, row.path_or_link, row.sprache or "255", row.art or "17", format_type_id])
    return export_rows


def build_video_export_rows(bundle: ExportBundle) -> list[list[str]]:
    export_rows: list[list[str]] = []
    for row in bundle.video_rows:
        normalized_video_link = normalize_youtube_url_for_embed(row.path_or_link)
        format_type_id = infer_attachment_format_type_id(normalized_video_link)
        if not format_type_id:
            raise ValueError(f"Kein TecDoc Anhangsformattyp für Video-Link ableitbar: {row.path_or_link}")
        export_rows.append([bundle.article_number, normalized_video_link, format_type_id])
    return export_rows


def build_web_export_rows(bundle: ExportBundle) -> list[list[str]]:
    export_rows: list[list[str]] = []
    for row in bundle.web_rows:
        format_type_id = infer_attachment_format_type_id(row.path_or_link)
        if not format_type_id:
            raise ValueError(f"Kein TecDoc Anhangsformattyp für Web-Link ableitbar: {row.path_or_link}")
        export_rows.append([bundle.article_number, row.path_or_link, format_type_id])
    return export_rows


def build_oe_export_rows(bundle: ExportBundle) -> list[list[str]]:
    export_rows: list[list[str]] = []
    for row in normalize_oe_number_rows(bundle.oe_number_rows):
        if not row.manufacturer_id:
            raise ValueError(f"Bitte Hersteller für OE-Nummer '{row.value}' auswählen.")
        export_rows.append([bundle.article_number, "4", row.manufacturer_id, row.value])
    return export_rows


def build_comparison_export_rows(bundle: ExportBundle) -> list[list[str]]:
    return [
        [bundle.article_number, "10", row.competitor_id, row.reference_number]
        for row in normalize_comparison_number_rows(bundle.comparison_number_rows)
    ]


def build_vehicle_link_export_rows(bundle: ExportBundle) -> list[list[str]]:
    """Je Fahrzeug eine Zeile pro KTyp-Nummer (Topmotive und TecDoc) und GenArt."""
    genart_values = [
        (selection.id.strip(), selection.bezeichnung.strip())
        for selection in bundle.genart_selections
        if selection.id.strip() or selection.bezeichnung.strip()
    ] or [("", "")]
    export_rows: list[list[str]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for row in normalize_vehicle_link_rows(bundle.vehicle_link_rows):
        for ktyp, system in ((row.ktyp_topmotive, "Topmotive"), (row.ktyp_tecdoc, "TecDoc")):
            ktyp = ktyp.strip()
            if not ktyp:
                continue
            for genart_id, genart_bezeichnung in genart_values:
                dedupe_key = (row.vehicle_type_id, ktyp, genart_id, genart_bezeichnung)
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)
                export_rows.append([bundle.article_number, row.vehicle_type_id, ktyp, system, genart_id, genart_bezeichnung])
    return export_rows


def build_attribute_export_rows(bundle: ExportBundle) -> list[list[str]]:
    export_rows: list[list[str]] = []
    seen_search_values: set[str] = set()
    for row in normalize_attribute_rows(bundle.attribute_rows):
        export_value = row.value
        if is_attribute_key_value_format(row.value_format):
            export_value = resolve_attribute_export_value(row, bundle.attribute_key_values_by_group)
        if row.criteria_id == SEARCH_TERM_CRITERIA_ID:
            seen_search_values.add(export_value.strip().casefold())
        export_rows.append(
            [
                bundle.article_number,
                row.criteria_id,
                row.label,
                row.value_format,
                export_value,
                row.value_to,
            ]
        )
    # Suchwörter als zusätzliche 'Zusatzbezeichnung'-Zeilen (eine pro Suchwort).
    for term in normalize_search_terms(bundle.search_terms):
        if term.casefold() in seen_search_values:
            continue
        seen_search_values.add(term.casefold())
        export_rows.append(
            [
                bundle.article_number,
                SEARCH_TERM_CRITERIA_ID,
                SEARCH_TERM_ATTRIBUTE_LABEL,
                SEARCH_TERM_ATTRIBUTE_FORMAT,
                term,
                "",
            ]
        )
    return export_rows


def append_written_at(row: list[str], written_at: str) -> list[str]:
    return [*row, written_at]


def export_bundle(bundle: ExportBundle, output_root: Path, use_timestamp_subdir: bool) -> Path:
    written_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if use_timestamp_subdir:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_dir = output_root / f"{safe_folder_name(bundle.article_number)}_{timestamp}"
    else:
        export_dir = output_root
    export_dir.mkdir(parents=True, exist_ok=True)

    short_row = append_written_at(build_short_text_export_row(bundle), written_at)
    long_row = append_written_at(build_long_text_export_row(bundle), written_at)
    short_mapping_row = append_written_at(build_short_mapping_export_row(bundle), written_at)
    genart_rows = [append_written_at(row, written_at) for row in build_genart_export_rows(bundle)]
    image_rows = [append_written_at(row, written_at) for row in build_image_export_rows(bundle)] if bundle.include_images else []
    document_rows = [append_written_at(row, written_at) for row in build_document_export_rows(bundle)] if bundle.include_documents else []
    video_rows = [append_written_at(row, written_at) for row in build_video_export_rows(bundle)] if bundle.include_videos else []
    web_rows = [append_written_at(row, written_at) for row in build_web_export_rows(bundle)] if bundle.include_web_links else []
    oe_rows = [append_written_at(row, written_at) for row in build_oe_export_rows(bundle)]
    comparison_rows = [append_written_at(row, written_at) for row in build_comparison_export_rows(bundle)]
    vehicle_link_rows = [append_written_at(row, written_at) for row in build_vehicle_link_export_rows(bundle)]
    attribute_rows = [append_written_at(row, written_at) for row in build_attribute_export_rows(bundle)]

    if use_timestamp_subdir:
        if bundle.include_short_text:
            write_workbook(export_dir / SHORT_TEXT_FILE[0], SHORT_TEXT_FILE[1], SHORT_TEXT_HEADERS, [short_row])
            write_workbook(export_dir / SHORT_MAPPING_FILE[0], SHORT_MAPPING_FILE[1], SHORT_MAPPING_HEADERS, [short_mapping_row])
        if bundle.include_long_text:
            write_workbook(export_dir / LONG_TEXT_FILE[0], LONG_TEXT_FILE[1], SHORT_TEXT_HEADERS, [long_row])
        if genart_rows:
            write_workbook(export_dir / GENART_FILE[0], GENART_FILE[1], GENART_HEADERS, genart_rows)
        if oe_rows:
            write_workbook(export_dir / OE_FILE[0], OE_FILE[1], OE_HEADERS, oe_rows)
        if comparison_rows:
            write_workbook(export_dir / COMPARISON_FILE[0], COMPARISON_FILE[1], COMPARISON_HEADERS, comparison_rows)
        if vehicle_link_rows:
            write_workbook(export_dir / VEHICLE_LINK_FILE[0], VEHICLE_LINK_FILE[1], VEHICLE_LINK_HEADERS, vehicle_link_rows)
        if attribute_rows:
            write_workbook(export_dir / ATTRIBUTE_FILE[0], ATTRIBUTE_FILE[1], ATTRIBUTE_HEADERS, attribute_rows)
        if bundle.include_images:
            write_workbook(export_dir / IMAGE_FILE[0], IMAGE_FILE[1], IMAGE_HEADERS, image_rows)
        if bundle.include_documents:
            write_workbook(export_dir / DOCUMENT_FILE[0], DOCUMENT_FILE[1], DOCUMENT_HEADERS, document_rows)
        if bundle.include_videos:
            write_workbook(export_dir / VIDEO_FILE[0], VIDEO_FILE[1], VIDEO_HEADERS, video_rows)
        if bundle.include_web_links:
            write_workbook(export_dir / WEB_LINK_FILE[0], WEB_LINK_FILE[1], WEB_HEADERS, web_rows)
    else:
        article_keys = {bundle.article_number}
        if bundle.include_short_text:
            write_workbook_with_upsert(
                export_dir / SHORT_TEXT_FILE[0],
                SHORT_TEXT_FILE[1],
                SHORT_TEXT_HEADERS,
                [short_row],
                replace_article_keys=article_keys,
            )
            write_workbook_with_upsert(
                export_dir / SHORT_MAPPING_FILE[0],
                SHORT_MAPPING_FILE[1],
                SHORT_MAPPING_HEADERS,
                [short_mapping_row],
                replace_article_keys=article_keys,
            )
        if bundle.include_long_text:
            write_workbook_with_upsert(
                export_dir / LONG_TEXT_FILE[0],
                LONG_TEXT_FILE[1],
                SHORT_TEXT_HEADERS,
                [long_row],
                replace_article_keys=article_keys,
            )
        write_workbook_with_upsert(
            export_dir / GENART_FILE[0],
            GENART_FILE[1],
            GENART_HEADERS,
            genart_rows,
            replace_article_keys=article_keys,
        )
        write_workbook_with_upsert(
            export_dir / OE_FILE[0],
            OE_FILE[1],
            OE_HEADERS,
            oe_rows,
            replace_article_keys=article_keys,
        )
        write_workbook_with_upsert(
            export_dir / COMPARISON_FILE[0],
            COMPARISON_FILE[1],
            COMPARISON_HEADERS,
            comparison_rows,
            replace_article_keys=article_keys,
        )
        write_workbook_with_upsert(
            export_dir / VEHICLE_LINK_FILE[0],
            VEHICLE_LINK_FILE[1],
            VEHICLE_LINK_HEADERS,
            vehicle_link_rows,
            replace_article_keys=article_keys,
        )
        write_workbook_with_upsert(
            export_dir / ATTRIBUTE_FILE[0],
            ATTRIBUTE_FILE[1],
            ATTRIBUTE_HEADERS,
            attribute_rows,
            replace_article_keys=article_keys,
        )
        if bundle.include_images:
            write_workbook_with_upsert(
                export_dir / IMAGE_FILE[0],
                IMAGE_FILE[1],
                IMAGE_HEADERS,
                image_rows,
                replace_article_keys=article_keys,
            )
        if bundle.include_documents:
            write_workbook_with_upsert(
                export_dir / DOCUMENT_FILE[0],
                DOCUMENT_FILE[1],
                DOCUMENT_HEADERS,
                document_rows,
                replace_article_keys=article_keys,
            )
        if bundle.include_videos:
            write_workbook_with_upsert(
                export_dir / VIDEO_FILE[0],
                VIDEO_FILE[1],
                VIDEO_HEADERS,
                video_rows,
                replace_article_keys=article_keys,
            )
        if bundle.include_web_links:
            write_workbook_with_upsert(
                export_dir / WEB_LINK_FILE[0],
                WEB_LINK_FILE[1],
                WEB_HEADERS,
                web_rows,
                replace_article_keys=article_keys,
            )

    return export_dir


def build_preview(bundle: ExportBundle) -> str:
    lines = [
        f"Artikelnummer: {bundle.article_number or '(leer)'}",
        "",
        "IDs",
        f"- Kurzbezeichnung: {bundle.short_module_id or '(noch nicht generiert)'}",
        f"- Text: {bundle.long_module_id or '(noch nicht generiert)'}",
        f"- GenArten: {summarize_genart_selections(bundle.genart_selections, empty_label='(nicht gesetzt)', limit=3)}",
        "",
        "Texte",
        f"- Kurzbezeichnung: {bundle.short_texts.populated_count(bundle.short_auto_uni)} / 7 Sprachfelder befüllt",
        f"- Text: {bundle.long_texts.populated_count(bundle.long_auto_uni)} / 7 Sprachfelder befüllt",
        "",
        "Referenzen",
        f"- OE-Nummern: {len(normalize_oe_number_rows(bundle.oe_number_rows))}",
        f"- Vergleichsnummern: {len(normalize_comparison_number_rows(bundle.comparison_number_rows))}",
        f"- Fahrzeugverknüpfungen: {len(normalize_vehicle_link_rows(bundle.vehicle_link_rows))} Fahrzeuge / {len(build_vehicle_link_export_rows(bundle))} Zeilen",
        f"- Attribute: {len(normalize_attribute_rows(bundle.attribute_rows))}",
        f"- Suchwörter: {len(normalize_search_terms(bundle.search_terms))} (als Attribut {SEARCH_TERM_CRITERIA_ID} {SEARCH_TERM_ATTRIBUTE_LABEL})",
        "",
        "Medien",
        f"- Bilder: {len(bundle.image_rows)}",
        f"- Dokumente: {len(bundle.document_rows)}",
        f"- Videos: {len(bundle.video_rows)}",
        f"- Web Links: {len(bundle.web_rows)}",
        "",
        "Exportdateien",
        f"- {SHORT_TEXT_FILE[0]}",
        f"- {SHORT_MAPPING_FILE[0]}",
        f"- {LONG_TEXT_FILE[0]}",
        f"- {GENART_FILE[0]}",
        f"- {OE_FILE[0]}",
        f"- {COMPARISON_FILE[0]}",
        f"- {VEHICLE_LINK_FILE[0]}",
        f"- {ATTRIBUTE_FILE[0]}",
        f"- {IMAGE_FILE[0]}",
        f"- {DOCUMENT_FILE[0]}",
        f"- {VIDEO_FILE[0]}",
        f"- {WEB_LINK_FILE[0]}",
    ]
    return "\n".join(lines)


def format_translation_set(label: str, translations: TranslationSet) -> str:
    values = translations.normalized()
    lines = [label]
    for code, language_label in UI_LANGUAGE_ORDER:
        text = getattr(values, code) or "-"
        lines.append(f"{language_label}:")
        lines.append(text)
        lines.append("")
    return "\n".join(lines).rstrip()


def format_media_rows(label: str, rows: list[MediaRow], include_meta: bool = True) -> str:
    lines = [label]
    if not rows:
        lines.append("-")
        return "\n".join(lines)

    for index, row in enumerate(rows, start=1):
        lines.append(f"{index}. {row.path_or_link}")
        if include_meta:
            lines.append(f"   Art: {row.art or '-'} | Sprache: {row.sprache or '-'}")
    return "\n".join(lines)


def format_vehicle_link_rows(label: str, rows: list[VehicleLinkRow]) -> str:
    lines = [label]
    normalized = normalize_vehicle_link_rows(rows)
    if not normalized:
        lines.append("-")
        return "\n".join(lines)
    for index, row in enumerate(normalized, start=1):
        vehicle = row.display_vehicle_label()
        meta_parts = [part for part in [format_year_range(row.year_from, row.year_to), row.power] if part]
        meta = f" ({', '.join(meta_parts)})" if meta_parts else ""
        ktyp_parts = []
        if row.ktyp_topmotive:
            ktyp_parts.append(f"TM {row.ktyp_topmotive}")
        if row.ktyp_tecdoc:
            ktyp_parts.append(f"TecDoc {row.ktyp_tecdoc}")
        lines.append(
            f"{index}. [{row.vehicle_type_label or row.vehicle_type_id}] {vehicle}{meta}"
            f" | Motorcode: {row.motorcode or '-'} | KTyp: {', '.join(ktyp_parts) or '-'}"
        )
    return "\n".join(lines)


def format_search_terms(label: str, terms: list[str]) -> str:
    normalized = normalize_search_terms(terms)
    lines = [label]
    if not normalized:
        lines.append("-")
        return "\n".join(lines)
    for index, term in enumerate(normalized, start=1):
        lines.append(f"{index}. {term}")
    return "\n".join(lines)


def format_article_snapshot(snapshot: StoredArticleSnapshot) -> str:
    sections = [
        f"Artikelnummer: {snapshot.article_number}",
        f"Quelle: {snapshot.source_label}",
        f"Ordner: {snapshot.source_folder}",
        "",
        f"Kurz-ID: {snapshot.short_module_id or '-'}",
        f"Text-ID: {snapshot.long_module_id or '-'}",
        format_genart_selections("GenArten", snapshot.genart_selections),
        "",
        format_translation_set("Kurzbezeichnung", snapshot.short_texts),
        "",
        format_translation_set("Text", snapshot.long_texts),
        "",
        format_oe_number_rows("OE-Nummern", snapshot.oe_number_rows),
        "",
        format_comparison_number_rows("Vergleichsnummern", snapshot.comparison_number_rows),
        "",
        format_vehicle_link_rows("Fahrzeugverknüpfungen", snapshot.vehicle_link_rows),
        "",
        format_attribute_rows("Attribute", snapshot.attribute_rows),
        "",
        format_search_terms("Suchwörter", snapshot.search_terms),
        "",
        format_media_rows("Bilder", snapshot.image_rows),
        "",
        format_media_rows("Dokumente", snapshot.document_rows),
        "",
        format_media_rows("Videos", snapshot.video_rows, include_meta=False),
        "",
        format_media_rows("Web Links", snapshot.web_rows, include_meta=False),
    ]
    return "\n".join(sections).strip()


def translation_set_from_export_row(row: list[str]) -> TranslationSet:
    padded = list(row) + [""] * max(0, len(SHORT_TEXT_HEADERS) - len(row))
    return TranslationSet(
        uni=padded[5].strip(),
        de=padded[7].strip(),
        en=padded[9].strip(),
        cz=padded[11].strip(),
        fr=padded[13].strip(),
        it=padded[15].strip(),
        nl=padded[17].strip(),
    )


def load_article_snapshots_from_folder(
    folder: Path,
    source_label: str,
    competitor_lookup: dict[str, CompetitorOption] | None = None,
    manufacturer_lookup: dict[str, CompetitorOption] | None = None,
    attribute_lookup: dict[str, AttributeOption] | None = None,
    attribute_key_values_by_group: dict[str, list[AttributeKeyValueOption]] | None = None,
    vehicle_catalog: VehicleCatalog | None = None,
) -> dict[str, StoredArticleSnapshot]:
    if not folder.exists():
        return {}

    snapshots: dict[str, StoredArticleSnapshot] = {}

    def ensure_snapshot(article_number: str) -> StoredArticleSnapshot:
        article_key = normalize_article_number(article_number)
        snapshot = snapshots.get(article_key)
        if snapshot is None:
            snapshot = StoredArticleSnapshot(article_number=article_key, source_label=source_label, source_folder=folder)
            snapshots[article_key] = snapshot
        return snapshot

    short_rows = read_workbook_rows(folder / SHORT_TEXT_FILE[0], SHORT_TEXT_FILE[1], len(SHORT_TEXT_HEADERS))
    for row in short_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        snapshot = ensure_snapshot(article_number)
        snapshot.short_module_id = row[1].strip()
        snapshot.short_texts = translation_set_from_export_row(row)
        snapshot.short_auto_uni = not snapshot.short_texts.uni or snapshot.short_texts.uni == snapshot.short_texts.de

    long_rows = read_workbook_rows(folder / LONG_TEXT_FILE[0], LONG_TEXT_FILE[1], len(SHORT_TEXT_HEADERS))
    for row in long_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        snapshot = ensure_snapshot(article_number)
        snapshot.long_module_id = row[1].strip()
        snapshot.long_texts = translation_set_from_export_row(row)
        snapshot.long_auto_uni = not snapshot.long_texts.uni or snapshot.long_texts.uni == snapshot.long_texts.de

    genart_rows = read_workbook_rows(folder / GENART_FILE[0], GENART_FILE[1], len(GENART_HEADERS))
    for row in genart_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        snapshot = ensure_snapshot(article_number)
        snapshot.genart_selections.append(GenArtSelection(id=row[1].strip(), bezeichnung=row[2].strip()))

    oe_rows = prepare_existing_rows_for_write(folder / OE_FILE[0], OE_FILE[1], OE_HEADERS)
    for row in oe_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        manufacturer_id = row[2].strip()
        manufacturer = manufacturer_lookup.get(manufacturer_id) if manufacturer_lookup is not None else None
        ensure_snapshot(article_number).oe_number_rows.append(
            OeNumberRow(
                value=row[3].strip(),
                manufacturer_id=manufacturer_id,
                manufacturer_code=manufacturer.code if manufacturer is not None else "",
                manufacturer_name=manufacturer.name if manufacturer is not None else "",
            )
        )

    comparison_rows = read_workbook_rows(folder / COMPARISON_FILE[0], COMPARISON_FILE[1], len(COMPARISON_HEADERS))
    for row in comparison_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        competitor_id = row[2].strip()
        competitor = competitor_lookup.get(competitor_id) if competitor_lookup is not None else None
        ensure_snapshot(article_number).comparison_number_rows.append(
            ComparisonNumberRow(
                competitor_id=competitor_id,
                competitor_code=competitor.code if competitor is not None else "",
                competitor_name=competitor.name if competitor is not None else "",
                reference_number=row[3].strip(),
            )
        )

    vehicle_link_rows = read_workbook_rows(folder / VEHICLE_LINK_FILE[0], VEHICLE_LINK_FILE[1], len(VEHICLE_LINK_HEADERS))
    for row in vehicle_link_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        vehicle_type_id = row[1].strip() or DEFAULT_VEHICLE_TYPE_ID
        ktyp = row[2].strip()
        if not ktyp:
            continue
        system = row[3].strip().casefold()
        match = None
        if vehicle_catalog is not None:
            if system.startswith("topmotive") or system.startswith("tm"):
                match = vehicle_catalog.lookup_by_ktyp(ktyp_topmotive=ktyp)
            elif system.startswith("tecdoc"):
                match = vehicle_catalog.lookup_by_ktyp(ktyp_tecdoc=ktyp)
            else:
                match = vehicle_catalog.lookup_by_ktyp(ktyp_topmotive=ktyp, ktyp_tecdoc=ktyp)
        if match is not None:
            link_row = VehicleLinkRow(
                vehicle_type_id=vehicle_type_id,
                vehicle_type_label=vehicle_type_label_for_id(vehicle_type_id),
                motorcode=match.motorcode,
                ktyp_topmotive=match.ktyp_topmotive,
                ktyp_tecdoc=match.ktyp_tecdoc,
                manufacturer=match.manufacturer,
                model=match.model,
                description=match.description,
                year_from=match.year_from,
                year_to=match.year_to,
                power=match.power,
            )
        else:
            link_row = VehicleLinkRow(
                vehicle_type_id=vehicle_type_id,
                vehicle_type_label=vehicle_type_label_for_id(vehicle_type_id),
                ktyp_topmotive=ktyp if system.startswith("topmotive") or system.startswith("tm") else "",
                ktyp_tecdoc=ktyp if system.startswith("tecdoc") else "",
            )
            if not link_row.has_ktyp():
                link_row = VehicleLinkRow(
                    vehicle_type_id=vehicle_type_id,
                    vehicle_type_label=vehicle_type_label_for_id(vehicle_type_id),
                    ktyp_tecdoc=ktyp,
                )
        ensure_snapshot(article_number).vehicle_link_rows.append(link_row)

    attribute_rows = prepare_existing_rows_for_write(folder / ATTRIBUTE_FILE[0], ATTRIBUTE_FILE[1], ATTRIBUTE_HEADERS)
    for row in attribute_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        criteria_id = row[1].strip()
        if criteria_id == SEARCH_TERM_CRITERIA_ID:
            # 'Zusatzbezeichnung'-Zeilen sind Suchwörter und gehören in den Suchwörter-Tab.
            search_term = row[4].strip()
            if search_term:
                ensure_snapshot(article_number).search_terms.append(search_term)
            continue
        attribute_option = attribute_lookup.get(criteria_id) if attribute_lookup is not None else None
        label = row[2].strip() or (attribute_option.label if attribute_option is not None else "")
        value_format = row[3].strip() or (attribute_option.value_format if attribute_option is not None else "")
        max_length = attribute_option.max_length if attribute_option is not None else None
        type_name = attribute_option.type_name if attribute_option is not None else ""
        attribute_row = AttributeRow(
            criteria_id=criteria_id,
            label=label,
            value_format=value_format,
            max_length=max_length,
            type_name=type_name,
            value=row[4].strip(),
            value_to=row[5].strip(),
        )
        if is_attribute_key_value_format(value_format) and attribute_key_values_by_group:
            attribute_row = AttributeRow(
                criteria_id=attribute_row.criteria_id,
                label=attribute_row.label,
                value_format=attribute_row.value_format,
                max_length=attribute_row.max_length,
                type_name=attribute_row.type_name,
                value=resolve_attribute_key_value_display_value(attribute_row, attribute_key_values_by_group),
                value_to=attribute_row.value_to,
            )
        ensure_snapshot(article_number).attribute_rows.append(
            attribute_row
        )

    image_rows = read_workbook_rows(folder / IMAGE_FILE[0], IMAGE_FILE[1], len(IMAGE_HEADERS))
    for row in image_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        ensure_snapshot(article_number).image_rows.append(MediaRow(path_or_link=row[1].strip(), art=row[2].strip(), sprache=row[3].strip()))

    document_rows = read_workbook_rows(folder / DOCUMENT_FILE[0], DOCUMENT_FILE[1], len(DOCUMENT_HEADERS))
    for row in document_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        ensure_snapshot(article_number).document_rows.append(MediaRow(path_or_link=row[1].strip(), art=row[3].strip(), sprache=row[2].strip()))

    video_rows = read_workbook_rows(folder / VIDEO_FILE[0], VIDEO_FILE[1], len(VIDEO_HEADERS))
    for row in video_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        ensure_snapshot(article_number).video_rows.append(MediaRow(path_or_link=row[1].strip()))

    web_rows = read_workbook_rows(folder / WEB_LINK_FILE[0], WEB_LINK_FILE[1], len(WEB_HEADERS))
    for row in web_rows:
        article_number = normalize_article_number(row[0])
        if not article_number:
            continue
        ensure_snapshot(article_number).web_rows.append(MediaRow(path_or_link=row[1].strip()))

    for snapshot in snapshots.values():
        snapshot.sync_genart_fields()
        snapshot.oe_number_rows = normalize_oe_number_rows(snapshot.oe_number_rows)
        snapshot.comparison_number_rows = normalize_comparison_number_rows(snapshot.comparison_number_rows)
        snapshot.vehicle_link_rows = normalize_vehicle_link_rows(snapshot.vehicle_link_rows)
        snapshot.attribute_rows = normalize_attribute_rows(snapshot.attribute_rows)
        snapshot.search_terms = normalize_search_terms(snapshot.search_terms)

    return snapshots


class SingleLineTranslationFrame(ttk.LabelFrame):
    def __init__(
        self,
        master: tk.Misc,
        title: str,
        on_change: Callable[[], None] | None = None,
        max_length: int | None = None,
    ) -> None:
        super().__init__(master, text=title, padding=14)
        self.on_change = on_change
        self.max_length = max_length
        self._suspend_limit_enforcement = False
        self.auto_uni_var = tk.BooleanVar(value=True)
        self.fields: dict[str, tk.StringVar] = {code: tk.StringVar() for code, _ in UI_LANGUAGE_ORDER}
        self.uni_entry: ttk.Entry | None = None
        self.columnconfigure(1, weight=1)

        auto_uni_check = ttk.Checkbutton(
            self,
            text="UNI automatisch aus Deutsch übernehmen",
            variable=self.auto_uni_var,
            command=self._handle_auto_uni_toggle,
        )
        auto_uni_check.grid(row=0, column=0, sticky="w", pady=(0, 10))

        if self.max_length is not None:
            ttk.Label(
                self,
                text=f"Maximal {self.max_length} Zeichen. Umlaute werden automatisch ersetzt.",
                foreground="#5E6472",
            ).grid(row=0, column=1, sticky="e", pady=(0, 10))

        for row_index, (code, label) in enumerate(UI_LANGUAGE_ORDER, start=1):
            ttk.Label(self, text=label).grid(row=row_index, column=0, sticky="w", padx=(0, 10), pady=4)
            entry_state = "readonly" if code == "uni" else "normal"
            entry = ttk.Entry(self, textvariable=self.fields[code], width=90, state=entry_state)
            entry.grid(row=row_index, column=1, sticky="ew", pady=4)
            entry.bind("<FocusOut>", self._handle_focus_out)
            self.fields[code].trace_add("write", lambda *_args, code=code: self._sanitize_field_value(code))
            if code == "uni":
                self.uni_entry = entry
            if code == "de":
                self.fields["de"].trace_add("write", self._sync_uni)

        self._toggle_uni_state()

    def _sync_uni(self, *_args: object) -> None:
        if self.auto_uni_var.get():
            self.fields["uni"].set(self.fields["de"].get())

    def _sanitize_short_value(self, value: str) -> str:
        sanitized = replace_short_text_umlauts(value)
        if self.max_length is not None:
            sanitized = sanitized[: self.max_length]
        return sanitized

    def _sanitize_field_value(self, code: str) -> None:
        if self._suspend_limit_enforcement:
            return
        value = self.fields[code].get()
        sanitized = self._sanitize_short_value(value)
        if sanitized == value:
            return
        self._suspend_limit_enforcement = True
        try:
            self.fields[code].set(sanitized)
        finally:
            self._suspend_limit_enforcement = False

    def _toggle_uni_state(self) -> None:
        if self.auto_uni_var.get():
            self.fields["uni"].set(self.fields["de"].get())
        state = "readonly" if self.auto_uni_var.get() else "normal"
        if self.uni_entry is not None:
            self.uni_entry.configure(state=state)

    def _handle_auto_uni_toggle(self) -> None:
        self._toggle_uni_state()
        self._emit_change()

    def _handle_focus_out(self, _event: tk.Event[tk.Misc]) -> None:
        self._emit_change()

    def _emit_change(self) -> None:
        if self.on_change is not None:
            self.on_change()

    def get_value(self) -> TranslationSet:
        values = {}
        for code, variable in self.fields.items():
            values[code] = self._sanitize_short_value(variable.get())
        return TranslationSet(**values)

    def get_german_text(self) -> str:
        return self.fields["de"].get().strip()

    def apply_translations(self, translations: dict[str, str]) -> None:
        for code, value in translations.items():
            if code in self.fields and code != "de":
                self.fields[code].set(self._sanitize_short_value(value))
        self._toggle_uni_state()

    def set_value(self, value: TranslationSet, auto_uni: bool | None = None) -> None:
        if auto_uni is not None:
            self.auto_uni_var.set(auto_uni)
        for code, variable in self.fields.items():
            variable.set(self._sanitize_short_value(getattr(value, code)))
        self._toggle_uni_state()


class MultiLineTranslationFrame(ttk.LabelFrame):
    def __init__(self, master: tk.Misc, title: str, on_change: Callable[[], None] | None = None) -> None:
        super().__init__(master, text=title, padding=14)
        self.on_change = on_change
        self.auto_uni_var = tk.BooleanVar(value=True)
        self.text_widgets: dict[str, ScrolledText] = {}
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        controls = ttk.Frame(self)
        controls.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        controls.columnconfigure(1, weight=1)

        ttk.Checkbutton(
            controls,
            text="UNI automatisch aus Deutsch übernehmen",
            variable=self.auto_uni_var,
            command=self._handle_auto_uni_toggle,
        ).grid(row=0, column=0, sticky="w")
        ttk.Label(
            controls,
            text="Jede Sprache hat ein eigenes grosses Textfeld.",
            foreground="#5E6472",
        ).grid(row=0, column=1, sticky="e")

        notebook = ttk.Notebook(self)
        notebook.grid(row=1, column=0, sticky="nsew")

        for code, label in UI_LANGUAGE_ORDER:
            tab = ttk.Frame(notebook, padding=12)
            tab.columnconfigure(0, weight=1)
            tab.rowconfigure(0, weight=1)
            widget = ScrolledText(tab, wrap="word", height=18, font=("Segoe UI", 10))
            widget.grid(row=0, column=0, sticky="nsew")
            if code == "uni":
                widget.configure(state="disabled")
            if code == "de":
                widget.bind("<<Modified>>", self._on_de_modified)
            widget.bind("<FocusOut>", self._handle_focus_out)
            self.text_widgets[code] = widget
            notebook.add(tab, text=label)

    def _set_text(self, code: str, value: str) -> None:
        widget = self.text_widgets[code]
        previous_state = str(widget.cget("state"))
        if previous_state == "disabled":
            widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", value)
        if previous_state == "disabled":
            widget.configure(state="disabled")

    def _on_de_modified(self, event: tk.Event[tk.Misc]) -> None:
        widget = event.widget
        if widget.edit_modified():
            if self.auto_uni_var.get():
                self._set_text("uni", self.get_text("de"))
            widget.edit_modified(False)

    def _toggle_uni_state(self) -> None:
        uni_widget = self.text_widgets["uni"]
        if self.auto_uni_var.get():
            self._set_text("uni", self.get_text("de"))
            uni_widget.configure(state="disabled")
        else:
            uni_widget.configure(state="normal")

    def _handle_auto_uni_toggle(self) -> None:
        self._toggle_uni_state()
        self._emit_change()

    def _handle_focus_out(self, _event: tk.Event[tk.Misc]) -> None:
        self._emit_change()

    def _emit_change(self) -> None:
        if self.on_change is not None:
            self.on_change()

    def get_text(self, code: str) -> str:
        return self.text_widgets[code].get("1.0", "end").strip()

    def get_german_text(self) -> str:
        return self.get_text("de")

    def apply_translations(self, translations: dict[str, str]) -> None:
        for code, value in translations.items():
            if code in self.text_widgets and code != "de":
                self._set_text(code, value)
        self._toggle_uni_state()

    def get_value(self) -> TranslationSet:
        return TranslationSet(**{code: self.get_text(code) for code, _ in UI_LANGUAGE_ORDER})

    def set_value(self, value: TranslationSet, auto_uni: bool | None = None) -> None:
        if auto_uni is not None:
            self.auto_uni_var.set(auto_uni)
        for code, _label in UI_LANGUAGE_ORDER:
            self._set_text(code, getattr(value, code))
        self._toggle_uni_state()


class MediaTableFrame(ttk.LabelFrame):
    def __init__(
        self,
        master: tk.Misc,
        title: str,
        path_label: str,
        path_dialog_title: str,
        default_art: str,
        default_sprache: str,
        browse_filetypes: list[tuple[str, str]],
        infer_art: bool = False,
        preview_kind: str = "document",
        on_change: Callable[[], None] | None = None,
    ) -> None:
        super().__init__(master, text=title, padding=14)
        self.on_change = on_change
        self.path_dialog_title = path_dialog_title
        self.default_art = default_art
        self.default_sprache = default_sprache
        self.browse_filetypes = browse_filetypes
        self.infer_art = infer_art
        self.preview_kind = preview_kind
        self.preview_photo: object | None = None
        self.preview_bytes_cache: dict[str, bytes] = {}
        self.pdf_preview_cache: dict[str, bytes] = {}
        self.inline_editor: ttk.Entry | None = None
        self.inline_editor_item = ""
        self.inline_editor_column = ""
        self.preview_visible = True
        self.preview_toggle_var = tk.StringVar(value="Vorschau ausblenden")
        self.compact_preview_layout: bool | None = None
        self._last_preview_visibility: bool | None = None
        self._last_path_column_width: int | None = None
        self._layout_after_id: str | None = None
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Zeilen kopieren", command=self.copy_selected_rows)
        self.context_menu.add_command(label="Zeilen löschen", command=self.remove_selected)

        self.columnconfigure(0, weight=1)
        self.columnconfigure(2, weight=0)
        self.rowconfigure(1, weight=1)

        form = ttk.Frame(self)
        form.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        form.columnconfigure(1, weight=1)

        self.path_var = tk.StringVar()
        self.art_var = tk.StringVar(value=default_art)
        self.sprache_var = tk.StringVar(value=default_sprache)

        ttk.Label(form, text=path_label).grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(form, textvariable=self.path_var).grid(row=0, column=1, sticky="ew", pady=4)
        ttk.Button(form, text="Dateien wählen", command=self.browse_files).grid(row=0, column=2, padx=(8, 0), pady=4)

        ttk.Label(form, text="Art").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(form, textvariable=self.art_var, width=12).grid(row=1, column=1, sticky="w", pady=4)
        ttk.Label(form, text="Sprache").grid(row=1, column=1, sticky="e", padx=(0, 80), pady=4)
        ttk.Entry(form, textvariable=self.sprache_var, width=12).grid(row=1, column=2, sticky="w", pady=4)

        actions = ttk.Frame(form)
        actions.grid(row=2, column=1, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Button(actions, text="Zeile hinzufügen", command=self.add_manual_row).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="Auswahl aktualisieren", command=self.update_selected_row).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(actions, text="Auswahl entfernen", command=self.remove_selected).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(actions, text="Formular leeren", command=self.clear_form).grid(row=0, column=3)
        ttk.Button(actions, textvariable=self.preview_toggle_var, command=self.toggle_preview).grid(row=0, column=4, padx=(8, 0))

        self.tree = ttk.Treeview(self, columns=("path", "art", "sprache"), show="headings", height=10, selectmode="extended")
        self.tree.grid(row=1, column=0, sticky="nsew")
        self.tree.heading("path", text=path_label)
        self.tree.heading("art", text="Art")
        self.tree.heading("sprache", text="Sprache")
        self.tree.column("path", width=840, anchor="w")
        self.tree.column("art", width=100, anchor="center")
        self.tree.column("sprache", width=100, anchor="center")

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.bind("<<TreeviewSelect>>", self._handle_selection)
        self.tree.bind("<Double-1>", self._begin_inline_edit)
        self.tree.bind("<Button-3>", self._open_context_menu)

        preview_frame = ttk.LabelFrame(self, text="Vorschau", padding=10)
        preview_frame.grid(row=1, column=2, sticky="nsew", padx=(12, 0))
        preview_frame.columnconfigure(0, weight=1)
        self.preview_frame = preview_frame

        self.preview_title_var = tk.StringVar(value="Keine Auswahl")
        self.preview_meta_var = tk.StringVar(value="Wähle eine Zeile aus, um mehr Details zu sehen.")
        self.preview_badge_var = tk.StringVar(value="DATEI")

        ttk.Label(preview_frame, textvariable=self.preview_title_var, font=("Segoe UI Semibold", 10)).grid(row=0, column=0, sticky="w")

        self.preview_visual = ttk.Label(preview_frame, text="", anchor="center", justify="center")
        self.preview_visual.grid(row=1, column=0, sticky="ew", pady=(10, 10))

        ttk.Label(preview_frame, textvariable=self.preview_meta_var, foreground="#5E6472", wraplength=250, justify="left").grid(
            row=2, column=0, sticky="w"
        )
        self._reset_preview()
        self.bind("<Configure>", self._handle_resize)
        self.after_idle(self._apply_responsive_layout)

    def browse_files(self) -> None:
        selected = filedialog.askopenfilenames(
            title=self.path_dialog_title,
            filetypes=self.browse_filetypes,
        )
        if not selected:
            return
        for path in selected:
            art = infer_document_art(path) if self.infer_art else self.art_var.get().strip() or self.default_art
            sprache = self.sprache_var.get().strip() or self.default_sprache
            self.tree.insert("", "end", values=(path, art, sprache))
        self.path_var.set(selected[-1])
        self._select_first_row()
        self._emit_change()

    def add_manual_row(self) -> None:
        path = self.path_var.get().strip()
        if not path:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst einen Pfad eingeben oder Dateien wählen.")
            return
        art = self.art_var.get().strip() or self.default_art
        sprache = self.sprache_var.get().strip() or self.default_sprache
        if self.infer_art and not self.art_var.get().strip():
            art = infer_document_art(path)
            self.art_var.set(art)
        self.tree.insert("", "end", values=(path, art, sprache))
        self._select_first_row(select_last=True)
        self._emit_change()

    def update_selected_row(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst eine Zeile zum Bearbeiten auswählen.")
            return

        path = self.path_var.get().strip()
        if not path:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst einen Pfad oder eine URL eingeben.")
            return

        art = self.art_var.get().strip() or self.default_art
        sprache = self.sprache_var.get().strip() or self.default_sprache
        self.tree.item(selected[0], values=(path, art, sprache))
        self.tree.focus(selected[0])
        self._handle_selection()
        self._emit_change()

    def remove_selected(self) -> None:
        for item_id in self.tree.selection():
            self.tree.delete(item_id)
        self.clear_form()
        self._select_first_row()
        self._emit_change()

    def copy_selected_rows(self) -> None:
        selected_items = list(self.tree.selection())
        if not selected_items:
            return
        rows = []
        for item_id in selected_items:
            values = [str(value).strip() for value in self.tree.item(item_id, "values")]
            rows.append("\t".join(values))
        self.clipboard_clear()
        self.clipboard_append("\n".join(rows))

    def clear_form(self) -> None:
        self.path_var.set("")
        self.art_var.set(self.default_art)
        self.sprache_var.set(self.default_sprache)

    def _emit_change(self) -> None:
        if self.on_change is not None:
            self.on_change()

    def get_rows(self) -> list[MediaRow]:
        rows = []
        for item_id in self.tree.get_children():
            path, art, sprache = self.tree.item(item_id, "values")
            rows.append(MediaRow(path_or_link=str(path).strip(), art=str(art).strip(), sprache=str(sprache).strip()))
        return rows

    def set_rows(self, rows: list[MediaRow]) -> None:
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        for row in rows:
            self.tree.insert("", "end", values=(row.path_or_link, row.art, row.sprache))
        self._select_first_row()

    def _select_first_row(self, select_last: bool = False) -> None:
        children = list(self.tree.get_children())
        if not children:
            self._reset_preview()
            return
        target = children[-1] if select_last else children[0]
        self.tree.selection_set(target)
        self.tree.focus(target)
        self._handle_selection()

    def _reset_preview(self) -> None:
        self.preview_photo = None
        self.preview_title_var.set("Keine Auswahl")
        self.preview_meta_var.set("Wähle eine Zeile aus, um mehr Details zu sehen.")
        self.preview_visual.configure(text="Keine Vorschau", image="")

    def toggle_preview(self) -> None:
        self.set_preview_visible(not self.preview_visible)

    def set_preview_visible(self, visible: bool) -> None:
        self.preview_visible = bool(visible)
        self.preview_toggle_var.set("Vorschau ausblenden" if self.preview_visible else "Vorschau einblenden")
        self._apply_responsive_layout()

    def _handle_resize(self, _event: tk.Event[tk.Misc]) -> None:
        if self._layout_after_id is not None:
            try:
                self.after_cancel(self._layout_after_id)
            except Exception:
                pass
        self._layout_after_id = self.after_idle(self._apply_responsive_layout)

    def _apply_responsive_layout(self) -> None:
        self._layout_after_id = None
        width = max(self.winfo_width(), self.winfo_reqwidth())
        compact = width < 1220
        layout_changed = self.compact_preview_layout is None or self.compact_preview_layout != compact or self._last_preview_visibility != self.preview_visible
        self.compact_preview_layout = compact
        self._last_preview_visibility = self.preview_visible

        if layout_changed:
            if self.preview_visible:
                if compact:
                    self.preview_frame.grid(row=2, column=0, columnspan=3, sticky="ew", padx=(0, 0), pady=(12, 0))
                else:
                    self.preview_frame.grid(row=1, column=2, columnspan=1, sticky="nsew", padx=(12, 0), pady=(0, 0))
            else:
                self.preview_frame.grid_remove()

        reserved_preview_width = 0 if (compact or not self.preview_visible) else 320
        path_width = max(360, width - reserved_preview_width - 260)
        if self._last_path_column_width != path_width:
            self.tree.column("path", width=path_width, anchor="w")
            self._last_path_column_width = path_width

    def _handle_selection(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        selected = self.tree.selection()
        if not selected:
            self._reset_preview()
            return
        path, art, sprache = self.tree.item(selected[0], "values")
        self.path_var.set(str(path).strip())
        self.art_var.set(str(art).strip())
        self.sprache_var.set(str(sprache).strip())
        self._update_preview(str(path).strip(), str(art).strip(), str(sprache).strip())

    def _open_context_menu(self, event: tk.Event[tk.Misc]) -> None:
        item_id = self.tree.identify_row(event.y)
        if item_id:
            if item_id not in self.tree.selection():
                self.tree.selection_set(item_id)
            self.tree.focus(item_id)
            self._handle_selection()
        has_selection = bool(self.tree.selection())
        self.context_menu.entryconfigure("Zeilen kopieren", state="normal" if has_selection else "disabled")
        self.context_menu.entryconfigure("Zeilen löschen", state="normal" if has_selection else "disabled")
        self.context_menu.tk_popup(event.x_root, event.y_root)
        self.context_menu.grab_release()

    def _begin_inline_edit(self, event: tk.Event[tk.Misc]) -> None:
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        item_id = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        if not item_id or column_id not in {"#1", "#2", "#3"}:
            return
        bbox = self.tree.bbox(item_id, column_id)
        if not bbox:
            return

        self._destroy_inline_editor()
        values = list(self.tree.item(item_id, "values"))
        column_index = int(column_id[1:]) - 1
        current_value = str(values[column_index]) if column_index < len(values) else ""
        x_pos, y_pos, width, height = bbox

        editor = ttk.Entry(self.tree)
        editor.place(x=x_pos, y=y_pos, width=width, height=height)
        editor.insert(0, current_value)
        editor.select_range(0, "end")
        editor.focus_set()
        editor.bind("<Return>", lambda _event: self._commit_inline_edit())
        editor.bind("<Escape>", lambda _event: self._destroy_inline_editor())
        editor.bind("<FocusOut>", lambda _event: self._commit_inline_edit())
        self.inline_editor = editor
        self.inline_editor_item = item_id
        self.inline_editor_column = column_id

    def _commit_inline_edit(self) -> None:
        if self.inline_editor is None or not self.inline_editor.winfo_exists():
            self.inline_editor = None
            return

        item_id = self.inline_editor_item
        column_id = self.inline_editor_column
        new_value = self.inline_editor.get().strip()
        self._destroy_inline_editor()

        if not item_id or column_id not in {"#1", "#2", "#3"}:
            return

        values = list(self.tree.item(item_id, "values"))
        column_index = int(column_id[1:]) - 1
        while len(values) < 3:
            values.append("")

        if column_index == 0 and not new_value:
            messagebox.showwarning(APP_TITLE, "Pfad oder URL darf nicht leer sein.")
            return
        if column_index == 1 and not new_value:
            new_value = self.default_art
        if column_index == 2 and not new_value:
            new_value = self.default_sprache

        values[column_index] = new_value
        if self.infer_art and column_index == 0 and not str(values[1]).strip():
            values[1] = infer_document_art(new_value)

        self.tree.item(item_id, values=values)
        self.tree.selection_set(item_id)
        self.tree.focus(item_id)
        self._handle_selection()
        self._emit_change()

    def _destroy_inline_editor(self) -> None:
        if self.inline_editor is not None and self.inline_editor.winfo_exists():
            self.inline_editor.destroy()
        self.inline_editor = None
        self.inline_editor_item = ""
        self.inline_editor_column = ""

    def _update_preview(self, path_or_link: str, art: str, sprache: str) -> None:
        parsed = urlparse(path_or_link)
        raw_name = unquote(parsed.path if parsed.scheme else path_or_link)
        name = Path(raw_name).name or path_or_link
        extension = Path(raw_name).suffix.lower().replace(".", "").upper() or "LINK"
        self.preview_title_var.set(name)

        if self.preview_kind == "image":
            self._update_image_preview(path_or_link, art, sprache, extension)
            return

        if self.preview_kind == "document":
            self._update_document_preview(path_or_link, art, sprache, extension)
            return

        self.preview_photo = None
        self.preview_visual.configure(
            text=extension,
            image="",
            font=("Segoe UI Semibold", 18),
            width=18,
            anchor="center",
            justify="center",
        )
        self.preview_meta_var.set(
            f"Art: {art or '-'}\nSprache: {sprache or '-'}\nTyp: {extension}\nPfad/URL: {path_or_link}"
        )

    def _update_image_preview(self, path_or_link: str, art: str, sprache: str, extension: str) -> None:
        self.preview_meta_var.set(
            f"Art: {art or '-'}\nSprache: {sprache or '-'}\nTyp: {extension}\nPfad/URL: {path_or_link}"
        )
        if Image is None or ImageTk is None:
            self.preview_photo = None
            self.preview_visual.configure(text="Pillow für Bildvorschau installieren", image="", wraplength=220, justify="center")
            return

        try:
            image = self._load_preview_image(path_or_link)
        except Exception as exc:  # pragma: no cover - user-specific preview issues
            self.preview_photo = None
            self.preview_visual.configure(text=f"Keine Bildvorschau verfügbar\n{exc}", image="", wraplength=220, justify="center")
            return

        image.thumbnail((260, 180))
        photo = ImageTk.PhotoImage(image)
        self.preview_photo = photo
        self.preview_visual.configure(image=photo, text="")

    def _update_document_preview(self, path_or_link: str, art: str, sprache: str, extension: str) -> None:
        self.preview_meta_var.set(
            f"Art: {art or '-'}\nSprache: {sprache or '-'}\nTyp: {extension}\nPfad/URL: {path_or_link}"
        )

        if extension == "PDF":
            self._update_pdf_preview(path_or_link)
            return

        self.preview_photo = None
        self.preview_visual.configure(
            text=extension,
            image="",
            font=("Segoe UI Semibold", 18),
            width=18,
            anchor="center",
            justify="center",
        )

    def _update_pdf_preview(self, path_or_link: str) -> None:
        if Image is None or ImageTk is None:
            self.preview_photo = None
            self.preview_visual.configure(text="Pillow für PDF-Vorschau installieren", image="", wraplength=220, justify="center")
            return
        if pymupdf is None:
            self.preview_photo = None
            self.preview_visual.configure(text="PyMuPDF für PDF-Vorschau installieren", image="", wraplength=220, justify="center")
            return

        try:
            image = self._render_pdf_preview(path_or_link)
        except Exception as exc:  # pragma: no cover - environment-specific PDF issues
            self.preview_photo = None
            self.preview_visual.configure(text=f"Keine PDF-Vorschau verfügbar\n{exc}", image="", wraplength=220, justify="center")
            return

        image.thumbnail((260, 180))
        photo = ImageTk.PhotoImage(image)
        self.preview_photo = photo
        self.preview_visual.configure(image=photo, text="")

    def _load_preview_image(self, path_or_link: str) -> object:
        data = self._load_preview_bytes(path_or_link)
        return Image.open(io.BytesIO(data))

    def _load_preview_bytes(self, path_or_link: str) -> bytes:
        cache_key = path_or_link.strip()
        cached = self.preview_bytes_cache.get(cache_key)
        if cached is not None:
            return cached

        parsed = urlparse(path_or_link)
        if parsed.scheme in {"http", "https"}:
            data = fetch_preview_bytes_from_url(path_or_link, timeout_seconds=12)
            self.preview_bytes_cache[cache_key] = data
            return data

        file_path = Path(path_or_link)
        if not file_path.exists():
            raise FileNotFoundError("Datei nicht gefunden")
        data = file_path.read_bytes()
        self.preview_bytes_cache[cache_key] = data
        return data

    def _render_pdf_preview(self, path_or_link: str) -> object:
        cache_key = path_or_link.strip()
        cached = self.pdf_preview_cache.get(cache_key)
        if cached is not None:
            return Image.open(io.BytesIO(cached))

        pdf_bytes = self._load_preview_bytes(path_or_link)
        document = pymupdf.open(stream=pdf_bytes, filetype="pdf")
        try:
            page = document[0]
            matrix = pymupdf.Matrix(1.5, 1.5)
            pixmap = page.get_pixmap(matrix=matrix, alpha=False)
            image = pixmap.pil_image()
        finally:
            document.close()

        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        self.pdf_preview_cache[cache_key] = buffer.getvalue()
        return image


class LinkTableFrame(ttk.LabelFrame):
    def __init__(
        self,
        master: tk.Misc,
        title: str,
        label_text: str,
        preview_kind: str = "web",
        normalize_link_fn: Callable[[str], str] | None = None,
        on_change: Callable[[], None] | None = None,
    ) -> None:
        super().__init__(master, text=title, padding=14)
        self.preview_kind = preview_kind
        self.normalize_link_fn = normalize_link_fn
        self.on_change = on_change
        self.inline_editor: ttk.Entry | None = None
        self.inline_editor_item = ""
        self.preview_photo: object | None = None
        self.remote_image_cache: dict[str, bytes] = {}
        self.web_preview_cache: dict[str, tuple[bytes, str]] = {}
        self.page_title_cache: dict[str, str] = {}
        self.preview_visible = True
        self.preview_toggle_var = tk.StringVar(value="Vorschau ausblenden")
        self.compact_preview_layout: bool | None = None
        self._last_preview_visibility: bool | None = None
        self._last_link_column_width: int | None = None
        self._layout_after_id: str | None = None
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Zeilen kopieren", command=self.copy_selected_rows)
        self.context_menu.add_command(label="Zeilen löschen", command=self.remove_selected)
        self.columnconfigure(0, weight=1)
        self.columnconfigure(2, weight=0)
        self.rowconfigure(1, weight=1)

        form = ttk.Frame(self)
        form.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        form.columnconfigure(1, weight=1)

        self.link_var = tk.StringVar()

        ttk.Label(form, text=label_text).grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Entry(form, textvariable=self.link_var).grid(row=0, column=1, sticky="ew")
        ttk.Button(form, text="Hinzufügen", command=self.add_row).grid(row=0, column=2, padx=(8, 0))
        ttk.Button(form, text="Auswahl aktualisieren", command=self.update_selected_row).grid(row=0, column=3, padx=(8, 0))
        ttk.Button(form, text="Auswahl entfernen", command=self.remove_selected).grid(row=0, column=4, padx=(8, 0))
        ttk.Button(form, text="Leeren", command=self.clear_form).grid(row=0, column=5, padx=(8, 0))
        ttk.Button(form, textvariable=self.preview_toggle_var, command=self.toggle_preview).grid(row=0, column=6, padx=(8, 0))

        self.tree = ttk.Treeview(self, columns=("link",), show="headings", height=7, selectmode="extended")
        self.tree.grid(row=1, column=0, sticky="nsew")
        self.tree.heading("link", text=label_text)
        self.tree.column("link", width=780, anchor="w")

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.bind("<<TreeviewSelect>>", self._handle_selection)
        self.tree.bind("<Double-1>", self._begin_inline_edit)
        self.tree.bind("<Button-3>", self._open_context_menu)

        preview_frame = ttk.LabelFrame(self, text="Vorschau", padding=10)
        preview_frame.grid(row=1, column=2, sticky="nsew", padx=(12, 0))
        preview_frame.columnconfigure(0, weight=1)
        self.preview_frame = preview_frame

        self.preview_title_var = tk.StringVar(value="Keine Auswahl")
        self.preview_meta_var = tk.StringVar(value="Wähle eine Zeile aus, um mehr Details zu sehen.")

        ttk.Label(preview_frame, textvariable=self.preview_title_var, font=("Segoe UI Semibold", 10)).grid(row=0, column=0, sticky="w")

        self.preview_visual = ttk.Label(preview_frame, text="", anchor="center", justify="center")
        self.preview_visual.grid(row=1, column=0, sticky="ew", pady=(10, 10))

        ttk.Label(preview_frame, textvariable=self.preview_meta_var, foreground="#5E6472", wraplength=250, justify="left").grid(
            row=2, column=0, sticky="w"
        )
        self._reset_preview()
        self.bind("<Configure>", self._handle_resize)
        self.after_idle(self._apply_responsive_layout)

    def add_row(self) -> None:
        link = self.link_var.get().strip()
        if not link:
            messagebox.showwarning(APP_TITLE, "Bitte einen Link eingeben.")
            return
        if self.normalize_link_fn is not None:
            link = self.normalize_link_fn(link)
        self.tree.insert("", "end", values=(link,))
        children = list(self.tree.get_children())
        if children:
            self.tree.selection_set(children[-1])
            self.tree.focus(children[-1])
            self._handle_selection()
        self._emit_change()

    def update_selected_row(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst eine Zeile zum Bearbeiten auswählen.")
            return

        link = self.link_var.get().strip()
        if not link:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst einen Link eingeben.")
            return
        if self.normalize_link_fn is not None:
            link = self.normalize_link_fn(link)
        self.tree.item(selected[0], values=(link,))
        self.tree.focus(selected[0])
        self._handle_selection()
        self._emit_change()

    def remove_selected(self) -> None:
        for item_id in self.tree.selection():
            self.tree.delete(item_id)
        children = list(self.tree.get_children())
        if children:
            self.tree.selection_set(children[0])
            self.tree.focus(children[0])
            self._handle_selection()
        else:
            self.clear_form()
            self._reset_preview()
        self._emit_change()

    def copy_selected_rows(self) -> None:
        selected_items = list(self.tree.selection())
        if not selected_items:
            return
        rows = []
        for item_id in selected_items:
            values = [str(value).strip() for value in self.tree.item(item_id, "values")]
            rows.append("\t".join(values))
        self.clipboard_clear()
        self.clipboard_append("\n".join(rows))

    def clear_form(self) -> None:
        self.link_var.set("")

    def _emit_change(self) -> None:
        if self.on_change is not None:
            self.on_change()

    def _handle_selection(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        selected = self.tree.selection()
        if not selected:
            self._reset_preview()
            return
        (link,) = self.tree.item(selected[0], "values")
        normalized_link = str(link).strip()
        self.link_var.set(normalized_link)
        self._update_preview(normalized_link)

    def _open_context_menu(self, event: tk.Event[tk.Misc]) -> None:
        item_id = self.tree.identify_row(event.y)
        if item_id:
            if item_id not in self.tree.selection():
                self.tree.selection_set(item_id)
            self.tree.focus(item_id)
            self._handle_selection()
        has_selection = bool(self.tree.selection())
        self.context_menu.entryconfigure("Zeilen kopieren", state="normal" if has_selection else "disabled")
        self.context_menu.entryconfigure("Zeilen löschen", state="normal" if has_selection else "disabled")
        self.context_menu.tk_popup(event.x_root, event.y_root)
        self.context_menu.grab_release()

    def _begin_inline_edit(self, event: tk.Event[tk.Misc]) -> None:
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        item_id = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        if not item_id or column_id != "#1":
            return
        bbox = self.tree.bbox(item_id, column_id)
        if not bbox:
            return

        self._destroy_inline_editor()
        (current_value,) = self.tree.item(item_id, "values")
        x_pos, y_pos, width, height = bbox
        editor = ttk.Entry(self.tree)
        editor.place(x=x_pos, y=y_pos, width=width, height=height)
        editor.insert(0, str(current_value))
        editor.select_range(0, "end")
        editor.focus_set()
        editor.bind("<Return>", lambda _event: self._commit_inline_edit())
        editor.bind("<Escape>", lambda _event: self._destroy_inline_editor())
        editor.bind("<FocusOut>", lambda _event: self._commit_inline_edit())
        self.inline_editor = editor
        self.inline_editor_item = item_id

    def _commit_inline_edit(self) -> None:
        if self.inline_editor is None or not self.inline_editor.winfo_exists():
            self.inline_editor = None
            return

        item_id = self.inline_editor_item
        link = self.inline_editor.get().strip()
        self._destroy_inline_editor()

        if not item_id:
            return
        if not link:
            messagebox.showwarning(APP_TITLE, "Link darf nicht leer sein.")
            return
        if self.normalize_link_fn is not None:
            link = self.normalize_link_fn(link)

        self.tree.item(item_id, values=(link,))
        self.tree.selection_set(item_id)
        self.tree.focus(item_id)
        self._handle_selection()
        self._emit_change()

    def _destroy_inline_editor(self) -> None:
        if self.inline_editor is not None and self.inline_editor.winfo_exists():
            self.inline_editor.destroy()
        self.inline_editor = None
        self.inline_editor_item = ""

    def _select_first_row(self, select_last: bool = False) -> None:
        children = list(self.tree.get_children())
        if not children:
            self._reset_preview()
            return
        target = children[-1] if select_last else children[0]
        self.tree.selection_set(target)
        self.tree.focus(target)
        self._handle_selection()

    def _reset_preview(self) -> None:
        self.preview_photo = None
        self.preview_title_var.set("Keine Auswahl")
        self.preview_meta_var.set("Wähle eine Zeile aus, um mehr Details zu sehen.")
        self.preview_visual.configure(text="Keine Vorschau", image="", wraplength=220, justify="center")

    def toggle_preview(self) -> None:
        self.set_preview_visible(not self.preview_visible)

    def set_preview_visible(self, visible: bool) -> None:
        self.preview_visible = bool(visible)
        self.preview_toggle_var.set("Vorschau ausblenden" if self.preview_visible else "Vorschau einblenden")
        self._apply_responsive_layout()

    def _handle_resize(self, _event: tk.Event[tk.Misc]) -> None:
        if self._layout_after_id is not None:
            try:
                self.after_cancel(self._layout_after_id)
            except Exception:
                pass
        self._layout_after_id = self.after_idle(self._apply_responsive_layout)

    def _apply_responsive_layout(self) -> None:
        self._layout_after_id = None
        width = max(self.winfo_width(), self.winfo_reqwidth())
        compact = width < 1160
        layout_changed = self.compact_preview_layout is None or self.compact_preview_layout != compact or self._last_preview_visibility != self.preview_visible
        self.compact_preview_layout = compact
        self._last_preview_visibility = self.preview_visible

        if layout_changed:
            if self.preview_visible:
                if compact:
                    self.preview_frame.grid(row=2, column=0, columnspan=3, sticky="ew", padx=(0, 0), pady=(12, 0))
                else:
                    self.preview_frame.grid(row=1, column=2, columnspan=1, sticky="nsew", padx=(12, 0), pady=(0, 0))
            else:
                self.preview_frame.grid_remove()

        reserved_preview_width = 0 if (compact or not self.preview_visible) else 320
        link_width = max(360, width - reserved_preview_width - 40)
        if self._last_link_column_width != link_width:
            self.tree.column("link", width=link_width, anchor="w")
            self._last_link_column_width = link_width

    def _update_preview(self, link: str) -> None:
        parsed = urlparse(link)
        host = parsed.netloc.lower().removeprefix("www.") if parsed.netloc else "-"
        path = unquote(parsed.path or "/")

        if self.preview_kind == "video":
            self._update_video_preview(link, host, path)
            return

        self._update_web_preview(link, host, path)

    def _update_video_preview(self, link: str, host: str, path: str) -> None:
        video_id = extract_youtube_video_id(link)
        self.preview_title_var.set(f"Video: {video_id}" if video_id else (host or "Video-Link"))
        self.preview_meta_var.set(f"Host: {host or '-'}\nPfad: {path or '/'}\nLink: {link}")

        if not video_id or Image is None or ImageTk is None:
            self.preview_photo = None
            self.preview_visual.configure(text="VIDEO", image="", font=("Segoe UI Semibold", 18), wraplength=220, justify="center")
            return

        thumbnail_url = f"https://img.youtube.com/vi/{video_id}/hqdefault.jpg"
        try:
            image = self._load_remote_image(thumbnail_url)
        except Exception:  # pragma: no cover - preview only
            self.preview_photo = None
            self.preview_visual.configure(text="VIDEO", image="", font=("Segoe UI Semibold", 18), wraplength=220, justify="center")
            return

        image.thumbnail((260, 180))
        photo = ImageTk.PhotoImage(image)
        self.preview_photo = photo
        self.preview_visual.configure(image=photo, text="")

    def _update_web_preview(self, link: str, host: str, path: str) -> None:
        title = self._fetch_page_title(link)
        self.preview_title_var.set(title or host or "Web-Link")
        self.preview_meta_var.set(f"Host: {host or '-'}\nPfad: {path or '/'}\nLink: {link}")
        if Image is not None and ImageTk is not None:
            try:
                image, page_title = self._render_web_preview(link)
                if page_title:
                    self.preview_title_var.set(page_title)
                image.thumbnail((260, 180))
                photo = ImageTk.PhotoImage(image)
                self.preview_photo = photo
                self.preview_visual.configure(image=photo, text="")
                return
            except Exception:  # pragma: no cover - preview only
                pass

        badge = (host or "LINK").upper()
        self.preview_photo = None
        self.preview_visual.configure(text=badge, image="", font=("Segoe UI Semibold", 16), wraplength=220, justify="center")

    def _load_remote_image(self, url: str) -> object:
        cached = self.remote_image_cache.get(url)
        if cached is None:
            cached = fetch_preview_bytes_from_url(url, timeout_seconds=10)
            self.remote_image_cache[url] = cached
        return Image.open(io.BytesIO(cached))

    def _render_web_preview(self, link: str) -> tuple[object, str]:
        parsed = urlparse(link)
        if parsed.scheme not in {"http", "https"}:
            raise ValueError("Nur http/https Links können als Webseite vorgeladen werden.")

        cached = self.web_preview_cache.get(link)
        if cached is not None:
            screenshot_bytes, title = cached
            return Image.open(io.BytesIO(screenshot_bytes)), title

        browser_path = detect_browser_path()
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                executable_path=str(browser_path) if browser_path else None,
                headless=True,
            )
            context = browser.new_context(
                viewport={"width": 1280, "height": 900},
                ignore_https_errors=True,
            )
            page = context.new_page()
            try:
                page.goto(link, wait_until="domcontentloaded", timeout=20000)
                try:
                    page.wait_for_load_state("networkidle", timeout=5000)
                except Exception:
                    pass
                page.wait_for_timeout(800)
                screenshot = page.screenshot(full_page=False, type="png")
                title = " ".join(page.title().split()).strip()
            finally:
                context.close()
                browser.close()

        self.web_preview_cache[link] = (screenshot, title)
        return Image.open(io.BytesIO(screenshot)), title

    def _fetch_page_title(self, link: str) -> str:
        parsed = urlparse(link)
        if parsed.scheme not in {"http", "https"}:
            return ""

        cached = self.page_title_cache.get(link)
        if cached is not None:
            return cached

        try:
            raw = fetch_preview_bytes_from_url(link, timeout_seconds=8, max_bytes=32768)
        except Exception:  # pragma: no cover - preview only
            return ""

        text = raw.decode("utf-8", errors="ignore")
        match = re.search(r"<title[^>]*>(.*?)</title>", text, flags=re.IGNORECASE | re.DOTALL)
        if not match:
            return ""
        title = " ".join(unescape(match.group(1)).split()).strip()
        self.page_title_cache[link] = title
        return title

    def get_rows(self) -> list[MediaRow]:
        rows = []
        for item_id in self.tree.get_children():
            (link,) = self.tree.item(item_id, "values")
            normalized = str(link).strip()
            if self.normalize_link_fn is not None:
                normalized = self.normalize_link_fn(normalized)
            rows.append(MediaRow(path_or_link=normalized))
        return rows

    def set_rows(self, rows: list[MediaRow]) -> None:
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        for row in rows:
            link = row.path_or_link
            if self.normalize_link_fn is not None:
                link = self.normalize_link_fn(link)
            self.tree.insert("", "end", values=(link,))
        self._select_first_row()


class SimpleValueTableFrame(ttk.LabelFrame):
    def __init__(
        self,
        master: tk.Misc,
        title: str,
        entry_label: str,
        on_change: Callable[[], None] | None = None,
    ) -> None:
        super().__init__(master, text=title, padding=14)
        self.on_change = on_change
        self.value_var = tk.StringVar()
        self.inline_editor: ttk.Entry | None = None
        self.inline_editor_item = ""
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Zeilen kopieren", command=self.copy_selected_rows)
        self.context_menu.add_command(label="Zeilen löschen", command=self.remove_selected)

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        form = ttk.Frame(self)
        form.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        form.columnconfigure(1, weight=1)

        ttk.Label(form, text=entry_label).grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Entry(form, textvariable=self.value_var).grid(row=0, column=1, sticky="ew")
        ttk.Button(form, text="Hinzufügen", command=self.add_row).grid(row=0, column=2, padx=(8, 0))
        ttk.Button(form, text="Auswahl aktualisieren", command=self.update_selected_row).grid(row=0, column=3, padx=(8, 0))
        ttk.Button(form, text="Auswahl entfernen", command=self.remove_selected).grid(row=0, column=4, padx=(8, 0))
        ttk.Button(form, text="Leeren", command=self.clear_form).grid(row=0, column=5, padx=(8, 0))

        self.tree = ttk.Treeview(self, columns=("value",), show="headings", height=12, selectmode="extended")
        self.tree.grid(row=1, column=0, sticky="nsew")
        self.tree.heading("value", text=entry_label)
        self.tree.column("value", width=980, anchor="w")

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.bind("<<TreeviewSelect>>", self._handle_selection)
        self.tree.bind("<Double-1>", self._begin_inline_edit)
        self.tree.bind("<Button-3>", self._open_context_menu)

    def _emit_change(self) -> None:
        if self.on_change is not None:
            self.on_change()

    def add_row(self) -> None:
        value = self.value_var.get().strip()
        if not value:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst einen Wert eingeben.")
            return
        self.tree.insert("", "end", values=(value,))
        self._select_first_row(select_last=True)
        self._emit_change()

    def update_selected_row(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst eine Zeile zum Bearbeiten auswählen.")
            return
        value = self.value_var.get().strip()
        if not value:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst einen Wert eingeben.")
            return
        self.tree.item(selected[0], values=(value,))
        self.tree.focus(selected[0])
        self._handle_selection()
        self._emit_change()

    def remove_selected(self) -> None:
        for item_id in self.tree.selection():
            self.tree.delete(item_id)
        self.clear_form()
        self._select_first_row()
        self._emit_change()

    def copy_selected_rows(self) -> None:
        selected_items = list(self.tree.selection())
        if not selected_items:
            return
        rows = []
        for item_id in selected_items:
            values = [str(value).strip() for value in self.tree.item(item_id, "values")]
            rows.append("\t".join(values))
        self.clipboard_clear()
        self.clipboard_append("\n".join(rows))

    def clear_form(self) -> None:
        self.value_var.set("")

    def get_rows(self) -> list[OeNumberRow]:
        rows = []
        for item_id in self.tree.get_children():
            (value,) = self.tree.item(item_id, "values")
            rows.append(OeNumberRow(value=str(value).strip()))
        return normalize_oe_number_rows(rows)

    def set_rows(self, rows: list[OeNumberRow]) -> None:
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        for row in normalize_oe_number_rows(rows):
            self.tree.insert("", "end", values=(row.value,))
        self._select_first_row()

    def _select_first_row(self, select_last: bool = False) -> None:
        children = list(self.tree.get_children())
        if not children:
            return
        target = children[-1] if select_last else children[0]
        self.tree.selection_set(target)
        self.tree.focus(target)
        self._handle_selection()

    def _handle_selection(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        (value,) = self.tree.item(selected[0], "values")
        self.value_var.set(str(value).strip())

    def _open_context_menu(self, event: tk.Event[tk.Misc]) -> None:
        item_id = self.tree.identify_row(event.y)
        if item_id:
            if item_id not in self.tree.selection():
                self.tree.selection_set(item_id)
            self.tree.focus(item_id)
            self._handle_selection()
        has_selection = bool(self.tree.selection())
        self.context_menu.entryconfigure("Zeilen kopieren", state="normal" if has_selection else "disabled")
        self.context_menu.entryconfigure("Zeilen löschen", state="normal" if has_selection else "disabled")
        self.context_menu.tk_popup(event.x_root, event.y_root)
        self.context_menu.grab_release()

    def _begin_inline_edit(self, event: tk.Event[tk.Misc]) -> None:
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        item_id = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        if not item_id or column_id != "#1":
            return
        bbox = self.tree.bbox(item_id, column_id)
        if not bbox:
            return
        self._destroy_inline_editor()
        (current_value,) = self.tree.item(item_id, "values")
        x_pos, y_pos, width, height = bbox
        editor = ttk.Entry(self.tree)
        editor.place(x=x_pos, y=y_pos, width=width, height=height)
        editor.insert(0, str(current_value))
        editor.select_range(0, "end")
        editor.focus_set()
        editor.bind("<Return>", lambda _event: self._commit_inline_edit())
        editor.bind("<Escape>", lambda _event: self._destroy_inline_editor())
        editor.bind("<FocusOut>", lambda _event: self._commit_inline_edit())
        self.inline_editor = editor
        self.inline_editor_item = item_id

    def _commit_inline_edit(self) -> None:
        if self.inline_editor is None or not self.inline_editor.winfo_exists():
            self.inline_editor = None
            return
        item_id = self.inline_editor_item
        value = self.inline_editor.get().strip()
        self._destroy_inline_editor()
        if not item_id:
            return
        if not value:
            messagebox.showwarning(APP_TITLE, "Der Wert darf nicht leer sein.")
            return
        self.tree.item(item_id, values=(value,))
        self.tree.selection_set(item_id)
        self.tree.focus(item_id)
        self._handle_selection()
        self._emit_change()

    def _destroy_inline_editor(self) -> None:
        if self.inline_editor is not None and self.inline_editor.winfo_exists():
            self.inline_editor.destroy()
        self.inline_editor = None
        self.inline_editor_item = ""


class SearchSuggestionPopup:
    IGNORED_KEYSYMS = {
        "Up",
        "Down",
        "Left",
        "Right",
        "Return",
        "Escape",
        "Tab",
        "Shift_L",
        "Shift_R",
        "Control_L",
        "Control_R",
        "Alt_L",
        "Alt_R",
    }

    def __init__(
        self,
        owner: tk.Misc,
        entry: ttk.Entry,
        values_provider: Callable[[str], list[str]],
        accept_callback: Callable[[str], None],
        *,
        on_focus_out: Callable[[], None] | None = None,
        on_missing_selection: Callable[[], None] | None = None,
        min_width: int = 360,
        max_visible_rows: int = 8,
    ) -> None:
        self.owner = owner
        self.entry = entry
        self.values_provider = values_provider
        self.accept_callback = accept_callback
        self.on_focus_out = on_focus_out
        self.on_missing_selection = on_missing_selection
        self.min_width = min_width
        self.max_visible_rows = max_visible_rows
        self.popup: tk.Toplevel | None = None
        self.listbox: tk.Listbox | None = None
        self.values: list[str] = []
        self.suppress_next_focus_show = False

        self.entry.bind("<KeyRelease>", self._handle_key_release)
        self.entry.bind("<Down>", self._open_event)
        self.entry.bind("<F4>", self._open_event)
        self.entry.bind("<Escape>", self._close_event)
        self.entry.bind("<Return>", self._handle_return)
        self.entry.bind("<FocusIn>", self._handle_focus_in)
        self.entry.bind("<FocusOut>", self._handle_focus_out)

    def refresh(self, show: bool = False) -> list[str]:
        values = self.values_provider(self.entry.get())
        self.values = values
        if show or (self.popup is not None and self.popup.winfo_exists() and self.popup.state() != "withdrawn"):
            self.show(values)
        return list(values)

    def show(self, values: list[str] | None = None) -> None:
        if values is None:
            values = self.values_provider(self.entry.get())
        self.values = list(values)
        if not self.values:
            self.hide()
            return

        self._ensure_popup()
        if self.popup is None or self.listbox is None:
            return

        self.listbox.delete(0, "end")
        for value in self.values:
            self.listbox.insert("end", value)

        visible_rows = min(max(len(self.values), 1), self.max_visible_rows)
        self.listbox.configure(height=visible_rows)
        self.listbox.selection_clear(0, "end")
        self.listbox.selection_set(0)
        self.listbox.activate(0)
        self.listbox.see(0)

        self.popup.update_idletasks()
        x_pos = self.entry.winfo_rootx()
        y_pos = self.entry.winfo_rooty() + self.entry.winfo_height()
        width = max(self.entry.winfo_width(), self.min_width)
        height = self.popup.winfo_reqheight()
        self.popup.geometry(f"{width}x{height}+{x_pos}+{y_pos}")
        self.popup.deiconify()
        self.popup.lift()

    def hide(self) -> None:
        if self.popup is None or not self.popup.winfo_exists():
            self.popup = None
            self.listbox = None
            return
        self.popup.withdraw()

    def _ensure_popup(self) -> None:
        if self.popup is not None and self.popup.winfo_exists():
            return

        popup = tk.Toplevel(self.owner)
        popup.withdraw()
        popup.overrideredirect(True)
        popup.transient(self.owner.winfo_toplevel())
        popup.configure(background=UI_BORDER, padx=1, pady=1)

        listbox = tk.Listbox(
            popup,
            activestyle="none",
            exportselection=False,
            borderwidth=0,
            highlightthickness=0,
            font=("Segoe UI", 10),
        )
        listbox.pack(fill="both", expand=True)
        listbox.bind("<ButtonRelease-1>", self._accept_listbox_click)
        listbox.bind("<Double-Button-1>", self._accept_listbox_click)
        listbox.bind("<Return>", self._accept_listbox_keyboard)
        listbox.bind("<Escape>", self._close_event)
        listbox.bind("<FocusOut>", self._handle_focus_out)

        self.popup = popup
        self.listbox = listbox

    def _selected_value(self) -> str | None:
        if self.listbox is not None and self.listbox.winfo_exists():
            selection = self.listbox.curselection()
            if selection:
                return str(self.listbox.get(selection[0]))
        if not self.values:
            self.values = self.values_provider(self.entry.get())
        return self.values[0] if self.values else None

    def _accept_value(self, value: str | None) -> None:
        if not value:
            if self.on_missing_selection is not None:
                self.on_missing_selection()
            self.hide()
            return
        self.accept_callback(value)
        self.hide()
        self.suppress_next_focus_show = True
        self.entry.focus_set()
        self.entry.icursor("end")

    def _accept_listbox_click(self, _event: tk.Event[tk.Misc]) -> str:
        self._accept_value(self._selected_value())
        return "break"

    def _accept_listbox_keyboard(self, _event: tk.Event[tk.Misc]) -> str:
        self._accept_value(self._selected_value())
        return "break"

    def _open_event(self, _event: tk.Event[tk.Misc]) -> str:
        values = self.values_provider(self.entry.get())
        self.show(values)
        if self.listbox is not None and self.values:
            self.listbox.focus_set()
        return "break"

    def _close_event(self, _event: tk.Event[tk.Misc]) -> str:
        self.hide()
        self.suppress_next_focus_show = True
        self.entry.focus_set()
        return "break"

    def _handle_return(self, _event: tk.Event[tk.Misc]) -> str:
        self._accept_value(self._selected_value())
        return "break"

    def _handle_focus_in(self, _event: tk.Event[tk.Misc]) -> None:
        if self.suppress_next_focus_show:
            self.suppress_next_focus_show = False
            return
        if self.entry.get().strip():
            self.show()

    def _handle_key_release(self, event: tk.Event[tk.Misc]) -> None:
        if event.keysym in self.IGNORED_KEYSYMS:
            return
        if self.entry.get().strip():
            self.show()
        else:
            self.hide()

    def _handle_focus_out(self, _event: tk.Event[tk.Misc]) -> None:
        self.owner.after(120, self._finalize_focus_out)

    def _finalize_focus_out(self) -> None:
        focus_widget = self.owner.focus_get()
        if self._is_popup_widget(focus_widget):
            return
        self.hide()
        if self.on_focus_out is not None:
            self.on_focus_out()

    def _is_popup_widget(self, widget: tk.Misc | None) -> bool:
        if widget is None:
            return False
        if widget is self.entry:
            return True
        if self.listbox is not None and widget is self.listbox:
            return True
        if self.popup is not None and widget is self.popup:
            return True
        master = getattr(widget, "master", None)
        while master is not None:
            if self.popup is not None and master is self.popup:
                return True
            master = getattr(master, "master", None)
        return False


class OeNumberTableFrame(ttk.LabelFrame):
    def __init__(self, master: tk.Misc, on_change: Callable[[], None] | None = None) -> None:
        super().__init__(master, text="OE-Nummern", padding=14)
        self.on_change = on_change
        self.catalog_options: list[CompetitorOption] = []
        self.catalog_by_id: dict[str, CompetitorOption] = {}
        self.catalog_by_code: dict[str, CompetitorOption] = {}
        self.catalog_by_label: dict[str, CompetitorOption] = {}
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Zeilen kopieren", command=self.copy_selected_rows)
        self.context_menu.add_command(label="Zeilen löschen", command=self.remove_selected)

        self.manufacturer_display_var = tk.StringVar()
        self.manufacturer_id_var = tk.StringVar()
        self.manufacturer_code_var = tk.StringVar()
        self.manufacturer_name_var = tk.StringVar()
        self.value_var = tk.StringVar()
        self.manufacturer_suggestions: SearchSuggestionPopup | None = None

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        header = ttk.Frame(self)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header.columnconfigure(1, weight=1)
        header.columnconfigure(5, weight=1)
        header.columnconfigure(7, weight=1)

        ttk.Label(
            header,
            text="Wähle den Hersteller aus und erfasse die zugehörige OE-Nummer.",
            foreground="#5E6472",
            wraplength=980,
        ).grid(row=0, column=0, columnspan=7, sticky="w", pady=(0, 10))

        ttk.Label(header, text="Hersteller").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        self.manufacturer_entry = ttk.Entry(header, textvariable=self.manufacturer_display_var)
        self.manufacturer_entry.grid(row=1, column=1, columnspan=6, sticky="ew", pady=4)
        self.manufacturer_suggestions = SearchSuggestionPopup(
            self,
            self.manufacturer_entry,
            self._update_manufacturer_combo_values,
            self._accept_manufacturer_suggestion,
            on_focus_out=self._apply_current_manufacturer_selection,
            min_width=520,
        )

        ttk.Label(header, text="KHerNr").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(header, textvariable=self.manufacturer_id_var, width=16).grid(row=2, column=1, sticky="w", pady=4)
        ttk.Label(header, text="Kürzel").grid(row=2, column=2, sticky="w", padx=(12, 8), pady=4)
        ttk.Entry(header, textvariable=self.manufacturer_code_var, width=18).grid(row=2, column=3, sticky="w", pady=4)
        ttk.Label(header, text="Name").grid(row=2, column=4, sticky="w", padx=(12, 8), pady=4)
        ttk.Entry(header, textvariable=self.manufacturer_name_var).grid(row=2, column=5, columnspan=2, sticky="ew", pady=4)

        ttk.Label(header, text="OE-Nummer").grid(row=3, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(header, textvariable=self.value_var).grid(row=3, column=1, columnspan=3, sticky="ew", pady=4)

        actions = ttk.Frame(header)
        actions.grid(row=3, column=4, columnspan=3, sticky="e", pady=4)
        ttk.Button(actions, text="Hinzufügen", command=self.add_row).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="Auswahl aktualisieren", command=self.update_selected_row).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(actions, text="Auswahl entfernen", command=self.remove_selected).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(actions, text="Leeren", command=self.clear_form).grid(row=0, column=3)

        columns = ("manufacturer_id", "manufacturer_code", "manufacturer_name", "value")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=11, selectmode="extended")
        self.tree.grid(row=1, column=0, sticky="nsew")
        self.tree.heading("manufacturer_id", text="KHerNr")
        self.tree.heading("manufacturer_code", text="Kürzel")
        self.tree.heading("manufacturer_name", text="Hersteller")
        self.tree.heading("value", text="OE-Nummer")
        self.tree.column("manufacturer_id", width=120, anchor="center")
        self.tree.column("manufacturer_code", width=120, anchor="center")
        self.tree.column("manufacturer_name", width=380, anchor="w")
        self.tree.column("value", width=300, anchor="w")
        self.tree.bind("<<TreeviewSelect>>", self._handle_selection)
        self.tree.bind("<Button-3>", self._open_context_menu)

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

    def set_manufacturer_catalog(self, options: list[CompetitorOption]) -> None:
        self.catalog_options = list(options)
        self.catalog_by_id = {option.competitor_id: option for option in self.catalog_options}
        self.catalog_by_code = {option.code.casefold(): option for option in self.catalog_options if option.code}
        self.catalog_by_label = {option.display_label(): option for option in self.catalog_options}
        if self.manufacturer_suggestions is not None:
            self.manufacturer_suggestions.refresh()

    def prefill_manufacturer(self, option: CompetitorOption) -> None:
        self.manufacturer_display_var.set(option.display_label())
        self.manufacturer_id_var.set(option.competitor_id)
        self.manufacturer_code_var.set(option.code)
        self.manufacturer_name_var.set(option.name)
        if self.manufacturer_suggestions is not None:
            self.manufacturer_suggestions.refresh()

    def _update_manufacturer_combo_values(self, query: str = "") -> list[str]:
        query_text = query.strip().casefold()
        filtered = [
            option.display_label()
            for option in self.catalog_options
            if not query_text or query_text in option.search_blob
        ]
        return filtered[:200]

    def _resolve_manufacturer(self, raw_value: str) -> CompetitorOption | None:
        value = raw_value.strip()
        if not value:
            return None
        if value in self.catalog_by_label:
            return self.catalog_by_label[value]
        if value in self.catalog_by_id:
            return self.catalog_by_id[value]
        by_code = self.catalog_by_code.get(value.casefold())
        if by_code is not None:
            return by_code

        exact_matches = [option for option in self.catalog_options if value.casefold() == option.name.casefold()]
        if len(exact_matches) == 1:
            return exact_matches[0]

        partial_matches = [option for option in self.catalog_options if value.casefold() in option.search_blob]
        if len(partial_matches) == 1:
            return partial_matches[0]
        return None

    def _apply_current_manufacturer_selection(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        option = self._resolve_manufacturer(self.manufacturer_display_var.get())
        if option is None:
            return
        self.prefill_manufacturer(option)

    def _accept_manufacturer_suggestion(self, value: str) -> None:
        self.manufacturer_display_var.set(value)
        self._apply_current_manufacturer_selection()

    def _on_manufacturer_key_release(self, _event: tk.Event[tk.Misc]) -> None:
        if self.manufacturer_suggestions is not None:
            self.manufacturer_suggestions.refresh(show=True)

    def _emit_change(self) -> None:
        if self.on_change is not None:
            self.on_change()

    def _build_current_row(self) -> OeNumberRow:
        return OeNumberRow(
            value=self.value_var.get().strip(),
            manufacturer_id=self.manufacturer_id_var.get().strip(),
            manufacturer_code=self.manufacturer_code_var.get().strip(),
            manufacturer_name=self.manufacturer_name_var.get().strip(),
        )

    def add_row(self) -> None:
        self._apply_current_manufacturer_selection()
        row = self._build_current_row()
        if not row.manufacturer_id or not row.value:
            messagebox.showwarning(APP_TITLE, "Bitte Hersteller und OE-Nummer eingeben.")
            return
        self.tree.insert("", "end", values=(row.manufacturer_id, row.manufacturer_code, row.manufacturer_name, row.value))
        self._select_first_row(select_last=True)
        self._emit_change()

    def update_selected_row(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst eine Zeile zum Bearbeiten auswählen.")
            return
        self._apply_current_manufacturer_selection()
        row = self._build_current_row()
        if not row.manufacturer_id or not row.value:
            messagebox.showwarning(APP_TITLE, "Bitte Hersteller und OE-Nummer eingeben.")
            return
        self.tree.item(selected[0], values=(row.manufacturer_id, row.manufacturer_code, row.manufacturer_name, row.value))
        self.tree.focus(selected[0])
        self._handle_selection()
        self._emit_change()

    def remove_selected(self) -> None:
        for item_id in self.tree.selection():
            self.tree.delete(item_id)
        self.clear_form()
        self._select_first_row()
        self._emit_change()

    def copy_selected_rows(self) -> None:
        selected_items = list(self.tree.selection())
        if not selected_items:
            return
        rows = []
        for item_id in selected_items:
            values = [str(value).strip() for value in self.tree.item(item_id, "values")]
            rows.append("\t".join(values))
        self.clipboard_clear()
        self.clipboard_append("\n".join(rows))

    def clear_form(self) -> None:
        self.manufacturer_display_var.set("")
        self.manufacturer_id_var.set("")
        self.manufacturer_code_var.set("")
        self.manufacturer_name_var.set("")
        self.value_var.set("")

    def get_rows(self) -> list[OeNumberRow]:
        rows = []
        for item_id in self.tree.get_children():
            manufacturer_id, manufacturer_code, manufacturer_name, value = self.tree.item(item_id, "values")
            rows.append(
                OeNumberRow(
                    value=str(value).strip(),
                    manufacturer_id=str(manufacturer_id).strip(),
                    manufacturer_code=str(manufacturer_code).strip(),
                    manufacturer_name=str(manufacturer_name).strip(),
                )
            )
        return normalize_oe_number_rows(rows)

    def set_rows(self, rows: list[OeNumberRow]) -> None:
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        for row in normalize_oe_number_rows(rows):
            self.tree.insert("", "end", values=(row.manufacturer_id, row.manufacturer_code, row.manufacturer_name, row.value))
        self._select_first_row()

    def _select_first_row(self, select_last: bool = False) -> None:
        children = list(self.tree.get_children())
        if not children:
            return
        target = children[-1] if select_last else children[0]
        self.tree.selection_set(target)
        self.tree.focus(target)
        self._handle_selection()

    def _handle_selection(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        manufacturer_id, manufacturer_code, manufacturer_name, value = self.tree.item(selected[0], "values")
        self.manufacturer_id_var.set(str(manufacturer_id).strip())
        self.manufacturer_code_var.set(str(manufacturer_code).strip())
        self.manufacturer_name_var.set(str(manufacturer_name).strip())
        self.value_var.set(str(value).strip())
        display_parts = [self.manufacturer_id_var.get(), self.manufacturer_code_var.get(), self.manufacturer_name_var.get()]
        self.manufacturer_display_var.set(" | ".join(part for part in display_parts if part))

    def _open_context_menu(self, event: tk.Event[tk.Misc]) -> None:
        item_id = self.tree.identify_row(event.y)
        if item_id:
            if item_id not in self.tree.selection():
                self.tree.selection_set(item_id)
            self.tree.focus(item_id)
            self._handle_selection()
        has_selection = bool(self.tree.selection())
        self.context_menu.entryconfigure("Zeilen kopieren", state="normal" if has_selection else "disabled")
        self.context_menu.entryconfigure("Zeilen löschen", state="normal" if has_selection else "disabled")
        self.context_menu.tk_popup(event.x_root, event.y_root)
        self.context_menu.grab_release()


class ComparisonTableFrame(ttk.LabelFrame):
    def __init__(self, master: tk.Misc, on_change: Callable[[], None] | None = None) -> None:
        super().__init__(master, text="Vergleichsnummern", padding=14)
        self.on_change = on_change
        self.catalog_options: list[CompetitorOption] = []
        self.catalog_by_id: dict[str, CompetitorOption] = {}
        self.catalog_by_code: dict[str, CompetitorOption] = {}
        self.catalog_by_label: dict[str, CompetitorOption] = {}
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Zeilen kopieren", command=self.copy_selected_rows)
        self.context_menu.add_command(label="Zeilen löschen", command=self.remove_selected)

        self.competitor_display_var = tk.StringVar()
        self.competitor_id_var = tk.StringVar()
        self.competitor_code_var = tk.StringVar()
        self.competitor_name_var = tk.StringVar()
        self.reference_number_var = tk.StringVar()
        self.competitor_suggestions: SearchSuggestionPopup | None = None

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        header = ttk.Frame(self)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header.columnconfigure(1, weight=1)
        header.columnconfigure(5, weight=1)

        ttk.Label(
            header,
            text="Mitbewerber kann direkt aus KHer.csv ausgewählt oder manuell über die ID gepflegt werden.",
            foreground="#5E6472",
            wraplength=980,
        ).grid(row=0, column=0, columnspan=7, sticky="w", pady=(0, 10))

        ttk.Label(header, text="Mitbewerber").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        self.competitor_entry = ttk.Entry(header, textvariable=self.competitor_display_var)
        self.competitor_entry.grid(row=1, column=1, columnspan=6, sticky="ew", pady=4)
        self.competitor_suggestions = SearchSuggestionPopup(
            self,
            self.competitor_entry,
            self._update_competitor_combo_values,
            self._accept_competitor_suggestion,
            on_focus_out=self._apply_current_competitor_selection,
            min_width=520,
        )

        ttk.Label(header, text="Mitbewerber ID").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(header, textvariable=self.competitor_id_var, width=16).grid(row=2, column=1, sticky="w", pady=4)
        ttk.Label(header, text="Kürzel").grid(row=2, column=2, sticky="w", padx=(12, 8), pady=4)
        ttk.Entry(header, textvariable=self.competitor_code_var, width=18).grid(row=2, column=3, sticky="w", pady=4)
        ttk.Label(header, text="Name").grid(row=2, column=4, sticky="w", padx=(12, 8), pady=4)
        ttk.Entry(header, textvariable=self.competitor_name_var).grid(row=2, column=5, columnspan=2, sticky="ew", pady=4)

        ttk.Label(header, text="Vergleichsnummer").grid(row=3, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(header, textvariable=self.reference_number_var).grid(row=3, column=1, columnspan=3, sticky="ew", pady=4)

        actions = ttk.Frame(header)
        actions.grid(row=3, column=4, columnspan=3, sticky="e", pady=4)
        ttk.Button(actions, text="Hinzufügen", command=self.add_row).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="Auswahl aktualisieren", command=self.update_selected_row).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(actions, text="Auswahl entfernen", command=self.remove_selected).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(actions, text="Leeren", command=self.clear_form).grid(row=0, column=3)

        columns = ("competitor_id", "competitor_code", "competitor_name", "reference_number")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=11, selectmode="extended")
        self.tree.grid(row=1, column=0, sticky="nsew")
        self.tree.heading("competitor_id", text="Mitbewerber ID")
        self.tree.heading("competitor_code", text="Kürzel")
        self.tree.heading("competitor_name", text="Mitbewerber")
        self.tree.heading("reference_number", text="Vergleichsnummer")
        self.tree.column("competitor_id", width=120, anchor="center")
        self.tree.column("competitor_code", width=120, anchor="center")
        self.tree.column("competitor_name", width=380, anchor="w")
        self.tree.column("reference_number", width=300, anchor="w")
        self.tree.bind("<<TreeviewSelect>>", self._handle_selection)
        self.tree.bind("<Button-3>", self._open_context_menu)

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

    def set_competitor_catalog(self, options: list[CompetitorOption]) -> None:
        self.catalog_options = list(options)
        self.catalog_by_id = {option.competitor_id: option for option in self.catalog_options}
        self.catalog_by_code = {option.code.casefold(): option for option in self.catalog_options if option.code}
        self.catalog_by_label = {option.display_label(): option for option in self.catalog_options}
        if self.competitor_suggestions is not None:
            self.competitor_suggestions.refresh()

    def prefill_competitor(self, option: CompetitorOption) -> None:
        self.competitor_display_var.set(option.display_label())
        self.competitor_id_var.set(option.competitor_id)
        self.competitor_code_var.set(option.code)
        self.competitor_name_var.set(option.name)
        if self.competitor_suggestions is not None:
            self.competitor_suggestions.refresh()

    def _update_competitor_combo_values(self, query: str = "") -> list[str]:
        query_text = query.strip().casefold()
        filtered = [
            option.display_label()
            for option in self.catalog_options
            if not query_text or query_text in option.search_blob
        ]
        return filtered[:200]

    def _resolve_competitor(self, raw_value: str) -> CompetitorOption | None:
        value = raw_value.strip()
        if not value:
            return None
        if value in self.catalog_by_label:
            return self.catalog_by_label[value]
        if value in self.catalog_by_id:
            return self.catalog_by_id[value]
        by_code = self.catalog_by_code.get(value.casefold())
        if by_code is not None:
            return by_code

        exact_matches = [option for option in self.catalog_options if value.casefold() == option.name.casefold()]
        if len(exact_matches) == 1:
            return exact_matches[0]

        partial_matches = [option for option in self.catalog_options if value.casefold() in option.search_blob]
        if len(partial_matches) == 1:
            return partial_matches[0]
        return None

    def _apply_current_competitor_selection(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        option = self._resolve_competitor(self.competitor_display_var.get())
        if option is None:
            return
        self.prefill_competitor(option)

    def _accept_competitor_suggestion(self, value: str) -> None:
        self.competitor_display_var.set(value)
        self._apply_current_competitor_selection()

    def _on_competitor_key_release(self, _event: tk.Event[tk.Misc]) -> None:
        if self.competitor_suggestions is not None:
            self.competitor_suggestions.refresh(show=True)

    def _emit_change(self) -> None:
        if self.on_change is not None:
            self.on_change()

    def _build_current_row(self) -> ComparisonNumberRow:
        return ComparisonNumberRow(
            competitor_id=self.competitor_id_var.get().strip(),
            competitor_code=self.competitor_code_var.get().strip(),
            competitor_name=self.competitor_name_var.get().strip(),
            reference_number=self.reference_number_var.get().strip(),
        )

    def add_row(self) -> None:
        self._apply_current_competitor_selection()
        row = self._build_current_row()
        if not row.competitor_id or not row.reference_number:
            messagebox.showwarning(APP_TITLE, "Bitte Mitbewerber-ID und Vergleichsnummer eingeben.")
            return
        self.tree.insert("", "end", values=(row.competitor_id, row.competitor_code, row.competitor_name, row.reference_number))
        self._select_first_row(select_last=True)
        self._emit_change()

    def update_selected_row(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst eine Zeile zum Bearbeiten auswählen.")
            return
        self._apply_current_competitor_selection()
        row = self._build_current_row()
        if not row.competitor_id or not row.reference_number:
            messagebox.showwarning(APP_TITLE, "Bitte Mitbewerber-ID und Vergleichsnummer eingeben.")
            return
        self.tree.item(selected[0], values=(row.competitor_id, row.competitor_code, row.competitor_name, row.reference_number))
        self.tree.focus(selected[0])
        self._handle_selection()
        self._emit_change()

    def remove_selected(self) -> None:
        for item_id in self.tree.selection():
            self.tree.delete(item_id)
        self.clear_form()
        self._select_first_row()
        self._emit_change()

    def copy_selected_rows(self) -> None:
        selected_items = list(self.tree.selection())
        if not selected_items:
            return
        rows = []
        for item_id in selected_items:
            values = [str(value).strip() for value in self.tree.item(item_id, "values")]
            rows.append("\t".join(values))
        self.clipboard_clear()
        self.clipboard_append("\n".join(rows))

    def clear_form(self) -> None:
        self.competitor_display_var.set("")
        self.competitor_id_var.set("")
        self.competitor_code_var.set("")
        self.competitor_name_var.set("")
        self.reference_number_var.set("")

    def get_rows(self) -> list[ComparisonNumberRow]:
        rows = []
        for item_id in self.tree.get_children():
            competitor_id, competitor_code, competitor_name, reference_number = self.tree.item(item_id, "values")
            rows.append(
                ComparisonNumberRow(
                    competitor_id=str(competitor_id).strip(),
                    competitor_code=str(competitor_code).strip(),
                    competitor_name=str(competitor_name).strip(),
                    reference_number=str(reference_number).strip(),
                )
            )
        return normalize_comparison_number_rows(rows)

    def set_rows(self, rows: list[ComparisonNumberRow]) -> None:
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        for row in normalize_comparison_number_rows(rows):
            self.tree.insert("", "end", values=(row.competitor_id, row.competitor_code, row.competitor_name, row.reference_number))
        self._select_first_row()

    def _select_first_row(self, select_last: bool = False) -> None:
        children = list(self.tree.get_children())
        if not children:
            return
        target = children[-1] if select_last else children[0]
        self.tree.selection_set(target)
        self.tree.focus(target)
        self._handle_selection()

    def _handle_selection(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        competitor_id, competitor_code, competitor_name, reference_number = self.tree.item(selected[0], "values")
        self.competitor_id_var.set(str(competitor_id).strip())
        self.competitor_code_var.set(str(competitor_code).strip())
        self.competitor_name_var.set(str(competitor_name).strip())
        self.reference_number_var.set(str(reference_number).strip())
        display_parts = [self.competitor_id_var.get(), self.competitor_code_var.get(), self.competitor_name_var.get()]
        self.competitor_display_var.set(" | ".join(part for part in display_parts if part))

    def _open_context_menu(self, event: tk.Event[tk.Misc]) -> None:
        item_id = self.tree.identify_row(event.y)
        if item_id:
            if item_id not in self.tree.selection():
                self.tree.selection_set(item_id)
            self.tree.focus(item_id)
            self._handle_selection()
        has_selection = bool(self.tree.selection())
        self.context_menu.entryconfigure("Zeilen kopieren", state="normal" if has_selection else "disabled")
        self.context_menu.entryconfigure("Zeilen löschen", state="normal" if has_selection else "disabled")
        self.context_menu.tk_popup(event.x_root, event.y_root)
        self.context_menu.grab_release()


class VehicleLinkTableFrame(ttk.LabelFrame):
    def __init__(self, master: tk.Misc, on_change: Callable[[], None] | None = None) -> None:
        super().__init__(master, text="Fahrzeugverknüpfungen", padding=14)
        self.on_change = on_change
        self.catalog: VehicleCatalog | None = None
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Zeilen kopieren", command=self.copy_selected_rows)
        self.context_menu.add_command(label="Zeilen löschen", command=self.remove_selected)

        self.motorcode_var = tk.StringVar()
        self.vehicle_type_var = tk.StringVar(value=format_vehicle_type_choice(*VEHICLE_TYPE_CHOICES[0]))
        self.result_var = tk.StringVar(value="")

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        header = ttk.Frame(self)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header.columnconfigure(1, weight=1)

        ttk.Label(
            header,
            text=(
                "Motorcode(s) eingeben - die passenden KTyp-Nummern (TopMotive und TecDoc) werden automatisch "
                "aus den KTyp-Stammdaten ergänzt. Für den Apollo-Import: Fahrzeugtyp -> TecDoc Verknuepfungstyp ID "
                "(PKW = 2), KTypNr -> TecDoc Verknuepfungs ID. Mehrere Motorcodes mit ; trennen."
            ),
            foreground="#5E6472",
            wraplength=980,
        ).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 10))

        ttk.Label(header, text="Motorcode(s)").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        self.motorcode_entry = ttk.Entry(header, textvariable=self.motorcode_var)
        self.motorcode_entry.grid(row=1, column=1, sticky="ew", pady=4)
        self.motorcode_suggestions = SearchSuggestionPopup(
            self,
            self.motorcode_entry,
            self._update_motorcode_suggestions,
            self._accept_motorcode_suggestion,
            min_width=340,
        )
        # Enter soll die Fahrzeuge hinzufügen (überschreibt die Popup-Bindung bewusst danach).
        self.motorcode_entry.bind("<Return>", self._on_motorcode_return)

        ttk.Label(header, text="Fahrzeugtyp").grid(row=1, column=2, sticky="w", padx=(12, 8), pady=4)
        self.vehicle_type_combo = ttk.Combobox(
            header,
            textvariable=self.vehicle_type_var,
            state="readonly",
            width=14,
            values=[format_vehicle_type_choice(type_id, label) for type_id, label in VEHICLE_TYPE_CHOICES],
        )
        self.vehicle_type_combo.grid(row=1, column=3, sticky="w", pady=4)

        actions = ttk.Frame(header)
        actions.grid(row=2, column=0, columnspan=4, sticky="w", pady=(6, 0))
        ttk.Button(actions, text="Fahrzeuge hinzufügen", command=self.add_from_motorcodes).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="Auswahl entfernen", command=self.remove_selected).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(actions, text="Alle entfernen", command=self.clear_rows).grid(row=0, column=2, padx=(0, 8))
        ttk.Label(actions, textvariable=self.result_var, foreground="#5E6472", wraplength=520).grid(row=0, column=3, padx=(8, 0), sticky="w")

        columns = (
            "vehicle_type_id",
            "vehicle_type",
            "motorcode",
            "manufacturer",
            "model",
            "description",
            "year",
            "power",
            "ktyp_topmotive",
            "ktyp_tecdoc",
        )
        display_columns = (
            "vehicle_type",
            "motorcode",
            "manufacturer",
            "model",
            "description",
            "year",
            "power",
            "ktyp_topmotive",
            "ktyp_tecdoc",
        )
        self.tree = ttk.Treeview(
            self,
            columns=columns,
            show="headings",
            height=13,
            selectmode="extended",
            displaycolumns=display_columns,
        )
        self.tree.grid(row=1, column=0, sticky="nsew")
        headings = {
            "vehicle_type": "Fahrzeugtyp",
            "motorcode": "Motorcode",
            "manufacturer": "Hersteller",
            "model": "Modell",
            "description": "Bezeichnung",
            "year": "Bauzeit",
            "power": "Leistung",
            "ktyp_topmotive": "KTyp TopMotive",
            "ktyp_tecdoc": "KTyp TecDoc",
        }
        widths = {
            "vehicle_type": 90,
            "motorcode": 110,
            "manufacturer": 120,
            "model": 210,
            "description": 150,
            "year": 120,
            "power": 120,
            "ktyp_topmotive": 120,
            "ktyp_tecdoc": 110,
        }
        for column, text in headings.items():
            self.tree.heading(column, text=text)
            self.tree.column(column, width=widths[column], anchor="w")
        self.tree.column("vehicle_type", anchor="center")
        self.tree.column("ktyp_topmotive", anchor="center")
        self.tree.column("ktyp_tecdoc", anchor="center")
        self.tree.bind("<Button-3>", self._open_context_menu)

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

    def set_vehicle_catalog(self, catalog: VehicleCatalog | None) -> None:
        self.catalog = catalog
        if self.motorcode_suggestions is not None:
            self.motorcode_suggestions.hide()

    @staticmethod
    def _split_last_motorcode_token(text: str) -> tuple[str, str]:
        separator_index = text.rfind(";")
        if separator_index >= 0:
            return text[: separator_index + 1], text[separator_index + 1 :]
        return "", text

    def _update_motorcode_suggestions(self, query: str = "") -> list[str]:
        if self.catalog is None or self.catalog.count == 0:
            return []
        _prefix, token = self._split_last_motorcode_token(query)
        return self.catalog.suggest(token.strip(), limit=12)

    def _accept_motorcode_suggestion(self, value: str) -> None:
        prefix, _token = self._split_last_motorcode_token(self.motorcode_var.get())
        new_value = f"{prefix} {value}".strip() if prefix else value
        self.motorcode_var.set(new_value)
        self.motorcode_entry.icursor("end")

    def _on_motorcode_return(self, _event: tk.Event[tk.Misc]) -> str:
        if self.motorcode_suggestions is not None:
            self.motorcode_suggestions.hide()
        self.add_from_motorcodes()
        return "break"

    def _current_vehicle_type(self) -> tuple[str, str]:
        text = self.vehicle_type_var.get().strip()
        match = re.search(r"\((\d+)\)\s*$", text)
        if match:
            type_id = match.group(1)
            label = text[: match.start()].strip() or vehicle_type_label_for_id(type_id)
            return type_id, label
        type_id = vehicle_type_id_for_label(text) or (text if text.isdigit() else DEFAULT_VEHICLE_TYPE_ID)
        return type_id, vehicle_type_label_for_id(type_id) or text

    def _emit_change(self) -> None:
        if self.on_change is not None:
            self.on_change()

    def add_from_motorcodes(self) -> None:
        raw = self.motorcode_var.get().strip()
        if not raw:
            messagebox.showwarning(APP_TITLE, "Bitte mindestens einen Motorcode eingeben.")
            return
        if self.catalog is None or self.catalog.count == 0:
            messagebox.showwarning(
                APP_TITLE,
                "Keine KTyp-Stammdaten geladen. Bitte die KTyp-Datei im Projekt-Tab unter 'Datenstämme' laden.",
            )
            return
        codes = split_code_list_input(raw)
        type_id, type_label = self._current_vehicle_type()
        existing = {(row.vehicle_type_id, row.ktyp_topmotive, row.ktyp_tecdoc) for row in self.get_rows()}
        added = 0
        matched_codes = 0
        not_found: list[str] = []
        for code in codes:
            matches = self.catalog.lookup(code)
            if not matches:
                not_found.append(code)
                continue
            matched_codes += 1
            for match in matches:
                key = (type_id, match.ktyp_topmotive, match.ktyp_tecdoc)
                if key in existing:
                    continue
                existing.add(key)
                self.tree.insert(
                    "",
                    "end",
                    values=(
                        type_id,
                        format_vehicle_type_choice(type_id, type_label),
                        match.motorcode or normalize_motorcode(code),
                        match.manufacturer,
                        match.model,
                        match.description,
                        format_year_range(match.year_from, match.year_to),
                        match.power,
                        match.ktyp_topmotive,
                        match.ktyp_tecdoc,
                    ),
                )
                added += 1
        summary = f"{added} Fahrzeug(e) aus {matched_codes} Motorcode(s) hinzugefügt."
        if not_found:
            shown = ", ".join(not_found[:8]) + (" ..." if len(not_found) > 8 else "")
            summary += f" Ohne Treffer: {shown}"
        if added == 0 and not not_found:
            summary = "Alle gefundenen Fahrzeuge sind bereits vorhanden."
        self.result_var.set(summary)
        if added:
            self.motorcode_var.set("")
            self._select_last_row()
            self._emit_change()

    def remove_selected(self) -> None:
        selection = self.tree.selection()
        if not selection:
            return
        for item_id in selection:
            self.tree.delete(item_id)
        self._emit_change()

    def clear_rows(self) -> None:
        if not self.tree.get_children():
            return
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        self.result_var.set("")
        self._emit_change()

    def copy_selected_rows(self) -> None:
        selected = list(self.tree.selection())
        if not selected:
            return
        rows = []
        for item_id in selected:
            values = [str(value).strip() for value in self.tree.item(item_id, "values")[1:]]
            rows.append("\t".join(values))
        self.clipboard_clear()
        self.clipboard_append("\n".join(rows))

    def _select_last_row(self) -> None:
        children = list(self.tree.get_children())
        if children:
            self.tree.selection_set(children[-1])
            self.tree.see(children[-1])

    def get_rows(self) -> list[VehicleLinkRow]:
        rows: list[VehicleLinkRow] = []
        for item_id in self.tree.get_children():
            values = list(self.tree.item(item_id, "values"))
            values += [""] * (10 - len(values))
            vehicle_type_id, _display, motorcode, manufacturer, model, description, year, power, ktyp_tm, ktyp_td = values[:10]
            year_text = str(year).strip()
            if " - " in year_text:
                year_from, year_to = [part.strip() for part in year_text.split(" - ", 1)]
            else:
                year_from, year_to = year_text, ""
            type_id = str(vehicle_type_id).strip() or DEFAULT_VEHICLE_TYPE_ID
            rows.append(
                VehicleLinkRow(
                    vehicle_type_id=type_id,
                    vehicle_type_label=vehicle_type_label_for_id(type_id),
                    motorcode=str(motorcode).strip(),
                    ktyp_topmotive=str(ktyp_tm).strip(),
                    ktyp_tecdoc=str(ktyp_td).strip(),
                    manufacturer=str(manufacturer).strip(),
                    model=str(model).strip(),
                    description=str(description).strip(),
                    year_from=year_from,
                    year_to=year_to,
                    power=str(power).strip(),
                )
            )
        return normalize_vehicle_link_rows(rows)

    def set_rows(self, rows: list[VehicleLinkRow]) -> None:
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        for row in normalize_vehicle_link_rows(rows):
            self.tree.insert(
                "",
                "end",
                values=(
                    row.vehicle_type_id,
                    format_vehicle_type_choice(row.vehicle_type_id, row.vehicle_type_label),
                    row.motorcode,
                    row.manufacturer,
                    row.model,
                    row.description,
                    format_year_range(row.year_from, row.year_to),
                    row.power,
                    row.ktyp_topmotive,
                    row.ktyp_tecdoc,
                ),
            )

    def _open_context_menu(self, event: tk.Event[tk.Misc]) -> None:
        item_id = self.tree.identify_row(event.y)
        if item_id and item_id not in self.tree.selection():
            self.tree.selection_set(item_id)
        has_selection = bool(self.tree.selection())
        self.context_menu.entryconfigure("Zeilen kopieren", state="normal" if has_selection else "disabled")
        self.context_menu.entryconfigure("Zeilen löschen", state="normal" if has_selection else "disabled")
        self.context_menu.tk_popup(event.x_root, event.y_root)
        self.context_menu.grab_release()


class SearchTermFrame(ttk.LabelFrame):
    def __init__(self, master: tk.Misc, on_change: Callable[[], None] | None = None) -> None:
        super().__init__(master, text="Suchwörter", padding=14)
        self.on_change = on_change
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Suchwörter kopieren", command=self.copy_selected_terms)
        self.context_menu.add_command(label="Suchwörter löschen", command=self.remove_selected)

        self.term_var = tk.StringVar()
        self.count_var = tk.StringVar(value="0 Suchwörter")

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        header = ttk.Frame(self)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header.columnconfigure(1, weight=1)

        ttk.Label(
            header,
            text=(
                "Apollo kennt keine nativen Suchbegriffe. Suchwörter werden deshalb beim Export als Attribut "
                f"'{SEARCH_TERM_ATTRIBUTE_LABEL}' (TecDoc Kriterien ID {SEARCH_TERM_CRITERIA_ID}) in die Attribute.xlsx geschrieben - "
                f"eine Zeile pro Suchwort, maximal {SEARCH_TERM_MAX_LENGTH} Zeichen. Mehrere Suchwörter mit ; trennen."
            ),
            foreground="#5E6472",
            wraplength=980,
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 10))

        ttk.Label(header, text="Suchwort(e)").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        self.term_entry = ttk.Entry(header, textvariable=self.term_var)
        self.term_entry.grid(row=1, column=1, sticky="ew", pady=4)
        self.term_entry.bind("<Return>", lambda _event: self.add_terms())

        actions = ttk.Frame(header)
        actions.grid(row=2, column=0, columnspan=3, sticky="w", pady=(6, 0))
        ttk.Button(actions, text="Hinzufügen", command=self.add_terms).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="Auswahl entfernen", command=self.remove_selected).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(actions, text="Alle entfernen", command=self.clear_terms).grid(row=0, column=2, padx=(0, 8))
        ttk.Label(actions, textvariable=self.count_var, foreground="#5E6472").grid(row=0, column=3, padx=(8, 0), sticky="w")

        columns = ("term", "length")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=13, selectmode="extended")
        self.tree.grid(row=1, column=0, sticky="nsew")
        self.tree.heading("term", text="Suchwort")
        self.tree.heading("length", text="Zeichen")
        self.tree.column("term", width=420, anchor="w")
        self.tree.column("length", width=90, anchor="center")
        self.tree.bind("<Button-3>", self._open_context_menu)

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

    def _emit_change(self) -> None:
        if self.on_change is not None:
            self.on_change()

    def _update_count(self) -> None:
        count = len(self.tree.get_children())
        self.count_var.set(f"{count} Suchwörter" if count != 1 else "1 Suchwort")

    def add_terms(self) -> None:
        raw = self.term_var.get().strip()
        if not raw:
            messagebox.showwarning(APP_TITLE, "Bitte mindestens ein Suchwort eingeben.")
            return
        candidates = normalize_search_terms(split_code_list_input(raw))
        if not candidates:
            return
        too_long = [term for term in candidates if len(term) > SEARCH_TERM_MAX_LENGTH]
        candidates = [term for term in candidates if len(term) <= SEARCH_TERM_MAX_LENGTH]
        existing = {term.casefold() for term in self.get_terms()}
        added = 0
        for term in candidates:
            if term.casefold() in existing:
                continue
            existing.add(term.casefold())
            self.tree.insert("", "end", values=(term, len(term)))
            added += 1
        self._update_count()
        if too_long:
            messagebox.showwarning(
                APP_TITLE,
                f"Diese Suchwörter sind länger als {SEARCH_TERM_MAX_LENGTH} Zeichen und wurden nicht übernommen:\n"
                + "\n".join(too_long),
            )
        if added:
            self.term_var.set("")
            self._emit_change()

    def remove_selected(self) -> None:
        selection = self.tree.selection()
        if not selection:
            return
        for item_id in selection:
            self.tree.delete(item_id)
        self._update_count()
        self._emit_change()

    def clear_terms(self) -> None:
        if not self.tree.get_children():
            return
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        self._update_count()
        self._emit_change()

    def copy_selected_terms(self) -> None:
        selected = list(self.tree.selection())
        if not selected:
            return
        terms = [str(self.tree.item(item_id, "values")[0]) for item_id in selected]
        self.clipboard_clear()
        self.clipboard_append("\n".join(terms))

    def get_terms(self) -> list[str]:
        terms = [str(self.tree.item(item_id, "values")[0]) for item_id in self.tree.get_children()]
        return normalize_search_terms(terms)

    def set_terms(self, terms: list[str]) -> None:
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        for term in normalize_search_terms(terms):
            self.tree.insert("", "end", values=(term, len(term)))
        self._update_count()

    def _open_context_menu(self, event: tk.Event[tk.Misc]) -> None:
        item_id = self.tree.identify_row(event.y)
        if item_id and item_id not in self.tree.selection():
            self.tree.selection_set(item_id)
        has_selection = bool(self.tree.selection())
        self.context_menu.entryconfigure("Suchwörter kopieren", state="normal" if has_selection else "disabled")
        self.context_menu.entryconfigure("Suchwörter löschen", state="normal" if has_selection else "disabled")
        self.context_menu.tk_popup(event.x_root, event.y_root)
        self.context_menu.grab_release()


class AttributeTableFrame(ttk.LabelFrame):
    def __init__(self, master: tk.Misc, on_change: Callable[[], None] | None = None) -> None:
        super().__init__(master, text="Attribute", padding=14)
        self.on_change = on_change
        self.catalog_options: list[AttributeOption] = []
        self.catalog_by_id: dict[str, AttributeOption] = {}
        self.catalog_by_label: dict[str, AttributeOption] = {}
        self.key_value_options_by_group: dict[str, list[AttributeKeyValueOption]] = {}
        self.row_type_names: dict[str, str] = {}
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Zeilen kopieren", command=self.copy_selected_rows)
        self.context_menu.add_command(label="Zeilen löschen", command=self.remove_selected)

        self.attribute_display_var = tk.StringVar()
        self.criteria_id_var = tk.StringVar()
        self.label_var = tk.StringVar()
        self.format_var = tk.StringVar()
        self.max_length_var = tk.StringVar()
        self.type_name_var = tk.StringVar()
        self.value_var = tk.StringVar()
        self.value_to_var = tk.StringVar()
        self.hint_var = tk.StringVar(value="Wähle ein Attribut aus oder gib die Kriterien-ID manuell ein.")
        self.attribute_suggestion_popup: tk.Toplevel | None = None
        self.attribute_suggestion_listbox: tk.Listbox | None = None
        self.attribute_suggestion_values: list[str] = []
        self.value_suggestion_popup: tk.Toplevel | None = None
        self.value_suggestion_listbox: tk.Listbox | None = None
        self.value_suggestion_values: list[str] = []

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        header = ttk.Frame(self)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header.columnconfigure(1, weight=1)
        header.columnconfigure(5, weight=1)

        ttk.Label(
            header,
            text="Attribute werden über die TecDoc Kriterien ID ausgewählt. Format und maximale Länge dienen als Pflegehilfe.",
            foreground="#5E6472",
            wraplength=980,
        ).grid(row=0, column=0, columnspan=9, sticky="w", pady=(0, 10))

        ttk.Label(header, text="Attribut").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        self.attribute_entry = ttk.Entry(header, textvariable=self.attribute_display_var)
        self.attribute_entry.grid(row=1, column=1, columnspan=8, sticky="ew", pady=4)
        self.attribute_entry.bind("<KeyRelease>", self._on_attribute_key_release)
        self.attribute_entry.bind("<Down>", self._open_attribute_dropdown_event)
        self.attribute_entry.bind("<F4>", self._open_attribute_dropdown_event)
        self.attribute_entry.bind("<Escape>", self._close_attribute_dropdown_event)
        self.attribute_entry.bind("<Return>", self._handle_attribute_return)
        self.attribute_entry.bind("<FocusOut>", self._handle_attribute_focus_out)

        ttk.Label(header, text="Kriterien ID").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(header, textvariable=self.criteria_id_var, width=16).grid(row=2, column=1, sticky="w", pady=4)
        ttk.Label(header, text="Bezeichnung").grid(row=2, column=2, sticky="w", padx=(12, 8), pady=4)
        ttk.Entry(header, textvariable=self.label_var).grid(row=2, column=3, columnspan=6, sticky="ew", pady=4)

        ttk.Label(header, text="Format").grid(row=3, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(header, textvariable=self.format_var, width=24).grid(row=3, column=1, sticky="w", pady=4)
        ttk.Label(header, text="Max. Länge").grid(row=3, column=2, sticky="w", padx=(12, 8), pady=4)
        ttk.Entry(header, textvariable=self.max_length_var, width=12).grid(row=3, column=3, sticky="w", pady=4)
        ttk.Label(header, text="Wert").grid(row=3, column=4, sticky="w", padx=(12, 8), pady=4)
        self.value_entry = ttk.Entry(header, textvariable=self.value_var)
        self.value_entry.grid(row=3, column=5, sticky="ew", pady=4)
        self.value_entry.bind("<KeyRelease>", self._on_value_key_release)
        self.value_entry.bind("<Down>", self._open_value_dropdown_event)
        self.value_entry.bind("<F4>", self._open_value_dropdown_event)
        self.value_entry.bind("<Escape>", self._close_value_dropdown_event)
        self.value_entry.bind("<Return>", self._handle_value_return)
        self.value_entry.bind("<FocusOut>", self._handle_value_focus_out)

        ttk.Label(header, text="Wert bis").grid(row=3, column=6, sticky="w", padx=(12, 8), pady=4)
        ttk.Entry(header, textvariable=self.value_to_var).grid(row=3, column=7, sticky="ew", pady=4)

        quick_row = ttk.Frame(header)
        quick_row.grid(row=3, column=8, sticky="e", padx=(12, 0), pady=4)
        ttk.Button(quick_row, text="Ja", command=lambda: self.value_var.set("Ja")).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(quick_row, text="Nein", command=lambda: self.value_var.set("Nein")).grid(row=0, column=1)

        ttk.Label(header, textvariable=self.hint_var, foreground="#5E6472", wraplength=980).grid(
            row=4,
            column=0,
            columnspan=9,
            sticky="w",
            pady=(6, 8),
        )

        actions = ttk.Frame(header)
        actions.grid(row=5, column=0, columnspan=9, sticky="w")
        ttk.Button(actions, text="Hinzufügen", command=self.add_row).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="Auswahl aktualisieren", command=self.update_selected_row).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(actions, text="Auswahl entfernen", command=self.remove_selected).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(actions, text="Leeren", command=self.clear_form).grid(row=0, column=3)

        columns = ("criteria_id", "label", "value_format", "value", "value_to")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=11, selectmode="extended")
        self.tree.grid(row=1, column=0, sticky="nsew")
        self.tree.heading("criteria_id", text="Kriterien ID")
        self.tree.heading("label", text="Attribut")
        self.tree.heading("value_format", text="Format")
        self.tree.heading("value", text="Wert")
        self.tree.heading("value_to", text="Wert bis")
        self.tree.column("criteria_id", width=120, anchor="center")
        self.tree.column("label", width=390, anchor="w")
        self.tree.column("value_format", width=160, anchor="center")
        self.tree.column("value", width=240, anchor="w")
        self.tree.column("value_to", width=160, anchor="w")
        self.tree.bind("<<TreeviewSelect>>", self._handle_selection)
        self.tree.bind("<Button-3>", self._open_context_menu)

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

    def set_attribute_catalog(self, options: list[AttributeOption]) -> None:
        self.catalog_options = list(options)
        self.catalog_by_id = {option.criteria_id: option for option in self.catalog_options}
        self.catalog_by_label = {option.display_label(): option for option in self.catalog_options}
        self._update_attribute_combo_values(self.attribute_display_var.get())
        self._sync_current_attribute_metadata()

    def set_attribute_key_value_catalog(
        self,
        options_by_group: dict[str, list[AttributeKeyValueOption]],
    ) -> None:
        self.key_value_options_by_group = {key: list(values) for key, values in options_by_group.items()}
        self._normalize_current_value_selection()

    def _update_attribute_combo_values(self, query: str = "") -> list[str]:
        query_text = query.strip().casefold()
        if not query_text:
            filtered_options = self.catalog_options[:120]
        else:
            ranked_matches = [
                (score_attribute_option_match(option, query_text), option)
                for option in self.catalog_options
                if all(token in option.search_blob for token in split_search_terms(query_text))
            ]
            ranked_matches = [entry for entry in ranked_matches if entry[0] > 0]
            ranked_matches.sort(
                key=lambda entry: (
                    -entry[0],
                    entry[1].label.casefold(),
                    entry[1].criteria_id.casefold(),
                )
            )
            filtered_options = [option for _score, option in ranked_matches[:150]]

        self.attribute_suggestion_values = [option.display_label() for option in filtered_options]
        return list(self.attribute_suggestion_values)

    def _resolve_attribute_option(self, raw_value: str) -> AttributeOption | None:
        value = raw_value.strip()
        if not value:
            return None
        if value in self.catalog_by_label:
            return self.catalog_by_label[value]
        if value in self.catalog_by_id:
            return self.catalog_by_id[value]
        exact_matches = [option for option in self.catalog_options if value.casefold() == option.label.casefold()]
        if len(exact_matches) == 1:
            return exact_matches[0]
        partial_matches = [option for option in self.catalog_options if value.casefold() in option.search_blob]
        if len(partial_matches) == 1:
            return partial_matches[0]
        return None

    def _sync_current_attribute_metadata(self) -> None:
        option = self.catalog_by_id.get(self.criteria_id_var.get().strip())
        if option is None:
            return
        self.label_var.set(option.label)
        self.format_var.set(option.value_format)
        self.max_length_var.set("" if option.max_length is None else str(option.max_length))
        self.type_name_var.set(option.type_name)
        if not self.attribute_display_var.get().strip():
            self.attribute_display_var.set(option.display_label())
        self._normalize_current_value_selection()
        self._update_hint()

    def _apply_current_attribute_selection(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        option = self._resolve_attribute_option(self.attribute_display_var.get())
        if option is None:
            self.type_name_var.set("")
            self._hide_value_suggestions()
            self._update_hint()
            return
        self.attribute_display_var.set(option.display_label())
        self.criteria_id_var.set(option.criteria_id)
        self.label_var.set(option.label)
        self.format_var.set(option.value_format)
        self.max_length_var.set("" if option.max_length is None else str(option.max_length))
        self.type_name_var.set(option.type_name)
        self._update_attribute_combo_values(option.display_label())
        self._hide_attribute_suggestions()
        self._normalize_current_value_selection()
        self._update_hint()

    def _ensure_attribute_suggestion_popup(self) -> None:
        if self.attribute_suggestion_popup is not None and self.attribute_suggestion_popup.winfo_exists():
            return

        popup = tk.Toplevel(self)
        popup.withdraw()
        popup.overrideredirect(True)
        popup.transient(self.winfo_toplevel())
        popup.configure(background=UI_BORDER, padx=1, pady=1)

        listbox = tk.Listbox(
            popup,
            activestyle="none",
            exportselection=False,
            borderwidth=0,
            highlightthickness=0,
            font=("Segoe UI", 10),
        )
        listbox.pack(fill="both", expand=True)
        listbox.bind("<ButtonRelease-1>", self._accept_attribute_listbox_click)
        listbox.bind("<Double-Button-1>", self._accept_attribute_listbox_click)
        listbox.bind("<Return>", self._accept_attribute_listbox_keyboard)
        listbox.bind("<Escape>", self._close_attribute_dropdown_event)
        listbox.bind("<FocusOut>", self._handle_attribute_popup_focus_out)

        self.attribute_suggestion_popup = popup
        self.attribute_suggestion_listbox = listbox

    def _show_attribute_suggestions(self, values: list[str]) -> None:
        if not values:
            self._hide_attribute_suggestions()
            return

        self._ensure_attribute_suggestion_popup()
        if self.attribute_suggestion_popup is None or self.attribute_suggestion_listbox is None:
            return

        listbox = self.attribute_suggestion_listbox
        listbox.delete(0, "end")
        for value in values:
            listbox.insert("end", value)

        visible_rows = min(max(len(values), 1), 8)
        listbox.configure(height=visible_rows)
        listbox.selection_clear(0, "end")
        listbox.selection_set(0)
        listbox.activate(0)
        listbox.see(0)

        self.attribute_suggestion_popup.update_idletasks()
        x_pos = self.attribute_entry.winfo_rootx()
        y_pos = self.attribute_entry.winfo_rooty() + self.attribute_entry.winfo_height()
        width = max(self.attribute_entry.winfo_width(), 420)
        height = self.attribute_suggestion_popup.winfo_reqheight()
        self.attribute_suggestion_popup.geometry(f"{width}x{height}+{x_pos}+{y_pos}")
        self.attribute_suggestion_popup.deiconify()
        self.attribute_suggestion_popup.lift()

    def _hide_attribute_suggestions(self) -> None:
        if self.attribute_suggestion_popup is None or not self.attribute_suggestion_popup.winfo_exists():
            self.attribute_suggestion_popup = None
            self.attribute_suggestion_listbox = None
            return
        self.attribute_suggestion_popup.withdraw()

    def _is_attribute_popup_widget(self, widget: tk.Misc | None) -> bool:
        if widget is None:
            return False
        if widget is self.attribute_entry:
            return True
        if self.attribute_suggestion_listbox is not None and widget is self.attribute_suggestion_listbox:
            return True
        if self.attribute_suggestion_popup is not None and widget is self.attribute_suggestion_popup:
            return True
        master = getattr(widget, "master", None)
        while master is not None:
            if self.attribute_suggestion_popup is not None and master is self.attribute_suggestion_popup:
                return True
            master = getattr(master, "master", None)
        return False

    def _selected_attribute_suggestion_value(self) -> str | None:
        if self.attribute_suggestion_listbox is None or not self.attribute_suggestion_listbox.winfo_exists():
            return None
        selection = self.attribute_suggestion_listbox.curselection()
        if selection:
            return str(self.attribute_suggestion_listbox.get(selection[0]))
        if self.attribute_suggestion_values:
            return self.attribute_suggestion_values[0]
        return None

    def _accept_attribute_suggestion(self, value: str | None) -> None:
        if not value:
            self._hide_attribute_suggestions()
            return
        self.attribute_display_var.set(value)
        self._apply_current_attribute_selection()
        self.attribute_entry.focus_set()
        self.attribute_entry.icursor("end")

    def _accept_attribute_listbox_click(self, _event: tk.Event[tk.Misc]) -> str:
        self._accept_attribute_suggestion(self._selected_attribute_suggestion_value())
        return "break"

    def _accept_attribute_listbox_keyboard(self, _event: tk.Event[tk.Misc]) -> str:
        self._accept_attribute_suggestion(self._selected_attribute_suggestion_value())
        return "break"

    def _open_attribute_dropdown_event(self, _event: tk.Event[tk.Misc]) -> str:
        values = self._update_attribute_combo_values(self.attribute_display_var.get())
        if values:
            self._show_attribute_suggestions(values)
            if self.attribute_suggestion_listbox is not None:
                self.attribute_suggestion_listbox.focus_set()
        else:
            self._hide_attribute_suggestions()
        return "break"

    def _close_attribute_dropdown_event(self, _event: tk.Event[tk.Misc]) -> str:
        self._hide_attribute_suggestions()
        self.attribute_entry.focus_set()
        return "break"

    def _handle_attribute_return(self, _event: tk.Event[tk.Misc]) -> str:
        suggestion = self._selected_attribute_suggestion_value()
        if suggestion is not None:
            self._accept_attribute_suggestion(suggestion)
        else:
            self._apply_current_attribute_selection()
        return "break"

    def _handle_attribute_focus_out(self, _event: tk.Event[tk.Misc]) -> None:
        self.after(120, self._finalize_attribute_focus_out)

    def _handle_attribute_popup_focus_out(self, _event: tk.Event[tk.Misc]) -> None:
        self.after(120, self._finalize_attribute_focus_out)

    def _finalize_attribute_focus_out(self) -> None:
        focus_widget = self.focus_get()
        if self._is_attribute_popup_widget(focus_widget):
            return
        self._hide_attribute_suggestions()
        self._apply_current_attribute_selection()

    def _on_attribute_key_release(self, event: tk.Event[tk.Misc]) -> None:
        ignored_keys = {
            "Up",
            "Down",
            "Left",
            "Right",
            "Return",
            "Escape",
            "Tab",
            "Shift_L",
            "Shift_R",
            "Control_L",
            "Control_R",
            "Alt_L",
            "Alt_R",
        }
        if event.keysym in ignored_keys:
            return

        values = self._update_attribute_combo_values(self.attribute_display_var.get())
        if self.attribute_display_var.get().strip():
            self._show_attribute_suggestions(values)
        else:
            self._hide_attribute_suggestions()

    def _current_attribute_row_context(self) -> AttributeRow:
        return AttributeRow(
            criteria_id=self.criteria_id_var.get().strip(),
            label=self.label_var.get().strip(),
            value_format=self.format_var.get().strip(),
            max_length=self._current_max_length(),
            type_name=self.type_name_var.get().strip(),
            value=self.value_var.get().strip(),
            value_to=self.value_to_var.get().strip(),
        )

    def _current_key_value_options(self) -> list[AttributeKeyValueOption]:
        row = self._current_attribute_row_context()
        if not is_attribute_key_value_format(row.value_format):
            return []
        for group_key in row.key_value_group_candidates():
            options = self.key_value_options_by_group.get(group_key, [])
            if options:
                return options
        return []

    def _resolve_value_option(self, raw_value: str) -> AttributeKeyValueOption | None:
        row = self._current_attribute_row_context()
        if not is_attribute_key_value_format(row.value_format):
            return None
        return resolve_attribute_key_value_option(row, raw_value, self.key_value_options_by_group)

    def _normalize_current_value_selection(self) -> None:
        if not is_attribute_key_value_format(self.format_var.get()):
            self._hide_value_suggestions()
            return
        raw_value = self.value_var.get().strip()
        if not raw_value:
            return
        option = self._resolve_value_option(raw_value)
        if option is not None and option.display_label() != raw_value:
            self.value_var.set(option.display_label())

    def _update_value_suggestion_values(self, query: str = "") -> list[str]:
        group_options = self._current_key_value_options()
        if not group_options:
            self.value_suggestion_values = []
            return []

        query_text = normalize_lookup_text(query)
        if not query_text:
            filtered_options = group_options[:120]
        else:
            tokens = split_search_terms(query_text)
            ranked_matches = []
            for option in group_options:
                normalized_blob = " ".join(
                    part
                    for part in [
                        normalize_lookup_text(option.key_value_id),
                        normalize_lookup_text(option.label),
                        normalize_lookup_text(option.display_label()),
                    ]
                    if part
                )
                if tokens and not all(token in normalized_blob for token in tokens):
                    continue
                score = score_attribute_key_value_match(option, query_text)
                if score > 0:
                    ranked_matches.append((score, option))
            ranked_matches.sort(
                key=lambda entry: (
                    -entry[0],
                    entry[1].label.casefold(),
                    entry[1].key_value_id.casefold(),
                )
            )
            filtered_options = [option for _score, option in ranked_matches[:150]]

        self.value_suggestion_values = [option.display_label() for option in filtered_options]
        return list(self.value_suggestion_values)

    def _ensure_value_suggestion_popup(self) -> None:
        if self.value_suggestion_popup is not None and self.value_suggestion_popup.winfo_exists():
            return

        popup = tk.Toplevel(self)
        popup.withdraw()
        popup.overrideredirect(True)
        popup.transient(self.winfo_toplevel())
        popup.configure(background=UI_BORDER, padx=1, pady=1)

        listbox = tk.Listbox(
            popup,
            activestyle="none",
            exportselection=False,
            borderwidth=0,
            highlightthickness=0,
            font=("Segoe UI", 10),
        )
        listbox.pack(fill="both", expand=True)
        listbox.bind("<ButtonRelease-1>", self._accept_value_listbox_click)
        listbox.bind("<Double-Button-1>", self._accept_value_listbox_click)
        listbox.bind("<Return>", self._accept_value_listbox_keyboard)
        listbox.bind("<Escape>", self._close_value_dropdown_event)
        listbox.bind("<FocusOut>", self._handle_value_popup_focus_out)

        self.value_suggestion_popup = popup
        self.value_suggestion_listbox = listbox

    def _show_value_suggestions(self, values: list[str]) -> None:
        if not values:
            self._hide_value_suggestions()
            return

        self._ensure_value_suggestion_popup()
        if self.value_suggestion_popup is None or self.value_suggestion_listbox is None:
            return

        listbox = self.value_suggestion_listbox
        listbox.delete(0, "end")
        for value in values:
            listbox.insert("end", value)

        visible_rows = min(max(len(values), 1), 8)
        listbox.configure(height=visible_rows)
        listbox.selection_clear(0, "end")
        listbox.selection_set(0)
        listbox.activate(0)
        listbox.see(0)

        self.value_suggestion_popup.update_idletasks()
        x_pos = self.value_entry.winfo_rootx()
        y_pos = self.value_entry.winfo_rooty() + self.value_entry.winfo_height()
        width = max(self.value_entry.winfo_width(), 360)
        height = self.value_suggestion_popup.winfo_reqheight()
        self.value_suggestion_popup.geometry(f"{width}x{height}+{x_pos}+{y_pos}")
        self.value_suggestion_popup.deiconify()
        self.value_suggestion_popup.lift()

    def _hide_value_suggestions(self) -> None:
        if self.value_suggestion_popup is None or not self.value_suggestion_popup.winfo_exists():
            self.value_suggestion_popup = None
            self.value_suggestion_listbox = None
            return
        self.value_suggestion_popup.withdraw()

    def _is_value_popup_widget(self, widget: tk.Misc | None) -> bool:
        if widget is None:
            return False
        if widget is self.value_entry:
            return True
        if self.value_suggestion_listbox is not None and widget is self.value_suggestion_listbox:
            return True
        if self.value_suggestion_popup is not None and widget is self.value_suggestion_popup:
            return True
        master = getattr(widget, "master", None)
        while master is not None:
            if self.value_suggestion_popup is not None and master is self.value_suggestion_popup:
                return True
            master = getattr(master, "master", None)
        return False

    def _selected_value_suggestion_value(self) -> str | None:
        if self.value_suggestion_listbox is None or not self.value_suggestion_listbox.winfo_exists():
            return None
        selection = self.value_suggestion_listbox.curselection()
        if selection:
            return str(self.value_suggestion_listbox.get(selection[0]))
        if self.value_suggestion_values:
            return self.value_suggestion_values[0]
        return None

    def _accept_value_suggestion(self, value: str | None) -> None:
        if not value:
            self._hide_value_suggestions()
            return
        self.value_var.set(value)
        self._normalize_current_value_selection()
        self._hide_value_suggestions()
        self.value_entry.focus_set()
        self.value_entry.icursor("end")

    def _accept_value_listbox_click(self, _event: tk.Event[tk.Misc]) -> str:
        self._accept_value_suggestion(self._selected_value_suggestion_value())
        return "break"

    def _accept_value_listbox_keyboard(self, _event: tk.Event[tk.Misc]) -> str:
        self._accept_value_suggestion(self._selected_value_suggestion_value())
        return "break"

    def _open_value_dropdown_event(self, _event: tk.Event[tk.Misc]) -> str:
        values = self._update_value_suggestion_values(self.value_var.get())
        if values:
            self._show_value_suggestions(values)
            if self.value_suggestion_listbox is not None:
                self.value_suggestion_listbox.focus_set()
        else:
            self._hide_value_suggestions()
        return "break"

    def _close_value_dropdown_event(self, _event: tk.Event[tk.Misc]) -> str:
        self._hide_value_suggestions()
        self.value_entry.focus_set()
        return "break"

    def _handle_value_return(self, _event: tk.Event[tk.Misc]) -> str:
        suggestion = self._selected_value_suggestion_value()
        if suggestion is not None:
            self._accept_value_suggestion(suggestion)
        else:
            self._normalize_current_value_selection()
        return "break"

    def _handle_value_focus_out(self, _event: tk.Event[tk.Misc]) -> None:
        self.after(120, self._finalize_value_focus_out)

    def _handle_value_popup_focus_out(self, _event: tk.Event[tk.Misc]) -> None:
        self.after(120, self._finalize_value_focus_out)

    def _finalize_value_focus_out(self) -> None:
        focus_widget = self.focus_get()
        if self._is_value_popup_widget(focus_widget):
            return
        self._hide_value_suggestions()
        self._normalize_current_value_selection()

    def _on_value_key_release(self, event: tk.Event[tk.Misc]) -> None:
        ignored_keys = {
            "Up",
            "Down",
            "Left",
            "Right",
            "Return",
            "Escape",
            "Tab",
            "Shift_L",
            "Shift_R",
            "Control_L",
            "Control_R",
            "Alt_L",
            "Alt_R",
        }
        if event.keysym in ignored_keys:
            return
        if not is_attribute_key_value_format(self.format_var.get()):
            self._hide_value_suggestions()
            return

        values = self._update_value_suggestion_values(self.value_var.get())
        if self.value_var.get().strip():
            self._show_value_suggestions(values)
        else:
            self._hide_value_suggestions()

    def _emit_change(self) -> None:
        if self.on_change is not None:
            self.on_change()

    def _current_max_length(self) -> int | None:
        raw_value = self.max_length_var.get().strip()
        if not raw_value:
            return None
        try:
            return int(raw_value)
        except ValueError:
            return None

    def _normalize_current_value(self, value_format: str, value: str) -> str:
        normalized = value.strip()
        format_key = value_format.strip().casefold()
        if format_key == "flag (ja / nein)":
            mapping = {
                "1": "Ja",
                "true": "Ja",
                "wahr": "Ja",
                "yes": "Ja",
                "ja": "Ja",
                "x": "Ja",
                "0": "Nein",
                "false": "Nein",
                "falsch": "Nein",
                "no": "Nein",
                "nein": "Nein",
            }
            return mapping.get(normalized.casefold(), normalized)
        if format_key == "kein wert":
            return ""
        return normalized

    def _build_current_row(self) -> AttributeRow:
        value_format = self.format_var.get().strip()
        return AttributeRow(
            criteria_id=self.criteria_id_var.get().strip(),
            label=self.label_var.get().strip(),
            value_format=value_format,
            max_length=self._current_max_length(),
            type_name=self.type_name_var.get().strip(),
            value=self._normalize_current_value(value_format, self.value_var.get()),
            value_to=self._normalize_current_value(value_format, self.value_to_var.get()),
        )

    def _validate_row(self, row: AttributeRow) -> None:
        if not row.criteria_id:
            raise ValueError("Bitte mindestens eine TecDoc Kriterien ID angeben.")
        if row.max_length is not None:
            for value_label, value in [("Wert", row.value), ("Wert bis", row.value_to)]:
                if value and len(value) > row.max_length:
                    raise ValueError(f"{value_label} für {row.display_label()} ist länger als erlaubt ({row.max_length}).")
        if row.value_format.strip().casefold() != "kein wert" and not any(
            value.strip() for value in [row.value, row.value_to]
        ):
            raise ValueError("Bitte für dieses Attribut einen Wert oder Wert bis eingeben.")

    def add_row(self) -> None:
        self._apply_current_attribute_selection()
        row = self._build_current_row()
        try:
            self._validate_row(row)
        except ValueError as exc:
            messagebox.showwarning(APP_TITLE, str(exc))
            return
        item_id = self.tree.insert(
            "",
            "end",
            values=(row.criteria_id, row.label, row.value_format, row.value, row.value_to),
        )
        self.row_type_names[item_id] = row.type_name
        self._select_first_row(select_last=True)
        self._emit_change()

    def update_selected_row(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst eine Zeile zum Bearbeiten auswählen.")
            return
        self._apply_current_attribute_selection()
        row = self._build_current_row()
        try:
            self._validate_row(row)
        except ValueError as exc:
            messagebox.showwarning(APP_TITLE, str(exc))
            return
        self.tree.item(
            selected[0],
            values=(row.criteria_id, row.label, row.value_format, row.value, row.value_to),
        )
        self.row_type_names[selected[0]] = row.type_name
        self.tree.focus(selected[0])
        self._handle_selection()
        self._emit_change()

    def remove_selected(self) -> None:
        for item_id in self.tree.selection():
            self.row_type_names.pop(item_id, None)
            self.tree.delete(item_id)
        self.clear_form()
        self._select_first_row()
        self._emit_change()

    def copy_selected_rows(self) -> None:
        selected_items = list(self.tree.selection())
        if not selected_items:
            return
        rows = []
        for item_id in selected_items:
            values = [str(value).strip() for value in self.tree.item(item_id, "values")]
            rows.append("\t".join(values))
        self.clipboard_clear()
        self.clipboard_append("\n".join(rows))

    def clear_form(self) -> None:
        self.attribute_display_var.set("")
        self.criteria_id_var.set("")
        self.label_var.set("")
        self.format_var.set("")
        self.max_length_var.set("")
        self.type_name_var.set("")
        self.value_var.set("")
        self.value_to_var.set("")
        self._hide_attribute_suggestions()
        self._hide_value_suggestions()
        self._update_hint()

    def _update_hint(self) -> None:
        format_label = self.format_var.get().strip()
        if not format_label:
            self.hint_var.set("Wähle ein Attribut aus oder gib die Kriterien-ID manuell ein.")
            return
        format_key = format_label.casefold()
        max_length = self.max_length_var.get().strip() or "-"
        if format_key == "kein wert":
            self.hint_var.set(f"Format: {format_label}. Dieses Attribut wird ohne Wert gespeichert. Max. Länge: {max_length}.")
            return
        if format_key == "flag (ja / nein)":
            self.hint_var.set(f"Format: {format_label}. Bitte bevorzugt 'Ja' oder 'Nein' verwenden. Max. Länge: {max_length}.")
            return
        if is_attribute_key_value_format(format_label):
            options = self._current_key_value_options()
            group_label = self.type_name_var.get().strip() or self.label_var.get().strip() or "-"
            self.hint_var.set(
                f"Format: {format_label}. Schlüsselwertgruppe: {group_label}. "
                f"{len(options)} mögliche Werte, Suche mit Live-Vorschlägen im Feld Wert. "
                f"Bereiche können über Wert und Wert bis gepflegt werden. Max. Länge: {max_length}."
            )
            return
        self.hint_var.set(f"Format: {format_label}. Wert oder Bereich über Wert und Wert bis pflegen. Max. Länge laut Katalog: {max_length}.")

    def get_rows(self) -> list[AttributeRow]:
        rows = []
        for item_id in self.tree.get_children():
            values = [str(value).strip() for value in self.tree.item(item_id, "values")]
            values.extend([""] * max(0, 5 - len(values)))
            criteria_id, label, value_format, value, value_to = values[:5]
            catalog_entry = self.catalog_by_id.get(str(criteria_id).strip())
            rows.append(
                AttributeRow(
                    criteria_id=str(criteria_id).strip(),
                    label=str(label).strip() or (catalog_entry.label if catalog_entry is not None else ""),
                    value_format=str(value_format).strip() or (catalog_entry.value_format if catalog_entry is not None else ""),
                    max_length=catalog_entry.max_length if catalog_entry is not None else None,
                    type_name=(
                        catalog_entry.type_name
                        if catalog_entry is not None
                        else self.row_type_names.get(item_id, "").strip()
                    ),
                    value=str(value).strip(),
                    value_to=str(value_to).strip(),
                )
            )
        return normalize_attribute_rows(rows)

    def set_rows(self, rows: list[AttributeRow]) -> None:
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        self.row_type_names = {}
        for row in normalize_attribute_rows(rows):
            item_id = self.tree.insert(
                "",
                "end",
                values=(row.criteria_id, row.label, row.value_format, row.value, row.value_to),
            )
            self.row_type_names[item_id] = row.type_name
        self._select_first_row()

    def _select_first_row(self, select_last: bool = False) -> None:
        children = list(self.tree.get_children())
        if not children:
            return
        target = children[-1] if select_last else children[0]
        self.tree.selection_set(target)
        self.tree.focus(target)
        self._handle_selection()

    def _handle_selection(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        values = [str(value).strip() for value in self.tree.item(selected[0], "values")]
        values.extend([""] * max(0, 5 - len(values)))
        criteria_id, label, value_format, value, value_to = values[:5]
        self.criteria_id_var.set(str(criteria_id).strip())
        self.label_var.set(str(label).strip())
        self.format_var.set(str(value_format).strip())
        self.value_var.set(str(value).strip())
        self.value_to_var.set(str(value_to).strip())
        catalog_entry = self.catalog_by_id.get(self.criteria_id_var.get())
        self.max_length_var.set("" if catalog_entry is None or catalog_entry.max_length is None else str(catalog_entry.max_length))
        self.type_name_var.set(
            catalog_entry.type_name
            if catalog_entry is not None
            else self.row_type_names.get(selected[0], "").strip()
        )
        display_parts = [self.criteria_id_var.get(), self.label_var.get()]
        self.attribute_display_var.set(" | ".join(part for part in display_parts if part))
        self._normalize_current_value_selection()
        self._update_hint()

    def _open_context_menu(self, event: tk.Event[tk.Misc]) -> None:
        item_id = self.tree.identify_row(event.y)
        if item_id:
            if item_id not in self.tree.selection():
                self.tree.selection_set(item_id)
            self.tree.focus(item_id)
            self._handle_selection()
        has_selection = bool(self.tree.selection())
        self.context_menu.entryconfigure("Zeilen kopieren", state="normal" if has_selection else "disabled")
        self.context_menu.entryconfigure("Zeilen löschen", state="normal" if has_selection else "disabled")
        self.context_menu.tk_popup(event.x_root, event.y_root)
        self.context_menu.grab_release()


class GenArtSearchDialog:
    def __init__(self, master: tk.Misc, registry: GenArtRegistry, initial_query: str = "") -> None:
        self.master = master
        self.registry = registry
        self.result: GenArtOption | None = None
        self.search_var = tk.StringVar(value=initial_query.strip())
        self.results_var = tk.StringVar(value="")

        self.window = tk.Toplevel(master)
        self.window.title("GenArt suchen")
        self.window.transient(master.winfo_toplevel())
        self.window.geometry("920x620")
        self.window.minsize(760, 460)
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(1, weight=1)
        self.window.protocol("WM_DELETE_WINDOW", self._cancel)

        header = ttk.Frame(self.window, padding=14)
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(1, weight=1)

        ttk.Label(header, text="Suche", font=("Segoe UI Semibold", 10)).grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.search_entry = ttk.Entry(header, textvariable=self.search_var)
        self.search_entry.grid(row=0, column=1, sticky="ew")
        ttk.Button(header, text="Auswählen", command=self._accept).grid(row=0, column=2, padx=(10, 0))
        ttk.Button(header, text="Abbrechen", command=self._cancel).grid(row=0, column=3, padx=(8, 0))

        ttk.Label(
            header,
            text="Suche in ID, GenArt und Bezeichnung. Mehrere Begriffe sind möglich.",
            foreground="#5E6472",
        ).grid(row=1, column=0, columnspan=4, sticky="w", pady=(8, 0))
        ttk.Label(header, textvariable=self.results_var, foreground="#5E6472").grid(
            row=2, column=0, columnspan=4, sticky="w", pady=(6, 0)
        )

        table_frame = ttk.Frame(self.window, padding=(14, 0, 14, 14))
        table_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        columns = ("id", "genart", "bezeichnung", "score")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.heading("id", text="ID")
        self.tree.heading("genart", text="GenArt")
        self.tree.heading("bezeichnung", text="Bezeichnung")
        self.tree.heading("score", text="Treffer")
        self.tree.column("id", width=110, anchor="w")
        self.tree.column("genart", width=220, anchor="w")
        self.tree.column("bezeichnung", width=420, anchor="w")
        self.tree.column("score", width=80, anchor="center")

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.search_var.trace_add("write", self._refresh_results)
        self.search_entry.bind("<Down>", self._focus_results)
        self.search_entry.bind("<Return>", self._accept_first_result)
        self.tree.bind("<Double-1>", lambda _event: self._accept())
        self.tree.bind("<Return>", lambda _event: self._accept())
        self.tree.bind("<Escape>", lambda _event: self._cancel())

        self._refresh_results()
        self.window.grab_set()
        self.search_entry.focus_set()
        self.search_entry.select_range(0, "end")

    def show(self) -> GenArtOption | None:
        self.window.wait_window()
        return self.result

    def _refresh_results(self, *_args) -> None:
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)

        query = self.search_var.get().strip()
        results = self.registry.search_options(query, limit=300)
        if not query:
            results = [(option, 0.0) for option in self.registry.options[:300]]

        for option, score in results:
            self.tree.insert(
                "",
                "end",
                iid=option.id,
                values=(
                    option.id,
                    option.genart,
                    option.bezeichnung,
                    "" if score <= 0 else f"{score:.0f}",
                ),
            )

        shown_count = len(results)
        total_count = len(self.registry.options)
        if query:
            self.results_var.set(f"{shown_count} Treffer angezeigt, {total_count} GenArts insgesamt.")
        else:
            self.results_var.set(f"{shown_count} GenArts angezeigt. Tippe oben, um die Liste zu filtern.")

        children = self.tree.get_children()
        if children:
            self.tree.selection_set(children[0])
            self.tree.focus(children[0])

    def _focus_results(self, _event: tk.Event[tk.Misc]) -> str:
        children = self.tree.get_children()
        if children:
            self.tree.focus(children[0])
            self.tree.selection_set(children[0])
            self.tree.focus_set()
        return "break"

    def _accept_first_result(self, _event: tk.Event[tk.Misc]) -> str:
        children = self.tree.get_children()
        if children:
            self.tree.selection_set(children[0])
            self.tree.focus(children[0])
            self._accept()
        return "break"

    def _accept(self) -> None:
        selection = self.tree.selection()
        if not selection:
            return
        self.result = self.registry.resolve(selection[0])
        self.window.destroy()

    def _cancel(self) -> None:
        self.result = None
        self.window.destroy()


class AttributeSuggestionDialog(tk.Toplevel):
    """Bestätigungsdialog für aus Produkttext erkannte Attribute."""

    def __init__(
        self,
        master: tk.Misc,
        suggestions: list[AttributeSuggestion],
        unmatched: list[ExtractedSpecLine],
        attribute_options: list[AttributeOption],
        on_apply: Callable[[list[AttributeRow], list[tuple[str, str]]], None],
    ) -> None:
        super().__init__(master)
        self.title("Attribute aus Text vorschlagen")
        self.on_apply = on_apply
        self.attribute_options = list(attribute_options)
        self.options_by_display = {option.display_label(): option for option in self.attribute_options}
        self.options_by_id = {option.criteria_id: option for option in self.attribute_options}
        self.suggestions_by_item: dict[str, AttributeSuggestion] = {}
        self.checked_items: set[str] = set()
        self.unmatched = list(unmatched)
        self.learned_entries: list[tuple[str, str]] = []
        self.selected_unmatched_label = ""

        self.transient(master.winfo_toplevel())
        self.configure(background=UI_BACKGROUND, padx=16, pady=14)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=3)
        self.rowconfigure(3, weight=2)

        ttk.Label(
            self,
            text=(
                "Gefundene Attribute aus dem Produkttext. Angehakte Zeilen werden in den Attribute-Tab übernommen. "
                "Klick auf eine Zeile schaltet das Häkchen um."
            ),
            foreground="#5E6472",
            wraplength=900,
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        columns = ("use", "attribute", "value", "value_to", "note", "source")
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=11, selectmode="browse")
        self.tree.grid(row=1, column=0, sticky="nsew")
        headings = {
            "use": "✓", "attribute": "Attribut", "value": "Wert",
            "value_to": "Wert bis", "note": "Hinweis", "source": "Quellzeile",
        }
        widths = {"use": 40, "attribute": 250, "value": 130, "value_to": 90, "note": 210, "source": 260}
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], anchor="center" if column == "use" else "w")
        self.tree.bind("<Button-1>", self._on_tree_click)
        self.tree.bind("<space>", self._on_tree_space)

        tree_scroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        tree_scroll.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=tree_scroll.set)

        for suggestion in suggestions:
            self._insert_suggestion(suggestion, checked=suggestion.confident)

        unmatched_frame = ttk.LabelFrame(self, text="Nicht erkannte Angaben", padding=10)
        unmatched_frame.grid(row=3, column=0, columnspan=2, sticky="nsew", pady=(12, 0))
        unmatched_frame.columnconfigure(1, weight=1)
        unmatched_frame.rowconfigure(0, weight=1)

        self.unmatched_list = tk.Listbox(unmatched_frame, height=5, exportselection=False)
        self.unmatched_list.grid(row=0, column=0, columnspan=4, sticky="nsew", pady=(0, 8))
        for spec in self.unmatched:
            self.unmatched_list.insert("end", spec.source_line)
        self.unmatched_list.bind("<<ListboxSelect>>", self._on_unmatched_select)

        ttk.Label(unmatched_frame, text="Attribut").grid(row=1, column=0, sticky="w", padx=(0, 8))
        self.assign_display_var = tk.StringVar()
        self.assign_entry = ttk.Entry(unmatched_frame, textvariable=self.assign_display_var)
        self.assign_entry.grid(row=1, column=1, sticky="ew")
        self.assign_suggestions = SearchSuggestionPopup(
            self,
            self.assign_entry,
            self._assign_suggestion_values,
            self._accept_assign_suggestion,
            min_width=420,
        )
        ttk.Label(unmatched_frame, text="Wert").grid(row=1, column=2, sticky="w", padx=(12, 8))
        self.assign_value_var = tk.StringVar()
        ttk.Entry(unmatched_frame, textvariable=self.assign_value_var, width=18).grid(row=1, column=3, sticky="w")

        self.remember_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            unmatched_frame,
            text="Zuordnung merken (Text-Label wird in die Zuordnungstabelle übernommen)",
            variable=self.remember_var,
        ).grid(row=2, column=0, columnspan=3, sticky="w", pady=(8, 0))
        ttk.Button(unmatched_frame, text="Zuweisen", command=self._assign_unmatched).grid(row=2, column=3, sticky="e", pady=(8, 0))

        buttons = ttk.Frame(self)
        buttons.grid(row=4, column=0, columnspan=2, sticky="e", pady=(12, 0))
        ttk.Button(buttons, text="Abbrechen", command=self.destroy).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(buttons, text="Übernehmen", style="Accent.TButton", command=self._apply).grid(row=0, column=1)

        self.geometry("1020x640")
        self.grab_set()
        self.tree.focus_set()

    def _insert_suggestion(self, suggestion: AttributeSuggestion, checked: bool) -> None:
        item_id = self.tree.insert(
            "",
            "end",
            values=(
                "☑" if checked else "☐",
                suggestion.option.display_label(),
                suggestion.value,
                suggestion.value_to,
                suggestion.note,
                suggestion.source_line,
            ),
        )
        self.suggestions_by_item[item_id] = suggestion
        if checked:
            self.checked_items.add(item_id)

    def _toggle_item(self, item_id: str) -> None:
        if item_id in self.checked_items:
            self.checked_items.discard(item_id)
        else:
            self.checked_items.add(item_id)
        values = list(self.tree.item(item_id, "values"))
        values[0] = "☑" if item_id in self.checked_items else "☐"
        self.tree.item(item_id, values=values)

    def _on_tree_click(self, event: tk.Event[tk.Misc]) -> None:
        item_id = self.tree.identify_row(event.y)
        if item_id:
            self._toggle_item(item_id)

    def _on_tree_space(self, _event: tk.Event[tk.Misc]) -> str:
        selected = self.tree.selection()
        if selected:
            self._toggle_item(selected[0])
        return "break"

    def _assign_suggestion_values(self, query: str = "") -> list[str]:
        query_text = query.strip().casefold()
        return [
            option.display_label()
            for option in self.attribute_options
            if not query_text or query_text in option.search_blob
        ][:200]

    def _accept_assign_suggestion(self, value: str) -> None:
        self.assign_display_var.set(value)

    def _on_unmatched_select(self, _event: tk.Event[tk.Misc]) -> None:
        selection = self.unmatched_list.curselection()
        if not selection:
            return
        spec = self.unmatched[selection[0]]
        self.selected_unmatched_label = spec.label
        parsed = parse_numeric_value(spec.raw_value)
        if parsed is not None:
            number, number_to, _unit = parsed
            value_text = format_german_number(number)
            if number_to is not None:
                value_text += f" - {format_german_number(number_to)}"
            self.assign_value_var.set(value_text)
        else:
            self.assign_value_var.set(spec.raw_value)

    def _resolve_assign_option(self) -> AttributeOption | None:
        value = self.assign_display_var.get().strip()
        if not value:
            return None
        if value in self.options_by_display:
            return self.options_by_display[value]
        if value in self.options_by_id:
            return self.options_by_id[value]
        return None

    def _assign_unmatched(self) -> None:
        selection = self.unmatched_list.curselection()
        if not selection:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst eine Zeile aus 'Nicht erkannte Angaben' auswählen.", parent=self)
            return
        option = self._resolve_assign_option()
        if option is None:
            messagebox.showwarning(APP_TITLE, "Bitte ein Attribut aus der Liste auswählen.", parent=self)
            return
        index = selection[0]
        spec = self.unmatched[index]
        raw_value = self.assign_value_var.get().strip()
        value, value_to = raw_value, ""
        range_match = re.match(r"^(.+?)\s*-\s*(.+)$", raw_value)
        if range_match and option.value_format.strip().casefold() == "numerisch":
            value, value_to = range_match.group(1).strip(), range_match.group(2).strip()
        suggestion = AttributeSuggestion(
            option=option, value=value, value_to=value_to,
            source_line=spec.source_line, confident=True, note="manuell zugeordnet",
        )
        self._insert_suggestion(suggestion, checked=True)
        if self.remember_var.get() and spec.label.strip():
            self.learned_entries.append((spec.label.strip(), option.criteria_id))
        self.unmatched_list.delete(index)
        del self.unmatched[index]
        self.assign_display_var.set("")
        self.assign_value_var.set("")
        self.selected_unmatched_label = ""

    def _apply(self) -> None:
        rows: list[AttributeRow] = []
        for item_id in self.tree.get_children():
            if item_id not in self.checked_items:
                continue
            suggestion = self.suggestions_by_item.get(item_id)
            if suggestion is None:
                continue
            values = self.tree.item(item_id, "values")
            option = suggestion.option
            rows.append(
                AttributeRow(
                    criteria_id=option.criteria_id,
                    label=option.label,
                    value_format=option.value_format,
                    max_length=option.max_length,
                    type_name=option.type_name,
                    value=str(values[2]).strip(),
                    value_to=str(values[3]).strip(),
                )
            )
        learned = list(self.learned_entries)
        self.destroy()
        self.on_apply(rows, learned)


class ApolloImportApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        default_width = min(1360, max(980, screen_width - 120))
        default_height = min(920, max(700, screen_height - 140))
        self.root.geometry(f"{default_width}x{default_height}")
        self.root.minsize(920, 620)
        self.app_icon_photo: object | None = None

        self.id_registry = IdRegistry()
        self.genart_registry = GenArtRegistry()
        self.kunzer_scraper = KunzerScraper()

        self.import_dir_var = tk.StringVar(value=str(DEFAULT_IMPORT_DIR))
        self.output_dir_var = tk.StringVar(value=str(DEFAULT_OUTPUT_DIR))
        self.genart_source_path_var = tk.StringVar(value=str(DEFAULT_GENART_SOURCE) if path_exists_safe(DEFAULT_GENART_SOURCE) else "")
        self.competitor_source_path_var = tk.StringVar(value=str(DEFAULT_COMPETITOR_SOURCE) if path_exists_safe(DEFAULT_COMPETITOR_SOURCE) else "")
        self.attribute_source_path_var = tk.StringVar(value=str(DEFAULT_ATTRIBUTE_SOURCE) if path_exists_safe(DEFAULT_ATTRIBUTE_SOURCE) else "")
        self.attribute_key_value_source_path_var = tk.StringVar(
            value=str(DEFAULT_ATTRIBUTE_KEY_VALUE_SOURCE) if path_exists_safe(DEFAULT_ATTRIBUTE_KEY_VALUE_SOURCE) else ""
        )
        self.vehicle_source_path_var = tk.StringVar(
            value=str(DEFAULT_VEHICLE_SOURCE) if path_exists_safe(DEFAULT_VEHICLE_SOURCE) else ""
        )
        self.attribute_mapping_source_path_var = tk.StringVar(
            value=str(DEFAULT_ATTRIBUTE_MAPPING_SOURCE) if path_exists_safe(DEFAULT_ATTRIBUTE_MAPPING_SOURCE) else ""
        )
        self.product_list_path_var = tk.StringVar()
        self.deepl_api_key_var = tk.StringVar(value=os.getenv("DEEPL_API_KEY", ""))
        self.deepl_base_url_var = tk.StringVar(value=os.getenv("DEEPL_API_BASE_URL", DEEPL_DEFAULT_BASE_URL))
        self.kunzer_product_url_var = tk.StringVar()
        self.genart_display_var = tk.StringVar()
        self.genart_suggestion_var = tk.StringVar(value="Keine GenArt geladen.")
        self.selected_genart_count_var = tk.StringVar(value="0 GenArten gesetzt")
        self.auto_translate_after_scrape_var = tk.BooleanVar(value=True)
        self.fixed_export_path_var = tk.BooleanVar(value=True)
        self.batch_short_text_var = tk.BooleanVar(value=True)
        self.batch_long_text_var = tk.BooleanVar(value=True)
        self.batch_image_var = tk.BooleanVar(value=True)
        self.batch_document_var = tk.BooleanVar(value=True)
        self.batch_video_var = tk.BooleanVar(value=True)
        self.batch_web_var = tk.BooleanVar(value=True)
        self.article_number_var = tk.StringVar()
        self.short_module_id_var = tk.StringVar()
        self.long_module_id_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Bereit.")
        self.update_status_var = tk.StringVar(value=f"Aktuelle Version: {APP_VERSION_TAG}")
        self.known_id_count_var = tk.StringVar(value="0 IDs geladen")
        self.genart_count_var = tk.StringVar(value="0 GenArts geladen")
        self.competitor_count_var = tk.StringVar(value="0 Hersteller / 0 Mitbewerber geladen")
        self.attribute_count_var = tk.StringVar(value="0 Attribute geladen")
        self.attribute_key_value_count_var = tk.StringVar(value="0 Schlüsselwerte geladen")
        self.vehicle_count_var = tk.StringVar(value="0 KTyp-Datensätze geladen")
        self.attribute_mapping_count_var = tk.StringVar(value="0 Zuordnungen geladen")
        self.article_browser_collapsed = False
        self.article_browser_toggle_var = tk.StringVar()
        self.settings_dialog: tk.Toplevel | None = None
        self.settings_notebook: ttk.Notebook | None = None
        self._settings_tab_canvases: list[tuple[str, tk.Canvas]] = []
        self.current_id_article_number = ""
        self.article_browser_records: dict[str, StoredArticleSnapshot] = {}
        self.selected_genart_selections: list[GenArtSelection] = []
        self.genart_suggestions: SearchSuggestionPopup | None = None
        self.manufacturer_options: list[CompetitorOption] = []
        self.manufacturer_options_by_id: dict[str, CompetitorOption] = {}
        self.competitor_options: list[CompetitorOption] = []
        self.competitor_options_by_id: dict[str, CompetitorOption] = {}
        self.attribute_options: list[AttributeOption] = []
        self.attribute_options_by_id: dict[str, AttributeOption] = {}
        self.attribute_key_value_options: list[AttributeKeyValueOption] = []
        self.attribute_key_values_by_group: dict[str, list[AttributeKeyValueOption]] = {}
        self.vehicle_catalog: VehicleCatalog = VehicleCatalog()
        self.vehicle_catalog_loading = False
        self.attribute_mapping: dict[str, str] = {}
        self.article_browser_cache_dir = ""
        self.article_browser_cache_signature: tuple[tuple[str, int, int], ...] | None = None
        self.live_write_suspended = False
        self.article_browser_context_menu: tk.Menu | None = None
        self.background_task_running = False
        self.pending_article_browser_selection = ""
        self.project_tab_compact_mode: bool | None = None
        self.article_browser_compact_mode: bool | None = None
        self._project_layout_after_id: str | None = None
        self._article_browser_layout_after_id: str | None = None
        self.wizard_active = False
        self.wizard_step_index = 0
        self.wizard_steps: list[tuple[str, str, tk.Misc, str]] = []
        self.wizard_step_var = tk.StringVar()
        self.wizard_instruction_var = tk.StringVar()
        self.wizard_error_var = tk.StringVar()

        self._configure_style()
        self._configure_window_icon()
        self._build_layout()
        self.root.protocol("WM_DELETE_WINDOW", self._handle_close)
        self._restore_session_state()
        self._load_known_ids(initial=True)
        self._load_genart_catalog(initial=True)
        self._load_competitor_catalog(initial=True)
        self._load_attribute_catalog(initial=True)
        self._load_attribute_key_value_catalog(initial=True)
        self._load_vehicle_catalog(initial=True)
        self._load_attribute_mapping(initial=True)
        self.refresh_preview()
        self._apply_pending_session_selection()

    def _configure_style(self) -> None:
        style = ttk.Style()
        if "clam" in style.theme_names():
            style.theme_use("clam")

        for font_name in ("TkDefaultFont", "TkTextFont", "TkMenuFont", "TkHeadingFont"):
            try:
                tkfont.nametofont(font_name).configure(family=UI_FONT_FAMILY, size=10)
            except tk.TclError:
                pass

        style.configure(".", background=UI_BACKGROUND, foreground=UI_TEXT, font=(UI_FONT_FAMILY, 10))
        style.configure("TFrame", background=UI_BACKGROUND)
        style.configure("TLabel", background=UI_BACKGROUND, foreground=UI_TEXT, font=(UI_FONT_FAMILY, 10))
        style.configure(
            "TLabelframe",
            background=UI_BACKGROUND,
            bordercolor=UI_BORDER,
            lightcolor=UI_BACKGROUND,
            darkcolor=UI_BACKGROUND,
            relief="solid",
            borderwidth=1,
        )
        style.configure(
            "TLabelframe.Label",
            background=UI_BACKGROUND,
            foreground=UI_ACCENT_DARK,
            font=(f"{UI_FONT_FAMILY} Semibold", 11),
        )

        style.configure(
            "TButton",
            font=(UI_FONT_FAMILY, 10),
            padding=(12, 6),
            background=UI_SURFACE,
            foreground=UI_TEXT,
            bordercolor=UI_BORDER,
            lightcolor=UI_SURFACE,
            darkcolor=UI_SURFACE,
        )
        style.map(
            "TButton",
            background=[("disabled", UI_HEADING_BG), ("pressed", UI_ACCENT_SOFT), ("active", "#F2F6FF")],
            bordercolor=[("focus", UI_ACCENT), ("active", UI_ACCENT)],
            foreground=[("disabled", "#9AA1AC")],
        )
        style.configure(
            "Accent.TButton",
            font=(f"{UI_FONT_FAMILY} Semibold", 10),
            background=UI_ACCENT,
            foreground="#FFFFFF",
            bordercolor=UI_ACCENT,
            lightcolor=UI_ACCENT,
            darkcolor=UI_ACCENT,
        )
        style.map(
            "Accent.TButton",
            background=[("disabled", "#A5B8E8"), ("pressed", UI_ACCENT_DARK), ("active", UI_ACCENT_DARK)],
            foreground=[("disabled", "#F4F6FA")],
        )

        style.configure("TNotebook", background=UI_BACKGROUND, borderwidth=0, tabmargins=(0, 4, 0, 0))
        style.configure(
            "TNotebook.Tab",
            font=(UI_FONT_FAMILY, 10),
            padding=(16, 7),
            background=UI_TAB_INACTIVE,
            foreground=UI_TEXT_MUTED,
            bordercolor=UI_BORDER,
            lightcolor=UI_TAB_INACTIVE,
            darkcolor=UI_TAB_INACTIVE,
        )
        style.map(
            "TNotebook.Tab",
            background=[("selected", UI_SURFACE), ("active", "#EFF3F8")],
            foreground=[("selected", UI_ACCENT_DARK), ("active", UI_TEXT)],
            lightcolor=[("selected", UI_SURFACE)],
            darkcolor=[("selected", UI_SURFACE)],
        )

        style.configure(
            "TEntry",
            fieldbackground=UI_SURFACE,
            foreground=UI_TEXT,
            bordercolor=UI_BORDER,
            lightcolor=UI_SURFACE,
            darkcolor=UI_SURFACE,
            insertcolor=UI_TEXT,
            padding=4,
        )
        style.map(
            "TEntry",
            bordercolor=[("focus", UI_ACCENT)],
            lightcolor=[("focus", UI_ACCENT_SOFT)],
            darkcolor=[("focus", UI_ACCENT_SOFT)],
        )
        style.configure(
            "TCombobox",
            fieldbackground=UI_SURFACE,
            background=UI_SURFACE,
            foreground=UI_TEXT,
            bordercolor=UI_BORDER,
            arrowcolor=UI_TEXT_MUTED,
            padding=4,
        )
        style.map("TCombobox", fieldbackground=[("readonly", UI_SURFACE)], bordercolor=[("focus", UI_ACCENT)])
        style.configure("TSpinbox", fieldbackground=UI_SURFACE, bordercolor=UI_BORDER, arrowcolor=UI_TEXT_MUTED)

        style.configure("TCheckbutton", background=UI_BACKGROUND, foreground=UI_TEXT)
        style.map("TCheckbutton", background=[("active", UI_BACKGROUND)])
        style.configure("TRadiobutton", background=UI_BACKGROUND, foreground=UI_TEXT)
        style.map("TRadiobutton", background=[("active", UI_BACKGROUND)])

        style.configure(
            "Treeview",
            background=UI_SURFACE,
            fieldbackground=UI_SURFACE,
            foreground=UI_TEXT,
            font=(UI_FONT_FAMILY, 10),
            rowheight=26,
            bordercolor=UI_BORDER,
            lightcolor=UI_SURFACE,
            darkcolor=UI_SURFACE,
        )
        style.map("Treeview", background=[("selected", UI_ACCENT_SOFT)], foreground=[("selected", UI_TEXT)])
        style.configure(
            "Treeview.Heading",
            font=(f"{UI_FONT_FAMILY} Semibold", 10),
            background=UI_HEADING_BG,
            foreground=UI_TEXT,
            bordercolor=UI_BORDER,
            relief="flat",
            padding=(8, 6),
        )
        style.map("Treeview.Heading", background=[("active", "#E2E7EE")])

        for orientation in ("Vertical", "Horizontal"):
            style.configure(
                f"{orientation}.TScrollbar",
                background=UI_HEADING_BG,
                troughcolor=UI_BACKGROUND,
                bordercolor=UI_BACKGROUND,
                arrowcolor=UI_TEXT_MUTED,
            )
        style.configure("TSeparator", background=UI_BORDER)

        self.root.option_add("*TCombobox*Listbox.background", UI_SURFACE)
        self.root.option_add("*TCombobox*Listbox.foreground", UI_TEXT)
        self.root.option_add("*TCombobox*Listbox.selectBackground", UI_ACCENT_SOFT)
        self.root.option_add("*TCombobox*Listbox.selectForeground", UI_TEXT)
        self.root.configure(background=UI_BACKGROUND)

    def _configure_window_icon(self) -> None:
        ico_path = resolve_application_asset_path(APP_ICON_ICO_RELATIVE_PATH)
        png_path = resolve_application_asset_path(APP_ICON_PNG_RELATIVE_PATH)

        if ico_path.exists():
            try:
                self.root.iconbitmap(default=str(ico_path))
            except tk.TclError:
                pass

        if png_path.exists():
            try:
                self.app_icon_photo = tk.PhotoImage(file=str(png_path))
                self.root.iconphoto(True, self.app_icon_photo)
            except tk.TclError:
                self.app_icon_photo = None

    def _build_layout(self) -> None:
        shell = ttk.Frame(self.root, padding=18)
        shell.pack(fill="both", expand=True)
        shell.columnconfigure(0, weight=1)
        shell.rowconfigure(2, weight=1)

        header = ttk.Frame(shell)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        header.columnconfigure(0, weight=1)

        ttk.Label(header, text="Apollo Import GUI", font=("Segoe UI Semibold", 20)).grid(row=0, column=0, sticky="w")
        ttk.Label(
            header,
            text="Prototyp für die Erfassung eines Artikels und den Export der zugehörigen Importdateien.",
            foreground="#5E6472",
        ).grid(row=1, column=0, sticky="w", pady=(4, 0))
        ttk.Label(
            header,
            text="⚙ = wird beim Kunzer-Abruf automatisch befüllt   ·   ✎ = manuell pflegen",
            foreground="#5E6472",
        ).grid(row=2, column=0, sticky="w", pady=(2, 0))
        header_actions = ttk.Frame(header)
        header_actions.grid(row=0, column=1, rowspan=2, sticky="e", padx=(12, 0))
        ttk.Button(header_actions, text="Geführter Modus", style="Accent.TButton", command=self.start_wizard).grid(
            row=0, column=0, padx=(0, 8)
        )
        ttk.Button(header_actions, text="⚙ Einstellungen", command=self.open_settings_dialog).grid(row=0, column=1)

        self._build_wizard_bar(shell)

        self.main_notebook = ttk.Notebook(shell)
        self.main_notebook.grid(row=2, column=0, sticky="nsew")

        self.project_tab = ttk.Frame(self.main_notebook, padding=18)
        self.short_tab = ttk.Frame(self.main_notebook, padding=18)
        self.long_tab = ttk.Frame(self.main_notebook, padding=18)
        self.genart_tab = ttk.Frame(self.main_notebook, padding=18)
        self.attribute_tab = ttk.Frame(self.main_notebook, padding=18)
        self.search_term_tab = ttk.Frame(self.main_notebook, padding=18)
        self.oe_tab = ttk.Frame(self.main_notebook, padding=18)
        self.comparison_tab = ttk.Frame(self.main_notebook, padding=18)
        self.vehicle_tab = ttk.Frame(self.main_notebook, padding=18)
        self.image_tab = ttk.Frame(self.main_notebook, padding=18)
        self.document_tab = ttk.Frame(self.main_notebook, padding=18)
        self.links_tab = ttk.Frame(self.main_notebook, padding=18)

        # Gruppierung: ⚙ = wird beim Kunzer-Abruf automatisch befuellt, ✎ = manuell pflegen.
        self.main_notebook.add(self.project_tab, text="Projekt")
        self.main_notebook.add(self.short_tab, text="⚙ Kurzbezeichnung")
        self.main_notebook.add(self.long_tab, text="⚙ Text")
        self.main_notebook.add(self.image_tab, text="⚙ Bilder")
        self.main_notebook.add(self.document_tab, text="⚙ Dokumente")
        self.main_notebook.add(self.links_tab, text="⚙ Links")
        self.main_notebook.add(self.genart_tab, text="✎ GenArten")
        self.main_notebook.add(self.attribute_tab, text="✎ Attribute")
        self.main_notebook.add(self.search_term_tab, text="✎ Suchwörter")
        self.main_notebook.add(self.oe_tab, text="✎ OE-Nummern")
        self.main_notebook.add(self.comparison_tab, text="✎ Vergleichsnummern")
        self.main_notebook.add(self.vehicle_tab, text="✎ Fahrzeuge")

        self._build_project_tab()
        self._build_short_tab()
        self._build_long_tab()
        self._build_genart_tab()
        self._build_attribute_tab()
        self._build_reference_tabs()
        self._build_media_tabs()

        self.wizard_steps = [
            (
                "artikel",
                "Artikel anlegen",
                self.project_tab,
                "Artikelnummer eingeben. Optional: 'Aus Kunzer per Artikelnummer laden' klicken – "
                "das füllt Texte, Bilder und Dokumente automatisch.",
            ),
            (
                "kurz",
                "Kurzbezeichnung",
                self.short_tab,
                "Deutsche Kurzbezeichnung eingeben. Weitere Sprachen sind optional – 'Übersetzen' füllt sie per DeepL.",
            ),
            (
                "lang",
                "Produkttext",
                self.long_tab,
                "Deutschen Produkttext eingeben oder prüfen (bei Kunzer-Abruf bereits gefüllt).",
            ),
            (
                "genart",
                "GenArt zuordnen",
                self.genart_tab,
                "Mindestens eine GenArt über das Suchfeld oder 'Suchen…' auswählen.",
            ),
            (
                "attribute",
                "Attribute (optional)",
                self.attribute_tab,
                "Technische Attribute pflegen. Tipp: 'Aus Text vorschlagen' liest Werte automatisch aus dem Produkttext.",
            ),
            (
                "suchwoerter",
                "Suchwörter (optional)",
                self.search_term_tab,
                "Suchbegriffe hinterlegen, unter denen der Artikel gefunden werden soll (max. 20 Zeichen pro Wort).",
            ),
            (
                "oe",
                "OE-Nummern (optional)",
                self.oe_tab,
                "OE-Nummern mit Hersteller erfassen. Ohne Hersteller kann nicht exportiert werden.",
            ),
            (
                "vergleich",
                "Vergleichsnummern (optional)",
                self.comparison_tab,
                "Vergleichsnummern von Mitbewerbern erfassen.",
            ),
            (
                "fahrzeuge",
                "Fahrzeuge (optional)",
                self.vehicle_tab,
                "Motorcode(s) eingeben (mehrere mit Semikolon trennen) und 'Fahrzeuge hinzufügen' klicken.",
            ),
            (
                "bilder",
                "Bilder (optional)",
                self.image_tab,
                "Bildpfade oder -links prüfen bzw. ergänzen.",
            ),
            (
                "dokumente",
                "Dokumente (optional)",
                self.document_tab,
                "Dokumente (z. B. PDF-Datenblätter) prüfen bzw. ergänzen.",
            ),
            (
                "links",
                "Videos & Links (optional)",
                self.links_tab,
                "Video- und Web-Links prüfen bzw. ergänzen.",
            ),
            (
                "fertig",
                "Zusammenfassung & Export",
                self.project_tab,
                "",
            ),
        ]

        status_bar = ttk.Label(shell, textvariable=self.status_var, anchor="w", foreground="#5E6472")
        status_bar.grid(row=3, column=0, sticky="ew", pady=(12, 0))

    def _build_wizard_bar(self, shell: ttk.Frame) -> None:
        self.wizard_frame = ttk.LabelFrame(shell, text="Geführter Modus", padding=(14, 10))
        self.wizard_frame.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        self.wizard_frame.columnconfigure(0, weight=1)

        ttk.Label(
            self.wizard_frame,
            textvariable=self.wizard_step_var,
            font=(f"{UI_FONT_FAMILY} Semibold", 11),
        ).grid(row=0, column=0, sticky="w")
        ttk.Label(
            self.wizard_frame,
            textvariable=self.wizard_instruction_var,
            foreground=UI_TEXT_MUTED,
            wraplength=940,
            justify="left",
        ).grid(row=1, column=0, sticky="w", pady=(2, 0))
        ttk.Label(
            self.wizard_frame,
            textvariable=self.wizard_error_var,
            foreground="#C2410C",
            wraplength=940,
            justify="left",
        ).grid(row=2, column=0, sticky="w", pady=(2, 0))

        wizard_buttons = ttk.Frame(self.wizard_frame)
        wizard_buttons.grid(row=0, column=1, rowspan=3, sticky="ne", padx=(12, 0))
        self.wizard_back_button = ttk.Button(wizard_buttons, text="Zurück", command=self._wizard_back)
        self.wizard_back_button.grid(row=0, column=0, padx=(0, 8))
        self.wizard_next_button = ttk.Button(wizard_buttons, text="Weiter", style="Accent.TButton", command=self._wizard_next)
        self.wizard_next_button.grid(row=0, column=1, padx=(0, 8))
        ttk.Button(wizard_buttons, text="Beenden", command=self._end_wizard).grid(row=0, column=2)

        self.wizard_frame.grid_remove()

    def start_wizard(self) -> None:
        self.wizard_active = True
        self.wizard_step_index = 0
        self.wizard_frame.grid()
        self._wizard_show_step()
        self.status_var.set("Geführter Modus gestartet.")

    def _end_wizard(self, status_message: str | None = None) -> None:
        self.wizard_active = False
        self.wizard_frame.grid_remove()
        self.status_var.set(status_message or "Geführter Modus beendet.")

    def _wizard_show_step(self) -> None:
        key, title, tab, instruction = self.wizard_steps[self.wizard_step_index]
        try:
            self.main_notebook.select(tab)
        except tk.TclError:  # pragma: no cover - defensive
            pass
        self.wizard_step_var.set(f"Schritt {self.wizard_step_index + 1} von {len(self.wizard_steps)}: {title}")
        if key == "fertig":
            instruction = self._wizard_summary_text()
        self.wizard_instruction_var.set(instruction)
        self.wizard_error_var.set("")
        self.wizard_back_button.state(["disabled"] if self.wizard_step_index == 0 else ["!disabled"])
        self.wizard_next_button.configure(text="Exportieren & Fertigstellen" if key == "fertig" else "Weiter")

    def _wizard_summary_text(self) -> str:
        article = normalize_article_number(self.article_number_var.get()) or "-"
        counts = [
            f"Artikel: {article}",
            f"GenArts: {len(self._get_selected_genart_selections())}",
            f"Attribute: {len(self.attribute_frame.get_rows())}",
            f"Suchwörter: {len(self.search_term_frame.get_terms())}",
            f"OE-Nummern: {len(self.oe_frame.get_rows())}",
            f"Vergleichsnummern: {len(self.comparison_frame.get_rows())}",
            f"Fahrzeuge: {len(self.vehicle_link_frame.get_rows())}",
            f"Bilder: {len(self.image_frame.get_rows())}",
            f"Dokumente: {len(self.document_frame.get_rows())}",
        ]
        return "Alles geprüft? 'Exportieren & Fertigstellen' schreibt die Importdateien.\n" + "  ·  ".join(counts)

    def _wizard_validation_error(self, key: str) -> str:
        if key == "artikel" and not normalize_article_number(self.article_number_var.get()):
            return "Bitte zuerst eine Artikelnummer eingeben."
        if key == "kurz" and not self.short_text_frame.get_value().de.strip():
            return "Bitte die deutsche Kurzbezeichnung eingeben."
        if key == "lang" and not self.long_text_frame.get_value().de.strip():
            return "Bitte den deutschen Produkttext eingeben."
        if key == "genart" and not self._get_selected_genart_selections():
            return "Bitte mindestens eine GenArt auswählen."
        if key == "oe":
            missing = [row.value for row in self.oe_frame.get_rows() if row.value.strip() and not row.manufacturer_id.strip()]
            if missing:
                shown = ", ".join(missing[:5]) + (" ..." if len(missing) > 5 else "")
                return f"Bitte für diese OE-Nummern einen Hersteller wählen: {shown}"
        if key == "fertig":
            for required_key in ("artikel", "kurz", "lang", "genart", "oe"):
                error = self._wizard_validation_error(required_key)
                if error:
                    return error
        return ""

    def _wizard_next(self) -> None:
        key = self.wizard_steps[self.wizard_step_index][0]
        error = self._wizard_validation_error(key)
        if error:
            self.wizard_error_var.set(error)
            return
        if key == "fertig":
            if self.export_current_bundle():
                self._end_wizard("Geführter Modus abgeschlossen – Export erstellt.")
            return
        self.wizard_step_index += 1
        self._wizard_show_step()

    def _wizard_back(self) -> None:
        if self.wizard_step_index > 0:
            self.wizard_step_index -= 1
            self._wizard_show_step()

    def _set_background_task_state(self, running: bool, status_message: str | None = None) -> None:
        self.background_task_running = running
        try:
            self.root.configure(cursor="watch" if running else "")
        except tk.TclError:
            pass
        if status_message:
            self.status_var.set(status_message)

    def _run_background_task(
        self,
        start_message: str,
        worker: Callable[[], object],
        on_success: Callable[[object], None],
        error_status_prefix: str,
    ) -> bool:
        if self.background_task_running:
            self.status_var.set("Es läuft bereits ein Hintergrundvorgang. Bitte kurz warten.")
            return False

        self._set_background_task_state(True, start_message)

        def finish_success(result: object) -> None:
            self._set_background_task_state(False)
            try:
                on_success(result)
            except Exception as exc:  # pragma: no cover - defensive GUI feedback
                messagebox.showwarning(APP_TITLE, str(exc))
                self.status_var.set(f"{error_status_prefix}: {exc}")

        def finish_error(exc: Exception) -> None:
            self._set_background_task_state(False)
            messagebox.showwarning(APP_TITLE, str(exc))
            self.status_var.set(f"{error_status_prefix}: {exc}")

        def task() -> None:
            try:
                result = worker()
            except Exception as exc:  # pragma: no cover - depends on user/network/runtime
                self.root.after(0, lambda exc=exc: finish_error(exc))
                return
            self.root.after(0, lambda result=result: finish_success(result))

        threading.Thread(target=task, daemon=True).start()
        return True

    def _translation_set_to_state(self, translations: TranslationSet) -> dict[str, str]:
        return {
            code: getattr(translations, code)
            for code, _label in UI_LANGUAGE_ORDER
        }

    def _translation_set_from_state(self, payload: object) -> TranslationSet:
        if not isinstance(payload, dict):
            return TranslationSet()
        values = {
            code: str(payload.get(code, ""))
            for code, _label in UI_LANGUAGE_ORDER
        }
        return TranslationSet(**values)

    def _media_rows_to_state(self, rows: list[MediaRow]) -> list[dict[str, str]]:
        return [
            {
                "path_or_link": row.path_or_link,
                "art": row.art,
                "sprache": row.sprache,
            }
            for row in rows
        ]

    def _media_rows_from_state(self, payload: object) -> list[MediaRow]:
        if not isinstance(payload, list):
            return []
        rows: list[MediaRow] = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            rows.append(
                MediaRow(
                    path_or_link=str(item.get("path_or_link", "")),
                    art=str(item.get("art", "")),
                    sprache=str(item.get("sprache", "")),
                )
            )
        return rows

    def _oe_number_rows_to_state(self, rows: list[OeNumberRow]) -> list[dict[str, str]]:
        return [
            {
                "value": row.value,
                "manufacturer_id": row.manufacturer_id,
                "manufacturer_code": row.manufacturer_code,
                "manufacturer_name": row.manufacturer_name,
            }
            for row in normalize_oe_number_rows(rows)
        ]

    def _oe_number_rows_from_state(self, payload: object) -> list[OeNumberRow]:
        if not isinstance(payload, list):
            return []
        rows: list[OeNumberRow] = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            rows.append(
                OeNumberRow(
                    value=str(item.get("value", "")),
                    manufacturer_id=str(item.get("manufacturer_id", "")),
                    manufacturer_code=str(item.get("manufacturer_code", "")),
                    manufacturer_name=str(item.get("manufacturer_name", "")),
                )
            )
        return normalize_oe_number_rows(rows)

    def _comparison_rows_to_state(self, rows: list[ComparisonNumberRow]) -> list[dict[str, str]]:
        return [
            {
                "competitor_id": row.competitor_id,
                "competitor_code": row.competitor_code,
                "competitor_name": row.competitor_name,
                "reference_number": row.reference_number,
            }
            for row in normalize_comparison_number_rows(rows)
        ]

    def _comparison_rows_from_state(self, payload: object) -> list[ComparisonNumberRow]:
        if not isinstance(payload, list):
            return []
        rows: list[ComparisonNumberRow] = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            rows.append(
                ComparisonNumberRow(
                    competitor_id=str(item.get("competitor_id", "")),
                    competitor_code=str(item.get("competitor_code", "")),
                    competitor_name=str(item.get("competitor_name", "")),
                    reference_number=str(item.get("reference_number", "")),
                )
            )
        return normalize_comparison_number_rows(rows)

    def _vehicle_link_rows_to_state(self, rows: list[VehicleLinkRow]) -> list[dict[str, str]]:
        return [
            {
                "vehicle_type_id": row.vehicle_type_id,
                "vehicle_type_label": row.vehicle_type_label,
                "motorcode": row.motorcode,
                "ktyp_topmotive": row.ktyp_topmotive,
                "ktyp_tecdoc": row.ktyp_tecdoc,
                "manufacturer": row.manufacturer,
                "model": row.model,
                "description": row.description,
                "year_from": row.year_from,
                "year_to": row.year_to,
                "power": row.power,
            }
            for row in normalize_vehicle_link_rows(rows)
        ]

    def _vehicle_link_rows_from_state(self, payload: object) -> list[VehicleLinkRow]:
        if not isinstance(payload, list):
            return []
        rows: list[VehicleLinkRow] = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            type_id = str(item.get("vehicle_type_id", DEFAULT_VEHICLE_TYPE_ID)) or DEFAULT_VEHICLE_TYPE_ID
            rows.append(
                VehicleLinkRow(
                    vehicle_type_id=type_id,
                    vehicle_type_label=str(item.get("vehicle_type_label", "")) or vehicle_type_label_for_id(type_id),
                    motorcode=str(item.get("motorcode", "")),
                    ktyp_topmotive=str(item.get("ktyp_topmotive", "")),
                    ktyp_tecdoc=str(item.get("ktyp_tecdoc", "")),
                    manufacturer=str(item.get("manufacturer", "")),
                    model=str(item.get("model", "")),
                    description=str(item.get("description", "")),
                    year_from=str(item.get("year_from", "")),
                    year_to=str(item.get("year_to", "")),
                    power=str(item.get("power", "")),
                )
            )
        return normalize_vehicle_link_rows(rows)

    def _attribute_rows_to_state(self, rows: list[AttributeRow]) -> list[dict[str, object]]:
        return [
            {
                "criteria_id": row.criteria_id,
                "label": row.label,
                "value_format": row.value_format,
                "max_length": row.max_length,
                "type_name": row.type_name,
                "value": row.value,
                "value_to": row.value_to,
            }
            for row in normalize_attribute_rows(rows)
        ]

    def _attribute_rows_from_state(self, payload: object) -> list[AttributeRow]:
        if not isinstance(payload, list):
            return []
        rows: list[AttributeRow] = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            max_length_raw = item.get("max_length")
            max_length: int | None
            if isinstance(max_length_raw, int):
                max_length = max_length_raw
            else:
                try:
                    max_length = int(str(max_length_raw).strip()) if str(max_length_raw).strip() else None
                except Exception:
                    max_length = None
            raw_value = str(item.get("value", ""))
            if not raw_value.strip():
                raw_value = str(item.get("value_from", ""))
            rows.append(
                AttributeRow(
                    criteria_id=str(item.get("criteria_id", "")),
                    label=str(item.get("label", "")),
                    value_format=str(item.get("value_format", "")),
                    max_length=max_length,
                    type_name=str(item.get("type_name", "")),
                    value=raw_value,
                    value_to=str(item.get("value_to", "")),
                )
            )
        return normalize_attribute_rows(rows)

    def _genart_selections_to_state(self, selections: list[GenArtSelection]) -> list[dict[str, str]]:
        return [
            {
                "id": selection.id,
                "bezeichnung": selection.bezeichnung,
            }
            for selection in normalize_genart_selections(selections)
        ]

    def _genart_selections_from_state(self, payload: object) -> list[GenArtSelection]:
        if not isinstance(payload, list):
            return []
        selections: list[GenArtSelection] = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            selections.append(
                GenArtSelection(
                    id=str(item.get("id", "")),
                    bezeichnung=str(item.get("bezeichnung", "")),
                )
            )
        return normalize_genart_selections(selections)

    def _collect_session_state(self) -> dict[str, object]:
        selected_tab = 0
        try:
            selected_tab = self.main_notebook.index(self.main_notebook.select())
        except Exception:
            selected_tab = 0

        selected_browser_article = ""
        if hasattr(self, "article_browser_tree") and self.article_browser_tree.selection():
            selected_browser_article = str(self.article_browser_tree.selection()[0]).strip()

        return {
            "version": 5,
            "window_geometry": self.root.winfo_geometry(),
            "selected_tab": selected_tab,
            "paths": {
                "import_dir": self.import_dir_var.get(),
                "output_dir": self.output_dir_var.get(),
                "genart_source_path": self.genart_source_path_var.get(),
                "competitor_source_path": self.competitor_source_path_var.get(),
                "attribute_source_path": self.attribute_source_path_var.get(),
                "attribute_key_value_source_path": self.attribute_key_value_source_path_var.get(),
                "vehicle_source_path": self.vehicle_source_path_var.get(),
                "attribute_mapping_source_path": self.attribute_mapping_source_path_var.get(),
                "product_list_path": self.product_list_path_var.get(),
            },
            "api": {
                "deepl_api_key": self.deepl_api_key_var.get(),
                "deepl_base_url": self.deepl_base_url_var.get(),
            },
            "options": {
                "fixed_export_path": self.fixed_export_path_var.get(),
                "batch_short_text": self.batch_short_text_var.get(),
                "batch_long_text": self.batch_long_text_var.get(),
                "batch_image": self.batch_image_var.get(),
                "batch_document": self.batch_document_var.get(),
                "batch_video": self.batch_video_var.get(),
                "batch_web": self.batch_web_var.get(),
            },
            "article": {
                "article_number": self.article_number_var.get(),
                "short_module_id": self.short_module_id_var.get(),
                "long_module_id": self.long_module_id_var.get(),
                "kunzer_product_url": self.kunzer_product_url_var.get(),
                "genart_selections": self._genart_selections_to_state(self.selected_genart_selections),
                "genart_display": self.genart_display_var.get(),
                "genart_suggestion": self.genart_suggestion_var.get(),
            },
            "translations": {
                "short_auto_uni": self.short_text_frame.auto_uni_var.get(),
                "short_values": self._translation_set_to_state(self.short_text_frame.get_value()),
                "long_auto_uni": self.long_text_frame.auto_uni_var.get(),
                "long_values": self._translation_set_to_state(self.long_text_frame.get_value()),
            },
            "references": {
                "attributes": {
                    "rows": self._attribute_rows_to_state(self.attribute_frame.get_rows()),
                    "form": {
                        "attribute_display": self.attribute_frame.attribute_display_var.get(),
                        "criteria_id": self.attribute_frame.criteria_id_var.get(),
                        "label": self.attribute_frame.label_var.get(),
                        "value_format": self.attribute_frame.format_var.get(),
                        "max_length": self.attribute_frame.max_length_var.get(),
                        "type_name": self.attribute_frame.type_name_var.get(),
                        "value": self.attribute_frame.value_var.get(),
                        "value_to": self.attribute_frame.value_to_var.get(),
                    },
                },
                "oe_numbers": {
                    "rows": self._oe_number_rows_to_state(self.oe_frame.get_rows()),
                    "form": {
                        "manufacturer_display": self.oe_frame.manufacturer_display_var.get(),
                        "manufacturer_id": self.oe_frame.manufacturer_id_var.get(),
                        "manufacturer_code": self.oe_frame.manufacturer_code_var.get(),
                        "manufacturer_name": self.oe_frame.manufacturer_name_var.get(),
                        "value": self.oe_frame.value_var.get(),
                    },
                },
                "comparison_numbers": {
                    "rows": self._comparison_rows_to_state(self.comparison_frame.get_rows()),
                    "form": {
                        "competitor_display": self.comparison_frame.competitor_display_var.get(),
                        "competitor_id": self.comparison_frame.competitor_id_var.get(),
                        "competitor_code": self.comparison_frame.competitor_code_var.get(),
                        "competitor_name": self.comparison_frame.competitor_name_var.get(),
                        "reference_number": self.comparison_frame.reference_number_var.get(),
                    },
                },
                "vehicle_links": {
                    "rows": self._vehicle_link_rows_to_state(self.vehicle_link_frame.get_rows()),
                    "form": {
                        "motorcode": self.vehicle_link_frame.motorcode_var.get(),
                        "vehicle_type": self.vehicle_link_frame.vehicle_type_var.get(),
                    },
                },
                "search_terms": {
                    "terms": self.search_term_frame.get_terms(),
                    "form": {
                        "term": self.search_term_frame.term_var.get(),
                    },
                },
            },
            "media": {
                "images": {
                    "rows": self._media_rows_to_state(self.image_frame.get_rows()),
                    "form": {
                        "path_or_link": self.image_frame.path_var.get(),
                        "art": self.image_frame.art_var.get(),
                        "sprache": self.image_frame.sprache_var.get(),
                    },
                },
                "documents": {
                    "rows": self._media_rows_to_state(self.document_frame.get_rows()),
                    "form": {
                        "path_or_link": self.document_frame.path_var.get(),
                        "art": self.document_frame.art_var.get(),
                        "sprache": self.document_frame.sprache_var.get(),
                    },
                },
                "videos": {
                    "rows": self._media_rows_to_state(self.video_frame.get_rows()),
                    "form": {
                        "path_or_link": self.video_frame.link_var.get(),
                    },
                },
                "web_links": {
                    "rows": self._media_rows_to_state(self.web_frame.get_rows()),
                    "form": {
                        "path_or_link": self.web_frame.link_var.get(),
                    },
                },
            },
            "ui": {
                "article_browser_collapsed": self.article_browser_collapsed,
                "image_preview_visible": getattr(self.image_frame, "preview_visible", True),
                "document_preview_visible": getattr(self.document_frame, "preview_visible", True),
                "video_preview_visible": getattr(self.video_frame, "preview_visible", True),
                "web_preview_visible": getattr(self.web_frame, "preview_visible", True),
            },
            "article_browser": {
                "selected_article": selected_browser_article,
            },
        }

    def _save_session_state(self) -> None:
        SESSION_STATE_FILE.write_text(
            json.dumps(self._collect_session_state(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _restore_session_state(self) -> None:
        if not SESSION_STATE_FILE.exists():
            return
        try:
            payload = json.loads(SESSION_STATE_FILE.read_text(encoding="utf-8"))
        except Exception:
            return
        if not isinstance(payload, dict):
            return

        with self._suspend_live_write():
            geometry = payload.get("window_geometry")
            if isinstance(geometry, str) and geometry.strip():
                try:
                    self.root.geometry(geometry.strip())
                except tk.TclError:
                    pass

            paths = payload.get("paths")
            if isinstance(paths, dict):
                self.import_dir_var.set(str(paths.get("import_dir", self.import_dir_var.get())))
                self.output_dir_var.set(str(paths.get("output_dir", self.output_dir_var.get())))
                self.genart_source_path_var.set(str(paths.get("genart_source_path", self.genart_source_path_var.get())))
                self.competitor_source_path_var.set(str(paths.get("competitor_source_path", self.competitor_source_path_var.get())))
                self.attribute_source_path_var.set(str(paths.get("attribute_source_path", self.attribute_source_path_var.get())))
                self.attribute_key_value_source_path_var.set(
                    str(paths.get("attribute_key_value_source_path", self.attribute_key_value_source_path_var.get()))
                )
                self.vehicle_source_path_var.set(str(paths.get("vehicle_source_path", self.vehicle_source_path_var.get())))
                self.attribute_mapping_source_path_var.set(
                    str(paths.get("attribute_mapping_source_path", self.attribute_mapping_source_path_var.get()))
                )
                self.product_list_path_var.set(str(paths.get("product_list_path", self.product_list_path_var.get())))

            api = payload.get("api")
            if isinstance(api, dict):
                self.deepl_api_key_var.set(str(api.get("deepl_api_key", self.deepl_api_key_var.get())))
                self.deepl_base_url_var.set(str(api.get("deepl_base_url", self.deepl_base_url_var.get())))

            options = payload.get("options")
            if isinstance(options, dict):
                self.fixed_export_path_var.set(bool(options.get("fixed_export_path", self.fixed_export_path_var.get())))
                self.batch_short_text_var.set(bool(options.get("batch_short_text", self.batch_short_text_var.get())))
                self.batch_long_text_var.set(bool(options.get("batch_long_text", self.batch_long_text_var.get())))
                self.batch_image_var.set(bool(options.get("batch_image", self.batch_image_var.get())))
                self.batch_document_var.set(bool(options.get("batch_document", self.batch_document_var.get())))
                self.batch_video_var.set(bool(options.get("batch_video", self.batch_video_var.get())))
                self.batch_web_var.set(bool(options.get("batch_web", self.batch_web_var.get())))

            article = payload.get("article")
            if isinstance(article, dict):
                article_number = normalize_article_number(str(article.get("article_number", "")))
                self.article_number_var.set(article_number)
                self.short_module_id_var.set(str(article.get("short_module_id", "")))
                self.long_module_id_var.set(str(article.get("long_module_id", "")))
                self.kunzer_product_url_var.set(str(article.get("kunzer_product_url", "")))
                restored_genart_selections = self._genart_selections_from_state(article.get("genart_selections"))
                if not restored_genart_selections:
                    legacy_selection = parse_genart_selection_label(str(article.get("genart_display", "")))
                    if legacy_selection is not None:
                        restored_genart_selections = [legacy_selection]
                self._set_selected_genart_selections(restored_genart_selections)
                self.genart_display_var.set(str(article.get("genart_display", "")))
                self.genart_suggestion_var.set(str(article.get("genart_suggestion", self.genart_suggestion_var.get())))
                self.current_id_article_number = article_number if article_number else ""

            translations = payload.get("translations")
            if isinstance(translations, dict):
                self.short_text_frame.set_value(
                    self._translation_set_from_state(translations.get("short_values")),
                    auto_uni=bool(translations.get("short_auto_uni", True)),
                )
                self.long_text_frame.set_value(
                    self._translation_set_from_state(translations.get("long_values")),
                    auto_uni=bool(translations.get("long_auto_uni", True)),
                )

            references = payload.get("references")
            if isinstance(references, dict):
                attribute_payload = references.get("attributes")
                if isinstance(attribute_payload, dict):
                    self.attribute_frame.set_rows(self._attribute_rows_from_state(attribute_payload.get("rows")))
                    form = attribute_payload.get("form")
                    if isinstance(form, dict):
                        self.attribute_frame.attribute_display_var.set(str(form.get("attribute_display", "")))
                        self.attribute_frame.criteria_id_var.set(str(form.get("criteria_id", "")))
                        self.attribute_frame.label_var.set(str(form.get("label", "")))
                        self.attribute_frame.format_var.set(str(form.get("value_format", "")))
                        self.attribute_frame.max_length_var.set(str(form.get("max_length", "")))
                        self.attribute_frame.type_name_var.set(str(form.get("type_name", "")))
                        form_value = str(form.get("value", ""))
                        if not form_value.strip():
                            form_value = str(form.get("value_from", ""))
                        self.attribute_frame.value_var.set(form_value)
                        self.attribute_frame.value_to_var.set(str(form.get("value_to", "")))
                        self.attribute_frame._update_hint()

                oe_payload = references.get("oe_numbers")
                if isinstance(oe_payload, dict):
                    self.oe_frame.set_rows(self._oe_number_rows_from_state(oe_payload.get("rows")))
                    form = oe_payload.get("form")
                    if isinstance(form, dict):
                        self.oe_frame.manufacturer_display_var.set(str(form.get("manufacturer_display", "")))
                        self.oe_frame.manufacturer_id_var.set(str(form.get("manufacturer_id", "")))
                        self.oe_frame.manufacturer_code_var.set(str(form.get("manufacturer_code", "")))
                        self.oe_frame.manufacturer_name_var.set(str(form.get("manufacturer_name", "")))
                        self.oe_frame.value_var.set(str(form.get("value", "")))

                comparison_payload = references.get("comparison_numbers")
                if isinstance(comparison_payload, dict):
                    self.comparison_frame.set_rows(self._comparison_rows_from_state(comparison_payload.get("rows")))
                    form = comparison_payload.get("form")
                    if isinstance(form, dict):
                        self.comparison_frame.competitor_display_var.set(str(form.get("competitor_display", "")))
                        self.comparison_frame.competitor_id_var.set(str(form.get("competitor_id", "")))
                        self.comparison_frame.competitor_code_var.set(str(form.get("competitor_code", "")))
                        self.comparison_frame.competitor_name_var.set(str(form.get("competitor_name", "")))
                        self.comparison_frame.reference_number_var.set(str(form.get("reference_number", "")))

                vehicle_payload = references.get("vehicle_links")
                if isinstance(vehicle_payload, dict):
                    self.vehicle_link_frame.set_rows(self._vehicle_link_rows_from_state(vehicle_payload.get("rows")))
                    form = vehicle_payload.get("form")
                    if isinstance(form, dict):
                        self.vehicle_link_frame.motorcode_var.set(str(form.get("motorcode", "")))
                        vehicle_type = str(form.get("vehicle_type", "")).strip()
                        if vehicle_type:
                            self.vehicle_link_frame.vehicle_type_var.set(vehicle_type)

                search_term_payload = references.get("search_terms")
                if isinstance(search_term_payload, dict):
                    terms_payload = search_term_payload.get("terms")
                    if isinstance(terms_payload, list):
                        self.search_term_frame.set_terms([str(term) for term in terms_payload])
                    form = search_term_payload.get("form")
                    if isinstance(form, dict):
                        self.search_term_frame.term_var.set(str(form.get("term", "")))

            media = payload.get("media")
            if isinstance(media, dict):
                image_payload = media.get("images")
                if isinstance(image_payload, dict):
                    self.image_frame.set_rows(self._media_rows_from_state(image_payload.get("rows")))
                    form = image_payload.get("form")
                    if isinstance(form, dict):
                        self.image_frame.path_var.set(str(form.get("path_or_link", "")))
                        self.image_frame.art_var.set(str(form.get("art", self.image_frame.default_art)))
                        self.image_frame.sprache_var.set(str(form.get("sprache", self.image_frame.default_sprache)))

                document_payload = media.get("documents")
                if isinstance(document_payload, dict):
                    self.document_frame.set_rows(self._media_rows_from_state(document_payload.get("rows")))
                    form = document_payload.get("form")
                    if isinstance(form, dict):
                        self.document_frame.path_var.set(str(form.get("path_or_link", "")))
                        self.document_frame.art_var.set(str(form.get("art", self.document_frame.default_art)))
                        self.document_frame.sprache_var.set(str(form.get("sprache", self.document_frame.default_sprache)))

                video_payload = media.get("videos")
                if isinstance(video_payload, dict):
                    self.video_frame.set_rows(self._media_rows_from_state(video_payload.get("rows")))
                    form = video_payload.get("form")
                    if isinstance(form, dict):
                        self.video_frame.link_var.set(str(form.get("path_or_link", "")))

                web_payload = media.get("web_links")
                if isinstance(web_payload, dict):
                    self.web_frame.set_rows(self._media_rows_from_state(web_payload.get("rows")))
                    form = web_payload.get("form")
                    if isinstance(form, dict):
                        self.web_frame.link_var.set(str(form.get("path_or_link", "")))

            ui = payload.get("ui")
            if isinstance(ui, dict):
                self.article_browser_collapsed = bool(ui.get("article_browser_collapsed", self.article_browser_collapsed))
                self.image_frame.set_preview_visible(bool(ui.get("image_preview_visible", getattr(self.image_frame, "preview_visible", True))))
                self.document_frame.set_preview_visible(bool(ui.get("document_preview_visible", getattr(self.document_frame, "preview_visible", True))))
                self.video_frame.set_preview_visible(bool(ui.get("video_preview_visible", getattr(self.video_frame, "preview_visible", True))))
                self.web_frame.set_preview_visible(bool(ui.get("web_preview_visible", getattr(self.web_frame, "preview_visible", True))))

            selected_tab = payload.get("selected_tab")
            if isinstance(selected_tab, int) and 0 <= selected_tab < len(self.main_notebook.tabs()):
                self.main_notebook.select(selected_tab)

            browser_state = payload.get("article_browser")
            if isinstance(browser_state, dict):
                self.pending_article_browser_selection = normalize_article_number(str(browser_state.get("selected_article", "")))

        self._apply_article_browser_visibility()
        self._schedule_project_tab_layout()

    def _apply_pending_session_selection(self) -> None:
        selected_article = normalize_article_number(self.pending_article_browser_selection)
        if not selected_article or selected_article not in self.article_browser_records:
            return
        self.article_browser_tree.selection_set(selected_article)
        self.article_browser_tree.focus(selected_article)
        self._update_article_browser_detail(self.article_browser_records[selected_article])

    def _handle_close(self) -> None:
        try:
            self._save_session_state()
        except Exception:
            pass
        self.root.destroy()

    def _open_github_releases_page(self) -> None:
        webbrowser.open(GITHUB_RELEASES_PAGE_URL)

    def _check_for_github_updates(self) -> None:
        self.update_status_var.set("Prüfe GitHub-Releases...")

        def worker() -> GitHubReleaseInfo:
            release = fetch_latest_github_release()
            if not release.tag_name:
                raise ValueError("GitHub Release enthält kein gültiges Versions-Tag.")
            return release

        def on_success(result: object) -> None:
            if not isinstance(result, GitHubReleaseInfo):
                raise ValueError("Unerwartete Antwort bei der Update-Prüfung.")

            release = result
            latest_label = release.name or release.tag_name
            if not is_newer_release_tag(APP_VERSION_TAG, release.tag_name):
                self.update_status_var.set(f"Aktuelle Version: {APP_VERSION_TAG} (kein neueres Release)")
                self.status_var.set(f"Kein neueres Release gefunden. Aktuell installiert: {APP_VERSION_TAG}")
                messagebox.showinfo(APP_TITLE, f"Du nutzt bereits die neueste Version.\n\nInstalliert: {APP_VERSION_TAG}")
                return

            asset = choose_release_asset(release)
            self.update_status_var.set(f"Update verfügbar: {release.tag_name}")

            body_lines = [line.strip("- ").strip() for line in release.body.splitlines() if line.strip()]
            notes_preview = "\n".join(body_lines[:4])
            message_parts = [
                f"Aktuell installiert: {APP_VERSION_TAG}",
                f"Neues Release: {release.tag_name}",
                f"Titel: {latest_label}",
            ]
            if asset is not None:
                message_parts.append(f"Gefundenes Paket: {asset.name}")
            if notes_preview:
                message_parts.append("")
                message_parts.append("Release-Hinweise:")
                message_parts.append(notes_preview)
            message_parts.append("")
            message_parts.append("Soll das Update jetzt heruntergeladen und gestartet werden?")
            wants_update = messagebox.askyesno(APP_TITLE, "\n".join(message_parts))
            if not wants_update:
                self.status_var.set(f"Update verfügbar: {release.tag_name}")
                return

            if asset is None:
                self.status_var.set(f"Kein direkt installierbares Paket in {release.tag_name} gefunden.")
                messagebox.showinfo(
                    APP_TITLE,
                    "Im neuesten Release wurde kein direkt installierbares Paket gefunden.\n\n"
                    "Die Release-Seite wird jetzt geöffnet.",
                )
                self._open_github_releases_page()
                return

            self._download_and_install_github_release(release, asset)

        self._run_background_task(
            "Suche auf GitHub nach neuen Releases...",
            worker,
            on_success,
            "Update-Prüfung fehlgeschlagen",
        )

    def _download_and_install_github_release(self, release: GitHubReleaseInfo, asset: GitHubReleaseAsset) -> None:
        target_dir = Path(tempfile.gettempdir()) / "ApolloImportUpdates" / safe_folder_name(release.tag_name)
        target_path = target_dir / asset.name

        def worker() -> Path:
            return download_release_asset(asset, target_path)

        def on_success(result: object) -> None:
            downloaded_path = Path(str(result))
            self._install_downloaded_release(release, asset, downloaded_path)

        self._run_background_task(
            f"Lade Update {release.tag_name} herunter...",
            worker,
            on_success,
            "Update-Download fehlgeschlagen",
        )

    def _install_downloaded_release(
        self,
        release: GitHubReleaseInfo,
        asset: GitHubReleaseAsset,
        downloaded_path: Path,
    ) -> None:
        if asset.suffix == ".zip":
            self.status_var.set(f"Update heruntergeladen: {downloaded_path.name}")
            messagebox.showinfo(
                APP_TITLE,
                "Das neueste Release wurde als ZIP heruntergeladen.\n\n"
                f"Datei: {downloaded_path}\n\n"
                "Ich öffne jetzt den Ordner, damit du das Paket direkt verwenden kannst.",
            )
            os.startfile(str(downloaded_path.parent))
            return

        if "setup" in asset.name.casefold() or "installer" in asset.name.casefold():
            os.startfile(str(downloaded_path))
            self.status_var.set(f"Installer gestartet: {release.tag_name}")
            messagebox.showinfo(
                APP_TITLE,
                "Der Update-Installer wurde gestartet.\n\n"
                "Nach Abschluss der Installation kannst du die neue Version direkt verwenden.",
            )
            return

        if not getattr(sys, "frozen", False):
            self.status_var.set(f"Update heruntergeladen: {downloaded_path.name}")
            messagebox.showinfo(
                APP_TITLE,
                "Das Update wurde heruntergeladen, aber die automatische Ersetzung funktioniert nur in der gebauten EXE.\n\n"
                f"Datei: {downloaded_path}\n\n"
                "Ich öffne jetzt den Ordner mit der heruntergeladenen Datei.",
            )
            os.startfile(str(downloaded_path.parent))
            return

        self._install_portable_exe_update(release, downloaded_path)

    def _install_portable_exe_update(self, release: GitHubReleaseInfo, downloaded_path: Path) -> None:
        current_executable = Path(sys.executable).resolve()
        if not current_executable.exists():
            raise ValueError("Die aktuelle EXE konnte nicht gefunden werden.")

        update_dir = Path(tempfile.gettempdir()) / "ApolloImportUpdates"
        update_dir.mkdir(parents=True, exist_ok=True)
        script_path = update_dir / f"apply_update_{os.getpid()}.ps1"

        def ps_quote(value: str) -> str:
            return value.replace("'", "''")

        source_path = ps_quote(str(downloaded_path))
        target_path = ps_quote(str(current_executable))
        working_dir = ps_quote(str(current_executable.parent))
        process_id = os.getpid()

        script_body = f"""$ErrorActionPreference = 'Stop'
$source = '{source_path}'
$target = '{target_path}'
$workingDir = '{working_dir}'
$oldPid = {process_id}
for ($i = 0; $i -lt 240; $i++) {{
    try {{
        Get-Process -Id $oldPid -ErrorAction Stop | Out-Null
        Start-Sleep -Milliseconds 500
    }} catch {{
        break
    }}
}}
$copied = $false
for ($i = 0; $i -lt 120; $i++) {{
    try {{
        Copy-Item -LiteralPath $source -Destination $target -Force
        $copied = $true
        break
    }} catch {{
        Start-Sleep -Milliseconds 500
    }}
}}
Start-Sleep -Milliseconds 300
if ($copied) {{
    Start-Process -FilePath $target -WorkingDirectory $workingDir -WindowStyle Normal
}} else {{
    Start-Process -FilePath $source -WorkingDirectory (Split-Path -LiteralPath $source -Parent) -WindowStyle Normal
}}
"""
        script_path.write_text(script_body, encoding="utf-8")

        try:
            self._save_session_state()
        except Exception:
            pass

        # Der PowerShell-Prozess erbt die Bootloader-Variablen der laufenden
        # Onefile-EXE (_PYI_*). Ohne Bereinigung haelt sich die danach neu
        # gestartete EXE fuer den Extraktions-Kindprozess der alten, bereits
        # beendeten Instanz und bricht mit einem Bootloader-Fehler ab.
        launch_env = {
            key: value
            for key, value in os.environ.items()
            if key != "_MEIPASS2" and not key.startswith("_PYI_")
        }
        launch_env["PYINSTALLER_RESET_ENVIRONMENT"] = "1"

        subprocess.Popen(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-WindowStyle",
                "Hidden",
                "-File",
                str(script_path),
            ],
            close_fds=True,
            env=launch_env,
        )
        self.status_var.set(f"Update wird installiert: {release.tag_name}")
        messagebox.showinfo(
            APP_TITLE,
            "Das Update wurde heruntergeladen.\n\n"
            "Die Anwendung wird jetzt geschlossen, die EXE ersetzt und danach automatisch neu gestartet.",
        )
        self.root.after(200, self.root.destroy)

    def open_settings_dialog(self, initial_tab: int = 0) -> None:
        if self.settings_dialog is not None and self.settings_dialog.winfo_exists():
            self.settings_dialog.deiconify()
            self.settings_dialog.lift()
            self.settings_dialog.focus_set()
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Einstellungen")
        dialog.transient(self.root)
        dialog.geometry("1080x680")
        dialog.minsize(880, 540)
        dialog.configure(background=UI_BACKGROUND)
        dialog.protocol("WM_DELETE_WINDOW", self._close_settings_dialog)
        self.settings_dialog = dialog
        self._settings_tab_canvases = []

        container = ttk.Frame(dialog, padding=14)
        container.pack(fill="both", expand=True)
        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)

        notebook = ttk.Notebook(container)
        notebook.grid(row=0, column=0, sticky="nsew")
        self.settings_notebook = notebook

        folders_tab = self._make_settings_scroll_tab(notebook, "Ordner")
        export_tab = self._make_settings_scroll_tab(notebook, "Export")
        api_tab = self._make_settings_scroll_tab(notebook, "API & Datenstämme")

        self._build_settings_folders_section(folders_tab)
        self._build_settings_export_section(export_tab)
        self._build_settings_api_section(api_tab)

        button_row = ttk.Frame(container)
        button_row.grid(row=1, column=0, sticky="e", pady=(12, 0))
        ttk.Button(button_row, text="Schließen", command=self._close_settings_dialog).grid(row=0, column=0)

        if not getattr(self, "_settings_mousewheel_bound", False):
            self.root.bind_all("<MouseWheel>", self._handle_settings_mousewheel, add="+")
            self.root.bind_all("<Button-4>", self._handle_settings_mousewheel, add="+")
            self.root.bind_all("<Button-5>", self._handle_settings_mousewheel, add="+")
            self._settings_mousewheel_bound = True

        try:
            notebook.select(initial_tab)
        except tk.TclError:  # pragma: no cover - defensive
            pass
        dialog.focus_set()

    def _close_settings_dialog(self) -> None:
        if self.settings_dialog is not None:
            try:
                self.settings_dialog.destroy()
            except tk.TclError:  # pragma: no cover - defensive
                pass
        self.settings_dialog = None
        self.settings_notebook = None
        self._settings_tab_canvases = []

    def _make_settings_scroll_tab(self, notebook: ttk.Notebook, title: str) -> ttk.Frame:
        tab = ttk.Frame(notebook)
        notebook.add(tab, text=title)
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(0, weight=1)

        canvas = tk.Canvas(tab, background=UI_BACKGROUND, borderwidth=0, highlightthickness=0)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=scrollbar.set)

        content = ttk.Frame(canvas, padding=18)
        window = canvas.create_window((0, 0), window=content, anchor="nw")
        content.bind(
            "<Configure>",
            lambda _event, canvas=canvas: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas.bind(
            "<Configure>",
            lambda event, canvas=canvas, window=window: canvas.itemconfigure(window, width=event.width),
        )
        self._settings_tab_canvases.append((str(tab), canvas))
        return content

    def _handle_settings_mousewheel(self, event: "tk.Event[tk.Misc]") -> str | None:
        dialog = self.settings_dialog
        if dialog is None or not dialog.winfo_exists():
            return None
        widget_path = str(event.widget)
        if not widget_path.startswith(str(dialog)):
            return None
        if getattr(event, "num", None) == 4:
            delta = -1
        elif getattr(event, "num", None) == 5:
            delta = 1
        else:
            delta = -1 if event.delta > 0 else 1
        for tab_path, canvas in self._settings_tab_canvases:
            if widget_path.startswith(tab_path):
                canvas.yview_scroll(delta, "units")
                return "break"
        return None

    def _build_settings_folders_section(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(1, weight=1)

        ttk.Label(parent, text="Ordner", font=("Segoe UI Semibold", 10)).grid(
            row=0, column=0, columnspan=4, sticky="w", pady=(0, 4)
        )

        ttk.Label(parent, text="Importordner").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(parent, textvariable=self.import_dir_var).grid(row=1, column=1, sticky="ew", pady=6)
        ttk.Button(parent, text="Wählen", command=self.choose_import_dir).grid(row=1, column=2, padx=(8, 0), pady=6)
        ttk.Button(parent, text="IDs laden", command=self._load_known_ids).grid(row=1, column=3, padx=(8, 0), pady=6)

        ttk.Label(parent, text="Ausgabeordner / Importpfad").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(parent, textvariable=self.output_dir_var).grid(row=2, column=1, sticky="ew", pady=6)
        ttk.Button(parent, text="Wählen", command=self.choose_output_dir).grid(row=2, column=2, padx=(8, 0), pady=6)
        ttk.Label(parent, textvariable=self.known_id_count_var, foreground="#5E6472").grid(
            row=2, column=3, sticky="w", padx=(8, 0), pady=6
        )

        ttk.Separator(parent, orient="horizontal").grid(row=3, column=0, columnspan=4, sticky="ew", pady=(14, 12))

        ttk.Label(parent, text="App-Update", font=("Segoe UI Semibold", 10)).grid(
            row=4, column=0, columnspan=4, sticky="w", pady=(0, 4)
        )
        ttk.Label(parent, textvariable=self.update_status_var, foreground="#5E6472").grid(
            row=5, column=0, columnspan=2, sticky="w", pady=6
        )
        update_actions = ttk.Frame(parent)
        update_actions.grid(row=5, column=2, columnspan=2, sticky="w", padx=(8, 0), pady=6)
        ttk.Button(update_actions, text="Nach Updates suchen", command=self._check_for_github_updates).grid(row=0, column=0)
        ttk.Button(update_actions, text="Releases", command=self._open_github_releases_page).grid(row=0, column=1, padx=(8, 0))

    def _build_settings_export_section(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(0, weight=1)

        ttk.Label(
            parent,
            text="Du kannst die zehn Exportdateien entweder in einen neuen Zeitstempel-Unterordner schreiben oder immer direkt im Ausgabeordner aktualisieren.",
            wraplength=760,
            foreground="#5E6472",
        ).grid(row=0, column=0, sticky="w")

        ttk.Checkbutton(
            parent,
            text="Immer direkt in den Ausgabeordner schreiben (fester Importpfad)",
            variable=self.fixed_export_path_var,
        ).grid(row=1, column=0, sticky="w", pady=(12, 0))

        ttk.Label(
            parent,
            text="Wenn aktiv, bleiben die Dateien bestehen: neue Artikel werden angehängt und bestehende Zeilen derselben Artikelnummer ersetzt.",
            wraplength=760,
            foreground="#5E6472",
        ).grid(row=2, column=0, sticky="w", pady=(8, 0))

        ttk.Label(
            parent,
            text="Der Massenimport schreibt unabhängig von dieser Einstellung immer direkt in den festen Output-Pfad.",
            wraplength=760,
            foreground="#5E6472",
        ).grid(row=3, column=0, sticky="w", pady=(8, 0))

    def _build_settings_api_section(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(1, weight=1)

        ttk.Label(parent, text="DeepL", font=("Segoe UI Semibold", 10)).grid(row=0, column=0, sticky="w", pady=(0, 4))

        ttk.Label(parent, text="API Key").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(parent, textvariable=self.deepl_api_key_var, show="*", width=60).grid(row=1, column=1, sticky="ew", pady=6)

        ttk.Label(parent, text="Base URL").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(parent, textvariable=self.deepl_base_url_var).grid(row=2, column=1, sticky="ew", pady=6)
        ttk.Label(
            parent,
            text="Pro: https://api.deepl.com  |  Free: https://api-free.deepl.com",
            foreground="#5E6472",
        ).grid(row=2, column=2, sticky="w", padx=(10, 0), pady=6)

        deepl_actions = ttk.Frame(parent)
        deepl_actions.grid(row=3, column=1, columnspan=2, sticky="w", pady=(10, 0))
        ttk.Button(deepl_actions, text="Kurzbezeichnung übersetzen", command=self.translate_short_texts).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(deepl_actions, text="Text übersetzen", command=self.translate_long_texts).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(deepl_actions, text="Alles übersetzen", command=self.translate_all_texts).grid(row=0, column=2)

        ttk.Label(
            parent,
            text="Die Übersetzung läuft jeweils aus dem deutschen Feld in EN, CZ, FR, IT und NL. UNI bleibt an Deutsch gekoppelt.",
            foreground="#5E6472",
            wraplength=760,
        ).grid(row=4, column=0, columnspan=3, sticky="w", pady=(10, 0))

        ttk.Separator(parent, orient="horizontal").grid(row=5, column=0, columnspan=3, sticky="ew", pady=(14, 12))

        ttk.Label(parent, text="Datenstämme", font=("Segoe UI Semibold", 10)).grid(
            row=9, column=0, sticky="w", pady=(0, 4)
        )
        ttk.Label(parent, text="GenArt XLSX").grid(row=10, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(parent, textvariable=self.genart_source_path_var).grid(row=10, column=1, sticky="ew", pady=6)
        genart_actions = ttk.Frame(parent)
        genart_actions.grid(row=10, column=2, sticky="w", padx=(10, 0), pady=6)
        ttk.Button(genart_actions, text="Datei wählen", command=self.choose_genart_source_file).grid(row=0, column=0)
        ttk.Button(genart_actions, text="Neu laden", command=self._reload_genart_catalog).grid(row=0, column=1, padx=(8, 0))
        ttk.Label(parent, textvariable=self.genart_count_var, foreground="#5E6472").grid(
            row=11, column=1, columnspan=2, sticky="w"
        )

        ttk.Label(parent, text="KHer CSV").grid(row=12, column=0, sticky="w", padx=(0, 10), pady=6)
        kher_actions = ttk.Frame(parent)
        kher_actions.grid(row=12, column=2, sticky="w", padx=(10, 0), pady=6)
        ttk.Button(kher_actions, text="Datei wählen", command=self.choose_competitor_source_file).grid(row=0, column=0)
        ttk.Button(kher_actions, text="Neu laden", command=self._reload_competitor_catalog).grid(row=0, column=1, padx=(8, 0))
        ttk.Entry(parent, textvariable=self.competitor_source_path_var).grid(row=12, column=1, sticky="ew", pady=6)
        ttk.Label(
            parent,
            text="Diese Datei wird für OE-Nummern und Vergleichsnummern verwendet. OE nutzt alle Hersteller, Vergleichsnummern nur Einträge mit VGL-Flag.",
            foreground="#5E6472",
            wraplength=760,
        ).grid(row=13, column=0, columnspan=3, sticky="w", pady=(6, 0))
        ttk.Label(parent, textvariable=self.competitor_count_var, foreground="#5E6472").grid(
            row=14, column=1, columnspan=2, sticky="w"
        )

        ttk.Label(parent, text="Attribute XLSX").grid(row=15, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(parent, textvariable=self.attribute_source_path_var).grid(row=15, column=1, sticky="ew", pady=6)
        attribute_actions = ttk.Frame(parent)
        attribute_actions.grid(row=15, column=2, sticky="w", padx=(10, 0), pady=6)
        ttk.Button(attribute_actions, text="Datei wählen", command=self.choose_attribute_source_file).grid(row=0, column=0)
        ttk.Button(attribute_actions, text="Neu laden", command=self._reload_attribute_catalog).grid(row=0, column=1, padx=(8, 0))

        ttk.Label(parent, text="Schlüsselwerte XLSX").grid(row=16, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(parent, textvariable=self.attribute_key_value_source_path_var).grid(
            row=16, column=1, sticky="ew", pady=6
        )
        key_value_actions = ttk.Frame(parent)
        key_value_actions.grid(row=16, column=2, sticky="w", padx=(10, 0), pady=6)
        ttk.Button(key_value_actions, text="Datei wählen", command=self.choose_attribute_key_value_source_file).grid(row=0, column=0)
        ttk.Button(key_value_actions, text="Neu laden", command=self._reload_attribute_key_value_catalog).grid(row=0, column=1, padx=(8, 0))
        ttk.Label(
            parent,
            text="Die Attributliste liefert TecDoc Kriterien, Formate und Bezeichnungen; Schlüsselwerte liefern die möglichen Auswahlwerte.",
            foreground="#5E6472",
            wraplength=760,
        ).grid(row=17, column=0, columnspan=3, sticky="w", pady=(6, 0))
        ttk.Label(parent, textvariable=self.attribute_count_var, foreground="#5E6472").grid(
            row=18, column=1, columnspan=2, sticky="w"
        )
        ttk.Label(parent, textvariable=self.attribute_key_value_count_var, foreground="#5E6472").grid(
            row=19, column=1, columnspan=2, sticky="w"
        )

        ttk.Label(parent, text="KTyp XLSX").grid(row=20, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(parent, textvariable=self.vehicle_source_path_var).grid(row=20, column=1, sticky="ew", pady=6)
        vehicle_actions = ttk.Frame(parent)
        vehicle_actions.grid(row=20, column=2, sticky="w", padx=(10, 0), pady=6)
        ttk.Button(vehicle_actions, text="Datei wählen", command=self.choose_vehicle_source_file).grid(row=0, column=0)
        ttk.Button(vehicle_actions, text="Neu laden", command=self._reload_vehicle_catalog).grid(row=0, column=1, padx=(8, 0))
        ttk.Label(
            parent,
            text=(
                "KTyp-Stammdaten (Motorcode -> KTyp-Nummer). Wird im Tab 'Fahrzeuge' verwendet, um über Motorcodes "
                "die KTyp-Nummern (TopMotive/TecDoc) zu finden. Datei bei Bedarf aktualisieren und 'Neu laden' drücken."
            ),
            foreground="#5E6472",
            wraplength=760,
        ).grid(row=21, column=0, columnspan=3, sticky="w", pady=(6, 0))
        ttk.Label(parent, textvariable=self.vehicle_count_var, foreground="#5E6472").grid(
            row=22, column=1, columnspan=2, sticky="w"
        )

        ttk.Label(parent, text="Attribut-Zuordnung XLSX").grid(row=23, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(parent, textvariable=self.attribute_mapping_source_path_var).grid(row=23, column=1, sticky="ew", pady=6)
        mapping_actions = ttk.Frame(parent)
        mapping_actions.grid(row=23, column=2, sticky="w", padx=(10, 0), pady=6)
        ttk.Button(mapping_actions, text="Datei wählen", command=self.choose_attribute_mapping_source_file).grid(row=0, column=0)
        ttk.Button(mapping_actions, text="Neu laden", command=self._reload_attribute_mapping).grid(row=0, column=1, padx=(8, 0))
        ttk.Label(
            parent,
            text=(
                "Zuordnung von Text-Labels (z.B. 'Gewicht') zu TecDoc Kriterien IDs für 'Attribute aus Text vorschlagen'. "
                "Manuell zugewiesene Labels können aus dem Vorschlagsdialog automatisch ergänzt werden."
            ),
            foreground="#5E6472",
            wraplength=760,
        ).grid(row=24, column=0, columnspan=3, sticky="w", pady=(6, 0))
        ttk.Label(parent, textvariable=self.attribute_mapping_count_var, foreground="#5E6472").grid(
            row=25, column=1, columnspan=2, sticky="w"
        )

    def _toggle_article_browser_section(self) -> None:
        self.article_browser_collapsed = not self.article_browser_collapsed
        self._apply_article_browser_visibility()

    def _apply_article_browser_visibility(self) -> None:
        self.article_browser_toggle_var.set("Einblenden" if self.article_browser_collapsed else "Einklappen")
        if self.article_browser_collapsed:
            self.browser_content_frame.grid_remove()
        else:
            self.browser_content_frame.grid()
            self._schedule_article_browser_layout()

    def _handle_project_scroll_content_configure(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        if not hasattr(self, "project_canvas"):
            return
        self.project_canvas.configure(scrollregion=self.project_canvas.bbox("all"))

    def _handle_project_canvas_configure(self, event: tk.Event[tk.Misc]) -> None:
        if hasattr(self, "project_canvas_window"):
            self.project_canvas.itemconfigure(self.project_canvas_window, width=event.width)
        self._schedule_project_tab_layout()

    def _is_project_widget(self, widget: tk.Misc) -> bool:
        current: tk.Misc | None = widget
        while current is not None:
            if current is self.project_tab:
                return True
            current = getattr(current, "master", None)
        return False

    def _bind_project_mousewheel(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        if not hasattr(self, "project_canvas") or getattr(self, "_project_mousewheel_bound", False):
            return
        self.root.bind_all("<MouseWheel>", self._handle_project_mousewheel, add="+")
        self.root.bind_all("<Button-4>", self._handle_project_mousewheel, add="+")
        self.root.bind_all("<Button-5>", self._handle_project_mousewheel, add="+")
        self._project_mousewheel_bound = True

    def _unbind_project_mousewheel(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        return

    def _handle_project_mousewheel(self, event: tk.Event[tk.Misc]) -> str | None:
        if not hasattr(self, "project_canvas") or not self._is_project_widget(event.widget):
            return None
        widget_class = event.widget.winfo_class()
        if widget_class in {"Treeview", "Text"}:
            return None
        if getattr(event, "num", None) == 4:
            delta = -1
        elif getattr(event, "num", None) == 5:
            delta = 1
        else:
            delta = -1 if event.delta > 0 else 1
        self.project_canvas.yview_scroll(delta, "units")
        return "break"

    def _schedule_project_tab_layout(self) -> None:
        if self._project_layout_after_id is not None:
            try:
                self.root.after_cancel(self._project_layout_after_id)
            except Exception:
                pass
        self._project_layout_after_id = self.root.after_idle(self._apply_project_tab_layout)

    def _handle_project_tab_resize(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        self._schedule_project_tab_layout()

    def _apply_project_tab_layout(self) -> None:
        self._project_layout_after_id = None
        project_parent = getattr(self, "project_content_frame", self.project_tab)
        if hasattr(self, "project_canvas"):
            width = max(self.project_canvas.winfo_width(), 1)
        else:
            width = max(self.project_tab.winfo_width(), self.project_tab.winfo_reqwidth())
        compact = width < 1380
        layout_changed = self.project_tab_compact_mode is None or self.project_tab_compact_mode != compact
        self.project_tab_compact_mode = compact

        project_parent.columnconfigure(0, weight=1, uniform="project_cols")
        project_parent.columnconfigure(1, weight=1 if not compact else 0, uniform="project_cols")
        for row_index in range(0, 6):
            project_parent.rowconfigure(row_index, weight=0)

        if layout_changed:
            if compact:
                self.single_import_frame.grid_configure(row=0, column=0, columnspan=2, sticky="ew", padx=(0, 0), pady=(0, 0))
                self.batch_import_frame.grid_configure(row=1, column=0, columnspan=2, sticky="ew", padx=(0, 0), pady=(14, 0))
                self.scope_frame.grid_configure(row=2, column=0, columnspan=2, sticky="ew", padx=(0, 0), pady=(14, 0))
                self.browser_frame.grid_configure(row=3, column=0, columnspan=2, sticky="nsew", padx=(0, 0), pady=(14, 0))
                project_parent.rowconfigure(3, weight=1)
            else:
                self.single_import_frame.grid_configure(row=0, column=0, columnspan=1, sticky="nsew", padx=(0, 9), pady=(0, 0))
                self.batch_import_frame.grid_configure(row=0, column=1, columnspan=1, sticky="nsew", padx=(9, 0), pady=(0, 0))
                self.scope_frame.grid_configure(row=1, column=0, columnspan=2, sticky="ew", padx=(0, 0), pady=(14, 0))
                self.browser_frame.grid_configure(row=2, column=0, columnspan=2, sticky="nsew", padx=(0, 0), pady=(14, 0))
                project_parent.rowconfigure(2, weight=1)

        self._handle_project_scroll_content_configure()
        self._schedule_article_browser_layout()

    def _schedule_article_browser_layout(self) -> None:
        if self._article_browser_layout_after_id is not None:
            try:
                self.root.after_cancel(self._article_browser_layout_after_id)
            except Exception:
                pass
        self._article_browser_layout_after_id = self.root.after_idle(self._apply_article_browser_layout)

    def _handle_article_browser_resize(self, _event: tk.Event[tk.Misc] | None = None) -> None:
        self._schedule_article_browser_layout()

    def _apply_article_browser_layout(self) -> None:
        self._article_browser_layout_after_id = None
        if self.article_browser_collapsed:
            return

        width = max(self.browser_content_frame.winfo_width(), self.browser_content_frame.winfo_reqwidth())
        compact = self.project_tab_compact_mode or width < 1320
        layout_changed = self.article_browser_compact_mode is None or self.article_browser_compact_mode != compact
        self.article_browser_compact_mode = compact

        self.browser_content_frame.columnconfigure(0, weight=1)
        self.browser_content_frame.columnconfigure(1, weight=1 if not compact else 0)
        self.browser_content_frame.rowconfigure(0, weight=1)
        self.browser_content_frame.rowconfigure(1, weight=0)

        if layout_changed:
            if compact:
                self.browser_table_frame.grid_configure(row=0, column=0, sticky="nsew", padx=(0, 0), pady=(0, 12))
                self.article_detail_frame.grid_configure(row=1, column=0, sticky="nsew", padx=(0, 0), pady=(0, 0))
            else:
                self.browser_table_frame.grid_configure(row=0, column=0, sticky="nsew", padx=(0, 14), pady=(0, 0))
                self.article_detail_frame.grid_configure(row=0, column=1, sticky="nsew", padx=(0, 0), pady=(0, 0))

    def _build_project_tab(self) -> None:
        self.project_tab.columnconfigure(0, weight=1)
        self.project_tab.rowconfigure(0, weight=1)

        self.project_canvas = tk.Canvas(self.project_tab, background=UI_BACKGROUND, borderwidth=0, highlightthickness=0)
        self.project_canvas.grid(row=0, column=0, sticky="nsew")
        project_scrollbar = ttk.Scrollbar(self.project_tab, orient="vertical", command=self.project_canvas.yview)
        project_scrollbar.grid(row=0, column=1, sticky="ns")
        self.project_canvas.configure(yscrollcommand=project_scrollbar.set)

        self.project_content_frame = ttk.Frame(self.project_canvas)
        self.project_canvas_window = self.project_canvas.create_window((0, 0), window=self.project_content_frame, anchor="nw")
        self.project_content_frame.bind("<Configure>", self._handle_project_scroll_content_configure)
        self.project_canvas.bind("<Configure>", self._handle_project_canvas_configure)
        self._bind_project_mousewheel()

        project_parent = self.project_content_frame
        project_parent.columnconfigure(0, weight=1, uniform="project_cols")
        project_parent.columnconfigure(1, weight=1, uniform="project_cols")
        project_parent.rowconfigure(2, weight=1)

        self.single_import_frame = ttk.LabelFrame(project_parent, text="Einzelimport", padding=14)
        single_frame = self.single_import_frame
        single_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 9))
        single_frame.columnconfigure(1, weight=1)

        ttk.Label(
            single_frame,
            text="Einen Artikel erfassen und optional direkt aus Kunzer befüllen.",
            foreground="#5E6472",
            wraplength=500,
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        ttk.Label(single_frame, text="Artikelnummer").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)
        article_entry = ttk.Entry(single_frame, textvariable=self.article_number_var)
        article_entry.grid(row=1, column=1, sticky="ew", pady=6)
        article_entry.bind("<FocusOut>", self._on_article_entry_focus_out)

        ttk.Label(
            single_frame,
            text="Kurz- und Text-ID werden automatisch im Hintergrund vergeben.",
            foreground="#5E6472",
        ).grid(row=2, column=0, columnspan=2, sticky="w", pady=(0, 6))

        ttk.Button(single_frame, text="Artikel anlegen", style="Accent.TButton", command=self.load_from_kunzer_article_number).grid(
            row=3, column=0, columnspan=2, sticky="w", pady=(10, 0)
        )

        self.batch_import_frame = ttk.LabelFrame(project_parent, text="Massenimport", padding=14)
        batch_import_frame = self.batch_import_frame
        batch_import_frame.grid(row=0, column=1, sticky="nsew", padx=(9, 0))
        batch_import_frame.columnconfigure(1, weight=1)

        ttk.Label(
            batch_import_frame,
            text="Produktliste als CSV/XLSX laden: alle Artikel werden von Kunzer abgerufen und direkt in die Output-Dateien geschrieben.",
            foreground="#5E6472",
            wraplength=500,
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 4))

        ttk.Label(
            batch_import_frame,
            text="Erwartete Spalte in der Kopfzeile: „Artikelnummer“ (auch ArtNr, Produktnummer, SKU) oder „Produkt-URL“ (auch URL, Link, Produktlink). Eine der beiden Spalten genügt.",
            foreground="#5E6472",
            wraplength=500,
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 8))

        ttk.Label(batch_import_frame, text="Produktliste CSV/XLSX").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(batch_import_frame, textvariable=self.product_list_path_var).grid(row=2, column=1, sticky="ew", pady=6)
        ttk.Button(batch_import_frame, text="Datei wählen", command=self.choose_product_list_file).grid(row=2, column=2, padx=(8, 0), pady=6)

        ttk.Button(batch_import_frame, text="Liste importieren", command=self.import_products_from_file).grid(
            row=3, column=0, columnspan=3, sticky="w", pady=(10, 0)
        )

        ttk.Label(
            batch_import_frame,
            text="Der Massenimport übersetzt ausgewählte Texte automatisch und schreibt immer direkt in den festen Output-Pfad.",
            foreground="#5E6472",
            wraplength=500,
        ).grid(row=4, column=0, columnspan=3, sticky="w", pady=(10, 0))

        self.scope_frame = ttk.LabelFrame(project_parent, text="Abrufumfang", padding=14)
        scope_frame = self.scope_frame
        scope_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(14, 0))
        scope_frame.columnconfigure(0, weight=1)
        scope_frame.columnconfigure(1, weight=1)

        ttk.Label(
            scope_frame,
            text="Diese Auswahl gilt für Einzelimport und Massenimport: nur angehakte Daten werden von Kunzer geholt.",
            wraplength=500,
            foreground="#5E6472",
        ).grid(row=0, column=0, columnspan=2, sticky="w")

        ttk.Checkbutton(scope_frame, text="Kurzbezeichnung", variable=self.batch_short_text_var).grid(row=1, column=0, sticky="w", pady=(10, 0))
        ttk.Checkbutton(scope_frame, text="Text", variable=self.batch_long_text_var).grid(row=1, column=1, sticky="w", pady=(10, 0))
        ttk.Checkbutton(scope_frame, text="Bilder", variable=self.batch_image_var).grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Checkbutton(scope_frame, text="Dokumente", variable=self.batch_document_var).grid(row=2, column=1, sticky="w", pady=(6, 0))
        ttk.Checkbutton(scope_frame, text="Videos", variable=self.batch_video_var).grid(row=3, column=0, sticky="w", pady=(6, 0))
        ttk.Checkbutton(scope_frame, text="Web Links", variable=self.batch_web_var).grid(row=3, column=1, sticky="w", pady=(6, 0))

        self.browser_frame = ttk.LabelFrame(project_parent, text="Artikelverzeichnis", padding=14)
        browser_frame = self.browser_frame
        browser_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=(14, 0))
        browser_frame.columnconfigure(0, weight=1)
        browser_frame.rowconfigure(1, weight=1)

        browser_header = ttk.Frame(browser_frame)
        browser_header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        browser_header.columnconfigure(0, weight=1)
        ttk.Label(
            browser_header,
            text="Hier werden nur Artikel angezeigt, die bereits in den Output-Excel-Dateien vorhanden sind. Du kannst sie von dort erneut in die Maske laden.",
            foreground="#5E6472",
            wraplength=900,
        ).grid(row=0, column=0, sticky="w")
        ttk.Button(browser_header, textvariable=self.article_browser_toggle_var, command=self._toggle_article_browser_section).grid(
            row=0, column=1, padx=(12, 8)
        )
        ttk.Button(browser_header, text="Ausgewählten Artikel laden", command=self.load_selected_article_from_browser).grid(
            row=0, column=2, padx=(0, 8)
        )
        ttk.Button(browser_header, text="Liste aktualisieren", command=self.refresh_preview).grid(row=0, column=3)

        columns = ("article", "source", "short_id", "long_id", "genart", "attr", "oe", "comparison", "images", "documents", "videos", "links")
        self.browser_content_frame = ttk.Frame(browser_frame)
        self.browser_content_frame.grid(row=1, column=0, sticky="nsew")
        self.browser_content_frame.columnconfigure(0, weight=1)
        self.browser_content_frame.columnconfigure(1, weight=1)
        self.browser_content_frame.rowconfigure(0, weight=1)

        browser_table_frame = ttk.Frame(self.browser_content_frame)
        self.browser_table_frame = browser_table_frame
        browser_table_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        browser_table_frame.columnconfigure(0, weight=1)
        browser_table_frame.rowconfigure(0, weight=1)

        self.article_browser_tree = ttk.Treeview(
            browser_table_frame,
            columns=columns,
            show="headings",
            height=10,
            selectmode="extended",
        )
        self.article_browser_tree.grid(row=0, column=0, sticky="nsew")
        self.article_browser_tree.heading("article", text="Artikelnummer")
        self.article_browser_tree.heading("source", text="Quelle")
        self.article_browser_tree.heading("short_id", text="Kurz-ID")
        self.article_browser_tree.heading("long_id", text="Text-ID")
        self.article_browser_tree.heading("genart", text="GenArten")
        self.article_browser_tree.heading("attr", text="Attr.")
        self.article_browser_tree.heading("oe", text="OE")
        self.article_browser_tree.heading("comparison", text="Vgl.")
        self.article_browser_tree.heading("images", text="Bilder")
        self.article_browser_tree.heading("documents", text="Dokumente")
        self.article_browser_tree.heading("videos", text="Videos")
        self.article_browser_tree.heading("links", text="Web Links")
        self.article_browser_tree.column("article", width=220, anchor="w")
        self.article_browser_tree.column("source", width=90, anchor="center")
        self.article_browser_tree.column("short_id", width=100, anchor="center")
        self.article_browser_tree.column("long_id", width=100, anchor="center")
        self.article_browser_tree.column("genart", width=180, anchor="w")
        self.article_browser_tree.column("attr", width=60, anchor="center")
        self.article_browser_tree.column("oe", width=60, anchor="center")
        self.article_browser_tree.column("comparison", width=60, anchor="center")
        self.article_browser_tree.column("images", width=70, anchor="center")
        self.article_browser_tree.column("documents", width=90, anchor="center")
        self.article_browser_tree.column("videos", width=70, anchor="center")
        self.article_browser_tree.column("links", width=80, anchor="center")
        self.article_browser_tree.bind("<<TreeviewSelect>>", self._on_article_browser_select)
        self.article_browser_tree.bind("<Double-1>", self._on_article_browser_double_click)
        self.article_browser_tree.bind("<Button-3>", self._open_article_browser_context_menu)

        self.article_browser_context_menu = tk.Menu(browser_frame, tearoff=0)
        self.article_browser_context_menu.add_command(label="Zeilen kopieren", command=self.copy_selected_articles_from_browser)
        self.article_browser_context_menu.add_command(label="Zeilen löschen", command=self.delete_selected_articles_from_browser)

        browser_scrollbar = ttk.Scrollbar(browser_table_frame, orient="vertical", command=self.article_browser_tree.yview)
        browser_scrollbar.grid(row=0, column=1, sticky="ns")
        self.article_browser_tree.configure(yscrollcommand=browser_scrollbar.set)

        detail_frame = ttk.LabelFrame(self.browser_content_frame, text="Exportdaten", padding=10)
        self.article_detail_frame = detail_frame
        detail_frame.grid(row=0, column=1, sticky="nsew")
        detail_frame.columnconfigure(0, weight=1)
        detail_frame.rowconfigure(1, weight=1)

        ttk.Label(
            detail_frame,
            text="Die Detailansicht zeigt alle Daten, die aktuell in den Export-Excel-Dateien für den markierten Artikel vorhanden sind.",
            foreground="#5E6472",
            wraplength=380,
        ).grid(row=0, column=0, sticky="w", pady=(0, 10))

        self.article_browser_detail = ScrolledText(detail_frame, wrap="word", height=16, font=("Consolas", 9))
        self.article_browser_detail.grid(row=1, column=0, sticky="nsew")
        self.article_browser_detail.configure(state="disabled")
        self.project_tab.bind("<Configure>", self._handle_project_tab_resize)
        self.browser_content_frame.bind("<Configure>", self._handle_article_browser_resize)
        self._apply_article_browser_visibility()
        self._schedule_project_tab_layout()

    def _build_short_tab(self) -> None:
        self.short_tab.columnconfigure(0, weight=1)
        self.short_tab.rowconfigure(0, weight=1)
        self.short_text_frame = SingleLineTranslationFrame(
            self.short_tab,
            "Kurzbezeichnung pro Sprache",
            on_change=lambda: self._write_live_section("short_text"),
            max_length=SHORT_TEXT_MAX_LENGTH,
        )
        self.short_text_frame.grid(row=0, column=0, sticky="nsew")

    def _build_long_tab(self) -> None:
        self.long_tab.columnconfigure(0, weight=1)
        self.long_tab.rowconfigure(0, weight=1)
        self.long_text_frame = MultiLineTranslationFrame(
            self.long_tab,
            "Langtext pro Sprache",
            on_change=lambda: self._write_live_section("long_text"),
        )
        self.long_text_frame.grid(row=0, column=0, sticky="nsew")

    def _build_genart_tab(self) -> None:
        self.genart_tab.columnconfigure(0, weight=1)
        self.genart_tab.rowconfigure(0, weight=1)

        genart_frame = ttk.LabelFrame(self.genart_tab, text="GenArten pro Artikel", padding=14)
        genart_frame.grid(row=0, column=0, sticky="nsew")
        genart_frame.columnconfigure(0, weight=1)
        genart_frame.rowconfigure(3, weight=1)

        header = ttk.Frame(genart_frame)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header.columnconfigure(0, weight=1)
        ttk.Label(
            header,
            text="Mehrere GenArten können pro Artikel gesetzt und gemeinsam exportiert werden.",
            foreground="#5E6472",
            wraplength=920,
        ).grid(row=0, column=0, sticky="w")
        ttk.Label(header, textvariable=self.selected_genart_count_var, foreground="#5E6472").grid(row=0, column=1, sticky="e", padx=(12, 0))

        input_row = ttk.Frame(genart_frame)
        input_row.grid(row=1, column=0, sticky="ew")
        input_row.columnconfigure(1, weight=1)
        ttk.Label(input_row, text="GenArt").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=6)
        self.genart_entry = ttk.Entry(input_row, textvariable=self.genart_display_var)
        self.genart_entry.grid(row=0, column=1, sticky="ew", pady=6)
        self.genart_entry.bind("<Control-f>", self._open_genart_search_dialog_event)
        self.genart_suggestions = SearchSuggestionPopup(
            self.root,
            self.genart_entry,
            self._genart_suggestion_values,
            self._accept_genart_suggestion,
            on_focus_out=self._normalize_genart_selection,
            on_missing_selection=self._handle_missing_genart_selection,
            min_width=520,
        )
        ttk.Button(input_row, text="Hinzufügen", command=self.add_current_genart_selection).grid(row=0, column=2, sticky="w", padx=(8, 0), pady=6)
        ttk.Button(input_row, text="Suchen...", command=self._open_genart_search_dialog).grid(row=0, column=3, sticky="w", padx=(8, 0), pady=6)

        hint_row = ttk.Frame(genart_frame)
        hint_row.grid(row=2, column=0, sticky="ew")
        hint_row.columnconfigure(0, weight=1)
        ttk.Label(hint_row, textvariable=self.genart_suggestion_var, foreground="#5E6472", wraplength=840).grid(row=0, column=0, sticky="w")

        table_frame = ttk.Frame(genart_frame)
        table_frame.grid(row=3, column=0, sticky="nsew", pady=(12, 0))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        self.selected_genart_tree = ttk.Treeview(
            table_frame,
            columns=("id", "bezeichnung"),
            show="headings",
            height=10,
            selectmode="extended",
        )
        self.selected_genart_tree.grid(row=0, column=0, sticky="nsew")
        self.selected_genart_tree.heading("id", text="GenArt ID")
        self.selected_genart_tree.heading("bezeichnung", text="GenArt Bezeichnung")
        self.selected_genart_tree.column("id", width=160, anchor="w")
        self.selected_genart_tree.column("bezeichnung", width=700, anchor="w")
        self.selected_genart_tree.bind("<Delete>", self._remove_selected_genarts_event)

        tree_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.selected_genart_tree.yview)
        tree_scrollbar.grid(row=0, column=1, sticky="ns")
        self.selected_genart_tree.configure(yscrollcommand=tree_scrollbar.set)

        action_row = ttk.Frame(genart_frame)
        action_row.grid(row=4, column=0, sticky="w", pady=(10, 0))
        ttk.Button(action_row, text="Auswahl entfernen", command=self.remove_selected_genarts).grid(row=0, column=0)
        ttk.Button(action_row, text="Alle GenArten entfernen", command=self.clear_selected_genarts).grid(row=0, column=1, padx=(8, 0))

    def _build_attribute_tab(self) -> None:
        self.attribute_tab.columnconfigure(0, weight=1)
        self.attribute_tab.rowconfigure(0, weight=1)

        self.attribute_frame = AttributeTableFrame(
            self.attribute_tab,
            on_change=lambda: self._write_live_section("attributes"),
        )
        self.attribute_frame.grid(row=0, column=0, sticky="nsew")

        attribute_actions = ttk.Frame(self.attribute_tab)
        attribute_actions.grid(row=1, column=0, sticky="w", pady=(10, 0))
        ttk.Button(
            attribute_actions,
            text="Attribute aus Text vorschlagen",
            command=self._open_attribute_suggestion_dialog,
        ).grid(row=0, column=0)
        ttk.Label(
            attribute_actions,
            text="Liest technische Angaben aus Kurzbezeichnung und Text und schlägt passende Attribute vor.",
            foreground="#5E6472",
        ).grid(row=0, column=1, padx=(10, 0))

    def _build_reference_tabs(self) -> None:
        self.oe_tab.columnconfigure(0, weight=1)
        self.oe_tab.rowconfigure(0, weight=1)
        self.comparison_tab.columnconfigure(0, weight=1)
        self.comparison_tab.rowconfigure(0, weight=1)
        self.vehicle_tab.columnconfigure(0, weight=1)
        self.vehicle_tab.rowconfigure(0, weight=1)
        self.search_term_tab.columnconfigure(0, weight=1)
        self.search_term_tab.rowconfigure(0, weight=1)

        self.search_term_frame = SearchTermFrame(
            self.search_term_tab,
            on_change=lambda: self._write_live_section("search_terms"),
        )
        self.search_term_frame.grid(row=0, column=0, sticky="nsew")

        self.oe_frame = OeNumberTableFrame(
            self.oe_tab,
            on_change=lambda: self._write_live_section("oe_numbers"),
        )
        self.oe_frame.grid(row=0, column=0, sticky="nsew")

        self.comparison_frame = ComparisonTableFrame(
            self.comparison_tab,
            on_change=lambda: self._write_live_section("comparison_numbers"),
        )
        self.comparison_frame.grid(row=0, column=0, sticky="nsew")

        self.vehicle_link_frame = VehicleLinkTableFrame(
            self.vehicle_tab,
            on_change=lambda: self._write_live_section("vehicle_links"),
        )
        self.vehicle_link_frame.grid(row=0, column=0, sticky="nsew")

    def _build_media_tabs(self) -> None:
        self.image_tab.columnconfigure(0, weight=1)
        self.image_tab.rowconfigure(0, weight=1)
        self.document_tab.columnconfigure(0, weight=1)
        self.document_tab.rowconfigure(0, weight=1)
        self.links_tab.columnconfigure(0, weight=1)
        self.links_tab.rowconfigure(0, weight=1)
        self.links_tab.rowconfigure(1, weight=1)

        self.image_frame = MediaTableFrame(
            self.image_tab,
            title="Bilder",
            path_label="Bild-URL",
            path_dialog_title="Bilddateien auswählen",
            default_art="5",
            default_sprache="255",
            browse_filetypes=[("Bilddateien", "*.png *.jpg *.jpeg *.webp *.tif *.tiff"), ("Alle Dateien", "*.*")],
            preview_kind="image",
            on_change=lambda: self._write_live_section("images"),
        )
        self.image_frame.grid(row=0, column=0, sticky="nsew")

        self.document_frame = MediaTableFrame(
            self.document_tab,
            title="Dokumente",
            path_label="Dokument-URL",
            path_dialog_title="Dokumente auswählen",
            default_art="17",
            default_sprache="255",
            browse_filetypes=[("PDF Dateien", "*.pdf"), ("Alle Dateien", "*.*")],
            infer_art=True,
            preview_kind="document",
            on_change=lambda: self._write_live_section("documents"),
        )
        self.document_frame.grid(row=0, column=0, sticky="nsew")

        self.video_frame = LinkTableFrame(
            self.links_tab,
            "Videos",
            "Video-Link",
            preview_kind="video",
            normalize_link_fn=normalize_youtube_url_for_embed,
            on_change=lambda: self._write_live_section("videos"),
        )
        self.video_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))

        self.web_frame = LinkTableFrame(
            self.links_tab,
            "Web Links",
            "Web-Link",
            preview_kind="web",
            on_change=lambda: self._write_live_section("web_links"),
        )
        self.web_frame.grid(row=1, column=0, sticky="nsew")

    def choose_import_dir(self) -> None:
        selected = filedialog.askdirectory(title="Importordner wählen", initialdir=self.import_dir_var.get())
        if selected:
            self.import_dir_var.set(selected)
            self._load_known_ids()
            self.refresh_preview()

    def choose_output_dir(self) -> None:
        selected = filedialog.askdirectory(title="Ausgabeordner wählen", initialdir=self.output_dir_var.get())
        if selected:
            self.output_dir_var.set(selected)
            self.status_var.set(f"Ausgabeordner gesetzt: {selected}")
            self._load_known_ids()
            self.refresh_preview()

    def choose_product_list_file(self) -> None:
        initial_dir = str(Path(self.product_list_path_var.get()).parent) if self.product_list_path_var.get().strip() else str(Path.cwd())
        selected = filedialog.askopenfilename(
            title="Produktliste CSV/XLSX wählen",
            initialdir=initial_dir,
            filetypes=[("Importlisten", "*.csv *.xlsx"), ("CSV Dateien", "*.csv"), ("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if selected:
            self.product_list_path_var.set(selected)
            self.status_var.set(f"Produktliste gesetzt: {selected}")

    def choose_genart_source_file(self) -> None:
        initial_dir = (
            str(Path(self.genart_source_path_var.get()).parent)
            if self.genart_source_path_var.get().strip()
            else str(DEFAULT_GENART_SOURCE.parent)
        )
        selected = filedialog.askopenfilename(
            title="GenArt XLSX wählen",
            initialdir=initial_dir,
            filetypes=[("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if selected:
            self.genart_source_path_var.set(selected)
            self._load_genart_catalog()

    def _reload_genart_catalog(self) -> None:
        self._load_genart_catalog()

    def choose_attribute_source_file(self) -> None:
        initial_dir = (
            str(Path(self.attribute_source_path_var.get()).parent)
            if self.attribute_source_path_var.get().strip()
            else str(DEFAULT_ATTRIBUTE_SOURCE.parent)
        )
        selected = filedialog.askopenfilename(
            title="Attribut XLSX wählen",
            initialdir=initial_dir,
            filetypes=[("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if selected:
            self.attribute_source_path_var.set(selected)
            self._load_attribute_catalog()

    def _reload_attribute_catalog(self) -> None:
        self._load_attribute_catalog()

    def choose_attribute_key_value_source_file(self) -> None:
        initial_dir = (
            str(Path(self.attribute_key_value_source_path_var.get()).parent)
            if self.attribute_key_value_source_path_var.get().strip()
            else str(DEFAULT_ATTRIBUTE_KEY_VALUE_SOURCE.parent)
        )
        selected = filedialog.askopenfilename(
            title="Schlüsselwert XLSX wählen",
            initialdir=initial_dir,
            filetypes=[("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if selected:
            self.attribute_key_value_source_path_var.set(selected)
            self._load_attribute_key_value_catalog()

    def _reload_attribute_key_value_catalog(self) -> None:
        self._load_attribute_key_value_catalog()

    def choose_vehicle_source_file(self) -> None:
        initial_dir = (
            str(Path(self.vehicle_source_path_var.get()).parent)
            if self.vehicle_source_path_var.get().strip()
            else str(DEFAULT_VEHICLE_SOURCE.parent)
        )
        selected = filedialog.askopenfilename(
            title="KTyp XLSX wählen",
            initialdir=initial_dir,
            filetypes=[("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if selected:
            self.vehicle_source_path_var.set(selected)
            self._load_vehicle_catalog()

    def _reload_vehicle_catalog(self) -> None:
        self._load_vehicle_catalog()

    def choose_attribute_mapping_source_file(self) -> None:
        initial_dir = (
            str(Path(self.attribute_mapping_source_path_var.get()).parent)
            if self.attribute_mapping_source_path_var.get().strip()
            else str(DEFAULT_ATTRIBUTE_MAPPING_SOURCE.parent)
        )
        selected = filedialog.askopenfilename(
            title="Attribut-Zuordnung XLSX wählen",
            initialdir=initial_dir,
            filetypes=[("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if selected:
            self.attribute_mapping_source_path_var.set(selected)
            self._load_attribute_mapping()

    def _reload_attribute_mapping(self) -> None:
        self._load_attribute_mapping()

    def _load_attribute_mapping(self, initial: bool = False) -> None:
        source_text = self.attribute_mapping_source_path_var.get().strip()
        source_path = Path(source_text) if source_text else None
        if source_path is None or not path_exists_safe(source_path):
            self.attribute_mapping = {}
            self.attribute_mapping_count_var.set("0 Zuordnungen geladen")
            if not initial and source_path is not None:
                self.status_var.set(f"Attribut-Zuordnung nicht gefunden: {source_path}")
            return
        try:
            self.attribute_mapping = load_attribute_mapping(source_path)
        except ValueError as exc:
            self.attribute_mapping = {}
            self.attribute_mapping_count_var.set("0 Zuordnungen geladen")
            if not initial:
                self.status_var.set(str(exc))
                messagebox.showwarning(APP_TITLE, str(exc))
            return
        self.attribute_mapping_count_var.set(f"{len(self.attribute_mapping)} Zuordnungen geladen")
        if not initial:
            self.status_var.set(f"Attribut-Zuordnung geladen: {len(self.attribute_mapping)} Einträge")

    def _open_attribute_suggestion_dialog(self, auto: bool = False) -> None:
        text = "\n".join(
            part
            for part in [self.short_text_frame.get_german_text(), self.long_text_frame.get_german_text()]
            if part.strip()
        )
        if not text.strip():
            if not auto:
                messagebox.showinfo(APP_TITLE, "Kein deutscher Text vorhanden, aus dem Attribute gelesen werden könnten.")
            return
        if not self.attribute_options_by_id:
            if not auto:
                messagebox.showwarning(APP_TITLE, "Es ist keine Attributliste geladen (Datenstämme im Projekt-Tab).")
            return
        suggestions, unmatched = build_attribute_suggestions_from_text(
            text,
            self.attribute_mapping,
            self.attribute_options_by_id,
            self.attribute_key_values_by_group,
        )
        if not suggestions and not unmatched:
            if not auto:
                messagebox.showinfo(APP_TITLE, "Im Text wurden keine technischen Angaben erkannt.")
            return
        if auto and not suggestions:
            return
        AttributeSuggestionDialog(
            self.root,
            suggestions=suggestions,
            unmatched=unmatched,
            attribute_options=self.attribute_options,
            on_apply=self._apply_attribute_suggestions,
        )

    def _apply_attribute_suggestions(self, rows: list[AttributeRow], learned: list[tuple[str, str]]) -> None:
        added = 0
        if rows:
            existing = self.attribute_frame.get_rows()
            seen = {(row.criteria_id, row.value.casefold(), row.value_to.casefold()) for row in existing}
            new_rows = []
            for row in rows:
                key = (row.criteria_id, row.value.casefold(), row.value_to.casefold())
                if key in seen:
                    continue
                seen.add(key)
                new_rows.append(row)
            if new_rows:
                self.attribute_frame.set_rows(existing + new_rows)
                added = len(new_rows)
                self._write_live_section("attributes", status_message=f"{added} Attribut(e) aus Text übernommen")

        if learned:
            source_text = self.attribute_mapping_source_path_var.get().strip() or str(DEFAULT_ATTRIBUTE_MAPPING_SOURCE)
            source_path = Path(source_text)
            entries = []
            for label, criteria_id in learned:
                key = normalize_mapping_label(label)
                if not key or self.attribute_mapping.get(key) == criteria_id:
                    continue
                option = self.attribute_options_by_id.get(criteria_id)
                entries.append((label, criteria_id, option.label if option is not None else "gelernt aus Vorschlagsdialog"))
                self.attribute_mapping[key] = criteria_id
            if entries:
                try:
                    append_attribute_mapping_entries(source_path, entries)
                    self.attribute_mapping_source_path_var.set(str(source_path))
                    self.attribute_mapping_count_var.set(f"{len(self.attribute_mapping)} Zuordnungen geladen")
                    self.status_var.set(
                        f"{added} Attribut(e) übernommen, {len(entries)} Zuordnung(en) gemerkt."
                    )
                except Exception as exc:  # pragma: no cover - defensive UI feedback
                    messagebox.showwarning(APP_TITLE, f"Zuordnung konnte nicht gespeichert werden:\n{exc}")
        elif added == 0:
            self.status_var.set("Keine neuen Attribute übernommen.")

    def _apply_vehicle_catalog(self, catalog: VehicleCatalog) -> None:
        self.vehicle_catalog = catalog
        if hasattr(self, "vehicle_link_frame"):
            self.vehicle_link_frame.set_vehicle_catalog(catalog)
        self.vehicle_count_var.set(f"{catalog.count} KTyp-Datensätze geladen")

    def _load_vehicle_catalog(self, initial: bool = False) -> None:
        source_text = self.vehicle_source_path_var.get().strip()
        source_path = Path(source_text) if source_text else None
        if source_path is None or not path_exists_safe(source_path):
            self._apply_vehicle_catalog(VehicleCatalog())
            if not initial and source_path is not None:
                self.status_var.set(f"KTyp-Datei nicht gefunden: {source_path}")
            return

        if self.vehicle_catalog_loading:
            return
        self.vehicle_catalog_loading = True
        self.vehicle_count_var.set("KTyp-Stammdaten werden geladen ...")
        if not initial:
            self.status_var.set(f"KTyp-Stammdaten werden geladen: {source_path}")

        def worker() -> None:
            error: Exception | None = None
            catalog = VehicleCatalog()
            try:
                catalog = load_vehicle_catalog(source_path)
            except Exception as exc:  # pragma: no cover - defensive UI feedback
                error = exc

            def finish() -> None:
                self.vehicle_catalog_loading = False
                if error is not None:
                    self._apply_vehicle_catalog(VehicleCatalog())
                    self.vehicle_count_var.set("KTyp-Datei konnte nicht geladen werden")
                    if not initial:
                        self.status_var.set(f"KTyp-Datei konnte nicht geladen werden: {error}")
                        messagebox.showwarning(APP_TITLE, f"KTyp-Datei konnte nicht geladen werden:\n{error}")
                    return
                self._apply_vehicle_catalog(catalog)
                self.article_browser_cache_signature = None
                if not initial:
                    self.status_var.set(f"KTyp-Stammdaten geladen: {catalog.count} Datensätze")

            self.root.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()

    def choose_competitor_source_file(self) -> None:
        initial_dir = (
            str(Path(self.competitor_source_path_var.get()).parent)
            if self.competitor_source_path_var.get().strip()
            else str(DEFAULT_COMPETITOR_SOURCE.parent)
        )
        selected = filedialog.askopenfilename(
            title="KHer CSV wählen",
            initialdir=initial_dir,
            filetypes=[("CSV Dateien", "*.csv"), ("Alle Dateien", "*.*")],
        )
        if selected:
            self.competitor_source_path_var.set(selected)
            self._load_competitor_catalog()

    def _reload_competitor_catalog(self) -> None:
        self._load_competitor_catalog()

    def _load_genart_catalog(self, initial: bool = False) -> None:
        source_path = Path(self.genart_source_path_var.get().strip()) if self.genart_source_path_var.get().strip() else None
        if source_path is None or not path_exists_safe(source_path):
            self.genart_registry = GenArtRegistry()
            self.genart_count_var.set("0 GenArts geladen")
            if self.selected_genart_selections:
                self.genart_suggestion_var.set(
                    f"Gesetzte GenArten: {summarize_genart_selections(self.selected_genart_selections, empty_label='-', limit=3)}"
                )
            else:
                self.genart_suggestion_var.set("Keine GenArt-Datei geladen.")
            if self.genart_suggestions is not None:
                self.genart_suggestions.values = []
                self.genart_suggestions.hide()
            if not initial and source_path is not None:
                self.status_var.set(f"GenArt-Datei nicht gefunden: {source_path}")
            return

        try:
            count = self.genart_registry.load_from_workbook(source_path)
        except Exception as exc:
            self.genart_count_var.set("0 GenArts geladen")
            if self.selected_genart_selections:
                self.genart_suggestion_var.set(
                    f"Gesetzte GenArten: {summarize_genart_selections(self.selected_genart_selections, empty_label='-', limit=3)}"
                )
            else:
                self.genart_suggestion_var.set("GenArt-Datei konnte nicht geladen werden.")
            if self.genart_suggestions is not None:
                self.genart_suggestions.values = []
                self.genart_suggestions.hide()
            if not initial:
                messagebox.showwarning(APP_TITLE, f"GenArt-Datei konnte nicht geladen werden:\n{exc}")
            return

        self.genart_count_var.set(f"{count} GenArts geladen")
        if self.genart_suggestions is not None:
            self._refresh_genart_combobox_values()
        if self.selected_genart_selections:
            self._set_selected_genart_selections(self.selected_genart_selections)
            self.genart_suggestion_var.set(
                f"Gesetzte GenArten: {summarize_genart_selections(self.selected_genart_selections, empty_label='-', limit=3)}"
            )
        else:
            self.genart_suggestion_var.set(
                "GenArt-Katalog geladen. Du kannst im Feld tippen oder über 'Suchen...' gezielt filtern."
            )
        if self.genart_display_var.get().strip():
            self._normalize_genart_selection()
        if not initial:
            self.status_var.set(f"GenArt-Katalog geladen: {count} Einträge")

    def _load_competitor_catalog(self, initial: bool = False) -> None:
        source_path = Path(self.competitor_source_path_var.get().strip()) if self.competitor_source_path_var.get().strip() else None
        if source_path is None or not path_exists_safe(source_path):
            self.manufacturer_options = []
            self.manufacturer_options_by_id = {}
            self.competitor_options = []
            self.competitor_options_by_id = {}
            self.competitor_count_var.set("0 Hersteller / 0 Mitbewerber geladen")
            if hasattr(self, "oe_frame"):
                self.oe_frame.set_manufacturer_catalog([])
            if hasattr(self, "comparison_frame"):
                self.comparison_frame.set_competitor_catalog([])
            if not initial and source_path is not None:
                self.status_var.set(f"KHer-CSV nicht gefunden: {source_path}")
            return

        try:
            manufacturer_options = load_competitor_options(source_path, comparison_only=False)
            competitor_options = load_competitor_options(source_path, comparison_only=True)
        except ValueError as exc:
            self.manufacturer_options = []
            self.manufacturer_options_by_id = {}
            self.competitor_options = []
            self.competitor_options_by_id = {}
            self.competitor_count_var.set("0 Hersteller / 0 Mitbewerber geladen")
            if hasattr(self, "oe_frame"):
                self.oe_frame.set_manufacturer_catalog([])
            if hasattr(self, "comparison_frame"):
                self.comparison_frame.set_competitor_catalog([])
            if not initial:
                self.status_var.set(str(exc))
                messagebox.showwarning(APP_TITLE, str(exc))
            return

        self.manufacturer_options = manufacturer_options
        self.manufacturer_options_by_id = {option.competitor_id: option for option in manufacturer_options}
        self.competitor_options = competitor_options
        self.competitor_options_by_id = {option.competitor_id: option for option in competitor_options}
        self.competitor_count_var.set(f"{len(manufacturer_options)} Hersteller / {len(competitor_options)} Mitbewerber geladen")
        if hasattr(self, "oe_frame"):
            self.oe_frame.set_manufacturer_catalog(manufacturer_options)
        if hasattr(self, "comparison_frame"):
            self.comparison_frame.set_competitor_catalog(competitor_options)
        self.article_browser_cache_signature = None
        if not initial:
            self.status_var.set(
                f"KHer-CSV geladen: {len(manufacturer_options)} Hersteller, {len(competitor_options)} Mitbewerber für Vergleichsnummern"
            )
            self.refresh_preview()

    def _load_attribute_catalog(self, initial: bool = False) -> None:
        source_path = Path(self.attribute_source_path_var.get().strip()) if self.attribute_source_path_var.get().strip() else None
        if source_path is None or not path_exists_safe(source_path):
            self.attribute_options = []
            self.attribute_options_by_id = {}
            self.attribute_count_var.set("0 Attribute geladen")
            if hasattr(self, "attribute_frame"):
                self.attribute_frame.set_attribute_catalog([])
            if not initial and source_path is not None:
                self.status_var.set(f"Attributdatei nicht gefunden: {source_path}")
            return

        try:
            options = load_attribute_options(source_path)
        except ValueError as exc:
            self.attribute_options = []
            self.attribute_options_by_id = {}
            self.attribute_count_var.set("0 Attribute geladen")
            if hasattr(self, "attribute_frame"):
                self.attribute_frame.set_attribute_catalog([])
            if not initial:
                self.status_var.set(str(exc))
                messagebox.showwarning(APP_TITLE, str(exc))
            return

        self.attribute_options = options
        self.attribute_options_by_id = {option.criteria_id: option for option in options}
        self.attribute_count_var.set(f"{len(options)} Attribute geladen")
        if hasattr(self, "attribute_frame"):
            self.attribute_frame.set_attribute_catalog(options)
        self.article_browser_cache_signature = None
        if not initial:
            self.status_var.set(f"Attributkatalog geladen: {len(options)} Einträge")
            self.refresh_preview()

    def _load_attribute_key_value_catalog(self, initial: bool = False) -> None:
        source_path = (
            Path(self.attribute_key_value_source_path_var.get().strip())
            if self.attribute_key_value_source_path_var.get().strip()
            else None
        )
        if source_path is None or not path_exists_safe(source_path):
            self.attribute_key_value_options = []
            self.attribute_key_values_by_group = {}
            self.attribute_key_value_count_var.set("0 Schlüsselwerte geladen")
            if hasattr(self, "attribute_frame"):
                self.attribute_frame.set_attribute_key_value_catalog({})
            if not initial and source_path is not None:
                self.status_var.set(f"Schlüsselwertdatei nicht gefunden: {source_path}")
            return

        try:
            options = load_attribute_key_value_options(source_path)
        except ValueError as exc:
            self.attribute_key_value_options = []
            self.attribute_key_values_by_group = {}
            self.attribute_key_value_count_var.set("0 Schlüsselwerte geladen")
            if hasattr(self, "attribute_frame"):
                self.attribute_frame.set_attribute_key_value_catalog({})
            if not initial:
                self.status_var.set(str(exc))
                messagebox.showwarning(APP_TITLE, str(exc))
            return

        self.attribute_key_value_options = options
        self.attribute_key_values_by_group = build_attribute_key_value_group_index(options)
        self.attribute_key_value_count_var.set(
            f"{len(options)} Schlüsselwerte in {len(self.attribute_key_values_by_group)} Gruppen geladen"
        )
        if hasattr(self, "attribute_frame"):
            self.attribute_frame.set_attribute_key_value_catalog(self.attribute_key_values_by_group)
        self.article_browser_cache_signature = None
        if not initial:
            self.status_var.set(f"Schlüsselwertkatalog geladen: {len(options)} Einträge")
            self.refresh_preview()

    def _refresh_genart_combobox_values(self) -> None:
        if self.genart_suggestions is None:
            return
        self.genart_suggestions.refresh()

    def _genart_suggestion_values(self, query: str = "") -> list[str]:
        return self.genart_registry.search_display_values(query, limit=250)

    def _canonicalize_genart_selection(self, selection: GenArtSelection) -> GenArtSelection:
        option = self.genart_registry.resolve(selection.id) or self.genart_registry.resolve(selection.bezeichnung)
        if option is not None:
            return GenArtSelection(id=option.id, bezeichnung=option.bezeichnung)
        normalized = normalize_genart_selections([selection])
        return normalized[0] if normalized else GenArtSelection()

    def _refresh_selected_genart_tree(self, selected_indices: set[int] | None = None) -> None:
        self.selected_genart_count_var.set(f"{len(self.selected_genart_selections)} GenArten gesetzt")
        if not hasattr(self, "selected_genart_tree"):
            return

        selected_iids = {str(index) for index in (selected_indices or set())}
        for item_id in self.selected_genart_tree.get_children():
            self.selected_genart_tree.delete(item_id)

        for index, selection in enumerate(self.selected_genart_selections):
            item_id = str(index)
            self.selected_genart_tree.insert(
                "",
                "end",
                iid=item_id,
                values=(selection.id, selection.bezeichnung),
            )

        if selected_iids:
            existing = [item_id for item_id in selected_iids if self.selected_genart_tree.exists(item_id)]
            if existing:
                self.selected_genart_tree.selection_set(existing)
                self.selected_genart_tree.focus(existing[0])

    def _set_selected_genart_selections(self, selections: list[GenArtSelection], focus_last: bool = False) -> None:
        canonicalized = [self._canonicalize_genart_selection(selection) for selection in selections]
        self.selected_genart_selections = normalize_genart_selections(canonicalized)
        selected_indices = {len(self.selected_genart_selections) - 1} if focus_last and self.selected_genart_selections else set()
        self._refresh_selected_genart_tree(selected_indices=selected_indices)

    def _get_selected_genart_selections(self) -> list[GenArtSelection]:
        return list(self.selected_genart_selections)

    def _resolve_current_genart_selection(self, prefer_first_suggestion: bool = False) -> GenArtSelection | None:
        current_value = self.genart_display_var.get().strip()
        option = self.genart_registry.resolve(current_value) if current_value else None
        if option is not None:
            return GenArtSelection(id=option.id, bezeichnung=option.bezeichnung)

        if current_value:
            parsed_selection = parse_genart_selection_label(current_value)
            if parsed_selection is not None and (not self.genart_registry.options or "|" in current_value):
                return self._canonicalize_genart_selection(parsed_selection)

        if prefer_first_suggestion and self.genart_suggestions is not None:
            values = self.genart_suggestions.values or self._genart_suggestion_values(current_value)
            for value in values:
                option = self.genart_registry.resolve(str(value))
                if option is not None:
                    return GenArtSelection(id=option.id, bezeichnung=option.bezeichnung)
        return None

    def _add_genart_selection(
        self,
        selection: GenArtSelection | None,
        *,
        write_live: bool = True,
        suggestion_message: str | None = None,
    ) -> bool:
        if selection is None:
            return False

        normalized = normalize_genart_selections([selection])
        if not normalized:
            return False

        canonical = self._canonicalize_genart_selection(normalized[0])
        if not canonical.display_label():
            return False

        self.genart_display_var.set(canonical.display_label())
        selection_key = canonical.id.casefold() or canonical.bezeichnung.casefold()
        existing_keys = {
            existing_selection.id.casefold() or existing_selection.bezeichnung.casefold()
            for existing_selection in self.selected_genart_selections
        }
        if selection_key in existing_keys:
            self.genart_suggestion_var.set(f"GenArt bereits gesetzt: {canonical.display_label()}")
            return False

        self._set_selected_genart_selections([*self.selected_genart_selections, canonical], focus_last=True)
        self.genart_suggestion_var.set(suggestion_message or f"GenArt hinzugefügt: {canonical.display_label()}")
        if write_live:
            self._write_live_section("genart")
        return True

    def add_current_genart_selection(self) -> None:
        if not self.genart_registry.options:
            self._load_genart_catalog()
        selection = self._resolve_current_genart_selection(prefer_first_suggestion=True)
        if selection is None:
            self.genart_suggestion_var.set("GenArt nicht erkannt. Bitte aus der Liste wählen oder Vorschlag nutzen.")
            return
        self._add_genart_selection(selection)

    def _accept_genart_suggestion(self, value: str) -> None:
        self.genart_display_var.set(value)
        self.add_current_genart_selection()

    def _handle_missing_genart_selection(self) -> None:
        self.genart_suggestion_var.set("GenArt nicht erkannt. Bitte aus der Liste wählen oder Vorschlag nutzen.")

    def _remove_selected_genarts_event(self, _event: tk.Event[tk.Misc]) -> str:
        self.remove_selected_genarts()
        return "break"

    def remove_selected_genarts(self) -> None:
        if not hasattr(self, "selected_genart_tree"):
            return
        selection = list(self.selected_genart_tree.selection())
        if not selection:
            return

        remove_indices = sorted(
            (int(item_id) for item_id in selection if str(item_id).isdigit()),
            reverse=True,
        )
        if not remove_indices:
            return

        remaining = [
            genart_selection
            for index, genart_selection in enumerate(self.selected_genart_selections)
            if index not in set(remove_indices)
        ]
        self._set_selected_genart_selections(remaining)
        self.genart_suggestion_var.set(f"{len(remove_indices)} GenArt(en) entfernt.")
        self._write_live_section("genart")

    def clear_selected_genarts(self) -> None:
        if not self.selected_genart_selections:
            return
        self._set_selected_genart_selections([])
        self.genart_display_var.set("")
        self.genart_suggestion_var.set("Keine GenArt für diesen Artikel gesetzt.")
        self._write_live_section("genart")

    def _open_genart_search_dialog_event(self, _event: tk.Event[tk.Misc]) -> str:
        self._open_genart_search_dialog()
        return "break"

    def _open_genart_search_dialog(self) -> None:
        if not self.genart_registry.options:
            self._load_genart_catalog()
        if not self.genart_registry.options:
            messagebox.showinfo(APP_TITLE, "Es ist noch keine GenArt-Datei geladen.")
            return

        initial_query = self.genart_display_var.get().strip()
        dialog = GenArtSearchDialog(self.root, self.genart_registry, initial_query=initial_query)
        selected_option = dialog.show()
        if selected_option is None:
            return

        self._add_genart_selection(
            GenArtSelection(id=selected_option.id, bezeichnung=selected_option.bezeichnung),
            suggestion_message=f"GenArt hinzugefügt: {selected_option.display_label()}",
        )

    def _on_genart_focus_in(self, _event: tk.Event[tk.Misc]) -> None:
        self._refresh_genart_combobox_values()

    def _on_genart_key_release(self, event: tk.Event[tk.Misc]) -> None:
        if event.keysym in {"Return", "Escape", "Up", "Down", "Left", "Right", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R"}:
            return
        self._refresh_genart_combobox_values()

    def _on_genart_return(self, _event: tk.Event[tk.Misc]) -> str:
        selection = self._resolve_current_genart_selection(prefer_first_suggestion=True)
        if selection is not None:
            self._add_genart_selection(selection)
        else:
            self.genart_suggestion_var.set("GenArt nicht erkannt. Bitte aus der Liste wählen oder Vorschlag nutzen.")
        return "break"

    def _normalize_genart_selection(self) -> bool:
        current_value = self.genart_display_var.get().strip()
        if not current_value:
            return True

        option = self.genart_registry.resolve(current_value)
        if option is not None:
            self.genart_display_var.set(option.display_label())
            return True

        parsed_selection = parse_genart_selection_label(current_value)
        if parsed_selection is None or (self.genart_registry.options and "|" not in current_value):
            self.genart_suggestion_var.set("GenArt nicht erkannt. Bitte aus der Liste wählen oder Vorschlag nutzen.")
            return False
        self.genart_display_var.set(self._canonicalize_genart_selection(parsed_selection).display_label())
        return True

    def _on_genart_focus_out(self, _event: tk.Event[tk.Misc]) -> None:
        self._normalize_genart_selection()

    def _resolve_output_root(self) -> Path:
        output_root = Path(self.output_dir_var.get().strip())
        if not output_root:
            raise ValueError("Bitte einen Ausgabeordner angeben.")
        return output_root

    @contextmanager
    def _suspend_live_write(self):
        previous = self.live_write_suspended
        self.live_write_suspended = True
        try:
            yield
        finally:
            self.live_write_suspended = previous

    def _snapshot_from_bundle(self, bundle: ExportBundle, source_folder: Path) -> StoredArticleSnapshot:
        return StoredArticleSnapshot(
            article_number=bundle.article_number,
            source_label="Output",
            source_folder=source_folder,
            short_module_id=bundle.short_module_id,
            long_module_id=bundle.long_module_id,
            genart_selections=[GenArtSelection(id=selection.id, bezeichnung=selection.bezeichnung) for selection in bundle.genart_selections],
            genart_id=bundle.genart_id,
            genart_bezeichnung=bundle.genart_bezeichnung,
            short_texts=bundle.short_texts,
            short_auto_uni=bundle.short_auto_uni,
            long_texts=bundle.long_texts,
            long_auto_uni=bundle.long_auto_uni,
            image_rows=bundle.image_rows,
            document_rows=bundle.document_rows,
            video_rows=bundle.video_rows,
            web_rows=bundle.web_rows,
            oe_number_rows=bundle.oe_number_rows,
            comparison_number_rows=bundle.comparison_number_rows,
            vehicle_link_rows=bundle.vehicle_link_rows,
            attribute_rows=bundle.attribute_rows,
            search_terms=list(bundle.search_terms),
        )

    def _copy_snapshot(self, snapshot: StoredArticleSnapshot) -> StoredArticleSnapshot:
        return StoredArticleSnapshot(
            article_number=snapshot.article_number,
            source_label=snapshot.source_label,
            source_folder=snapshot.source_folder,
            short_module_id=snapshot.short_module_id,
            long_module_id=snapshot.long_module_id,
            genart_selections=[GenArtSelection(id=selection.id, bezeichnung=selection.bezeichnung) for selection in snapshot.genart_selections],
            genart_id=snapshot.genart_id,
            genart_bezeichnung=snapshot.genart_bezeichnung,
            short_texts=TranslationSet(**vars(snapshot.short_texts)),
            short_auto_uni=snapshot.short_auto_uni,
            long_texts=TranslationSet(**vars(snapshot.long_texts)),
            long_auto_uni=snapshot.long_auto_uni,
            image_rows=[MediaRow(row.path_or_link, art=row.art, sprache=row.sprache) for row in snapshot.image_rows],
            document_rows=[MediaRow(row.path_or_link, art=row.art, sprache=row.sprache) for row in snapshot.document_rows],
            video_rows=[MediaRow(row.path_or_link, art=row.art, sprache=row.sprache) for row in snapshot.video_rows],
            web_rows=[MediaRow(row.path_or_link, art=row.art, sprache=row.sprache) for row in snapshot.web_rows],
            oe_number_rows=[
                OeNumberRow(
                    value=row.value,
                    manufacturer_id=row.manufacturer_id,
                    manufacturer_code=row.manufacturer_code,
                    manufacturer_name=row.manufacturer_name,
                )
                for row in snapshot.oe_number_rows
            ],
            comparison_number_rows=[
                ComparisonNumberRow(
                    competitor_id=row.competitor_id,
                    competitor_code=row.competitor_code,
                    competitor_name=row.competitor_name,
                    reference_number=row.reference_number,
                )
                for row in snapshot.comparison_number_rows
            ],
            vehicle_link_rows=[
                VehicleLinkRow(
                    vehicle_type_id=row.vehicle_type_id,
                    vehicle_type_label=row.vehicle_type_label,
                    motorcode=row.motorcode,
                    ktyp_topmotive=row.ktyp_topmotive,
                    ktyp_tecdoc=row.ktyp_tecdoc,
                    manufacturer=row.manufacturer,
                    model=row.model,
                    description=row.description,
                    year_from=row.year_from,
                    year_to=row.year_to,
                    power=row.power,
                )
                for row in snapshot.vehicle_link_rows
            ],
            attribute_rows=[
                AttributeRow(
                    criteria_id=row.criteria_id,
                    label=row.label,
                    value_format=row.value_format,
                    max_length=row.max_length,
                    type_name=row.type_name,
                    value=row.value,
                    value_to=row.value_to,
                )
                for row in snapshot.attribute_rows
            ],
            search_terms=list(snapshot.search_terms),
        )

    def _render_article_browser(self, current_article: str | None = None) -> None:
        selected_article = current_article or (self.article_browser_tree.selection()[0] if self.article_browser_tree.selection() else "")
        selected_article = normalize_article_number(selected_article) or normalize_article_number(self.article_number_var.get())

        for item_id in self.article_browser_tree.get_children():
            self.article_browser_tree.delete(item_id)

        for article_number in sorted(self.article_browser_records):
            snapshot = self.article_browser_records[article_number]
            self.article_browser_tree.insert(
                "",
                "end",
                iid=article_number,
                values=(
                    snapshot.article_number,
                    snapshot.source_label,
                    snapshot.short_module_id,
                    snapshot.long_module_id,
                    summarize_genart_selections(snapshot.genart_selections, empty_label="-", limit=2),
                    len(snapshot.attribute_rows),
                    len(snapshot.oe_number_rows),
                    len(snapshot.comparison_number_rows),
                    len(snapshot.image_rows),
                    len(snapshot.document_rows),
                    len(snapshot.video_rows),
                    len(snapshot.web_rows),
                ),
            )

        if selected_article and selected_article in self.article_browser_records:
            self.article_browser_tree.selection_set(selected_article)
            self.article_browser_tree.focus(selected_article)
            self._update_article_browser_detail(self.article_browser_records[selected_article])
        else:
            self._update_article_browser_detail(None)

    def _build_output_folder_signature(self, output_folder: Path) -> tuple[tuple[str, int, int], ...]:
        signature: list[tuple[str, int, int]] = []
        for file_name in [
            SHORT_TEXT_FILE[0],
            SHORT_MAPPING_FILE[0],
            LONG_TEXT_FILE[0],
            GENART_FILE[0],
            OE_FILE[0],
            COMPARISON_FILE[0],
            VEHICLE_LINK_FILE[0],
            ATTRIBUTE_FILE[0],
            IMAGE_FILE[0],
            DOCUMENT_FILE[0],
            VIDEO_FILE[0],
            WEB_LINK_FILE[0],
        ]:
            path = output_folder / file_name
            if path.exists():
                stat = path.stat()
                signature.append((file_name, stat.st_mtime_ns, stat.st_size))
            else:
                signature.append((file_name, -1, -1))
        return tuple(signature)

    def _upsert_article_browser_from_bundle(self, bundle: ExportBundle, output_root: Path) -> None:
        self.article_browser_records[bundle.article_number] = self._snapshot_from_bundle(bundle, output_root)
        self.article_browser_cache_dir = str(output_root)
        self.article_browser_cache_signature = None
        self._render_article_browser(bundle.article_number)

    def _upsert_article_browser_section(self, section: str, bundle: ExportBundle, output_root: Path) -> None:
        existing_snapshot = self.article_browser_records.get(bundle.article_number)
        snapshot = self._copy_snapshot(existing_snapshot) if existing_snapshot is not None else StoredArticleSnapshot(
            article_number=bundle.article_number,
            source_label="Output",
            source_folder=output_root,
        )
        snapshot.article_number = bundle.article_number
        snapshot.source_label = "Output"
        snapshot.source_folder = output_root
        if bundle.short_module_id:
            snapshot.short_module_id = bundle.short_module_id
        if bundle.long_module_id:
            snapshot.long_module_id = bundle.long_module_id
        snapshot.genart_selections = [GenArtSelection(id=selection.id, bezeichnung=selection.bezeichnung) for selection in bundle.genart_selections]
        snapshot.genart_id = bundle.genart_id
        snapshot.genart_bezeichnung = bundle.genart_bezeichnung
        snapshot.sync_genart_fields()

        if section in {"short_text", "all"}:
            snapshot.short_texts = bundle.short_texts
            snapshot.short_auto_uni = bundle.short_auto_uni
            snapshot.short_module_id = bundle.short_module_id
        if section in {"long_text", "all"}:
            snapshot.long_texts = bundle.long_texts
            snapshot.long_auto_uni = bundle.long_auto_uni
            snapshot.long_module_id = bundle.long_module_id
        if section in {"genart", "all"}:
            snapshot.genart_selections = [GenArtSelection(id=selection.id, bezeichnung=selection.bezeichnung) for selection in bundle.genart_selections]
            snapshot.genart_id = bundle.genart_id
            snapshot.genart_bezeichnung = bundle.genart_bezeichnung
            snapshot.sync_genart_fields()
        if section in {"images", "all"}:
            snapshot.image_rows = [MediaRow(row.path_or_link, art=row.art, sprache=row.sprache) for row in bundle.image_rows]
        if section in {"documents", "all"}:
            snapshot.document_rows = [MediaRow(row.path_or_link, art=row.art, sprache=row.sprache) for row in bundle.document_rows]
        if section in {"videos", "all"}:
            snapshot.video_rows = [MediaRow(row.path_or_link, art=row.art, sprache=row.sprache) for row in bundle.video_rows]
        if section in {"web_links", "all"}:
            snapshot.web_rows = [MediaRow(row.path_or_link, art=row.art, sprache=row.sprache) for row in bundle.web_rows]
        if section in {"oe_numbers", "all"}:
            snapshot.oe_number_rows = [
                OeNumberRow(
                    value=row.value,
                    manufacturer_id=row.manufacturer_id,
                    manufacturer_code=row.manufacturer_code,
                    manufacturer_name=row.manufacturer_name,
                )
                for row in bundle.oe_number_rows
            ]
        if section in {"comparison_numbers", "all"}:
            snapshot.comparison_number_rows = [
                ComparisonNumberRow(
                    competitor_id=row.competitor_id,
                    competitor_code=row.competitor_code,
                    competitor_name=row.competitor_name,
                    reference_number=row.reference_number,
                )
                for row in bundle.comparison_number_rows
            ]
        if section in {"vehicle_links", "all"}:
            snapshot.vehicle_link_rows = [
                VehicleLinkRow(
                    vehicle_type_id=row.vehicle_type_id,
                    vehicle_type_label=row.vehicle_type_label,
                    motorcode=row.motorcode,
                    ktyp_topmotive=row.ktyp_topmotive,
                    ktyp_tecdoc=row.ktyp_tecdoc,
                    manufacturer=row.manufacturer,
                    model=row.model,
                    description=row.description,
                    year_from=row.year_from,
                    year_to=row.year_to,
                    power=row.power,
                )
                for row in bundle.vehicle_link_rows
            ]
        if section in {"attributes", "all"}:
            snapshot.attribute_rows = [
                AttributeRow(
                    criteria_id=row.criteria_id,
                    label=row.label,
                    value_format=row.value_format,
                    max_length=row.max_length,
                    type_name=row.type_name,
                    value=row.value,
                    value_to=row.value_to,
                )
                for row in bundle.attribute_rows
            ]
        if section in {"search_terms", "all"}:
            snapshot.search_terms = list(bundle.search_terms)

        self.article_browser_records[bundle.article_number] = snapshot
        self.article_browser_cache_dir = str(output_root)
        self.article_browser_cache_signature = None
        self._render_article_browser(bundle.article_number)

    def _write_short_text_live(self, bundle: ExportBundle, output_root: Path) -> None:
        written_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        article_keys = {bundle.article_number}
        write_workbook_with_upsert(
            output_root / SHORT_TEXT_FILE[0],
            SHORT_TEXT_FILE[1],
            SHORT_TEXT_HEADERS,
            [append_written_at(build_short_text_export_row(bundle), written_at)],
            replace_article_keys=article_keys,
        )
        write_workbook_with_upsert(
            output_root / SHORT_MAPPING_FILE[0],
            SHORT_MAPPING_FILE[1],
            SHORT_MAPPING_HEADERS,
            [append_written_at(build_short_mapping_export_row(bundle), written_at)],
            replace_article_keys=article_keys,
        )

    def _write_long_text_live(self, bundle: ExportBundle, output_root: Path) -> None:
        written_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        write_workbook_with_upsert(
            output_root / LONG_TEXT_FILE[0],
            LONG_TEXT_FILE[1],
            SHORT_TEXT_HEADERS,
            [append_written_at(build_long_text_export_row(bundle), written_at)],
            replace_article_keys={bundle.article_number},
        )

    def _write_genart_live(self, bundle: ExportBundle, output_root: Path) -> None:
        written_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        write_workbook_with_upsert(
            output_root / GENART_FILE[0],
            GENART_FILE[1],
            GENART_HEADERS,
            [append_written_at(row, written_at) for row in build_genart_export_rows(bundle)],
            replace_article_keys={bundle.article_number},
        )

    def _write_oe_live(self, bundle: ExportBundle, output_root: Path) -> None:
        replace_article_rows_preserving_timestamps(
            output_root / OE_FILE[0],
            OE_FILE[1],
            OE_HEADERS,
            bundle.article_number,
            build_oe_export_rows(bundle),
        )

    def _write_comparison_live(self, bundle: ExportBundle, output_root: Path) -> None:
        replace_article_rows_preserving_timestamps(
            output_root / COMPARISON_FILE[0],
            COMPARISON_FILE[1],
            COMPARISON_HEADERS,
            bundle.article_number,
            build_comparison_export_rows(bundle),
        )

    def _write_vehicle_link_live(self, bundle: ExportBundle, output_root: Path) -> None:
        replace_article_rows_preserving_timestamps(
            output_root / VEHICLE_LINK_FILE[0],
            VEHICLE_LINK_FILE[1],
            VEHICLE_LINK_HEADERS,
            bundle.article_number,
            build_vehicle_link_export_rows(bundle),
        )

    def _write_attribute_live(self, bundle: ExportBundle, output_root: Path) -> None:
        replace_article_rows_preserving_timestamps(
            output_root / ATTRIBUTE_FILE[0],
            ATTRIBUTE_FILE[1],
            ATTRIBUTE_HEADERS,
            bundle.article_number,
            build_attribute_export_rows(bundle),
        )

    def _write_media_section_live(
        self,
        bundle: ExportBundle,
        output_root: Path,
        file_spec: tuple[str, str],
        headers: list[str],
        rows: list[list[str]],
    ) -> None:
        replace_article_rows_preserving_timestamps(
            output_root / file_spec[0],
            file_spec[1],
            headers,
            bundle.article_number,
            rows,
        )

    def _write_live_section(self, section: str, status_message: str | None = None) -> bool:
        if self.live_write_suspended:
            return False

        try:
            output_root = self._resolve_output_root()
            bundle = self.collect_bundle()
        except ValueError:
            return False

        try:
            output_root.mkdir(parents=True, exist_ok=True)
            if section == "short_text":
                self._write_short_text_live(bundle, output_root)
            elif section == "long_text":
                self._write_long_text_live(bundle, output_root)
            elif section == "genart":
                self._write_genart_live(bundle, output_root)
            elif section == "oe_numbers":
                self._write_oe_live(bundle, output_root)
            elif section == "comparison_numbers":
                self._write_comparison_live(bundle, output_root)
            elif section == "vehicle_links":
                self._write_vehicle_link_live(bundle, output_root)
            elif section == "attributes":
                self._write_attribute_live(bundle, output_root)
            elif section == "search_terms":
                # Suchwörter liegen mit in der Attribute-Datei.
                self._write_attribute_live(bundle, output_root)
            elif section == "images":
                self._write_media_section_live(bundle, output_root, IMAGE_FILE, IMAGE_HEADERS, build_image_export_rows(bundle))
            elif section == "documents":
                self._write_media_section_live(bundle, output_root, DOCUMENT_FILE, DOCUMENT_HEADERS, build_document_export_rows(bundle))
            elif section == "videos":
                self._write_media_section_live(bundle, output_root, VIDEO_FILE, VIDEO_HEADERS, build_video_export_rows(bundle))
            elif section == "web_links":
                self._write_media_section_live(bundle, output_root, WEB_LINK_FILE, WEB_HEADERS, build_web_export_rows(bundle))
            else:
                raise ValueError(f"Unbekannter Live-Bereich: {section}")
        except Exception as exc:  # pragma: no cover - defensive UI feedback
            self.status_var.set(f"Live-Speichern fehlgeschlagen: {exc}")
            return False

        self.id_registry.remember_article_ids(bundle.article_number, bundle.short_module_id, bundle.long_module_id)
        self._upsert_article_browser_section(section, bundle, output_root)
        self.status_var.set(status_message or f"Live gespeichert: {bundle.article_number}")
        return True

    def _write_live_database(self, status_message: str | None = None) -> bool:
        if self.live_write_suspended:
            return False

        try:
            output_root = self._resolve_output_root()
            bundle = self.collect_bundle()
        except ValueError:
            return False

        try:
            export_bundle(bundle, output_root, use_timestamp_subdir=False)
        except Exception as exc:  # pragma: no cover - defensive UI feedback
            self.status_var.set(f"Live-Speichern fehlgeschlagen: {exc}")
            return False

        self.id_registry.remember_article_ids(bundle.article_number, bundle.short_module_id, bundle.long_module_id)
        self._upsert_article_browser_from_bundle(bundle, output_root)
        self.status_var.set(status_message or f"Live gespeichert: {bundle.article_number}")
        return True

    def _on_live_data_changed(self) -> None:
        self._write_live_database()

    def _on_article_entry_focus_out(self, _event: tk.Event[tk.Misc]) -> None:
        article_number = normalize_article_number(self.article_number_var.get())
        if not article_number:
            return
        self.article_number_var.set(article_number)
        self._ensure_ids_for_article(article_number)
        self._write_live_database()

    def _resolve_ids_for_article(self, article_number: str) -> tuple[str, str]:
        article_key = normalize_article_number(article_number)
        if not article_key:
            return "", ""

        short_id, long_id = self.id_registry.get_ids_for_article(article_key)
        if not short_id:
            short_id = self.id_registry.generate_unique()
        if not long_id:
            long_id = self.id_registry.generate_unique()

        self.id_registry.remember_article_ids(article_key, short_id, long_id)
        return short_id, long_id

    def _ensure_ids_for_article(self, article_number: str) -> tuple[str, str]:
        article_key = normalize_article_number(article_number)
        if not article_key:
            self.current_id_article_number = ""
            self.short_module_id_var.set("")
            self.long_module_id_var.set("")
            return "", ""

        current_short_id = self.short_module_id_var.get().strip()
        current_long_id = self.long_module_id_var.get().strip()
        if article_key == self.current_id_article_number and current_short_id and current_long_id:
            self.id_registry.remember_article_ids(article_key, current_short_id, current_long_id)
            return current_short_id, current_long_id

        short_id, long_id = self._resolve_ids_for_article(article_key)
        self.short_module_id_var.set(short_id)
        self.long_module_id_var.set(long_id)
        self.current_id_article_number = article_key
        return short_id, long_id

    def _get_scrape_options(self) -> dict[str, bool]:
        options = {
            "short_text": self.batch_short_text_var.get(),
            "long_text": self.batch_long_text_var.get(),
            "images": self.batch_image_var.get(),
            "documents": self.batch_document_var.get(),
            "videos": self.batch_video_var.get(),
            "web_links": self.batch_web_var.get(),
        }
        if not any(options.values()):
            raise ValueError("Bitte im Abrufumfang mindestens eine Datenart auswählen.")
        return options

    def _build_translation_set(
        self,
        german_text: str,
        client: DeepLClient | None,
        sanitize_short_text: bool = False,
    ) -> TranslationSet:
        values = {"de": german_text.strip(), "uni": german_text.strip()}
        if client and german_text.strip():
            values.update(client.translate_from_german(german_text))
        translations = TranslationSet(**values)
        if sanitize_short_text:
            return sanitize_short_translation_set(translations)
        return translations

    def _build_bundle_from_kunzer_result(
        self,
        result: KunzerScrapeResult,
        client: DeepLClient | None = None,
        options: dict[str, bool] | None = None,
    ) -> ExportBundle:
        article_number = normalize_article_number(result.article_number)
        short_module_id, long_module_id = self._resolve_ids_for_article(article_number)
        scrape_options = options or {
            "short_text": True,
            "long_text": True,
            "images": True,
            "documents": True,
            "videos": True,
            "web_links": True,
        }
        short_texts = self._build_translation_set(result.short_text_de, client, sanitize_short_text=True) if scrape_options["short_text"] else TranslationSet()
        long_texts = self._build_translation_set(result.long_text_de, client) if scrape_options["long_text"] else TranslationSet()
        image_rows = [MediaRow(link, art="5", sprache="255") for link in result.image_links] if scrape_options["images"] else []
        return ExportBundle(
            article_number=article_number,
            short_module_id=short_module_id,
            long_module_id=long_module_id,
            short_texts=short_texts,
            short_auto_uni=True,
            long_texts=long_texts,
            long_auto_uni=True,
            image_rows=image_rows,
            document_rows=[MediaRow(link, art=infer_document_art(link), sprache="255") for link in result.document_links] if scrape_options["documents"] else [],
            video_rows=[MediaRow(link) for link in result.video_links] if scrape_options["videos"] else [],
            web_rows=[MediaRow(result.product_url)] if scrape_options["web_links"] else [],
            include_short_text=scrape_options["short_text"],
            include_long_text=scrape_options["long_text"],
            include_images=scrape_options["images"],
            include_documents=scrape_options["documents"],
            include_videos=scrape_options["videos"],
            include_web_links=scrape_options["web_links"],
        )

    def _apply_kunzer_result(
        self,
        result: KunzerScrapeResult,
        translate_after_load: bool = True,
        write_live: bool = True,
        short_translations: dict[str, str] | None = None,
        long_translations: dict[str, str] | None = None,
        options: dict[str, bool] | None = None,
    ) -> None:
        scrape_options = options or {
            "short_text": True,
            "long_text": True,
            "images": True,
            "documents": True,
            "videos": True,
            "web_links": True,
        }
        with self._suspend_live_write():
            if not self.article_number_var.get().strip():
                self.article_number_var.set(result.article_number)
            else:
                self.article_number_var.set(normalize_article_number(result.article_number))
            self._ensure_ids_for_article(self.article_number_var.get())
            self.kunzer_product_url_var.set(result.product_url)

            short_text_de = result.short_text_de if scrape_options["short_text"] else ""
            long_text_de = result.long_text_de if scrape_options["long_text"] else ""
            self.short_text_frame.set_value(
                TranslationSet(de=short_text_de, uni=short_text_de),
                auto_uni=True,
            )
            self.long_text_frame.set_value(
                TranslationSet(de=long_text_de, uni=long_text_de),
                auto_uni=True,
            )
            self._set_selected_genart_selections([])
            self.genart_display_var.set("")
            if short_translations and scrape_options["short_text"]:
                self.short_text_frame.apply_translations(short_translations)
            if long_translations and scrape_options["long_text"]:
                self.long_text_frame.apply_translations(long_translations)

            image_rows = (
                [MediaRow(link, art="5", sprache="255") for link in result.image_links]
                if scrape_options["images"]
                else []
            )
            document_rows = (
                [MediaRow(link, art=infer_document_art(link), sprache="255") for link in result.document_links]
                if scrape_options["documents"]
                else []
            )
            video_rows = [MediaRow(link) for link in result.video_links] if scrape_options["videos"] else []
            web_rows = [MediaRow(result.product_url)] if scrape_options["web_links"] else []

            self.image_frame.set_rows(image_rows)
            self.document_frame.set_rows(document_rows)
            self.video_frame.set_rows(video_rows)
            self.web_frame.set_rows(web_rows)
            self.attribute_frame.set_rows([])
            self.search_term_frame.set_terms([])
            self.oe_frame.set_rows([])
            self.comparison_frame.set_rows([])
            self.vehicle_link_frame.set_rows([])
            self.genart_suggestion_var.set("GenArt bitte über das Suchfeld oder 'Suchen...' auswählen.")

        if translate_after_load and self.auto_translate_after_scrape_var.get() and self.deepl_api_key_var.get().strip():
            self._translate_loaded_texts()

        self.refresh_preview()
        if write_live:
            self._write_live_database(status_message=f"Live gespeichert: {normalize_article_number(result.article_number)}")
        # Nach dem Laden automatisch Attributvorschläge aus dem Text anbieten.
        self.root.after(200, lambda: self._open_attribute_suggestion_dialog(auto=True))

    def _translate_loaded_texts(self) -> None:
        client = self._build_deepl_client()
        short_translations = client.translate_from_german(self.short_text_frame.get_german_text())
        long_translations = client.translate_from_german(self.long_text_frame.get_german_text())
        self.short_text_frame.apply_translations(short_translations)
        self.long_text_frame.apply_translations(long_translations)

    def _load_from_kunzer(self, article_number_or_url: str) -> None:
        identifier = article_number_or_url.strip()
        if not identifier:
            messagebox.showwarning(APP_TITLE, "Bitte eine Artikelnummer oder Kunzer Produkt-URL eingeben.")
            return

        try:
            options = self._get_scrape_options()
        except ValueError as exc:
            messagebox.showwarning(APP_TITLE, str(exc))
            return

        deepl_client: DeepLClient | None = None
        if (
            self.auto_translate_after_scrape_var.get()
            and self.deepl_api_key_var.get().strip()
            and (options["short_text"] or options["long_text"])
        ):
            try:
                deepl_client = self._build_deepl_client()
            except ValueError as exc:
                messagebox.showwarning(APP_TITLE, str(exc))
                return

        def worker() -> tuple[KunzerScrapeResult, dict[str, str] | None, dict[str, str] | None]:
            result = self.kunzer_scraper.scrape_product(identifier)
            if deepl_client is None:
                return result, None, None
            short_translations = deepl_client.translate_from_german(result.short_text_de) if options["short_text"] else None
            long_translations = deepl_client.translate_from_german(result.long_text_de) if options["long_text"] else None
            return result, short_translations, long_translations

        def on_success(payload: object) -> None:
            result, short_translations, long_translations = payload  # type: ignore[misc]
            self._apply_kunzer_result(
                result,
                translate_after_load=False,
                write_live=True,
                short_translations=short_translations,
                long_translations=long_translations,
                options=options,
            )
            self.status_var.set(f"Kunzer-Daten geladen: {result.article_number}")

        self._run_background_task(
            "Kunzer-Daten werden geladen ...",
            worker,
            on_success,
            "Kunzer-Import fehlgeschlagen",
        )

    def load_from_kunzer_article_number(self) -> None:
        self._load_from_kunzer(self.article_number_var.get())

    def import_products_from_file(self) -> None:
        source_path = Path(self.product_list_path_var.get().strip())
        if not source_path:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst eine Produktliste als CSV oder XLSX auswählen.")
            return
        if not source_path.exists():
            messagebox.showwarning(APP_TITLE, f"Produktliste nicht gefunden:\n{source_path}")
            return

        try:
            output_root = self._resolve_output_root()
            options = self._get_scrape_options()
            items = read_product_import_items(source_path)
        except ValueError as exc:
            messagebox.showwarning(APP_TITLE, str(exc))
            return

        deepl_client: DeepLClient | None = None
        try:
            if options["short_text"] or options["long_text"]:
                deepl_client = self._build_deepl_client()
        except ValueError as exc:
            messagebox.showwarning(APP_TITLE, str(exc))
            return

        if output_root.exists():
            self.id_registry.load_from_folder(output_root, clear_existing=False)

        success_count = 0
        error_messages: list[str] = []
        last_result: KunzerScrapeResult | None = None
        last_bundle: ExportBundle | None = None
        total = len(items)

        for index, item in enumerate(items, start=1):
            identifier = item.product_url or item.article_number
            try:
                self.status_var.set(f"Massenimport {index}/{total}: {identifier}")
                self.root.update_idletasks()
                result = self.kunzer_scraper.scrape_product(identifier)
                bundle = self._build_bundle_from_kunzer_result(result, client=deepl_client, options=options)
                export_bundle(bundle, output_root, use_timestamp_subdir=False)
                success_count += 1
                last_result = result
                last_bundle = bundle
            except Exception as exc:  # pragma: no cover - user facing batch feedback
                error_messages.append(f"{identifier}: {exc}")

        if last_result and last_bundle:
            self._apply_kunzer_result(last_result, translate_after_load=False, write_live=False)
            self.short_module_id_var.set(last_bundle.short_module_id)
            self.long_module_id_var.set(last_bundle.long_module_id)
            if last_bundle.include_short_text:
                self.short_text_frame.set_value(last_bundle.short_texts, auto_uni=True)
            else:
                self.short_text_frame.set_value(TranslationSet(), auto_uni=True)
            if last_bundle.include_long_text:
                self.long_text_frame.set_value(last_bundle.long_texts, auto_uni=True)
            else:
                self.long_text_frame.set_value(TranslationSet(), auto_uni=True)
            self.attribute_frame.set_rows(last_bundle.attribute_rows)
            self.search_term_frame.set_terms(last_bundle.search_terms)
            self.oe_frame.set_rows(last_bundle.oe_number_rows)
            self.comparison_frame.set_rows(last_bundle.comparison_number_rows)
            self.vehicle_link_frame.set_rows(last_bundle.vehicle_link_rows)
            if not last_bundle.include_images:
                self.image_frame.set_rows([])
            if not last_bundle.include_documents:
                self.document_frame.set_rows([])
            if not last_bundle.include_videos:
                self.video_frame.set_rows([])
            if not last_bundle.include_web_links:
                self.web_frame.set_rows([])

        self.refresh_preview()
        if error_messages:
            preview = "\n".join(error_messages[:8])
            if len(error_messages) > 8:
                preview += f"\n... und {len(error_messages) - 8} weitere"
            self.status_var.set(f"Massenimport abgeschlossen mit Fehlern: {success_count}/{total} erfolgreich")
            messagebox.showwarning(
                APP_TITLE,
                f"Massenimport abgeschlossen und direkt in den Output-Pfad geschrieben: {success_count} von {total} Produkten erfolgreich.\n\nFehler:\n{preview}",
            )
        else:
            self.status_var.set(f"Massenimport abgeschlossen: {success_count}/{total} Produkte erfolgreich")
            messagebox.showinfo(
                APP_TITLE,
                f"Massenimport erfolgreich abgeschlossen:\n{success_count} Produkte importiert und direkt in den Output-Pfad geschrieben.",
            )

    def _build_deepl_client(self) -> DeepLClient:
        auth_key = self.deepl_api_key_var.get().strip()
        if not auth_key:
            raise ValueError("Bitte einen DeepL API Key eingeben oder die Umgebungsvariable DEEPL_API_KEY setzen.")

        base_url = self.deepl_base_url_var.get().strip() or DEEPL_DEFAULT_BASE_URL
        return DeepLClient(auth_key=auth_key, base_url=base_url)

    def _translate_frame_from_german(self, frame: SingleLineTranslationFrame | MultiLineTranslationFrame, label: str) -> None:
        german_text = frame.get_german_text()
        if not german_text:
            raise ValueError(f"Bitte zuerst den deutschen Text für {label} eingeben.")

        client = self._build_deepl_client()

        def worker() -> dict[str, str]:
            return client.translate_from_german(german_text)

        def on_success(payload: object) -> None:
            translations = payload  # type: ignore[assignment]
            frame.apply_translations(translations)
            self.refresh_preview()
            self._write_live_database(status_message=f"Live gespeichert: {normalize_article_number(self.article_number_var.get())}")
            self.status_var.set(f"{label} mit DeepL aus Deutsch übersetzt.")

        self._run_background_task(
            f"{label} wird mit DeepL übersetzt ...",
            worker,
            on_success,
            f"{label} Übersetzung fehlgeschlagen",
        )

    def translate_short_texts(self) -> None:
        try:
            self._translate_frame_from_german(self.short_text_frame, "Kurzbezeichnung")
        except (ValueError, DeepLTranslationError) as exc:
            messagebox.showwarning(APP_TITLE, str(exc))

    def translate_long_texts(self) -> None:
        try:
            self._translate_frame_from_german(self.long_text_frame, "Text")
        except (ValueError, DeepLTranslationError) as exc:
            messagebox.showwarning(APP_TITLE, str(exc))

    def translate_all_texts(self) -> None:
        try:
            short_german = self.short_text_frame.get_german_text()
            long_german = self.long_text_frame.get_german_text()
            if not short_german and not long_german:
                raise ValueError("Bitte zuerst einen deutschen Kurztext oder Langtext eingeben.")
            client = self._build_deepl_client()
        except ValueError as exc:
            messagebox.showwarning(APP_TITLE, str(exc))
            return

        def worker() -> tuple[dict[str, str], dict[str, str]]:
            short_translations = client.translate_from_german(short_german) if short_german else {}
            long_translations = client.translate_from_german(long_german) if long_german else {}
            return short_translations, long_translations

        def on_success(payload: object) -> None:
            short_translations, long_translations = payload  # type: ignore[misc]
            if short_translations:
                self.short_text_frame.apply_translations(short_translations)
            if long_translations:
                self.long_text_frame.apply_translations(long_translations)
            self.refresh_preview()
            self._write_live_database(status_message=f"Live gespeichert: {normalize_article_number(self.article_number_var.get())}")
            self.status_var.set("Kurzbezeichnung und Text mit DeepL aus Deutsch übersetzt.")

        self._run_background_task(
            "Kurzbezeichnung und Text werden mit DeepL übersetzt ...",
            worker,
            on_success,
            "DeepL Übersetzung fehlgeschlagen",
        )

    def _load_known_ids(self, initial: bool = False) -> None:
        import_folder = Path(self.import_dir_var.get().strip())
        output_folder = Path(self.output_dir_var.get().strip())
        warnings: list[str] = []
        loaded_from: list[Path] = []
        total_count = 0

        self.id_registry.used_ids.clear()
        self.id_registry.short_ids_by_article.clear()
        self.id_registry.long_ids_by_article.clear()

        if import_folder.exists():
            total_count, import_warnings = self.id_registry.load_from_folder(import_folder, clear_existing=True)
            warnings.extend(import_warnings)
            loaded_from.append(import_folder)

        if output_folder.exists() and output_folder != import_folder:
            total_count, output_warnings = self.id_registry.load_from_folder(output_folder, clear_existing=not loaded_from)
            warnings.extend(output_warnings)
            loaded_from.append(output_folder)

        if not loaded_from:
            self.known_id_count_var.set("0 IDs geladen")
            self.article_browser_records = {}
            for item_id in self.article_browser_tree.get_children():
                self.article_browser_tree.delete(item_id)
            self._update_article_browser_detail(None)
            if not initial:
                self.status_var.set(f"Keine ID-Quelle gefunden: {import_folder}")
            return

        current_article = normalize_article_number(self.article_number_var.get())
        if current_article:
            self.current_id_article_number = ""
            self._ensure_ids_for_article(current_article)

        self.known_id_count_var.set(f"{total_count} IDs geladen")
        if warnings:
            self.status_var.set(f"IDs geladen mit Hinweisen: {warnings[0]}")
        else:
            sources = ", ".join(str(path) for path in loaded_from)
            self.status_var.set(f"IDs erfolgreich geladen aus: {sources}")

    def _refresh_article_browser(self) -> None:
        records: dict[str, StoredArticleSnapshot] = {}
        output_folder = Path(self.output_dir_var.get().strip())
        signature: tuple[tuple[str, int, int], ...] = ()

        if output_folder.exists():
            signature = self._build_output_folder_signature(output_folder)
            if (
                self.article_browser_cache_dir == str(output_folder)
                and self.article_browser_cache_signature == signature
            ):
                self._render_article_browser()
                return
            records = load_article_snapshots_from_folder(
                output_folder,
                "Output",
                competitor_lookup=self.competitor_options_by_id,
                manufacturer_lookup=self.manufacturer_options_by_id,
                attribute_lookup=self.attribute_options_by_id,
                attribute_key_values_by_group=self.attribute_key_values_by_group,
                vehicle_catalog=self.vehicle_catalog,
            )

        self.article_browser_records = records
        self.article_browser_cache_dir = str(output_folder)
        self.article_browser_cache_signature = signature
        self._render_article_browser()

    def _update_article_browser_detail(self, snapshot: StoredArticleSnapshot | None) -> None:
        self.article_browser_detail.configure(state="normal")
        self.article_browser_detail.delete("1.0", "end")
        if snapshot is None:
            self.article_browser_detail.insert(
                "1.0",
                "Wähle einen Artikel im Verzeichnis aus, um alle vorhandenen Exportdaten anzuzeigen.\n\n"
                "Mit Doppelklick oder dem Button kannst du den Artikel anschliessend in die Bearbeitungsmaske laden.",
            )
        else:
            self.article_browser_detail.insert("1.0", format_article_snapshot(snapshot))
        self.article_browser_detail.configure(state="disabled")

    def _update_article_browser_multi_detail(self, article_numbers: list[str]) -> None:
        self.article_browser_detail.configure(state="normal")
        self.article_browser_detail.delete("1.0", "end")
        self.article_browser_detail.insert(
            "1.0",
            "Mehrere Artikel markiert.\n\n"
            f"Anzahl: {len(article_numbers)}\n"
            "Aktionen:\n"
            "- Rechtsklick für Kopieren oder Löschen\n"
            "- Doppelklick oder 'Ausgewählten Artikel laden' ist für einen einzelnen Artikel gedacht\n\n"
            "Markierte Artikel:\n"
            + "\n".join(article_numbers),
        )
        self.article_browser_detail.configure(state="disabled")

    def _apply_article_snapshot(self, snapshot: StoredArticleSnapshot) -> None:
        self.article_number_var.set(snapshot.article_number)
        self.current_id_article_number = snapshot.article_number
        self.short_module_id_var.set(snapshot.short_module_id)
        self.long_module_id_var.set(snapshot.long_module_id)
        self.id_registry.remember_article_ids(snapshot.article_number, snapshot.short_module_id, snapshot.long_module_id)
        self._set_selected_genart_selections(snapshot.genart_selections)
        self.genart_display_var.set("")
        if snapshot.genart_selections:
            self.genart_suggestion_var.set(
                f"Gespeicherte GenArten: {summarize_genart_selections(snapshot.genart_selections, empty_label='-', limit=3)}"
            )
        else:
            self.genart_suggestion_var.set("Keine GenArt für diesen Artikel gespeichert.")
        self.short_text_frame.set_value(snapshot.short_texts, auto_uni=snapshot.short_auto_uni)
        self.long_text_frame.set_value(snapshot.long_texts, auto_uni=snapshot.long_auto_uni)
        self.attribute_frame.set_rows(snapshot.attribute_rows)
        self.search_term_frame.set_terms(snapshot.search_terms)
        self.oe_frame.set_rows(snapshot.oe_number_rows)
        self.comparison_frame.set_rows(snapshot.comparison_number_rows)
        self.vehicle_link_frame.set_rows(snapshot.vehicle_link_rows)
        self.image_frame.set_rows(snapshot.image_rows)
        self.document_frame.set_rows(snapshot.document_rows)
        self.video_frame.set_rows(snapshot.video_rows)
        self.web_frame.set_rows(snapshot.web_rows)
        if snapshot.web_rows:
            self.kunzer_product_url_var.set(snapshot.web_rows[0].path_or_link)
        else:
            self.kunzer_product_url_var.set("")
        self.status_var.set(f"Artikel geladen: {snapshot.article_number} ({snapshot.source_label})")

    def load_selected_article_from_browser(self) -> None:
        selection = self.article_browser_tree.selection()
        if not selection:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst einen Artikel im Verzeichnis auswählen.")
            return
        if len(selection) > 1:
            messagebox.showwarning(APP_TITLE, "Bitte zum Laden genau einen Artikel im Verzeichnis auswählen.")
            return

        snapshot = self.article_browser_records.get(selection[0])
        if snapshot is None:
            messagebox.showwarning(APP_TITLE, "Der ausgewählte Artikel konnte nicht geladen werden.")
            return

        self._apply_article_snapshot(snapshot)
        self.refresh_preview()

    def _on_article_browser_select(self, _event: tk.Event[tk.Misc]) -> None:
        selection = self.article_browser_tree.selection()
        if not selection:
            self._update_article_browser_detail(None)
            return
        if len(selection) > 1:
            self._update_article_browser_multi_detail(list(selection))
            return
        snapshot = self.article_browser_records.get(selection[0])
        self._update_article_browser_detail(snapshot)

    def _on_article_browser_double_click(self, _event: tk.Event[tk.Misc]) -> None:
        self.load_selected_article_from_browser()

    def _open_article_browser_context_menu(self, event: tk.Event[tk.Misc]) -> None:
        item_id = self.article_browser_tree.identify_row(event.y)
        if item_id:
            if item_id not in self.article_browser_tree.selection():
                self.article_browser_tree.selection_set(item_id)
            self.article_browser_tree.focus(item_id)
            self._on_article_browser_select(event)

        has_selection = bool(self.article_browser_tree.selection())
        if self.article_browser_context_menu is None:
            return
        self.article_browser_context_menu.entryconfigure("Zeilen kopieren", state="normal" if has_selection else "disabled")
        self.article_browser_context_menu.entryconfigure("Zeilen löschen", state="normal" if has_selection else "disabled")
        self.article_browser_context_menu.tk_popup(event.x_root, event.y_root)
        self.article_browser_context_menu.grab_release()

    def copy_selected_articles_from_browser(self) -> None:
        selection = list(self.article_browser_tree.selection())
        if not selection:
            return

        rows = []
        for item_id in selection:
            values = [str(value).strip() for value in self.article_browser_tree.item(item_id, "values")]
            rows.append("\t".join(values))
        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(rows))
        self.status_var.set(f"{len(selection)} Artikelzeile(n) in die Zwischenablage kopiert.")

    def delete_selected_articles_from_browser(self) -> None:
        selection = list(self.article_browser_tree.selection())
        if not selection:
            messagebox.showwarning(APP_TITLE, "Bitte zuerst mindestens einen Artikel im Verzeichnis auswählen.")
            return

        article_numbers = [normalize_article_number(item_id) for item_id in selection if normalize_article_number(item_id)]
        if not article_numbers:
            return

        article_preview = "\n".join(article_numbers[:8])
        if len(article_numbers) > 8:
            article_preview += f"\n... und {len(article_numbers) - 8} weitere"
        confirmed = messagebox.askyesno(
            APP_TITLE,
            "Sollen die markierten Artikel wirklich aus allen Output-Dateien gelöscht werden?\n\n" + article_preview,
        )
        if not confirmed:
            return

        try:
            output_root = self._resolve_output_root()
            for file_spec, headers in [
                (SHORT_TEXT_FILE, SHORT_TEXT_HEADERS),
                (SHORT_MAPPING_FILE, SHORT_MAPPING_HEADERS),
                (LONG_TEXT_FILE, SHORT_TEXT_HEADERS),
                (GENART_FILE, GENART_HEADERS),
                (ATTRIBUTE_FILE, ATTRIBUTE_HEADERS),
                (OE_FILE, OE_HEADERS),
                (COMPARISON_FILE, COMPARISON_HEADERS),
                (VEHICLE_LINK_FILE, VEHICLE_LINK_HEADERS),
                (IMAGE_FILE, IMAGE_HEADERS),
                (DOCUMENT_FILE, DOCUMENT_HEADERS),
                (VIDEO_FILE, VIDEO_HEADERS),
                (WEB_LINK_FILE, WEB_HEADERS),
            ]:
                remove_article_rows_from_workbook(output_root / file_spec[0], file_spec[1], headers, set(article_numbers))
        except Exception as exc:  # pragma: no cover - defensive UI feedback
            messagebox.showerror(APP_TITLE, f"Artikel konnten nicht gelöscht werden:\n{exc}")
            self.status_var.set(f"Löschen fehlgeschlagen: {exc}")
            return

        for article_number in article_numbers:
            self.article_browser_records.pop(article_number, None)
        self._load_known_ids(initial=True)
        self._refresh_article_browser()
        self.status_var.set(f"{len(article_numbers)} Artikel aus den Output-Dateien gelöscht.")

    def collect_bundle(self) -> ExportBundle:
        article_number = normalize_article_number(self.article_number_var.get())
        if not article_number:
            raise ValueError("Bitte eine Artikelnummer eingeben.")

        short_module_id, long_module_id = self._ensure_ids_for_article(article_number)

        return ExportBundle(
            article_number=article_number,
            short_module_id=short_module_id,
            long_module_id=long_module_id,
            short_texts=self.short_text_frame.get_value(),
            short_auto_uni=self.short_text_frame.auto_uni_var.get(),
            long_texts=self.long_text_frame.get_value(),
            long_auto_uni=self.long_text_frame.auto_uni_var.get(),
            genart_selections=self._get_selected_genart_selections(),
            attribute_rows=self.attribute_frame.get_rows(),
            search_terms=self.search_term_frame.get_terms(),
            attribute_key_values_by_group=self.attribute_key_values_by_group,
            oe_number_rows=self.oe_frame.get_rows(),
            comparison_number_rows=self.comparison_frame.get_rows(),
            vehicle_link_rows=self.vehicle_link_frame.get_rows(),
            image_rows=self.image_frame.get_rows(),
            document_rows=self.document_frame.get_rows(),
            video_rows=self.video_frame.get_rows(),
            web_rows=self.web_frame.get_rows(),
        )

    def refresh_preview(self) -> None:
        current_article = normalize_article_number(self.article_number_var.get())
        if current_article:
            self._ensure_ids_for_article(current_article)
        self._refresh_article_browser()

    def export_current_bundle(self) -> bool:
        try:
            bundle = self.collect_bundle()
            output_root = self._resolve_output_root()
        except ValueError as exc:
            messagebox.showwarning(APP_TITLE, str(exc))
            return False

        try:
            export_dir = export_bundle(bundle, output_root, use_timestamp_subdir=not self.fixed_export_path_var.get())
        except Exception as exc:  # pragma: no cover - defensive user feedback
            messagebox.showerror(APP_TITLE, f"Export fehlgeschlagen:\n{exc}")
            self.status_var.set(f"Export fehlgeschlagen: {exc}")
            return False

        self.refresh_preview()
        if self.fixed_export_path_var.get():
            self.status_var.set(f"Exportdateien im festen Pfad aktualisiert: {export_dir}")
            messagebox.showinfo(APP_TITLE, f"Exportdateien erfolgreich aktualisiert:\n{export_dir}")
        else:
            self.status_var.set(f"Export erstellt: {export_dir}")
            messagebox.showinfo(APP_TITLE, f"Export erfolgreich erstellt:\n{export_dir}")
        return True


def main() -> None:
    root = tk.Tk()
    app = ApolloImportApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
